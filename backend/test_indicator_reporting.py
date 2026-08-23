import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

from backend.indicator_reporting import build_indicator_report, build_indicator_workbook, build_narrative_workbook


def test_indicator_parent_carryover_and_reporting_mappings():
    assessments = pd.DataFrame({
        "Assessment ID": ["A1", "A2", "A3", "A4", "A5"],
        "Projects - المشروع": ["UNHCR 2026 - Gov", "UNHCR 2026 - Gov", "UNHCR 2026 - Baghdad", "UNHCR 2026 - SULI", "UNHCR 2026 - SULI"],
        "Project Location": ["Anbar أنبار", "Anbar أنبار", "Diyala ديالى", "Pshdar Urban (Refugees)", "Rania"],
        "Date of Assessment تاريخ التقييم": pd.to_datetime(["2026-01-10"] * 5),
        "Type of Legal Service Needed / الخدمات القانونية الازمةs": ["Legal Representation", "Legal Counselling", "Legal Assistance", "Legal Assistance", "Legal Assistance"],
        "Legal Service (Type of Service Provided / نوع الخدمة)": ["Legal Representation", "Legal Counselling", "Legal Assistance", "Legal Assistance", "Legal Assistance"],
        "Is the beneficiary detained هل المستفيد موقوف": ["Yes نعم", "No لا", "No لا", "No لا", "No لا"],
        "Is it an immigration related charge? هل هو معتقل على اساس قانون الاقامة ؟": ["No", "No", "No", "No", "No"],
        "Gender النوع الاجتماعي": ["Male ذکر", "Female أنثی", "Male ذکر", "Female أنثی", "Female أنثی"],
        "UNHCR Age Group": ["(18-39)", "(18-39)", "(18-39)", "(18-39)", "(18-39)"],
        "Community Type": ["Syrian Refugee لاجيء-سوري"] * 5,
    })
    services = pd.DataFrame({
        "Service ID": ["S1", "S2", "S3", "S4"],
        "Beneficiary ID": ["B1", "B2", "B2", "B2"],
        "Assessment ID": ["A1", "OLD", "OLD", "OLD"],
        "Projects - المشروع": ["UNHCR 2026 - Gov"] * 4,
        "Project Location": ["Anbar أنبار"] * 4,
        "Date of Service Provision /  تاريخ بدء الخدمة": pd.to_datetime(["2026-01-10", "2025-05-01", "2025-06-01", "2024-06-01"]),
        "Date Service Completed تاريخ انجاز الخدمة": pd.to_datetime(["2026-01-10", "2026-01-05", "2026-01-20", "2026-02-02"]),
        "Type of Service Provided / نوع الخدمة": ["Legal Representation", "Legal Assistance", "Legal Representation", "Legal Assistance"],
        "Service Status حالة الخدمة": ["Completed اکتملت"] * 4,
        "Is the beneficiary detained هل المستفيد موقوف": ["Yes نعم"] * 4,
        "Gender النوع الاجتماعي": ["Male ذکر"] * 4,
        "UNHCR Age Group": ["(18-39)"] * 4,
        "Community Type": ["Syrian Refugee لاجيء-سوري"] * 4,
        "Type of Document نوع الوثيقة": [""] * 4,
    })
    report = build_indicator_report({"assessments": assessments, "legalservices": services}, "2026-01-01", "2026-02-28")
    indicators = {item["id"]: item for item in report["groups"][0]["indicators"]}
    representation = indicators["legal-representation"]
    assert representation["total"] == 3
    assert representation["children"][0]["total"] == 3
    assert representation["children"][1]["total"] == 0
    assert representation["contributions"] == {"assessmentPeriod": 1, "carryOver": 2, "monthlyCarryOver": {"2026-01": 1, "2026-02": 1}}
    assert representation["sections"][0]["totals"] == [0, 0, 0, 3, 0, 0, 0, 0, 0, 0, 0, 0, 3]

    reached = next(item for group in report["groups"] for item in group["indicators"] if item["id"] == "individuals-reached")
    assert reached["narrative"]["rows"]
    assert [row["population"] for row in reached["narrative"]["rows"]] == ["Refugees", "IDP"]
    assert reached["narrative"]["rows"][0]["totalAchievement"] == sum(section["total"] for section in reached["sections"][:2])
    assert "Anbar Ø" not in reached["narrative"]["remark"]
    narrative_sheet = load_workbook(BytesIO(build_narrative_workbook(report)))["Narrative Report"]
    assert [narrative_sheet.cell(1, column).value for column in range(1, 5)] == ["Indicators", "Population", "Total Achievement - Reporting Period", "Remarks"]
    assert [section["id"] for section in reached["sections"]] == ["syrian-refugee", "non-syrian-refugee", "idp"]
    all_rows = reached["sections"][0]["rows"]
    diyala = next(row for row in all_rows if row["location"] == "Diyala ديالى")
    combined = next(row for row in all_rows if row["location"] == "Pshdar Urban (Refugees) + Rania")
    assert diyala["project"] == "UNHCR 2026 - Gov" and diyala["values"][-1] == 1
    assert combined["values"][-1] == 2
    assert any(row["project"] == "UNHCR 2026 - AMAL CAMP" for row in all_rows)
    assert all(row["project"] != "UNHCR 2026 - AMAL CAMP" for section in representation["sections"] for row in section["rows"])


def test_indicator_filters_hide_rows_and_validate_dates():
    assessments = pd.DataFrame({
        "Assessment ID": ["A1"], "Projects - المشروع": ["UNHCR 2026 - Gov"], "Project Location": ["Anbar أنبار"],
        "Date of Assessment تاريخ التقييم": pd.to_datetime(["2026-01-10"]), "Type of Legal Service Needed / الخدمات القانونية الازمةs": ["Legal Assistance"],
        "Gender النوع الاجتماعي": ["Male ذکر"], "UNHCR Age Group": ["(18-39)"], "Community Type": ["IDP"],
    })
    services = pd.DataFrame(columns=["Service ID", "Assessment ID", "Projects - المشروع", "Project Location", "Type of Service Provided / نوع الخدمة"])
    report = build_indicator_report({"assessments": assessments, "legalservices": services}, projects=["UNHCR 2026 - SULI"])
    reached = next(item for group in report["groups"] for item in group["indicators"] if item["id"] == "individuals-reached")
    assert len(reached["sections"][0]["rows"]) == 3
    january = build_indicator_report({"assessments": assessments, "legalservices": services}, years=["2026"], quarters=["2026-Q1"], months=["2026-01"])
    january_reached = next(item for group in january["groups"] for item in group["indicators"] if item["id"] == "individuals-reached")
    assert january_reached["total"] == 1
    assert load_workbook(BytesIO(build_narrative_workbook(january)))["Narrative Report"]["C1"].value == "Total Achievement - January 2026"
    assert january["filterOptions"]["years"] == ["2026"]
    assert january["filterOptions"]["quarters"] == ["2026-Q1"]
    assert january["filterOptions"]["months"] == ["2026-01"]
    february = build_indicator_report({"assessments": assessments, "legalservices": services}, months=["2026-02"])
    february_reached = next(item for group in february["groups"] for item in group["indicators"] if item["id"] == "individuals-reached")
    assert february_reached["total"] == 0
    assert february_reached["narrative"]["remark"] == ""
    assert all(row["remarks"] == "" for row in february_reached["narrative"]["rows"])

    amal = build_indicator_report({"assessments": assessments, "legalservices": services}, projects=["UNHCR 2026 - AMAL CAMP"])
    amal_reached = next(item for group in amal["groups"] for item in group["indicators"] if item["id"] == "individuals-reached")
    assert [section["id"] for section in amal_reached["sections"]] == ["idp"]
    non_amal = build_indicator_report({"assessments": assessments, "legalservices": services}, projects=["UNHCR 2026 - Gov"])
    non_amal_reached = next(item for group in non_amal["groups"] for item in group["indicators"] if item["id"] == "individuals-reached")
    assert [section["id"] for section in non_amal_reached["sections"]] == ["syrian-refugee", "non-syrian-refugee"]

    workbook = load_workbook(BytesIO(build_indicator_workbook(january)))
    assert workbook.sheetnames == ["Refugee", "IDP", "Individual Beneficiaries"]
    sheet = workbook["Refugee"]
    assert all(export_sheet.freeze_panes is None for export_sheet in workbook.worksheets)
    assert sheet["A1"].value.startswith("# of persons identified")
    assert sheet.max_row > 10 and sheet.max_column > 10
    total_header = next(cell for cell in sheet[4] if cell.value == "Total")
    assert sheet["A2"].value is None and sheet["B2"].value is None and sheet.cell(2, total_header.column).value is None
    assert total_header.fill.fgColor.rgb == "00E7EEF4"
    assert sheet.cell(5, total_header.column).fill.fgColor.rgb == "00E7EEF4"
    activity_headers = [cell.column for row in sheet.iter_rows() for cell in row if cell.value == "Activity"]
    assert activity_headers
    age_header_row = next(cell.row for row in sheet.iter_rows() for cell in row if cell.value == "00-04")
    assert sheet.cell(age_header_row, 3).fill.fgColor.rgb == "00D9EAF7"
    assert sheet.cell(age_header_row, 9).fill.fgColor.rgb == "00FCE8C3"
    pshdar_row = next(cell.row for row in sheet.iter_rows() for cell in row if cell.value == "Pshdar Urban (Refugees) + Rania")
    mosul_row = next(cell.row for row in sheet.iter_rows() for cell in row if cell.value == "UNHCR 2026 - Mosul & Kirkuk")
    assert sheet.row_dimensions[pshdar_row].height == 21
    assert sheet.row_dimensions[mosul_row].height == 21
    assert sheet.cell(pshdar_row, 2).alignment.shrink_to_fit is True
    assert sheet.cell(mosul_row, 1).alignment.shrink_to_fit is True
    zero_cells = [cell for row in sheet.iter_rows() for cell in row if cell.value == 0]
    assert zero_cells and all(cell.column == sheet.max_column for cell in zero_cells)
    amal_workbook = load_workbook(BytesIO(build_indicator_workbook(amal)))
    assert amal_workbook.sheetnames == ["IDP", "Individual Beneficiaries"]

    analysis_workbook = load_workbook(BytesIO(build_indicator_workbook(january, [("2026-01", january)])))
    assert "Analysis" in analysis_workbook.sheetnames
    analysis_sheet = analysis_workbook["Analysis"]
    assert analysis_sheet["A1"].value == "Monthly indicator analysis"
    assert analysis_sheet._charts


def test_indicator_report_uses_renamed_mosul_and_kirkuk_locations():
    assessments = pd.DataFrame({
        "Assessment ID": ["new-ninewa", "old-mosul", "new-kirkuk", "old-kirkuk"],
        "Projects - project": ["UNHCR 2026 - Mosul & Kirkuk"] * 4,
        "Project Location": ["Ninewa نينوى", "UNHCR Mosul", "Kirkuk كركوك", "UNHCR Kirkuk"],
        "Date of Assessment": pd.to_datetime(["2026-01-10"] * 4),
        "Type of Legal Service Needed": ["Legal Assistance"] * 4,
        "Gender": ["Male"] * 4,
        "UNHCR Age Group": ["(18-39)"] * 4,
        "Community Type": ["Syrian Refugee"] * 4,
    })
    services = pd.DataFrame(columns=assessments.columns)

    report = build_indicator_report({"assessments": assessments, "legalservices": services})
    reached = next(item for group in report["groups"] for item in group["indicators"] if item["id"] == "individuals-reached")
    rows = reached["sections"][0]["rows"]
    location_counts = {row["location"]: row["values"][-1] for row in rows if row["project"] == "UNHCR 2026 - Mosul & Kirkuk"}

    assert location_counts == {"Ninewa نينوى": 2, "Kirkuk كركوك": 2}


def test_representation_uses_unique_assessment_and_monthly_beneficiary_ids():
    common_assessment = {
        "Projects - المشروع": "UNHCR 2026 - Gov", "Project Location": "Anbar أنبار",
        "Date of Assessment تاريخ التقييم": pd.Timestamp("2026-01-10"),
        "Gender النوع الاجتماعي": "Male ذکر", "UNHCR Age Group": "(18-39)",
        "Community Type": "Syrian Refugee لاجئ-سوري",
    }
    assessment_rows = [
        {**common_assessment, "Assessment ID": "A1", "Type of Legal Service Needed / الخدمات القانونية الازمةs": "Legal Representation - تمثيل", "Legal Service (Type of Service Provided / نوع الخدمة)": "Legal Representation - تمثيل", "Is the beneficiary detained هل المستفيد موقوف": "Yes نعم", "Is it an immigration related charge? هل هو معتقل على اساس قانون الاقامة ؟": "Yes"},
        {**common_assessment, "Assessment ID": "A1", "Type of Legal Service Needed / الخدمات القانونية الازمةs": "Legal Representation - تمثيل", "Legal Service (Type of Service Provided / نوع الخدمة)": "Legal Representation - تمثيل", "Is the beneficiary detained هل المستفيد موقوف": "Yes نعم", "Is it an immigration related charge? هل هو معتقل على اساس قانون الاقامة ؟": "Yes"},
        {**common_assessment, "Assessment ID": "A2", "Type of Legal Service Needed / الخدمات القانونية الازمةs": "Legal Representation - تمثيل", "Legal Service (Type of Service Provided / نوع الخدمة)": "Legal Assistance - مساعدة", "Is the beneficiary detained هل المستفيد موقوف": "Yes نعم", "Is it an immigration related charge? هل هو معتقل على اساس قانون الاقامة ؟": "Yes", "Detainee current status حالة المعتقل الحالية": "Released تم الافراج عنه", "Date of the released or deported تاريخ الافراج او الترحيل": pd.Timestamp("2026-01-15")},
        {**common_assessment, "Assessment ID": "A3", "Type of Legal Service Needed / الخدمات القانونية الازمةs": "Legal Representation - تمثيل", "Legal Service (Type of Service Provided / نوع الخدمة)": "Legal Assistance - مساعدة", "Is the beneficiary detained هل المستفيد موقوف": "No لا", "Is it an immigration related charge? هل هو معتقل على اساس قانون الاقامة ؟": "Yes", "Detainee current status حالة المعتقل الحالية": "Released تم الافراج عنه", "Date of the released or deported تاريخ الافراج او الترحيل": pd.Timestamp("2026-01-15")},
        {**common_assessment, "Assessment ID": "A4", "Type of Legal Service Needed / الخدمات القانونية الازمةs": "Legal Counselling - استشارة", "Legal Service (Type of Service Provided / نوع الخدمة)": "Legal Counselling - استشارة", "Is the beneficiary detained هل المستفيد موقوف": "Yes نعم"},
        {**common_assessment, "Assessment ID": "A5", "Type of Legal Service Needed / الخدمات القانونية الازمةs": "Legal Counselling - استشارة,Legal Representation - تمثيل", "Legal Service (Type of Service Provided / نوع الخدمة)": "Legal Counselling - استشارة", "Is the beneficiary detained هل المستفيد موقوف": "Yes نعم"},
        {**common_assessment, "Assessment ID": "A6", "Type of Legal Service Needed / الخدمات القانونية الازمةs": "Legal Counselling - استشارة", "Legal Service (Type of Service Provided / نوع الخدمة)": "Legal Counselling - استشارة", "Is the beneficiary detained هل المستفيد موقوف": "No لا"},
        {**common_assessment, "Assessment ID": "A7", "Type of Legal Service Needed / الخدمات القانونية الازمةs": "Legal Assistance - مساعدة,Legal Representation - تمثيل", "Legal Service (Type of Service Provided / نوع الخدمة)": "Legal Assistance - مساعدة", "Is the beneficiary detained هل المستفيد موقوف": "No لا"},
    ]
    common_service = {
        "Projects - المشروع": "UNHCR 2026 - AMAL CAMP", "Project Location": "AMAL Camp",
        "Gender النوع الاجتماعي": "Female أنثى", "UNHCR Age Group": "(18-39)", "Community Type": "IDP",
        "Date of Service Provision /  تاريخ بدء الخدمة": pd.Timestamp("2026-01-01"),
        "Date Service Completed تاريخ انجاز الخدمة": pd.Timestamp("2026-01-20"),
        "Is the beneficiary detained هل المستفيد موقوف": "No لا",
    }
    carry = {**common_service, "Projects - المشروع": "UNHCR 2026 - Gov", "Project Location": "Anbar أنبار", "Gender النوع الاجتماعي": "Male ذکر", "Community Type": "Syrian Refugee لاجئ-سوري", "Beneficiary ID": "B-CARRY", "Assessment ID": "OLD", "Date of Service Provision /  تاريخ بدء الخدمة": pd.Timestamp("2025-12-01"), "Type of Service Provided / نوع الخدمة": "Legal Representation - تمثيل", "Service Status حالة الخدمة": "Completed اکتملت", "Is the beneficiary detained هل المستفيد موقوف": "Yes نعم", "Is Civil Documents": "No", "Type of Document نوع الوثيقة": ""}
    services = pd.DataFrame([
        {**carry, "Service ID": "C1"}, {**carry, "Service ID": "C2"},
        {**common_service, "Service ID": "S1", "Assessment ID": "S1A", "Type of Service Provided / نوع الخدمة": "Legal Assistance - مساعدة", "Service Status حالة الخدمة": "Completed اکتملت", "Is Civil Documents": "Yes", "Type of Document نوع الوثيقة": "Birth Certificate"},
        {**common_service, "Service ID": "S2", "Assessment ID": "S2A", "Type of Service Provided / نوع الخدمة": "Legal Assistance - مساعدة", "Service Status حالة الخدمة": "Completed اکتملت", "Is Civil Documents": "No", "Type of Document نوع الوثيقة": "Birth Certificate"},
        {**common_service, "Service ID": "S3", "Assessment ID": "S3A", "Type of Service Provided / نوع الخدمة": "Legal Assistance - مساعدة", "Service Status حالة الخدمة": "In-Process قيد المعالجة", "Is Civil Documents": "Yes", "Type of Document نوع الوثيقة": "Unified National Card البطاقة الوطنية الموحدة"},
        {**common_service, "Service ID": "S4", "Assessment ID": "S4A", "Type of Service Provided / نوع الخدمة": "Legal Counselling - استشارة", "Service Status حالة الخدمة": "Completed اکتملت", "Is Civil Documents": "Yes", "Type of Document نوع الوثيقة": "Unified National Card البطاقة الوطنية الموحدة"},
        {**common_service, "Service ID": "S5", "Assessment ID": "S5A", "Type of Service Provided / نوع الخدمة": "Legal Representation - تمثيل", "Service Status حالة الخدمة": "In-Process قيد المعالجة", "Is Civil Documents": "No", "Type of Document نوع الوثيقة": "Unified National Card البطاقة الوطنية الموحدة"},
    ])
    deportation = pd.DataFrame([
        {"Projects - المشروع": "UNHCR 2026 - Gov", "Project Location": "Anbar أنبار", "PN ID": "P1", "Gender النوع الاجتماعي": "Male ذکر", "UNHCR Age Group": "(18-39)", "Community Type / نوع المجتمع": "Syrian Refugee لاجئ-سوري", "Date of Deportation Knowledge - تاريخ العلم بالترحيل": pd.Timestamp("2026-01-12")},
        {"Projects - المشروع": "UNHCR 2026 - Gov", "Project Location": "Anbar أنبار", "PN ID": "P1", "Gender النوع الاجتماعي": "Male ذکر", "UNHCR Age Group": "(18-39)", "Community Type / نوع المجتمع": "Syrian Refugee لاجئ-سوري", "Date of Deportation Knowledge - تاريخ العلم بالترحيل": pd.Timestamp("2026-01-12")},
    ])
    awareness = pd.DataFrame([
        {"Projects - المشروع": "UNHCR 2026 - AMAL CAMP", "Project location": "AMAL Camp", "Awareness ID": "AW-1", "Session Topic": "Civil documentation / الوثائق المدنية", "Participant Name": "Person", "Gender الجنس": "Female أنثى", "UNHCR Age Group": "(18-39)", "Community Type / نوع المجتمع": "IDP", "Date of Session تاريخ الجلسة": pd.Timestamp("2026-01-05")},
        {"Projects - المشروع": "UNHCR 2026 - AMAL CAMP", "Project location": "AMAL Camp", "Awareness ID": "AW-1", "Session Topic": "Civil documentation / الوثائق المدنية", "Participant Name": "", "Gender الجنس": "Female أنثى", "UNHCR Age Group": "(18-39)", "Community Type / نوع المجتمع": "IDP", "Date of Session تاريخ الجلسة": pd.Timestamp("2026-01-05")},
    ])
    report = build_indicator_report({"assessments": pd.DataFrame(assessment_rows), "legalservices": services, "deportationrecords": deportation, "awareness": awareness}, "2026-01-01", "2026-01-31")
    indicators = {item["id"]: item for group in report["groups"] for item in group["indicators"]}
    assert [item["id"] for item in report["groups"][0]["indicators"]] == [
        "detention-immigration", "deported", "released-immigration", "06-1-1-legal-assistance",
        "legal-counselling", "legal-representation",
    ]
    assert [item["id"] for item in report["groups"][1]["indicators"]] == ["civil-counselling", "secured-civil-documentation", "uid-secured", "civil-representation", "legal-awareness-participants"]
    assert [item["id"] for item in report["groups"][2]["indicators"]] == ["individuals-reached"]
    amal_report = build_indicator_report({"assessments": pd.DataFrame(assessment_rows), "legalservices": services, "deportationrecords": deportation, "awareness": awareness}, "2026-01-01", "2026-01-31", projects=["UNHCR 2026 - AMAL CAMP"])
    assert [item["id"] for item in amal_report["groups"][0]["indicators"]] == ["civil-counselling", "secured-civil-documentation", "uid-secured", "civil-representation", "legal-awareness-participants"]
    amal_indicators = {item["id"]: item for group in amal_report["groups"] for item in group["indicators"]}
    assert indicators["detention-immigration"]["total"] == 3
    assert indicators["released-immigration"]["total"] == 1
    assert indicators["deported"]["total"] == 2
    assert indicators["legal-counselling"]["children"][0]["total"] == 1
    assert indicators["legal-counselling"]["children"][1]["total"] == 1
    assert indicators["legal-representation"]["children"][0]["total"] == 3
    assert indicators["legal-representation"]["children"][1]["total"] == 2
    assistance_remark = indicators["06-1-1-legal-assistance"]["narrative"]["remark"]
    assert "age, gender, location, and nationality breakdown" in assistance_remark
    assert "Service shares were" in assistance_remark
    assert "0 girls" not in assistance_remark and "0 women" not in assistance_remark
    assert amal_indicators["secured-civil-documentation"]["total"] == 1
    assert amal_indicators["uid-secured"]["total"] == 2
    assert amal_indicators["legal-awareness-participants"]["total"] == 2
    awareness_remark = amal_indicators["legal-awareness-participants"]["narrative"]["remark"]
    assert "PARTICIPANT PROFILE  |  Women: 2 (100.0%)" in awareness_remark and "Girls: 0" not in awareness_remark
    assert "Awareness sessions: 1" in awareness_remark
    assert "SESSION-TOPIC BREAKDOWN" in awareness_remark
    assert "• Civil documentation - 1 awareness session | Participants: 2 | Women: 2" in awareness_remark
    assert "الوثائق" not in awareness_remark
