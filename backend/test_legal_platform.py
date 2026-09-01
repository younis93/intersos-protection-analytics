import io
import zipfile
from datetime import date, timedelta

import pandas as pd
import pytest
from openpyxl import load_workbook

from backend.legal_platform import DETENTION_ASSESSMENT_RULES, LegalStore, normalize_name, phone_digits, versioned_dataset_name
from backend.duplicate_exclusions import DuplicateExclusionRegistry


def csv(**columns):
    return pd.DataFrame(columns).to_csv(index=False).encode("utf-8")


def required_payload():
    return {
        "beneficiaries": csv(**{"Case ID":["B1","B2"],"Name (Filter Color Red)":["أحمد علي","احمدعلي"],"Name / الأسم: First":["Ahmed","Ahmed"],"Age":[17,120],"# total assessments":[0,1],"Contact Number":["1","123"],"Date of Identification / تاريخ التحديد":["31/01/2026","01/02/2026"]}),
        "assessments": csv(**{"Assessment ID":["A1","A2"],"Beneficiary ID":["B1","B1"],"# Total Services":[0,1],"Date of Assessment تاريخ التقييم":["01/01/2026","01/02/2026"]}),
        "legalservices": csv(**{"Service ID":["S1"],"Assessment ID":["A2"],"Beneficiary ID":["B1"],"Secured documents Files  الرجاء ارفاق الوثيقة الصادرة":["blob"],"Secured documents Files  الرجاء ارفاق الوثيقة الصادرة: URL":["https://example.test/a.jpg"],"Date of Service Provision / تاريخ بدء الخدمة":["02/02/2026"]}),
    }


def test_optional_files_are_not_required_and_cleanup_is_applied():
    store=LegalStore.from_files(required_payload(),"test")
    assert store.metadata()["availability"]["awareness"] is False
    assert len(store.warnings)==4
    assert not any(column.endswith(": First") for column in store.frames["beneficiaries"].columns)
    assert "Secured documents Files  الرجاء ارفاق الوثيقة الصادرة" not in store.frames["legalservices"].columns
    dates=store.explorer("beneficiaries")["rows"]
    assert dates[0]["Date of Identification / تاريخ التحديد"]=="2026-01-31"
    assert dates[1]["Date of Identification / تاريخ التحديد"]=="2026-02-01"


def test_representation_case_load_uses_service_status_and_the_correct_event_month():
    payload=required_payload()
    payload["legalservices"]=csv(**{
        "Service ID":["S1","S2","S3","S4","S5"],
        "Lawyers":["Lawyer A","Lawyer A","Lawyer A","Lawyer B","Lawyer B"],
        "Type of Service Provided":["Legal Representation"]*4+["Legal Counselling"],
        "Service Status":["In-Process","Completed","Closed","In-Process","In-Process"],
        "Type of Document":["Court Verdict","ID Card","Passport","Court Verdict","ID Card"],
        "Date of Service Provision":["05/01/2026","06/02/2026","07/03/2026","invalid","08/01/2026"],
        "Date Service Completed":["","10/02/2026","11/03/2026","12/04/2026",""],
    })
    store=LegalStore.from_files(payload,"test")
    open_load=store.representation_case_load(status="open")
    assert open_load["months"]==["2026-01"]
    assert [{key:value for key,value in row.items() if key!="services"} for row in open_load["rows"]]==[{"lawyer":"Lawyer A","document":"Court Verdict","month":"2026-01","count":1}]
    assert open_load["rows"][0]["services"][0]["serviceId"]=="S1"
    closed_load=store.representation_case_load(status="closed")
    assert closed_load["months"]==["2026-02","2026-03"]
    assert [{key:value for key,value in row.items() if key!="services"} for row in closed_load["rows"]]==[
        {"lawyer":"Lawyer A","document":"ID Card","month":"2026-02","count":1},
        {"lawyer":"Lawyer A","document":"Passport","month":"2026-03","count":1},
    ]


def test_review_export_neutralizes_spreadsheet_formulas():
    payload=required_payload()
    frame=pd.read_csv(io.BytesIO(payload["beneficiaries"]),dtype=object)
    frame.loc[0,"Name (Filter Color Red)"]="=1+1"
    payload["beneficiaries"]=frame.to_csv(index=False).encode("utf-8")
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("beneficiaries")),data_only=False)
    cells=[cell for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row]
    assert any(cell.value=="'=1+1" for cell in cells)
    assert not any(cell.data_type=="f" for cell in cells)


def test_review_export_starts_with_page_fields_and_uses_source_fields_once():
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(required_payload(),"test").review_export("beneficiaries")),read_only=True,data_only=True)
    headers=[cell.value for cell in next(workbook["Unclassified"].iter_rows(min_row=1,max_row=1))]
    assert headers[:7]==["Review Finding","Finding detail","Recommended action","Lawyer","Project","Project location","Case ID"]
    assert "Priority" not in headers
    assert headers.index("Case ID") < headers.index("Name")
    assert headers.count("Name")==1
    assert headers.count("Case ID")==1
    assert "Name (Filter Color Red)" not in headers


def test_review_export_includes_selected_rule_page_fields_before_source_fields():
    payload=required_payload();payload["legalservices"]=csv(**{
        "Service ID":["S1","S2"],"Assessment ID":["A1","A2"],"Beneficiary ID":["B1","B1"],
        "Type of Service Provided":["Legal Assistance","Legal Assistance"],"Type of Document":["National ID","National ID"],
    })
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("legalservices",selected_rules=["Duplicate service without Assessment ID"])),read_only=True,data_only=True)
    headers=[cell.value for cell in next(workbook["Unclassified"].iter_rows(min_row=1,max_row=1))]
    assert headers.index("Service") < headers.index("Name")
    assert headers.count("Type of Service Provided")==1
    assert headers.count("Type of Document")==1


def test_review_export_places_dataset_identifiers_before_name():
    payload=required_payload();payload["awareness"]=csv(**{"Awareness ID":["W1"],"Participant Name":["Participant"]})
    store=LegalStore.from_files(payload,"test")
    assessment_book=load_workbook(io.BytesIO(store.review_export("assessments")),read_only=True,data_only=True)
    service_book=load_workbook(io.BytesIO(store.review_export("legalservices")),read_only=True,data_only=True)
    assessment_headers=[cell.value for cell in next(assessment_book.worksheets[0].iter_rows(min_row=1,max_row=1))]
    service_headers=[cell.value for cell in next(service_book.worksheets[0].iter_rows(min_row=1,max_row=1))]
    awareness_headers=[cell.value for cell in next(load_workbook(io.BytesIO(store.review_export("awareness")),read_only=True,data_only=True)["Review findings"].iter_rows(min_row=1,max_row=1))]
    assert assessment_headers.index("Case ID") < assessment_headers.index("Assessment") < assessment_headers.index("Name")
    assert service_headers.index("Case ID") < service_headers.index("Assessment") < service_headers.index("Service") < service_headers.index("Name")
    assert awareness_headers.index("Awareness ID") < awareness_headers.index("Name")


def test_review_export_colors_each_project_cell_consistently():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2"],"Name (Filter Color Red)":["One","Two"],"Age":[17,120],
        "Project":["UNHCR 2026 - Erbil","UNHCR 2026 - Gov"],
    })
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("beneficiaries")))
    colors={}
    for sheet in workbook.worksheets:
        headers=[cell.value for cell in sheet[1]];project_index=headers.index("Project")+1
        for row in range(2,sheet.max_row+1):
            project=sheet.cell(row,project_index).value
            if project: colors[project]=sheet.cell(row,project_index).fill.fgColor.rgb
    assert colors["UNHCR 2026 - Erbil"] != colors["UNHCR 2026 - Gov"]


def test_review_export_separates_selected_issues_into_regional_tables():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1"],"Name (Filter Color Red)":["Person One"],"Contact Number":["123"],
        "# total assessments":[0],"Project":["UNHCR 2026 - Erbil"],
    })
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("beneficiaries",selected_rules=["Invalid contact number","Case without assessment"])))
    sheet=workbook["North Iraq"]
    assert sheet["A1"].value=="Review Finding"
    assert sheet["A4"].value=="Review Finding"
    assert sheet.freeze_panes is None
    assert len(sheet.tables)==2
    assert all(sheet.cell(int(table.ref.split(":")[0][1:]),1).value=="Review Finding" for table in sheet.tables.values())


def test_detention_workbook_rejects_extreme_xlsx_compression_ratio():
    output=io.BytesIO()
    with zipfile.ZipFile(output,"w",zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml",b"0"*1_000_000)
    with pytest.raises(ValueError,match="valid .xlsx workbook"):
        LegalStore.detention_workbook_sheets(output.getvalue())


def test_amal_only_hides_detention_and_deportation_but_awareness_depends_on_file():
    payload=required_payload()
    for name in ("beneficiaries","assessments","legalservices"):
        frame=pd.read_csv(io.BytesIO(payload[name]),dtype=object)
        frame["Project"]="UNHCR 2026 - AMAL CAMP"
        payload[name]=frame.to_csv(index=False).encode("utf-8")
    payload["awareness"]=csv(**{"Awareness ID":["W1"],"Participant Name":["Person"],"Project":["UNHCR 2026 - AMAL CAMP"]})
    payload["deportationrecords"]=csv(**{"PN ID":["D1"],"Project":["UNHCR 2026 - AMAL CAMP"]})
    metadata=LegalStore.from_files(payload,"test").metadata()
    assert metadata["availability"]["awareness"] is True
    assert metadata["features"]=={"awareness":True,"detention":False,"deportation":False}


def test_deportation_dashboard_uses_deportationrecords_csv():
    payload=required_payload()
    payload["deportationrecords"]=csv(**{"PN ID":["D1","D2"],"Date of deporting":["05/01/2026","10/01/2026"],"Destination":["Country A","Country B"],"Nationality":["Iraqi","Syrian"]})
    dashboard=LegalStore.from_files(payload,"test").deportation_dashboard()
    assert dashboard["total"]==2
    assert dashboard["trend"]==[{"label":"2026-01","count":2,"percent":1.0}]
    assert next(chart for chart in dashboard["charts"] if chart["title"]=="Deportations by destination")["rows"][0]["count"]==1


def test_deportation_dashboard_filters_and_charts_use_distinct_pn_ids():
    payload=required_payload()
    payload["deportationrecords"]=csv(**{"PN ID":["D1","D1","D2"],"Date of deporting":["05/01/2026","06/01/2026","10/02/2026"],"Destination":["Country A","Country A","Country B"]})
    store=LegalStore.from_files(payload,"test")
    dashboard=store.deportation_dashboard({"Month":["2026-01"]})
    destination=next(chart for chart in dashboard["charts"] if chart["title"]=="Deportations by destination")
    assert dashboard["total"]==1
    assert destination["rows"]==[{"label":"Country A","count":1,"percent":1.0}]
    assert dashboard["filterOptions"]["Month"]==["2026-01","2026-02"]


def test_missing_mandatory_file_rejects_import():
    payload=required_payload();del payload["assessments"]
    with pytest.raises(ValueError,match="assessments.csv"): LegalStore.from_files(payload,"test")


def test_blank_consolidated_names_are_built_before_split_columns_are_hidden():
    payload=required_payload()
    beneficiaries=pd.read_csv(io.BytesIO(payload["beneficiaries"]),dtype=object)
    beneficiaries=beneficiaries.drop(columns=[column for column in beneficiaries.columns if column.endswith(": First")])
    beneficiaries["Name (Filter Color Red)"]=""
    beneficiaries["Name / Arabic: First"]=["Ahmed","Ahmad"]
    beneficiaries["Name / Arabic: Middle"]=["Ali","Ali"]
    beneficiaries["Name / Arabic: Last"]=["Hassani","Hassani"]
    beneficiaries["Project"]=["UNHCR 2026 - Erbil","UNHCR 2026 - Erbil"]
    payload["beneficiaries"]=beneficiaries.to_csv(index=False).encode("utf-8")
    store=LegalStore.from_files(payload,"test")
    assert store.frames["beneficiaries"]["Name (Filter Color Red)"].tolist()==["Ahmed Ali Hassani","Ahmad Ali Hassani"]
    assert not any(column.endswith(": First") for column in store.frames["beneficiaries"].columns)
    assert store.review("beneficiaries",rule="Possible duplicate name")["total"]==0
    assert store.review("beneficiaries",rule="Possible duplicate name",allow_name_variations=True)["total"]==2


def test_name_matching_requires_at_least_the_selected_number_of_characters():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2","B3","B4"],
        "Name (Filter Color Red)":["short","short","A"*31,"A"*31],
        "Project":["UNHCR 2026 - Erbil"]*4,
    })
    store=LegalStore.from_files(payload,"test")
    assert store.review("beneficiaries",rule="Possible duplicate name",name_compare_chars=30)["total"]==2
    assert store.review("beneficiaries",rule="Possible duplicate name",name_compare_chars=30)["nameCompareCharsApplied"]==30


def test_name_character_sensitivity_changes_duplicate_table_results():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2"],
        "Name (Filter Color Red)":["A"*15+"X"*15,"A"*15+"Y"*15],
        "Project":["UNHCR 2026 - Erbil"]*2,
    })
    store=LegalStore.from_files(payload,"test")
    assert store.review("beneficiaries",rule="Possible duplicate name",name_compare_chars=10)["total"]==2
    assert store.review("beneficiaries",rule="Possible duplicate name",name_compare_chars=15)["total"]==2
    assert store.review("beneficiaries",rule="Possible duplicate name",name_compare_chars=30)["total"]==0


def test_small_spelling_differences_are_optional_and_do_not_match_unrelated_names():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2","B3"],
        "Name (Filter Color Red)":["Ahmed Ali Hassani","Ahmad Ali Hassani","Completely Other"],
        "Project":["UNHCR 2026 - Erbil"]*3,
    })
    store=LegalStore.from_files(payload,"test")
    assert store.review("beneficiaries",rule="Possible duplicate name")["total"]==0
    result=store.review("beneficiaries",rule="Possible duplicate name",allow_name_variations=True)
    assert result["total"]==2
    assert {row["caseId"] for row in result["rows"]}=={"B1","B2"}
    assert result["allowNameVariationsApplied"] is True
    assert {row["nameMatchMode"] for row in result["rows"]}=={"variation"}
    assert all(90 <= row["duplicateSimilarity"] < 100 for row in result["rows"])


def test_contact_and_name_duplicate_requires_matching_contact_markers_and_project_group():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2","B3","B4","B5"],
        "Name (Filter Color Red)":["Ahmed Ali Hassan","Ahmad Ali Hassan","Ahmed Ali Hassan","Ahmed Ali Hassan","Ahmed Ali Hassan"],
        "Contact Number":["07701234567","0770-123-4567","07709999999","07701234567","07701234567"],
        "# UNHCR":["U-1","","U-3","","U-5"],
        "Spouse name":["","Spouse Two","","","Spouse Five"],
        "Project":["UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - SULI"],
    })
    store=LegalStore.from_files(payload,"test")
    result=store.review("beneficiaries",rule="Possible duplicate contact and name",page_size=100)
    assert {row["caseId"] for row in result["rows"]}=={"B1","B2"}
    assert all(row["duplicateSimilarity"] >= 90 for row in result["rows"])
    assert all("contact number 07701234567 matches" in row["detail"] for row in result["rows"])
    store.set_review_exclusions({("Possible duplicate contact and name","B1")})
    assert store.review("beneficiaries",rule="Possible duplicate contact and name")["total"]==0
    workbook=load_workbook(io.BytesIO(store.review_export("beneficiaries")),read_only=True,data_only=True)
    for sheet in workbook.worksheets:
        headers=[cell.value for cell in next(sheet.iter_rows(min_row=1,max_row=1))]
        finding_index=headers.index("Review Finding")
        assert all(row[finding_index]!="Possible duplicate contact and name" for row in sheet.iter_rows(min_row=2,values_only=True))


def test_invalid_age_includes_invalid_and_future_spouse_dates_without_flagging_blank_dates():
    payload=required_payload();future=(date.today()+timedelta(days=1)).strftime("%d/%m/%Y")
    payload["beneficiaries"]=csv(**{
        "Case ID":["Valid","Malformed","Future","Blank"],
        "Name (Filter Color Red)":["Valid Person","Malformed Person","Future Person","Blank Person"],
        "Spouse DoB":["01/01/1990","not a date",future,""],
    })
    rows=LegalStore.from_files(payload,"test").review("beneficiaries",rule="Invalid age",page_size=100)["rows"]
    assert {row["caseId"] for row in rows}=={"Malformed","Future"}
    assert {row["detail"] for row in rows}=={"Spouse DoB is not a valid date","Spouse DoB is later than the current date"}


def test_exact_duplicate_mode_remains_exact_when_variations_are_enabled():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2"],"Name (Filter Color Red)":["Identical Beneficiary Name"]*2,
        "Project":["UNHCR 2026 - Baghdad","UNHCR 2026 - Gov"],
    })
    rows=LegalStore.from_files(payload,"test").review("beneficiaries",rule="Possible duplicate name",allow_name_variations=True)["rows"]
    assert {row["nameMatchMode"] for row in rows}=={"exact"}
    assert {row["duplicateSimilarity"] for row in rows}=={100}


def test_duplicate_name_selector_count_uses_only_non_excluded_exact_matches():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["Exact1","Exact2","Similar1","Similar2"],
        "Name (Filter Color Red)":["Identical Person Long","Identical Person Long","Ahmed Ali Hassani","Ahmad Ali Hassani"],
        "Project":["UNHCR 2026 - Erbil"]*4,
    })
    store=LegalStore.from_files(payload,"test")
    all_matches=store.review("beneficiaries",rule="Possible duplicate name",name_compare_chars=15,allow_name_variations=True,exact_matches_only=False)
    exact_table=store.review("beneficiaries",rule="Possible duplicate name",name_compare_chars=15,allow_name_variations=True,exact_matches_only=True)
    assert all_matches["ruleCounts"]["Possible duplicate name"]==2
    assert exact_table["ruleCounts"]["Possible duplicate name"]==2
    assert all_matches["total"] > exact_table["total"]
    store.set_review_exclusions({("Possible duplicate name","Exact1")})
    assert store.review("beneficiaries",name_compare_chars=15,allow_name_variations=True)["ruleCounts"]["Possible duplicate name"]==0


def test_beneficiaries_navigation_count_uses_exact_non_excluded_name_matches():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["Exact1","Exact2","Similar1","Similar2"],
        "Name (Filter Color Red)":["Identical Person Long","Identical Person Long","Ahmed Ali Hassani","Ahmad Ali Hassani"],
        "Project":["UNHCR 2026 - Erbil"]*4,
    })
    store=LegalStore.from_files(payload,"test")
    before=store.metadata()["reviewCounts"]["beneficiaries"]
    assert before==sum(store.review("beneficiaries")["ruleCounts"].values())
    store.set_review_exclusions({("Possible duplicate name","Exact1")})
    assert store.metadata()["reviewCounts"]["beneficiaries"]==before-2


def test_duplicate_exclusion_registry_persists_and_restores(tmp_path):
    path=tmp_path/"duplicate-name-exclusions.json"
    registry=DuplicateExclusionRegistry(path)
    registry.exclude("B1","Possible duplicate name","Identical Name","UNHCR 2026 - Erbil","Beneficiaries Review")
    reloaded=DuplicateExclusionRegistry(path)
    assert reloaded.case_ids()=={"B1"}
    assert reloaded.entries()[0]["name"]=="Identical Name"
    assert reloaded.restore("B1","Possible duplicate name") is True
    assert DuplicateExclusionRegistry(path).entries()==[]


def test_duplicate_exclusion_registry_bulk_adds_once_and_skips_existing_records(tmp_path):
    registry=DuplicateExclusionRegistry(tmp_path/"exclusions.json")
    records=[
        {"dataset":"assessments","rule":"Missing document","identifierType":"assessmentId","identifierValue":"A1","name":"First","source":"Assessments Review"},
        {"dataset":"assessments","rule":"Missing document","identifierType":"assessmentId","identifierValue":"A2","name":"Second","source":"Assessments Review"},
    ]
    created, created_count, duplicates=registry.exclude_records(records)
    assert created_count==2
    assert duplicates==0
    assert {row["identifierValue"] for row in created}=={"A1","A2"}
    _, created_count, duplicates=registry.exclude_records(records)
    assert created_count==0
    assert duplicates==2
    assert {row["identifierValue"] for row in DuplicateExclusionRegistry(registry.path).entries()}=={"A1","A2"}


def test_excluding_case_recalculates_duplicate_groups_and_export():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2","B3"],"Name (Filter Color Red)":["Identical Beneficiary Name"]*3,
        "Project":["UNHCR 2026 - Erbil"]*3,
    })
    store=LegalStore.from_files(payload,"test")
    store.set_review_exclusions({("Possible duplicate name","B2")})
    rows=store.review("beneficiaries",rule="Possible duplicate name")["rows"]
    assert {row["caseId"] for row in rows}=={"B1","B3"}
    store.set_review_exclusions({("Possible duplicate name","B1"),("Possible duplicate name","B2")})
    assert store.review("beneficiaries",rule="Possible duplicate name")["total"]==0
    workbook=load_workbook(io.BytesIO(store.review_export("beneficiaries")),read_only=True,data_only=True)
    sheet=workbook["North Iraq"]
    headers=[cell.value for cell in next(sheet.iter_rows(min_row=1,max_row=1))]
    rows=list(sheet.iter_rows(min_row=2,values_only=True));rule_index=headers.index("Review Finding");case_index=headers.index("Case ID")
    assert not any(row[rule_index]=="Possible duplicate name" and row[case_index] in {"B1","B2"} for row in rows)


def test_rule_specific_exclusion_keeps_other_beneficiary_findings_visible():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1"],"Name (Filter Color Red)":["Beneficiary Name"],"Project":["UNHCR 2026 - Erbil"],
        "Age":[16],"Marital Status":["Married"],"Contact Number":["123"],
    })
    store=LegalStore.from_files(payload,"test")
    store.set_review_exclusions({("Invalid contact number","B1")})
    assert store.review("beneficiaries",rule="Invalid contact number")["total"]==0
    assert store.review("beneficiaries",rule="Marital status below 18")["total"]==1


def test_duplicate_similarity_uses_each_rows_strongest_peer():
    payload=required_payload();base="A"*20
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2","B3"],"Name (Filter Color Red)":[base,base,"A"*18+"BB"],
        "Project":["UNHCR 2026 - Erbil"]*3,
    })
    rows=LegalStore.from_files(payload,"test").review("beneficiaries",rule="Possible duplicate name",name_compare_chars=20,allow_name_variations=True)["rows"]
    by_case={row["caseId"]:row for row in rows}
    assert by_case["B1"]["duplicateSimilarity"]==100
    assert by_case["B2"]["duplicateSimilarity"]==100
    assert by_case["B3"]["duplicateSimilarity"]==90


def test_duplicate_name_review_can_show_only_100_percent_matches():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2","B3","B4","B5","B6","B7"],
        "Name (Filter Color Red)":["Identical Beneficiary Name","Identical Beneficiary Name","Identical Beneficiary Namo","Li","Li","Li","Li"],
        "Project":["UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - Baghdad","UNHCR 2026 - Gov","UNHCR 2026 - SULI","UNHCR 2026 - Mosul & Kirkuk"],
    })
    store=LegalStore.from_files(payload,"test")
    rows=store.review("beneficiaries",rule="Possible duplicate name",name_compare_chars=30,allow_name_variations=True,exact_matches_only=True)["rows"]
    assert {row["caseId"] for row in rows}=={"B1","B2","B4","B5"}
    assert {row["duplicateSimilarity"] for row in rows}=={100}


def test_exact_match_export_keeps_other_findings_and_excludes_variations():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2","B3","B4"],
        "Name (Filter Color Red)":["Identical Beneficiary Name","Identical Beneficiary Name","Identical Beneficiary Namo","Unrelated Person"],
        "Project":["UNHCR 2026 - Erbil"]*4,
        "Age":[20,20,20,"invalid"],
        "# total assessments":[1,1,1,1],
    })
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("beneficiaries",exact_matches_only=True)),read_only=True,data_only=True)
    rows=list(workbook["North Iraq"].iter_rows(values_only=True));header=list(rows[0])
    finding_index=header.index("Review Finding");case_index=header.index("Case ID")
    duplicate_cases={row[case_index] for row in rows[1:] if row[finding_index]=="Possible duplicate name"}
    assert duplicate_cases=={"B1","B2"}
    assert any(row[finding_index]=="Invalid age" and row[case_index]=="B4" for row in rows[1:])


def test_duplicate_names_follow_north_south_and_amal_project_boundaries():
    payload=required_payload()
    duplicate="Identical Beneficiary Name"
    payload["beneficiaries"]=csv(**{
        "Case ID":["E1","E2","M1","S1","B1","G1","B2","A1","A2","U1","U2"],
        "Name (Filter Color Red)":[duplicate]*11,
        "Project":[
            "UNHCR 2026 - Erbil","UNHCR 2026 - Erbil",
            "UNHCR 2026 - Mosul & Kirkuk","UNHCR 2026 - SULI",
            "UNHCR 2026 - Baghdad","UNHCR 2026 - Gov","UNHCR 2026 - Baghdad",
            "UNHCR 2026 - AMAL CAMP","UNHCR 2026 - AMAL CAMP",
            "Unknown Project","",
        ],
    })
    store=LegalStore.from_files(payload,"test")
    result=store.review("beneficiaries",rule="Possible duplicate name")
    flagged={row["caseId"] for row in result["rows"]}
    assert flagged=={"E1","E2","B1","G1","B2","A1","A2"}
    details={row["caseId"]:row["detail"] for row in result["rows"]}
    assert "North - Erbil" in details["E1"]
    assert "South (Baghdad + Gov)" in details["G1"]
    assert "AMAL" in details["A1"]


def test_name_counts_distinguish_loaded_names_from_threshold_eligible_names():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2"],
        "Name (Filter Color Red)":["Short Name","Another Short"],
        "Project":["UNHCR 2026 - Baghdad","UNHCR 2026 - Gov"],
    })
    result=LegalStore.from_files(payload,"test").review("beneficiaries",name_compare_chars=30)
    assert result["nameRecordCount"]==2
    assert result["eligibleNameRecordCount"]==0


def test_review_export_uses_the_same_south_duplicate_group():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","G1","E1"],
        "Name (Filter Color Red)":["Shared Beneficiary Name"]*3,
        "Project":["UNHCR 2026 - Baghdad","UNHCR 2026 - Gov","UNHCR 2026 - Erbil"],
    })
    exported=LegalStore.from_files(payload,"test").review_export("beneficiaries")
    workbook=load_workbook(io.BytesIO(exported),read_only=True,data_only=True)
    assert workbook.sheetnames==["North Iraq","AMAL Camp","South Iraq"]
    rows=list(workbook["South Iraq"].iter_rows(values_only=True))
    header=list(rows[0]);case_id_index=header.index("Case ID");detail_index=header.index("Review Detail")
    duplicate_rows=[row for row in rows[1:] if row[header.index("Review Finding")]=="Possible duplicate name"]
    assert {row[case_id_index] for row in duplicate_rows}=={"B1","G1"}
    assert all("South (Baghdad + Gov)" in row[detail_index] for row in duplicate_rows)


def test_review_export_uses_consistent_rows_and_duplicate_name_colors():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","G1"],"Name (Filter Color Red)":["Shared Beneficiary Name"]*2,
        "Project":["UNHCR 2026 - Baghdad","UNHCR 2026 - Gov"],
    })
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("beneficiaries")))
    sheet=workbook["South Iraq"]
    header=[cell.value for cell in sheet[1]];name_column=header.index("Name")+1
    duplicate_rows=[row for row in range(2,sheet.max_row+1) if sheet.cell(row,2).value=="Possible duplicate name"]
    assert {sheet.row_dimensions[row].height for row in duplicate_rows}=={24.0}
    assert len({sheet.cell(row,name_column).fill.fgColor.rgb for row in duplicate_rows})==1
    assert {sheet.cell(row,name_column).fill.fgColor.rgb for row in duplicate_rows}=={"00FDE8E8"}
    assert {sheet.cell(row,name_column).font.color.rgb for row in duplicate_rows}=={"00991B1B"}


def test_beneficiary_review_export_groups_findings_by_region():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["E","S","M","A","B","G"],
        "Name (Filter Color Red)":["Erbil","Suli","Mosul","Amal","Baghdad","Gov"],
        "Project":["UNHCR 2026 - Erbil","UNHCR 2026 - SULI","UNHCR 2026 - Mosul & Kirkuk","UNHCR 2026 - AMAL CAMP","UNHCR 2026 - Baghdad","UNHCR 2026 - Gov"],
        "Age":["bad"]*6,
    })
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("beneficiaries")),read_only=True,data_only=True)
    assert workbook.sheetnames==["North Iraq","AMAL Camp","South Iraq"]
    def case_ids(sheet_name):
        rows=list(workbook[sheet_name].iter_rows(values_only=True));case_index=list(rows[0]).index("Case ID")
        return {row[case_index] for row in rows[1:]}
    assert case_ids("North Iraq")=={"E","S","M"}
    assert case_ids("AMAL Camp")=={"A"}
    assert case_ids("South Iraq")=={"B","G"}


def test_marital_and_spouse_age_use_current_date_of_birth_age():
    payload=required_payload();current_year=date.today().year
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2"],"Name (Filter Color Red)":["Adult Beneficiary","Minor Beneficiary"],
        "Age":[17,25],"DoB / Date of Birth":[f"01/01/{current_year-19}",f"01/01/{current_year-17}"],
        "Marital Status":["Married","Married"],"Spouse name":["Young Spouse","Adult Spouse"],
        "Spouse DoB":[f"01/01/{current_year-17}",f"01/01/{current_year-20}"],
    })
    result=LegalStore.from_files(payload,"test").review("beneficiaries",page_size=100)
    marital=[row for row in result["rows"] if row["rule"]=="Marital status below 18"]
    assert {row["caseId"] for row in marital}=={"B2"}
    spouses=[row for row in result["rows"] if row["rule"]=="Spouse below 18"]
    assert {row["caseId"] for row in spouses}=={"B1"}
    assert spouses[0]["spouseName"]=="Young Spouse"
    assert spouses[0]["spouseAge"]==17


def test_normalization_and_review_rules():
    store=LegalStore.from_files(required_payload(),"test")
    assert normalize_name(" أَحْمَد ة ")==normalize_name("احمده")
    assert phone_digits("٠٧٧٠-١٢٣-٤٥٦٧")=="07701234567"
    rules={row["rule"] for row in store.review("beneficiaries")["rows"]}
    assert {"Case without assessment","Invalid age","Invalid contact number"}<=rules
    assert "Possible duplicate name" not in rules
    assessment_rules={row["rule"] for row in store.review("assessments")["rows"]}
    assert "Beneficiary has multiple assessments" not in assessment_rules


def test_contact_numbers_ignore_one_digit_and_prefix_ten_digits():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2","B3"],
        "Name (Filter Color Red)":["Alpha Person","Beta Person","Gamma Person"],
        "Contact Number":["7","7701234567","12345"],
    })
    store=LegalStore.from_files(payload,"test")
    assert store.frames["beneficiaries"]["Contact Number"].tolist()==["7","07701234567","12345"]
    flagged={row["caseId"] for row in store.review("beneficiaries",rule="Invalid contact number")["rows"]}
    assert flagged=={"B3"}


def test_contact_numbers_ignore_requested_prefixes():
    payload=required_payload()
    prefixes=["6939","4915","2376","9054","2951","4916"]
    payload["beneficiaries"]=csv(**{
        "Case ID":[f"B{index}" for index in range(len(prefixes)+1)],
        "Name (Filter Color Red)":[f"Person {index}" for index in range(len(prefixes)+1)],
        "Contact Number":[f"{prefix}123" for prefix in prefixes]+["12345"],
    })
    flagged={row["caseId"] for row in LegalStore.from_files(payload,"test").review("beneficiaries",rule="Invalid contact number")["rows"]}
    assert flagged=={"B6"}


def test_community_type_vs_nationality_flags_only_invalid_combinations():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["I1","S1","N1","I2","S2","N2","B1","U1"],
        "Name (Filter Color Red)":[f"Person {index}" for index in range(8)],
        "Community Type":["IDPs","Syrian Refugee","Non-Syrian Refugee","IDPs","Syrian Refugee","Non-Syrian Refugee","","Host Community"],
        "Nationality":["Iraq","Syria","Iran","Syria","Iraq","Syria","Iraq","Iraq"],
        "Project":["UNHCR 2026 - AMAL CAMP","UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - AMAL CAMP","UNHCR 2026 - AMAL CAMP"],
    })
    rows=LegalStore.from_files(payload,"test").review("beneficiaries",rule="Check Community Type vs Nationality",page_size=100)["rows"]
    assert {row["caseId"] for row in rows}=={"I2","S2","N2","B1","U1"}
    assert "blank" in next(row["detail"] for row in rows if row["caseId"]=="B1")


def test_under_18_review_rows_sort_youngest_first_and_include_beneficiary_age():
    current_year=date.today().year
    payload=required_payload()
    payload["beneficiaries"]=csv(**{
        "Case ID":["Older","Younger"], "Name (Filter Color Red)":["Older Person","Younger Person"],
        "Age":[17,15], "DoB":[f"01/01/{current_year-17}",f"01/01/{current_year-15}"],
        "Marital Status":["Married","Married"], "Spouse DoB":[f"01/01/{current_year-17}",f"01/01/{current_year-15}"],
    })
    store=LegalStore.from_files(payload,"test")
    marital=store.review("beneficiaries",rule="Marital status below 18",page_size=100)["rows"]
    spouses=store.review("beneficiaries",rule="Spouse below 18",page_size=100)["rows"]
    assert [row["caseId"] for row in marital]==["Younger","Older"]
    assert [row["caseId"] for row in spouses]==["Younger","Older"]
    assert marital[0]["beneficiaryAge"]==15


def connected_case_payload():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{"Case ID":["B1","B2"],"Name (Filter Color Red)":["First Person","Second Person"],"Lawyer":["Beneficiary Lawyer","Other Lawyer"],"Project":["North","South"]})
    payload["assessments"]=csv(**{"Assessment ID":["A1","A2","A3"],"Beneficiary ID":["B1","B1","B2"],"Assessment Status":["Closed","Open","Open"],"Lawyer":["Assessment Lawyer","Second Assessment Lawyer","Other Lawyer"]})
    payload["legalservices"]=csv(**{"Service ID":["S1","S2","S3"],"Assessment ID":["A1","A2","A3"],"Beneficiary ID":["B1","B1","B2"],"Service Status":["Closed after review","Open","Open"],"Lawyer":["Service Lawyer","","Other Lawyer"]})
    payload["followupslogbooks"]=csv(**{"Follow-up ID":["F1","F2"],"Service ID":["S1","S2"],"Follow-up Status":["Close","Pending"]})
    payload["legalfees"]=csv(**{"Fee ID":["L1"],"Legal Service ID":["S1"],"Payment Status":["Paid"]})
    return payload


def test_lawyer_summary_monthly_assessments_only_uses_2026_and_later():
    payload=required_payload()
    payload["assessments"]=csv(**{
        "Assessment ID":["OLD","A1","A2","A3","A3"],
        "Beneficiary ID":["B1","B1","B2","B3","B3"],
        "Date of Assessment تاريخ التقييم":["31/12/2025","10/01/2026","20/02/2026","21/02/2026","21/02/2026"],
        "Lawyers":["Lawyer One","Lawyer One","Lawyer One","Lawyer Two","Lawyer Two"],
    })
    result=LegalStore.from_files(payload,"test").lawyer_summary()
    monthly={(row["lawyer"],row["month"]):(row["count"],row["average"]) for row in result["monthlyAssessments"]}
    assert monthly[("Lawyer One","2026-01")]==(1,1.0)
    assert monthly[("Lawyer One","2026-02")]==(1,1.0)
    assert monthly[("Lawyer Two","2026-02")]==(1,0.5)
    assert all(month>="2026-01" for _,month in monthly)


def test_overview_representation_trend_includes_legal_assistance():
    payload=required_payload()
    payload["legalservices"]=csv(**{
        "Service ID":["S1","S2","S3","S3"],
        "Assessment ID":["A1","A1","A2","A2"],
        "Beneficiary ID":["B1","B1","B1","B1"],
        "Type of Service Provided":["Legal Representation","Legal Assistance","Legal Counselling","Legal Representation"],
        "Date of Service Provision":["10/01/2026","15/01/2026","20/02/2026","20/02/2026"],
    })
    trend=LegalStore.from_files(payload,"test").metadata()["overview"]["representationTrend"]
    assert trend==[{"month":"2026-01","representation":2},{"month":"2026-02","representation":1}]


def test_intelligence_integrates_distinct_records_and_keeps_awareness_separate():
    payload=required_payload()
    payload["assessments"]=csv(**{"Assessment ID":["OLD","A1","A2"],"Beneficiary ID":["B1","B1","B2"],"Date of Assessment":["31/12/2025","10/01/2026","12/01/2026"],"Lawyers":["One","One","Two"]})
    payload["legalservices"]=csv(**{"Service ID":["S1","S1","S2"],"Assessment ID":["A1","A1","A2"],"Beneficiary ID":["B1","B1","B2"],"Service Status":["Completed","Completed","Open"],"Lawyers":["One","One","Two"]})
    payload["followupslogbooks"]=csv(**{"Follow-ups & Logbook ID":["F1"],"Service ID":["S1"],"Date of follow-up":["15/01/2026"],"Lawyer":["One"]})
    payload["legalfees"]=csv(**{"Fee ID":["L1"],"Legal Service ID":["S1"],"Amount Spent (IQD)":["10,000"],"Paid date":["16/01/2026"],"Created by":["One"]})
    payload["deportationrecords"]=csv(**{"PN ID":["D1"],"Date of deporting":["17/01/2026"]})
    payload["awareness"]=csv(**{"Awareness ID":["W1","W2"],"Date of Session":["18/01/2026","18/01/2026"],"Lawyer":["One","One"]})
    result=LegalStore.from_files(payload,"test").intelligence("command-center")
    kpis={item["label"]:item["value"] for item in result["kpis"]}
    assert kpis["Assessments"]==2
    assert kpis["Legal services"]==2
    assert kpis["Completed services"]==1
    assert kpis["Awareness participants"]==2
    assert kpis["Legal fees"]==10000
    assert result["period"]=="2026 onward"


def test_case_search_resolves_case_assessment_and_service_ids():
    store=LegalStore.from_files(connected_case_payload(),"test")
    assert {item["beneficiary"]["Case ID"] for item in store.case("B1")["cases"]}=={"B1"}
    assert {item["beneficiary"]["Case ID"] for item in store.case("A2")["cases"]}=={"B1"}
    assert {item["beneficiary"]["Case ID"] for item in store.case("S1")["cases"]}=={"B1"}


def test_case_filters_cover_connected_datasets_with_contains_semantics():
    store=LegalStore.from_files(connected_case_payload(),"test")
    result=store.case("",{"assessments::Assessment Status":["Close"],"legalservices::Service Status":["review"]})
    assert {item["beneficiary"]["Case ID"] for item in result["cases"]}=={"B1"}
    options=store.case_filters();labels={group["label"] for group in options["groups"]}
    assert {"Beneficiary","Assessment","Legal service","Follow-up","Legal fee"}<=labels
    assessment_group=next(group for group in options["groups"] if group["dataset"]=="assessments")
    status=next(column for column in assessment_group["columns"] if column["name"]=="Assessment Status")
    assert status["values"].count("Closed")==1


def test_connected_case_table_returns_one_expandable_hierarchy_per_case():
    store=LegalStore.from_files(connected_case_payload(),"test")
    result=store.case("B1",view_mode="table",page=1,page_size=1)
    assert result["totalCases"]==result["totalRows"]==1 and len(result["cases"])==1
    case=result["cases"][0]
    assert len(case["assessments"])==2
    assert sum(len(node["services"]) for node in case["assessments"])==2
    assert case["counts"]=={"assessments":2,"services":2,"followups":2,"fees":1}
    assert case["lawyers"]==["Service Lawyer","Assessment Lawyer","Second Assessment Lawyer","Beneficiary Lawyer"]


def test_assessment_review_excludes_requested_service_rule():
    store=LegalStore.from_files(required_payload(),"test")
    result=store.review("assessments")
    assert "Requested service not provided" not in result["ruleCounts"]
    assert all(row["rule"]!="Requested service not provided" for row in result["rows"])


def test_awareness_duplicate_priority_uses_name_and_session_and_minor_is_hidden_from_overview():
    payload=required_payload()
    payload["awareness"]=csv(**{
        "Awareness ID":["W1","W2","W3"],
        "Participant Name":["Same Person","Same Person","Same Person"],
        "Session Topic":["Housing","Housing","Documentation"],
    })
    store=LegalStore.from_files(payload,"test")
    rows=store.review("awareness",page_size=100)["rows"]
    high=[row for row in rows if row["rule"]=="Duplicate participant in session"]
    minor=[row for row in rows if row["rule"]=="Possible duplicate participant name"]
    assert {row["awarenessId"] for row in high}=={"W1","W2"}
    assert {row["awarenessId"] for row in minor}=={"W3"}
    assert len({row["duplicateGroup"] for row in high})==1
    assert all(row["severity"]=="High" for row in high)
    assert all(row["severity"]=="Minor" for row in minor)
    assert store.metadata()["reviewCounts"]["awareness"]==2


def test_duplicate_participants_in_session_are_sorted_by_name():
    payload=required_payload();payload["awareness"]=csv(**{
        "Awareness ID":["W1","W2","W3","W4"],
        "Participant Name":["Zain","Zain","Ahmed","Ahmed"],
        "Session Topic":["Housing","Housing","Housing","Housing"],
    })
    rows=LegalStore.from_files(payload,"test").review("awareness",rule="Duplicate participant in session",page_size=100)["rows"]
    assert [row["name"] for row in rows]==["Ahmed","Ahmed","Zain","Zain"]


def test_assessment_review_hides_detention_rules_for_amal_only_projects():
    payload=required_payload();payload["assessments"]=csv(**{
        "Assessment ID":["A1"], "Beneficiary ID":["B1"], "Projects - المشروع":["UNHCR 2026 - AMAL CAMP"],
        "Community Type":["Syrian Refugee"], "Date of Assessment":["01/01/2026"], "Created On":["05/01/2026"],
        "Is the beneficiary detained":["No"], "Is it an immigration related charge?":[""],
    })
    payload["legalservices"]=csv(**{"Service ID":["S1"],"Assessment ID":["A1"],"Beneficiary ID":["B1"],"Type of Service Provided":["Legal Representation"]})
    result=LegalStore.from_files(payload,"test").review("assessments",page_size=100)
    assert not DETENTION_ASSESSMENT_RULES.intersection(result["ruleCounts"])
    assert not any(row["rule"] in DETENTION_ASSESSMENT_RULES for row in result["rows"])
    assert "Representation while not detained" not in result["ruleCounts"]


def test_assessment_review_keeps_detention_rules_when_project_scope_is_not_amal_only():
    payload=required_payload();payload["assessments"]=csv(**{
        "Assessment ID":["A1"], "Beneficiary ID":["B1"], "Projects - المشروع":["UNHCR 2026 - Erbil"],
        "Community Type":["Syrian Refugee"], "Date of Assessment":["01/01/2026"],
        "Is the beneficiary detained":["Yes"], "Is it an immigration related charge?":[""],
    })
    result=LegalStore.from_files(payload,"test").review("assessments",page_size=100)
    assert "Detention/immigration inconsistency" in result["ruleCounts"]


def test_metadata_hides_detention_for_amal_assessments_despite_auxiliary_project_data():
    payload=required_payload()
    payload["assessments"]=csv(**{"Assessment ID":["A1"],"Beneficiary ID":["B1"],"Projects - المشروع":["UNHCR 2026 - AMAL CAMP"]})
    payload["followupslogbooks"]=csv(**{"Follow-ups & Logbook ID":["F1"],"Projects - المشروع":["UNHCR 2026 - Baghdad"]})
    assert LegalStore.from_files(payload,"test").metadata()["features"]["detention"] is False


def test_review_flags_assessment_and_service_dates_after_today():
    tomorrow=(date.today()+timedelta(days=1)).strftime("%d/%m/%Y")
    payload=required_payload()
    payload["assessments"]=csv(**{
        "Assessment ID":["A1"], "Beneficiary ID":["B1"], "Date of Assessment":[tomorrow],
        "Date of Detention":[tomorrow], "Date of Assessment Closure":["01/01/2026"],
    })
    payload["legalservices"]=csv(**{
        "Service ID":["S1"], "Assessment ID":["A1"], "Beneficiary ID":["B1"],
        "Date of Service Provision":[tomorrow], "Date of Issuance":[tomorrow],
    })
    store=LegalStore.from_files(payload,"test")
    assessment_rows=store.review("assessments",rule="Assessment date after today",page_size=100)["rows"]
    service_rows=store.review("legalservices",rule="Legal service date after today",page_size=100)["rows"]
    assert len(assessment_rows)==1 and "Date of Assessment" in assessment_rows[0]["detail"] and "Date of Detention" in assessment_rows[0]["detail"]
    assert len(service_rows)==1 and "Date of Service Provision" in service_rows[0]["detail"] and "Date of Issuance" in service_rows[0]["detail"]


def test_assessment_date_of_request_imports_as_month_day_year_and_displays_day_month_year():
    payload=required_payload();payload["assessments"]=csv(**{
        "Assessment ID":["A1"], "Beneficiary ID":["B1"], "Date of the Request":["03/04/2026"],
    })
    store=LegalStore.from_files(payload,"test")
    request_column="Date of the Request"
    assert store.frames["assessments"].loc[0,request_column]==pd.Timestamp("2026-03-04")
    assert store.explorer("assessments",page_size=10)["rows"][0][request_column]=="04/03/2026"


def test_case_hierarchy_works_without_optional_files():
    store=LegalStore.from_files(required_payload(),"test")
    case=store.case("B1")["cases"][0]
    assert len(case["assessments"])==2
    assert sum(len(assessment["services"]) for assessment in case["assessments"])==1
    service=case["assessments"][1]["services"][0]
    assert service["followups"]==[] and service["fees"]==[]
    workbook=load_workbook(io.BytesIO(store.case_export("B1")))
    assert workbook.sheetnames==["Beneficiaries","Assessments","Services"]
    sheet=workbook["Beneficiaries"]
    assert sheet.freeze_panes=="A2"
    assert "Case ID" in [cell.value for cell in sheet[1]]


def test_connected_case_table_sorts_and_paginates_by_beneficiary_case():
    store=LegalStore.from_files(connected_case_payload(),"test")
    result=store.case("",{},view_mode="table",page=1,page_size=1,sort_column="beneficiaries::Case ID",sort_direction="desc")
    assert result["totalCases"]==2 and len(result["cases"])==1
    assert result["cases"][0]["beneficiary"]["Case ID"]=="B2"
    second=store.case("",{},view_mode="table",page=2,page_size=1,sort_column="beneficiaries::Case ID",sort_direction="desc")
    assert second["cases"][0]["beneficiary"]["Case ID"]=="B1"


def test_explorer_excel_export_respects_search_and_checkbox_filters():
    store=LegalStore.from_files(required_payload(),"test")
    exported=store.explorer_export("beneficiaries","B1",{"Age":["17"]},"xlsx")
    assert exported[:2]==b"PK"
    workbook=pd.read_excel(io.BytesIO(exported))
    assert workbook["Case ID"].tolist()==["B1"]


def test_refugee_2026_detention_scope_and_selectable_month():
    payload=required_payload()
    payload["assessments"]=csv(**{
        "Assessment ID":["A1","A2","A3","A4"],"Beneficiary ID":["B1","B1","B2","B2"],
        "Date of Assessment تاريخ التقييم":["15/12/2025","10/01/2026","12/01/2026","15/02/2026"],
        "Community Type":["Syrian Refugee لاجيء-سوري","Syrian Refugee لاجيء-سوري","IDPs النازحين","Non-Syrian Refugee لاجئ غير سوري"],
        "Is the beneficiary detained هل المستفيد موقوف":["Yes","Yes","Yes","No"],
        "Is it an immigration related charge? هل هو معتقل على اساس قانون الاقامة ؟":["","","","Yes"],
        "# Total Services":[1,1,1,1],
    })
    store=LegalStore.from_files(payload,"test")
    january=store.review("assessments",comparison_month="2026-01",page_size=100)
    detention=[row for row in january["rows"] if row["rule"]=="Detention/immigration inconsistency"]
    assert {row["assessmentId"] for row in detention}=={"A2","A4"}
    repeated=[row for row in january["rows"] if row["rule"]=="Selected month with previous assessment"]
    assert repeated==[]
    assert january["activeComparisonMonth"]=="2026-01"
    assert "Marital status below 18" in store.review("beneficiaries")["ruleCounts"]
    assert "Spouse below 18" in store.review("beneficiaries")["ruleCounts"]


def test_multiple_assessments_flags_same_month_or_two_open_assessments_only():
    payload=required_payload();payload["assessments"]=csv(**{
        "Assessment ID":["M1","M2","O1","O2","H1","H2"],
        "Beneficiary ID":["SameMonth","SameMonth","TwoOpen","TwoOpen","History","History"],
        "Date of Assessment":["10/08/2026","20/08/2026","10/07/2026","10/08/2026","10/06/2026","10/07/2026"],
        "Created On":["10/08/2026","20/08/2026","10/07/2026","10/08/2026","10/06/2026","10/07/2026"],
        "Assessment Status":["Closed","Open","Open","Open","Closed","Closed"],
    })
    rows=LegalStore.from_files(payload,"test").review("assessments",rule="Beneficiary has multiple assessments",page_size=100)["rows"]
    assert {row["assessmentId"] for row in rows}=={"M1","M2","O1","O2"}
    assert "2 assessments in 2026-08" in next(row["detail"] for row in rows if row["assessmentId"]=="M1")
    assert "2 Open assessments" in next(row["detail"] for row in rows if row["assessmentId"]=="O1")


def test_selected_month_previous_assessment_uses_created_on_grace_from_august_2026():
    payload=required_payload();payload["assessments"]=csv(**{
        "Assessment ID":["A1","A2","B1","B2","C1","C2","C3","SameMonth1","SameMonth2"],
        "Beneficiary ID":["Allowed","Allowed","Late","Late","OlderHistory","OlderHistory","OlderHistory","SameMonth","SameMonth"],
        "Date of Assessment":["15/06/2026","15/07/2026","15/06/2026","15/07/2026","15/05/2026","15/06/2026","15/07/2026","15/06/2026","15/07/2026"],
        "Created On":["15/06/2026","04/08/2026","15/06/2026","05/08/2026","15/05/2026","15/06/2026","04/08/2026","15/06/2026","20/07/2026"],
    })
    rows=LegalStore.from_files(payload,"test").review("assessments",comparison_month="2026-07",rule="Selected month with previous assessment",page_size=100)["rows"]
    assert {row["assessmentId"] for row in rows}=={"B2","C3"}
    assert {str(row["createdOn"])[:10] for row in rows}=={"04/08/2026","05/08/2026"}


def test_detention_governorate_project_mapping_overrides_conflicting_location():
    payload=required_payload();payload["assessments"]=csv(**{
        "Assessment ID":["SuliCorrect","ErbilCorrect","SuliMismatch"],
        "Beneficiary ID":["B1","B2","B3"],
        "Projects - المشروع":["UNHCR   2026 - SULI","UNHCR   2026 - Erbil","UNHCR 2026 - SULI"],
        "Project Location":["Baghdad","Baghdad","Baghdad"],
        "Detention Governorate":["Sulaymaniyah - السليمانية","Erbil اربيل","Baghdad"],
        "Is the beneficiary detained":["Yes","Yes","Yes"],
    })
    rows=LegalStore.from_files(payload,"test").review("assessments",rule="Detention Governorate mismatch",page_size=100)["rows"]
    assert {row["assessmentId"] for row in rows}=={"SuliMismatch"}


def test_assessment_review_export_groups_findings_by_region():
    payload=required_payload();payload["assessments"]=csv(**{
        "Assessment ID":["ERB","AMAL","GOV"],
        "Beneficiary ID":["B1","B2","B3"],
        "Projects - المشروع":["UNHCR 2026 - Erbil","UNHCR 2026 - AMAL CAMP","UNHCR 2026 - Gov"],
        "Assessment Status":["Pending","Pending","Pending"],
    })
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("assessments")),read_only=True,data_only=True)
    assert workbook.sheetnames==["North Iraq","AMAL Camp","South Iraq"]
    for sheet_name,assessment_id in (("North Iraq","ERB"),("AMAL Camp","AMAL"),("South Iraq","GOV")):
        rows=list(workbook[sheet_name].iter_rows(values_only=True));assessment_index=list(rows[0]).index("Assessment ID")
        assert assessment_id in {row[assessment_index] for row in rows[1:]}


def test_legal_services_review_export_groups_findings_by_region():
    payload=required_payload();payload["legalservices"]=csv(**{
        "Service ID":["ERB","AMAL","GOV"],
        "Assessment ID":["A1","A2","A3"],
        "Beneficiary ID":["B1","B2","B3"],
        "Project":["UNHCR 2026 - Erbil","UNHCR 2026 - AMAL CAMP","UNHCR 2026 - Gov"],
        "Type of Document نوع الوثيقة":["","",""]
    })
    workbook=load_workbook(io.BytesIO(LegalStore.from_files(payload,"test").review_export("legalservices")),read_only=True,data_only=True)
    assert workbook.sheetnames==["North Iraq","AMAL Camp","South Iraq"]
    for sheet_name,service_id in (("North Iraq","ERB"),("AMAL Camp","AMAL"),("South Iraq","GOV")):
        rows=list(workbook[sheet_name].iter_rows(values_only=True));service_index=list(rows[0]).index("Service ID")
        assert service_id in {row[service_index] for row in rows[1:]}


def test_legal_services_review_export_applies_filters_and_ignores_court_verdict_other():
    payload=required_payload();payload["legalservices"]=csv(**{
        "Service ID":["Court1","Court2","Card1","Card2"],
        "Assessment ID":["A1","A1","A2","A2"],
        "Beneficiary ID":["B1","B1","B2","B2"],
        "Project":["UNHCR 2026 - Erbil","UNHCR 2026 - Erbil","UNHCR 2026 - SULI","UNHCR 2026 - SULI"],
        "Type of Service Provided":["Legal Representation"]*4,
        "Type of Document":["Court Verdict","Court Verdict","ID Card","ID Card"],
    })
    store=LegalStore.from_files(payload,"test")
    exported=load_workbook(io.BytesIO(store.review_export("legalservices",selected_rules=["Duplicate service"],project="UNHCR 2026 - Erbil")),read_only=True,data_only=True)
    rows=list(exported["North Iraq"].iter_rows(values_only=True));service_index=list(rows[0]).index("Service ID")
    assert {row[service_index] for row in rows[1:]}=={"Court1","Court2"}
    ignored=load_workbook(io.BytesIO(store.review_export("legalservices",selected_rules=["Duplicate service"],project="UNHCR 2026 - Erbil",ignore_court_verdict=True)),read_only=True,data_only=True)
    assert ignored["North Iraq"].max_row==1


def test_duplicate_service_without_assessment_id_compares_across_assessments():
    payload=required_payload();payload["legalservices"]=csv(**{
        "Service ID":["S1","S2","S3"],
        "Assessment ID":["A1","A2","A1"],
        "Beneficiary ID":["B1","B1","B1"],
        "Type of Service Provided":["Legal Assistance","Legal Assistance","Legal Representation"],
        "Type of Document":["National ID","National ID","National ID"],
    })
    store=LegalStore.from_files(payload,"test")
    cross_assessment=store.review("legalservices",rule="Duplicate service without Assessment ID",page_size=100)["rows"]
    within_assessment=store.review("legalservices",rule="Duplicate service",page_size=100)["rows"]
    assert {row["serviceId"] for row in cross_assessment}=={"S1","S2"}
    assert not within_assessment


def test_duplicate_service_without_assessment_id_excludes_existing_duplicate_service():
    payload=required_payload();payload["legalservices"]=csv(**{
        "Service ID":["S1","S2"],
        "Assessment ID":["A1","A1"],
        "Beneficiary ID":["B1","B1"],
        "Type of Service Provided":["Legal Assistance","Legal Assistance"],
        "Type of Document":["National ID","National ID"],
    })
    store=LegalStore.from_files(payload,"test")
    assert {row["serviceId"] for row in store.review("legalservices",rule="Duplicate service",page_size=100)["rows"]}=={"S1","S2"}
    assert not store.review("legalservices",rule="Duplicate service without Assessment ID",page_size=100)["rows"]


def test_contact_name_duplicates_exclude_existing_name_duplicates():
    payload=required_payload();payload["beneficiaries"]=csv(**{
        "Case ID":["B1","B2"],
        "Name (Filter Color Red)":["Ahmed Ali Hassan Ali","Ahmed Ali Hassan Ali"],
        "Project":["UNHCR 2026 - Erbil","UNHCR 2026 - Erbil"],
        "Contact Number":["07701234567","07701234567"],
        "ID Number":["ID1","ID2"],
    })
    store=LegalStore.from_files(payload,"test")
    assert {row["caseId"] for row in store.review("beneficiaries",rule="Possible duplicate name",page_size=100)["rows"]}=={"B1","B2"}
    assert not store.review("beneficiaries",rule="Possible duplicate contact and name",page_size=100)["rows"]


def test_current_previous_month_service_duplicate_uses_created_on_grace_from_august_2026():
    payload=required_payload();payload["legalservices"]=csv(**{
        "Service ID":["S1","S2","S3","S4","S5","S6","S7","SameMonth1","SameMonth2"],
        "Beneficiary ID":["Allowed","Allowed","Late","Late","OlderHistory","OlderHistory","OlderHistory","SameMonth","SameMonth"],
        "Assessment ID":["A1","A2","A3","A4","A5","A6","A7","A8","A9"],
        "Date of Service Provision":["15/06/2026","15/07/2026","15/06/2026","15/07/2026","15/05/2026","15/06/2026","15/07/2026","15/06/2026","15/07/2026"],
        "Created On":["15/06/2026","04/08/2026","15/06/2026","05/08/2026","15/05/2026","15/06/2026","04/08/2026","15/06/2026","20/07/2026"],
    })
    rows=LegalStore.from_files(payload,"test").review("legalservices",comparison_month="2026-07",rule="Current and previous month duplicate",page_size=100)["rows"]
    assert {row["serviceId"] for row in rows}=={"S4","S7"}
    assert {str(row["createdOn"])[:10] for row in rows}=={"04/08/2026","05/08/2026"}


def test_representation_rules_use_created_on_and_require_requested_counselling():
    payload=required_payload();payload["assessments"]=csv(**{
        "Assessment ID":["OldAdult","MissingCounselling","RepresentationOnly","NoLinkedService","OldNotDetained","NewNotDetained"],
        "Beneficiary ID":["B1","B2","B3","B4","B5","B6"],
        "Age":[30,30,30,30,30,30],
        "Created On":["31/12/2025","05/01/2026","05/01/2026","05/01/2026","31/12/2025","05/01/2026"],
        "Date of Assessment":["01/01/2026"]*6,
        "Type of Legal Service Needed":["Legal Representation, Legal Counselling","Legal Representation, Legal Counselling","Legal Representation","Legal Representation, Legal Counselling","",""],
        "Community Type":["IDP","IDP","IDP","IDP","Syrian Refugee","Syrian Refugee"],
        "Is the beneficiary detained":["Yes","Yes","Yes","Yes","No","No"],
    })
    payload["legalservices"]=csv(**{
        "Service ID":["S1","S2","S3","S4"],
        "Assessment ID":["OldAdult","MissingCounselling","RepresentationOnly","NewNotDetained"],
        "Beneficiary ID":["B1","B2","B3","B6"],
        "Type of Service Provided":["Legal Representation","Legal Representation","Legal Representation","Legal Assistance"],
    })
    store=LegalStore.from_files(payload,"test")
    adult=store.review("assessments",rule="Adult representation without counselling",page_size=100)["rows"]
    assert {row["assessmentId"] for row in adult}=={"MissingCounselling","NoLinkedService"}
    not_detained=store.review("assessments",rule="Representation while not detained",page_size=100)["rows"]
    assert {row["assessmentId"] for row in not_detained}=={"NewNotDetained"}


def test_assessment_document_and_service_type_reconciliation():
    payload=required_payload()
    payload["assessments"]=csv(**{
        "Assessment ID":["A1","A2","A3","A4"],"Beneficiary ID":["B1","B2","B3","B4"],
        "Date of Assessment":["01/01/2026","01/01/2026","01/01/2026","01/01/2026"],
        "Type of Documents to be issued":["Birth Certificate بيان ولادة, Marriage Certificate عقد زواج","","","Passport جواز السفر"],
        "Type of Legal Service Needed":["Legal Counselling - استشارة, Legal Representation - تمثيل","","","Legal Assistance - مساعدة"],
    })
    payload["legalservices"]=csv(**{
        "Service ID":["S1","S2","S3"],"Assessment ID":["A1","A1","A2"],"Beneficiary ID":["B1","B1","B2"],
        "Type of Document":["Birth Certificate","Divorce Certificate شهادة الطلاق","Proof of Marriage اثبات الزواج"],
        "Type of Service Provided":["Legal Counselling - استشارة","Legal Representation - تمثيل","Legal Counselling - استشارة"],
    })
    store=LegalStore.from_files(payload,"test")
    documents=store.review("assessments",rule="Type of document in Assessments vs Services",page_size=100)["rows"]
    assert {(row["assessmentId"],row["comparisonFinding"],row["missingValues"]) for row in documents}=={
        ("A1","Missing Type of Document in Services","Marriage Certificate عقد زواج"),
        ("A2","Missing Type of Document in Assessment","Proof of Marriage اثبات الزواج"),
        ("A4","Missing Type of Document in Services","Passport جواز السفر"),
    }
    service_types=store.review("assessments",rule="Type of Legal Service in Assessment vs Services",page_size=100)["rows"]
    assert [(row["assessmentId"],row["missingValues"]) for row in service_types]==[("A4","Legal Representation - تمثيل")]


def test_assessment_reconciliation_ignores_dates_before_2026():
    payload=required_payload()
    payload["assessments"]=csv(**{"Assessment ID":["Old"],"Beneficiary ID":["B1"],"Date of Assessment":["31/12/2025"],"Type of Documents to be issued":["Passport"],"Type of Legal Service Needed":["Legal Counselling"]})
    payload["legalservices"]=csv(**{"Service ID":["S1"],"Assessment ID":["Old"],"Beneficiary ID":["B1"],"Type of Document":["Birth Certificate"],"Type of Service Provided":["Legal Representation"]})
    review=LegalStore.from_files(payload,"test").review("assessments",page_size=100)
    assert review["ruleCounts"]["Type of document in Assessments vs Services"]==0
    assert review["ruleCounts"]["Type of Legal Service in Assessment vs Services"]==0


def test_versioned_csv_names_and_detention_page():
    assert versioned_dataset_name("assessments.csv") == ("assessments", 0)
    assert versioned_dataset_name("Assessments (2).csv") == ("assessments", 2)
    assert versioned_dataset_name("legal services (12).CSV") == ("legalservices", 12)
    assert versioned_dataset_name("notes (2).csv") is None
    payload=required_payload()
    payload["assessments"]=csv(**{
        "Assessment ID":["A1","A2","A3"],"Beneficiary ID":["B1","B2","B3"],
        "Is the beneficiary detained":["Yes","No","Yes"],"Date of Assessment":["05/01/2026","01/02/2026","06/01/2026"],"Date of Detention":["02/01/2026","","03/01/2026"] ,
        "Detention Governorate":["Ninewa","","Baghdad"] ,"Detainee current status":["Detained","Released","Released after review"],
        "Date of the released or deported":["10/02/2026","","12/02/2026"] ,
    })
    store=LegalStore.from_files(payload,"test")
    result=store.detention_cases()
    assert result["total"]==2
    assert result["rows"][0]["caseId"]=="B1"
    assert result["filterOptions"]["Detention governorate"]==["Baghdad","Ninewa"]
    assert result["map"]=={"items":[{"label":"Ninawa","count":1,"detained":1,"released":0,"values":["Ninewa"]},{"label":"Baghdad","count":1,"detained":1,"released":1,"values":["Baghdad"]}]}
    assert result["trend"]==[{"month":"2026-01","detainedAssessments":2,"released":0},{"month":"2026-02","detainedAssessments":0,"released":1}]
    february=store.detention_cases(filters={"month":["2026-02"]})
    assert february["total"]==0
    assert february["trend"]==[{"month":"2026-02","detainedAssessments":0,"released":1}]
    assert february["map"]=={"items":[{"label":"Baghdad","count":0,"detained":0,"released":1,"values":["Baghdad"]}]}
    empty_month=store.detention_cases(filters={"month":["2026-03"]})
    assert empty_month["trend"]==[]


def test_detention_detail_columns_and_monthly_excel_reconciliation():
    payload=required_payload()
    shared={
        "DoB / تأريخ الولاده":"01/01/1990","Age":36,"Gender النوع الاجتماعي":"Male ذكر",
        "Date of Detention تاريخ الاحتجاز":"02/01/2026","Detention Governorate / محافظة الاحتجاز":"Ninewa",
        "Detaining Authority جهة الاحتجاز":"Police","Reasons for Detention أسباب الاحتجاز":"Immigration",
        "Possible Charges التهم المحتملة":"Overstay","Nationality الجنسية":"Iraqi",
        "Name of the reporting person اسم الشخص المبلغ":"Reporter","Relationship to the detainee العلاقة بالمحتجز":"Brother",
        "Phone number of the reporter رقم هاتف المبلغ":"07501234567","Type of Legal Service Needed":"Legal Representation - Court follow-up",
        "Detainee current status حالة المعتقل الحالية":"Detained","Type of Released نوع الافراج":"",
        "Date of the released or deported تاريخ الافراج او الترحيل":"",
    }
    payload["assessments"]=csv(**{
        "Assessment ID":["A1","A3"],"Beneficiary ID":["B1","B3"],"Lawyer":["Assessment Lawyer",""],"Projects":["P1","P2"],"Is the beneficiary detained":["Yes","Yes"],
        "Date of Assessment تاريخ التقييم":["05/01/2026","06/01/2026"],"Name / الأسم":["أحمد علي","سارة حسن"],
        **{column:[value,value] for column,value in shared.items()},
    })
    store=LegalStore.from_files(payload,"test")
    detail=store.detention_cases()
    assert "Name of the reporting person اسم الشخص المبلغ" in detail["columns"]
    assert "Phone number of the reporter رقم هاتف المبلغ" in detail["columns"]
    external=pd.DataFrame({
        "Lawyer":["Excel Lawyer","L2","L3"],"Identification Date | تاريخ تحديد الحالة":["05/01/2026","06/01/2026","07/01/2026"],
        "Beneficiary ID (Platform Case ID)":["B1","B3","B9"],"Registration Number | رقم التسجيل":["R1","R3","R9"],
        "English Name | الاسم بالإنكليزي":["Ahmed Ali","Sara Hassan","Other"],"Arabic Name | الاسم بالعربي":["أحمد علي","سارة حسن","شخص آخر"],
        "Date of Birth | تاريخ الميلاد":["01/01/1990"]*3,"Age | العمر":[99]*3,"Sex | الجنس":["Male"]*3,
        "Date of Arrest | تاريخ الاعتقال":["02/01/2026"]*3,"Detention Governorate | محافظة الاحتجاز":["Ninewa"]*3,
        "Detaining Authority | الجهة المحتجزة":["Police"]*3,"Place of Detention | مكان الاحتجاز":[""]*3,
        "Reason of Arrest | سبب الاعتقال":["Immigration"]*3,"Charges | التهم":["Different charge","Overstay","Overstay"],
        "Nationality | الجنسية":["Iraqi"]*3,"Case Reporter | مُبلغ القضية":["Different reporter"]*3,
        "Relationship with Detainee | العلاقة مع المعتقل":["Different relationship"]*3,"Phone Number | رقم الهاتف":["Different phone"]*3,
        "Type of Service by INTERSOS | نوع الخدمة المقدمة من قبل محامي انترسوس":["Legal Representation"]*3,
        "Detainee Current Status | الحالة الحالية للمعتقل":["Detained"]*3,"Type of Release | نوع الإفراج":[""]*3,
        "Date of Release/Deportation | تاريخ الإفراج أو الترحيل":[""]*3,"ملاحظات | Note":["Check charge","",""]
    })
    workbook=io.BytesIO();external.to_excel(workbook,index=False)
    result=store.detention_reconciliation(workbook.getvalue(),"detention.xlsx","2026-01")
    assert result["platformRecords"]==2
    assert result["comparisonRecords"]==3
    assert result["matched"]==1
    assert result["unmatched"]==2
    assert "Date of birth" in result["comparedFields"]
    assert "Detention governorate" in result["comparedFields"]
    assert "Age" not in result["comparedFields"]
    assert result["rows"][0]["beneficiaryId"]=="B1"
    assert result["rows"][0]["differences"]==[{"field":"Possible charges","assessment":"Overstay","excel":"Different charge"}]
    assert result["rows"][1]["beneficiaryId"]=="B9"
    exported=store.detention_reconciliation_export(workbook.getvalue(),"detention.xlsx","2026-01")
    issue_sheet=load_workbook(io.BytesIO(exported))["Comparison issues"]
    assert [cell.value for cell in issue_sheet[4]]==["Lawyer","Note group","Case ID","Name","Different field","Assessment value","Excel value"]
    assert issue_sheet["A5"].value=="Assessment Lawyer"
    assert issue_sheet["A5"].fill.fgColor.rgb.endswith("E8F1FB")
    multiple=store.detention_reconciliation(workbook.getvalue(),"detention.xlsx","2026-01,2026-02")
    assert multiple["months"]==["2026-01","2026-02"]
    assert multiple["platformRecords"]==2
    project_result=store.detention_reconciliation(workbook.getvalue(),"detention.xlsx","2026-01","P1")
    assert project_result["project"]=="P1"
    assert project_result["platformRecords"]==1
    assert project_result["comparisonRecords"]==3
    assert not any("no Project column" in warning for warning in project_result["warnings"])
    excel_only=next(row for row in project_result["rows"] if row["beneficiaryId"]=="B9")
    assert excel_only["note"]=="Case ID available in Excel but missing from Assessments"
    assert excel_only["caseAvailable"] is False
    projects_result=store.detention_reconciliation(workbook.getvalue(),"detention.xlsx","2026-01",["P1","P2"])
    assert projects_result["projects"]==["P1","P2"]
    assert projects_result["platformRecords"]==2
    multi_sheet=io.BytesIO()
    with pd.ExcelWriter(multi_sheet,engine="openpyxl") as writer:
        external.to_excel(writer,index=False,sheet_name="January cases")
        external.to_excel(writer,index=False,sheet_name="Reviewed cases")
    assert LegalStore.detention_workbook_sheets(multi_sheet.getvalue())==["January cases","Reviewed cases"]
    selected_sheet=store.detention_reconciliation(multi_sheet.getvalue(),"detention.xlsx","2026-01","","Reviewed cases")
    assert selected_sheet["sheet"]=="Reviewed cases"


def test_project_reconciliation_compares_dob_and_normalized_detention_governorate():
    payload=required_payload()
    payload["assessments"]=csv(**{
        "Assessment ID":["A1","A2"],"Beneficiary ID":["B1","B2"],"Projects":["P1","P2"],
        "Is the beneficiary detained":["Yes","Yes"],"Date of Assessment":["05/01/2026","05/01/2026"],
        "Name":["Person One","Person Two"],"DoB":["01/01/1990","01/01/1991"],
        "Detention Governorate":["Ninewa","Baghdad"],
    })
    store=LegalStore.from_files(payload,"test")
    external=pd.DataFrame({
        "Identification Date":["05/01/2026","05/01/2026"],"Beneficiary ID (Platform Case ID)":["B1","B2"],
        "Arabic Name":["Person One","Person Two"],"Date of Birth":["02/01/1990","01/01/1991"],
        "Detention Governorate":["Baghdad بغداد","Baghdad"],
    })
    workbook=io.BytesIO();external.to_excel(workbook,index=False)
    result=store.detention_reconciliation(workbook.getvalue(),"detention.xlsx","2026-01","P1")
    assert result["platformRecords"]==1
    assert result["comparisonRecords"]==2
    assert result["rows"][0]["beneficiaryId"]=="B1"
    assert [difference["field"] for difference in result["rows"][0]["differences"]]==["Date of birth","Detention governorate"]

    external.loc[0,"Date of Birth"]=32874
    external.loc[0,"Detention Governorate"]="Ninawa نينوى"
    matched_workbook=io.BytesIO();external.to_excel(matched_workbook,index=False)
    matched=store.detention_reconciliation(matched_workbook.getvalue(),"detention.xlsx","2026-01","P1")
    assert matched["matched"]==1
    assert matched["rows"]==[{"beneficiaryId":"B2","caseAvailable":True,"name":"Person Two","lawyer":"","note":"Case ID available in Excel but missing from Assessments","differences":[{"field":"Case ID","assessment":"Missing","excel":"Present"}]}]


def test_detention_reconciliation_reports_blank_case_ids_on_both_sides():
    payload=required_payload()
    payload["assessments"]=csv(**{
        "Assessment ID":["A1","A2"],"Beneficiary ID":["B1",""],"Projects":["P1","P1"],
        "Is the beneficiary detained":["Yes","Yes"],"Date of Assessment":["05/01/2026","05/01/2026"],
        "Name":["Matched person","Assessment without ID"],
    })
    store=LegalStore.from_files(payload,"test")
    workbook=io.BytesIO()
    pd.DataFrame({
        "Identification Date":["05/01/2026","05/01/2026"],
        "Beneficiary ID (Platform Case ID)":["B1",""],
        "Arabic Name":["Matched person","Excel without ID"],
    }).to_excel(workbook,index=False)
    result=store.detention_reconciliation(workbook.getvalue(),"detention.xlsx","2026-01","P1")
    assert result["missingCaseIds"]=={"assessments":1,"excel":1}
    assert {row["note"] for row in result["rows"]}=={"Case ID missing in Assessments","Case ID missing in Excel"}


def test_service_missing_document_and_generic_review_exclusions(tmp_path):
    payload=required_payload()
    payload["legalservices"]=csv(**{"Service ID":["S1","S2"],"Assessment ID":["A1","A2"],"Beneficiary ID":["B1","B2"],"Type of Document نوع الوثيقة":["","National ID"]})
    store=LegalStore.from_files(payload,"test")
    assert {row["serviceId"] for row in store.review("legalservices",rule="Missing Type of Document")["rows"]}=={"S1"}
    registry=DuplicateExclusionRegistry(tmp_path/"exclusions.json")
    _, created=registry.exclude_record("legalservices","Missing Type of Document","serviceId","S1")
    assert created is True
    store.set_review_exclusions(registry.exclusion_rows())
    assert store.review("legalservices",rule="Missing Type of Document")["total"]==0


def test_awareness_name_exclusion_normalizes_spaces_and_case(tmp_path):
    payload=required_payload();payload["awareness"]=csv(**{"Awareness ID":["W1"],"Participant Name":["  Person   Name "],"Phone Number":["12345"]})
    store=LegalStore.from_files(payload,"test")
    registry=DuplicateExclusionRegistry(tmp_path/"exclusions.json")
    registry.exclude_record("awareness","Invalid contact number","awarenessName","person name")
    store.set_review_exclusions(registry.exclusion_rows())
    assert store.review("awareness",rule="Invalid contact number")["total"]==0


def test_analytics_dashboard_filters_sorting_pagination_and_chart_counts():
    payload=required_payload()
    payload["assessments"]=csv(**{
        "Assessment ID":["A1","A2","A3"],"Beneficiary ID":["B1","B2","B3"],
        "Date of Assessment":["05/01/2026","06/01/2026","05/02/2026"],
        "Projects":["P1","P1","P2"],"Project Location":["L1","L2","L1"],
        "Assessment Status":["Open","Closed","Open"],
    })
    store=LegalStore.from_files(payload,"test")
    result=store.analytics_dashboard("assessments",filters={"Projects":["P1"]},page=1,page_size=1,sort_column="Assessment ID",sort_direction="desc")
    assert result["total"]==2
    assert result["matchedRows"]==2
    assert result["rows"][0]["Assessment ID"]=="A2"
    assert next(chart for chart in result["charts"] if chart["title"]=="Project")["rows"]==[{"label":"P1","count":2,"percent":1.0}]
    assert result["trend"]==[{"label":"2026-01","count":2,"percent":1.0}]


def test_analytics_dashboard_includes_detained_immigration_and_uncompleted_service_kpis():
    payload=required_payload()
    payload["assessments"]=csv(**{"Assessment ID":["A1","A2"],"Beneficiary ID":["B1","B2"],"Is the beneficiary detained":["Yes","Yes"],"Is it an immigration related charge":["Yes","No"]})
    payload["legalservices"]=csv(**{"Service ID":["S1","S2"],"Assessment ID":["A1","A2"],"Beneficiary ID":["B1","B2"],"Service Status":["Completed","Uncompleted"]})
    store=LegalStore.from_files(payload,"test")
    assessment_kpis={item["label"]:item["value"] for item in store.analytics_dashboard("assessments")["kpis"]}
    service_kpis={item["label"]:item["value"] for item in store.analytics_dashboard("legalservices")["kpis"]}
    assert assessment_kpis["Detention cases with immigration charges"]==1
    assert service_kpis["Uncompleted services"]==1
    assert "Projects" not in assessment_kpis and "Locations" not in assessment_kpis


def test_services_analytics_joins_assessment_need_and_handles_missing_optional_columns():
    payload=required_payload()
    payload["assessments"]=csv(**{"Assessment ID":["A1","A2"],"Beneficiary ID":["B1","B2"],"Type of Legal Service Needed":["Documentation","Counselling"]})
    payload["legalservices"]=csv(**{"Service ID":["S1","S2"],"Assessment ID":["A1","A2"],"Beneficiary ID":["B1","B2"],"Service Status":["Completed","In process"]})
    store=LegalStore.from_files(payload,"test")
    result=store.analytics_dashboard("legalservices",filters={"_assessment_need":["Documentation"]})
    assert result["total"]==1
    assert result["rows"][0]["Service ID"]=="S1"
    assert any(chart["title"]=="Assessment legal-service need" for chart in result["charts"])
    assert any("source column not available" in warning for warning in result["warnings"])


def test_analytics_export_applies_derived_month_filter():
    payload=required_payload()
    payload["beneficiaries"]=csv(**{"Case ID":["B1","B2"],"Name (Filter Color Red)":["Person 1","Person 2"],"Date of Identification":["05/01/2026","05/02/2026"],"Projects":["P1","P2"]})
    store=LegalStore.from_files(payload,"test")
    exported=store.explorer_export("beneficiaries",filters={"Month":["2026-02"]})
    rows=pd.read_excel(io.BytesIO(exported))
    assert rows["Case ID"].tolist()==["B2"]
