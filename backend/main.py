from __future__ import annotations

import json
import io
import ipaddress
import os
import re
import secrets
import socket
import sys
from pathlib import Path
from urllib.parse import unquote, urlparse
from urllib.request import HTTPRedirectHandler, Request, build_opener

from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from fastapi.staticfiles import StaticFiles
from starlette.concurrency import run_in_threadpool
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .legal_platform import LegalStore, FILES, versioned_dataset_name
from .file_security import safe_spreadsheet_value, validate_xlsx_archive
from .duplicate_exclusions import DuplicateExclusionRegistry
from .indicator_reporting import build_indicator_report, build_indicator_workbook, build_narrative_workbook
from . import updater


ROOT = Path(__file__).resolve().parents[1]
BUNDLE_ROOT = Path(getattr(sys, "_MEIPASS", ROOT))
STATIC_DIR = Path(os.getenv("UNHCR_STATIC_DIR", BUNDLE_ROOT / "frontend" / "dist"))
UPLOAD_ONLY = os.getenv("UNHCR_UPLOAD_ONLY", "").lower() in {"1", "true", "yes"}
MAX_UPLOAD_BYTES = 100 * 1024 * 1024
LOCAL_SESSION_TOKEN = os.getenv("INTERSOS_LOCAL_SESSION_TOKEN", "")
SESSION_COOKIE = "intersos_session"
LEGAL_SAMPLE = ROOT / "Legal Platform Data"
REMEMBERED_LEGAL_FOLDER = Path(os.getenv("INTERSOS_LEGAL_FOLDER", "")) if os.getenv("INTERSOS_LEGAL_FOLDER") else None
REMEMBERED_LEGAL_SOURCE = os.getenv("INTERSOS_LEGAL_SOURCE", "folder")
REMEMBERED_LEGAL_SOURCE_CONFIGURED = os.getenv("INTERSOS_LEGAL_SOURCE_CONFIGURED", "").lower() in {"1", "true", "yes"}
try:
    REMEMBERED_LEGAL_FILES = [Path(value) for value in json.loads(os.getenv("INTERSOS_LEGAL_FILES", "[]"))]
except (json.JSONDecodeError, TypeError):
    REMEMBERED_LEGAL_FILES = []


def validate_attachment_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("Only HTTP(S) attachment links can be downloaded.")
    if parsed.username or parsed.password or parsed.fragment:
        raise ValueError("Attachment links cannot contain credentials or fragments.")
    try:
        port = parsed.port or (443 if parsed.scheme == "https" else 80)
    except ValueError as exc:
        raise ValueError("Attachment link contains an invalid port.") from exc
    try:
        addresses = socket.getaddrinfo(parsed.hostname, port, type=socket.SOCK_STREAM)
    except OSError as exc:
        raise ValueError("Attachment host could not be resolved.") from exc
    if not addresses:
        raise ValueError("Attachment host could not be resolved.")
    for address in addresses:
        host = str(address[4][0]).split("%", 1)[0]
        try:
            resolved = ipaddress.ip_address(host)
        except ValueError as exc:
            raise ValueError("Attachment host resolved to an invalid address.") from exc
        if not resolved.is_global:
            raise ValueError("Attachment links cannot target local or private network addresses.")
    return url


class ValidatedAttachmentRedirectHandler(HTTPRedirectHandler):
    def redirect_request(self, request, file_pointer, code, message, headers, new_url):
        validate_attachment_url(new_url)
        return super().redirect_request(request, file_pointer, code, message, headers, new_url)


ATTACHMENT_OPENER = build_opener(ValidatedAttachmentRedirectHandler())


async def read_upload_limited(file: UploadFile, limit: int = MAX_UPLOAD_BYTES) -> bytes:
    if file.size is not None and file.size > limit:
        raise HTTPException(413, "Workbook must be 100 MB or smaller.")
    chunks: list[bytes] = []
    total = 0
    while chunk := await file.read(1024 * 1024):
        total += len(chunk)
        if total > limit:
            raise HTTPException(413, "Workbook must be 100 MB or smaller.")
        chunks.append(chunk)
    return b"".join(chunks)


def remembered_legal_file_payload(paths: list[Path]) -> dict[str, bytes]:
    selected: dict[str, tuple[int, Path]] = {}
    for path in paths:
        parsed = versioned_dataset_name(path.name)
        if not parsed: continue
        name, version = parsed
        current = selected.get(name)
        if current is None or version > current[0] or (version == current[0] and path.stat().st_mtime > current[1].stat().st_mtime): selected[name] = (version, path)
    return {name: path.read_bytes() for name, (_, path) in selected.items()}


duplicate_exclusions = DuplicateExclusionRegistry()
legal_store: LegalStore | None = None
legal_store_loading = False
legal_store_restore_error = ""

def load_initial_legal_store() -> None:
    """Restore the remembered Legal data outside the desktop window startup path."""
    global legal_store, legal_store_loading, legal_store_restore_error
    legal_store_loading = True
    try:
        use_files=REMEMBERED_LEGAL_SOURCE=="files"
        if REMEMBERED_LEGAL_SOURCE_CONFIGURED:
            if use_files:
                candidate = LegalStore.from_files(remembered_legal_file_payload(REMEMBERED_LEGAL_FILES), "Selected Legal Platform CSV files") if REMEMBERED_LEGAL_FILES and all(path.is_file() for path in REMEMBERED_LEGAL_FILES) else None
                unavailable = "The last selected Legal Platform CSV files are unavailable. Choose a new source."
            else:
                candidate = LegalStore.from_folder(REMEMBERED_LEGAL_FOLDER) if REMEMBERED_LEGAL_FOLDER and REMEMBERED_LEGAL_FOLDER.is_dir() else None
                unavailable = "The last selected Legal Platform folder is unavailable. Choose a new source."
        else:
            candidate = LegalStore.from_folder(LEGAL_SAMPLE) if LEGAL_SAMPLE.exists() else None
            unavailable = ""
        if candidate is None and unavailable:
            legal_store = None
            legal_store_restore_error = unavailable
            return
        if candidate:
            candidate.set_review_exclusions(duplicate_exclusions.exclusion_rows())
        legal_store = candidate
        legal_store_restore_error = ""
    except Exception as exc:
        legal_store = None
        legal_store_restore_error = f"Unable to restore the last Legal Platform source: {exc}"
    finally:
        legal_store_loading = False

if os.getenv("INTERSOS_DEFER_LEGAL_LOAD", "").lower() in {"1", "true", "yes"}:
    legal_store_loading = True
else:
    load_initial_legal_store()

app = FastAPI(title="Iraq Data Analysis API", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


@app.middleware("http")
async def prevent_stale_api_state(request, call_next):
    if request.url.path.startswith("/api/") and LOCAL_SESSION_TOKEN:
        supplied_token = request.cookies.get(SESSION_COOKIE, "")
        if not secrets.compare_digest(supplied_token, LOCAL_SESSION_TOKEN):
            return Response(status_code=403, content="Local application session required.")
    response = await call_next(request)
    if request.url.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    elif LOCAL_SESSION_TOKEN:
        response.set_cookie(SESSION_COOKIE, LOCAL_SESSION_TOKEN, httponly=True, samesite="strict", path="/")
    return response


class LegalQuery(BaseModel):
    dataset: str = "beneficiaries"
    search: str = ""
    rule: str = ""
    page: int = 1
    pageSize: int = 100
    nameCompareChars: int = 15
    allowNameVariations: bool = False
    exactMatchesOnly: bool = False
    filterColumn: str = ""
    filterValue: str = ""
    severity: str = ""
    lawyer: str = ""
    project: str = ""
    location: str = ""
    date: str = ""
    filters: dict[str, list[str]] = {}
    comparisonMonth: str = ""
    sortColumn: str = ""
    sortDirection: str = "asc"

class LegalStudioRequest(BaseModel):
    dataset: str
    rowDimension: str
    columnDimension: str = ""
    filters: dict[str, list[str]] = {}
    measure: str = "records"

class LegalAnalyticsDashboardRequest(BaseModel):
    dataset: str
    filters: dict[str, list[str]] = {}
    search: str = ""
    page: int = 1
    pageSize: int = 100
    sortColumn: str = ""
    sortDirection: str = "asc"


class DuplicateExclusionRequest(BaseModel):
    caseId: str = ""
    rule: str
    dataset: str = "beneficiaries"
    identifierType: str = "caseId"
    identifierValue: str = ""
    name: str = ""
    project: str = ""
    source: str = ""


class IndicatorReportRequest(BaseModel):
    fromDate: str = ""
    toDate: str = ""
    projects: list[str] = []
    projectLocations: list[str] = []
    years: list[str] = []
    quarters: list[str] = []
    months: list[str] = []
    communityTypes: list[str] = []


class CaseQuery(BaseModel):
    query: str
    filters: dict[str, list[str]] = {}
    caseIds: list[str] = []
    viewMode: str = "cards"
    page: int = 1
    pageSize: int = 100
    sortColumn: str = ""
    sortDirection: str = "asc"
    columns: list[str] = []


class TableWorkbookSheet(BaseModel):
    title: str
    columns: list[str]
    rows: list[dict[str, object]]


class TableWorkbookRequest(BaseModel):
    filename: str = "table-export.xlsx"
    columns: list[str] = []
    rows: list[dict[str, object]] = []
    sheets: list[TableWorkbookSheet] = []


@app.get("/api/health")
def health(): return {"status": "ready" if legal_store else "awaiting_source", "source": legal_store.source if legal_store else None, "loading": legal_store_loading}


@app.get("/api/update/check")
def update_check(): return updater.check()


@app.get("/api/update/status")
def update_status(): return updater.status()


@app.post("/api/update/install")
def update_install():
    try: return updater.install()
    except ValueError as exc: raise HTTPException(409, str(exc)) from exc


@app.get("/api/legal/metadata")
def legal_metadata():
    return legal_store.metadata() if legal_store else {"ready": False, "loading":legal_store_loading, "source": None, "warnings": [legal_store_restore_error] if legal_store_restore_error else [], "availability": {name: False for name in FILES}, "features":{"detention":False,"deportation":False}, "sheets": [], "months": [], "reviewCounts": {}}

@app.get("/api/legal/deportation-dashboard")
def legal_deportation_dashboard():
    try:return require_legal_store().deportation_dashboard()
    except ValueError as exc: raise HTTPException(400,str(exc)) from exc

@app.post("/api/legal/deportation-dashboard")
def legal_deportation_dashboard_filtered(request:LegalQuery):
    try:return require_legal_store().deportation_dashboard(request.filters)
    except ValueError as exc: raise HTTPException(400,str(exc)) from exc


@app.post("/api/legal/upload")
async def legal_upload(files: list[UploadFile] = File(...)):
    global legal_store
    payload: dict[str, bytes] = {}
    versions: dict[str, int] = {}
    total = 0
    try:
        for file in files:
            filename = Path(file.filename or "").name
            parsed = versioned_dataset_name(filename)
            if not parsed: continue
            key, version = parsed
            raw = await file.read(MAX_UPLOAD_BYTES + 1)
            total += len(raw)
            if total > MAX_UPLOAD_BYTES: raise HTTPException(413, "Legal Platform files must total 100 MB or smaller.")
            if version >= versions.get(key, -1):payload[key]=raw;versions[key]=version
        candidate = await run_in_threadpool(LegalStore.from_files, payload, "Selected Legal Platform folder")
        candidate.set_review_exclusions(duplicate_exclusions.exclusion_rows())
        legal_store = candidate
        return candidate.metadata()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, str(exc)) from exc
    finally:
        for file in files: await file.close()


def require_legal_store() -> LegalStore:
    if legal_store is None: raise HTTPException(409, "Select a Legal Platform data folder to begin.")
    return legal_store


@app.post("/api/legal/review")
def legal_review(request: LegalQuery):
    if request.page < 1 or request.pageSize < 1 or request.pageSize > 5000: raise HTTPException(400, "Invalid pagination")
    return require_legal_store().review(request.dataset, request.search, request.rule, request.page, request.pageSize, request.severity,request.lawyer,request.project,request.location,request.date,request.comparisonMonth,request.nameCompareChars,request.allowNameVariations,request.exactMatchesOnly)


@app.get("/api/legal/review-export/{dataset}")
def legal_review_export(dataset:str,comparison_month:str="",name_compare_chars:int=15,allow_name_variations:bool=False,exact_matches_only:bool=False,rules:str="",severity:str="",lawyer:str="",project:str="",location:str="",date:str="",search:str="",ignore_court_verdict:bool=False):
    selected_rules=[rule.strip() for rule in rules.split(",") if rule.strip()]
    payload=require_legal_store().review_export(dataset,comparison_month,name_compare_chars,allow_name_variations,exact_matches_only,selected_rules,severity,lawyer,project,location,date,search,ignore_court_verdict)
    return Response(payload,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="{dataset}-review-findings.xlsx"'})


def duplicate_exclusion_payload() -> dict[str, object]:
    rows = duplicate_exclusions.entries()
    return {"rows": rows, "count": len(rows)}


@app.get("/api/legal/duplicate-exclusions")
def list_duplicate_exclusions():
    return duplicate_exclusion_payload()


@app.post("/api/legal/duplicate-exclusions")
def create_duplicate_exclusion(request: DuplicateExclusionRequest):
    try:
        value=request.identifierValue or request.caseId
        record, _ = duplicate_exclusions.exclude_record(request.dataset, request.rule, request.identifierType, value, request.name, request.project, request.source)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    if legal_store:
        legal_store.set_review_exclusions(duplicate_exclusions.exclusion_rows())
    return {"record": record, **duplicate_exclusion_payload()}


@app.delete("/api/legal/duplicate-exclusions/{case_id}")
def restore_duplicate_exclusion(case_id: str, rule: str, dataset: str = "", identifier_type: str = ""):
    if not duplicate_exclusions.restore(case_id, rule, dataset, identifier_type):
        raise HTTPException(404, "Excluded Case ID was not found.")
    if legal_store:
        legal_store.set_review_exclusions(duplicate_exclusions.exclusion_rows())
    return duplicate_exclusion_payload()


@app.post("/api/legal/duplicate-exclusions/import")
async def import_duplicate_exclusions(file: UploadFile = File(...), dataset: str = Form(...), identifier_type: str = Form(...), rules: str = Form(...)):
    if dataset not in {"assessments", "legalservices", "awareness", "beneficiaries"}:
        raise HTTPException(400, "Unsupported review page.")
    selected_rules=[item.strip() for item in rules.split(",") if item.strip()]
    if not selected_rules: raise HTTPException(400, "Choose at least one finding table.")
    raw=await file.read(MAX_UPLOAD_BYTES + 1)
    if len(raw)>MAX_UPLOAD_BYTES: raise HTTPException(413, "Exclusion file must be 100 MB or smaller.")
    try:
        import pandas as pd
        source=io.BytesIO(raw)
        suffix=Path(file.filename or "").suffix.lower()
        if suffix==".xlsx": validate_xlsx_archive(raw)
        frame=pd.read_excel(source, dtype=object) if suffix in {".xlsx", ".xls"} else pd.read_csv(source, dtype=object, encoding="utf-8-sig")
        if frame.empty or not len(frame.columns): raise ValueError("The exclusion file has no identifier column.")
        expected_columns={
            "beneficiaries": ("case id", "beneficiary id"),
            "assessments": ("assessment id",),
            "legalservices": ("service id",),
            "awareness": ("awareness id",),
        }
        normalized_columns={re.sub(r"\s+", " ", str(column).strip().casefold()): column for column in frame.columns}
        column=next((normalized_columns[name] for name in expected_columns[dataset] if name in normalized_columns), None)
        if column is None:
            raise ValueError(f"No {expected_columns[dataset][0].title()} column was found in the uploaded file.")
        values=frame[column].dropna().astype(str).map(str.strip).tolist()
        imported=duplicates=invalid=0
        for value in values:
            if not value: invalid+=1; continue
            for rule in selected_rules:
                _, created=duplicate_exclusions.exclude_record(dataset, rule, identifier_type, value, source=f"Imported from {Path(file.filename or 'file').name}")
                imported += int(created); duplicates += int(not created)
        if legal_store: legal_store.set_review_exclusions(duplicate_exclusions.exclusion_rows())
        return {"imported":imported,"duplicates":duplicates,"invalid":invalid,"column":str(column),"rows":duplicate_exclusions.entries(),"count":len(duplicate_exclusions.entries())}
    except ValueError as exc: raise HTTPException(400, str(exc)) from exc
    finally: await file.close()


@app.get("/api/legal/duplicate-exclusions-export")
def export_duplicate_exclusions():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Excluded findings"
    sheet.append(["Review page", "Finding", "Identifier type", "Identifier value", "Name", "Project", "Excluded on", "Source context"])
    for row in duplicate_exclusions.entries():
        sheet.append([safe_spreadsheet_value(value) for value in [row.get("dataset", "beneficiaries"), row.get("rule", ""), row.get("identifierType", "caseId"), row.get("identifierValue", row.get("caseId", "")), row.get("name", ""), row.get("project", ""), row.get("excludedAt", ""), row.get("source", "")]])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2454C6")
    for column, width in zip("ABCDEF", (32, 20, 32, 34, 28, 28)):
        sheet.column_dimensions[column].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    return Response(buffer.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="beneficiary-finding-exclusions.xlsx"'})


@app.post("/api/legal/explorer")
def legal_explorer(request: LegalQuery):
    if request.page < 1 or request.pageSize < 1 or request.pageSize > 500: raise HTTPException(400, "Invalid pagination")
    if request.sortDirection not in {"asc","desc"}: raise HTTPException(400,"Invalid sort direction")
    try: return require_legal_store().explorer(request.dataset, request.search, request.page, request.pageSize, request.filterColumn, request.filterValue,request.filters,request.sortColumn,request.sortDirection)
    except ValueError as exc: raise HTTPException(400, str(exc)) from exc

@app.post("/api/legal/studio")
def legal_studio(request: LegalStudioRequest):
    try: return require_legal_store().studio(request.dataset,request.rowDimension,request.columnDimension,request.filters,request.measure)
    except ValueError as exc: raise HTTPException(400,str(exc)) from exc

@app.post("/api/legal/analytics-dashboard")
def legal_analytics_dashboard(request: LegalAnalyticsDashboardRequest):
    if request.page<1 or request.pageSize<1 or request.pageSize>10000 or request.sortDirection not in {"asc","desc"}: raise HTTPException(400,"Invalid dashboard query.")
    try:return require_legal_store().analytics_dashboard(request.dataset,request.filters,request.search,request.page,request.pageSize,request.sortColumn,request.sortDirection)
    except ValueError as exc:raise HTTPException(400,str(exc)) from exc


@app.get("/api/legal/explorer-filters/{dataset}")
def legal_explorer_filters(dataset:str):
    try:return require_legal_store().explorer_filters(dataset)
    except ValueError as exc:raise HTTPException(400,str(exc)) from exc


@app.post("/api/legal/explorer-export/{export_format}")
def legal_explorer_export(export_format:str,request:LegalQuery):
    try:payload=require_legal_store().explorer_export(request.dataset,request.search,request.filters,export_format)
    except ValueError as exc:raise HTTPException(400,str(exc)) from exc
    media="text/csv" if export_format=="csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return Response(payload,media_type=media,headers={"Content-Disposition":f'attachment; filename="{request.dataset}-filtered.{export_format}"'})


@app.post("/api/legal/case")
def legal_case(request: CaseQuery): return require_legal_store().case(request.query,request.filters,20,request.viewMode,request.page,request.pageSize,request.sortColumn,request.sortDirection,request.columns)


@app.get("/api/legal/case-filters")
def legal_case_filters(): return require_legal_store().case_filters()


@app.get("/api/legal/attachment-download")
def legal_attachment_download(url: str = Query(..., min_length=8)):
    try:
        validate_attachment_url(url)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    parsed=urlparse(url)
    try:
        with ATTACHMENT_OPENER.open(Request(url,headers={"User-Agent":"INTERSOS-Legal-Platform"}),timeout=30) as remote:
            validate_attachment_url(remote.geturl())
            length=int(remote.headers.get("Content-Length","0") or 0)
            if length>50*1024*1024: raise HTTPException(413,"Attachment is larger than 50 MB.")
            payload=remote.read(50*1024*1024+1)
            if len(payload)>50*1024*1024: raise HTTPException(413,"Attachment is larger than 50 MB.")
            media_type=remote.headers.get_content_type() or "application/octet-stream"
    except HTTPException: raise
    except ValueError as exc: raise HTTPException(400,str(exc)) from exc
    except Exception as exc: raise HTTPException(502,"Unable to download the secured document.") from exc
    filename=Path(unquote(parsed.path)).name or "secured-document"
    filename=re.sub(r"[^A-Za-z0-9._ -]","_",filename)
    return Response(payload,media_type=media_type,headers={"Content-Disposition":f'attachment; filename="{filename}"'})


@app.post("/api/legal/case-export")
def legal_case_export(request: CaseQuery):
    payload=require_legal_store().case_export(request.query,request.filters,request.caseIds)
    name="beneficiary-case" if request.query else "filtered-beneficiary-cases"
    return Response(payload,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="{name}.xlsx"'})


@app.post("/api/table-workbook")
def table_workbook(request: TableWorkbookRequest):
    """Create a local Excel workbook for selected visible rows and pivots."""
    sheets=request.sheets or [TableWorkbookSheet(title="Data",columns=request.columns,rows=request.rows)]
    if any(not item.columns for item in sheets): raise HTTPException(400,"Choose at least one table column.")
    if sum(len(item.rows) for item in sheets)>10000: raise HTTPException(400,"Excel export is limited to 10,000 selected rows.")
    output=io.BytesIO();book=Workbook();book.remove(book.active)
    for item in sheets:
        sheet=book.create_sheet(re.sub(r"[\\/*?:\[\]]","_",item.title)[:31] or "Data")
        sheet.append([safe_spreadsheet_value(column) for column in item.columns])
        for cell in sheet[1]:
            cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="2563EB");cell.alignment=Alignment(wrap_text=True,vertical="center")
        for row in item.rows: sheet.append([safe_spreadsheet_value(row.get(column,"")) for column in item.columns])
        sheet.freeze_panes="A2";sheet.auto_filter.ref=sheet.dimensions
        for index,column in enumerate(item.columns,1):
            values=[str(row.get(column,"") or "") for row in item.rows[:500]]
            sheet.column_dimensions[get_column_letter(index)].width=min(42,max(12,len(column)+2,*(len(value)+2 for value in values)))
    book.save(output)
    name=re.sub(r"[^A-Za-z0-9._ -]","_",request.filename or "table-export.xlsx")
    if not name.lower().endswith(".xlsx"): name+= ".xlsx"
    return Response(output.getvalue(),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="{name}"'})


@app.post("/api/legal/lawyers")
def legal_lawyers(request:LegalQuery): return require_legal_store().lawyer_summary(request.filters)


@app.post("/api/legal/representation-case-load/{status}")
def legal_representation_case_load(status:str,request:LegalQuery):
    if status not in {"open","closed"}:raise HTTPException(404,"Unknown case load status")
    return require_legal_store().representation_case_load(request.filters,status)


@app.post("/api/legal/intelligence/{page}")
def legal_intelligence(page:str,request:LegalQuery):
    if page not in {"command-center","lawyer-intelligence","donor-impact"}:raise HTTPException(404,"Unknown intelligence page")
    return require_legal_store().intelligence(page,request.filters)


@app.post("/api/legal/indicators")
def legal_indicators(request:IndicatorReportRequest):
    store=require_legal_store()
    key=(request.fromDate,request.toDate,tuple(sorted(request.projects)),tuple(sorted(request.projectLocations)),tuple(sorted(request.years)),tuple(sorted(request.quarters)),tuple(sorted(request.months)),tuple(sorted(request.communityTypes)))
    with store._cache_lock:
        cached=store._indicator_cache.get(key)
    if cached is not None:return cached
    try:result=build_indicator_report(store.frames,request.fromDate,request.toDate,request.projects,request.projectLocations,request.years,request.quarters,request.months,request.communityTypes)
    except ValueError as exc:raise HTTPException(400,str(exc)) from exc
    with store._cache_lock:
        store._indicator_cache[key]=result
    return result


@app.post("/api/legal/indicators/export")
def legal_indicators_export(request:IndicatorReportRequest):
    report=legal_indicators(request)
    # The detailed report is the export's primary deliverable. Generating every
    # monthly report and chart here can take several minutes on full datasets;
    # the interactive Analysis tab already prepares that view in the background.
    payload=build_indicator_workbook(report)
    return Response(payload,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":'attachment; filename="professional-indicator-report.xlsx"'})


@app.post("/api/legal/indicators/narrative-export")
def legal_indicators_narrative_export(request:IndicatorReportRequest):
    payload=build_narrative_workbook(legal_indicators(request))
    return Response(payload,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":'attachment; filename="indicator-narrative-report.xlsx"'})


@app.post("/api/legal/detention")
def legal_detention(request:LegalQuery):
    if request.page < 1 or request.pageSize < 1 or request.pageSize > 500: raise HTTPException(400,"Invalid pagination")
    if request.sortDirection not in {"asc","desc"}: raise HTTPException(400,"Invalid sort direction")
    return require_legal_store().detention_cases(request.search,request.page,request.pageSize,request.filters,request.sortColumn,request.sortDirection)


@app.post("/api/legal/detention/reconcile")
async def legal_detention_reconcile(month: str, project: list[str] = Query([]), sheet: str = "", file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"): raise HTTPException(400,"Upload an .xlsx comparison workbook")
    if file.size is not None and file.size > MAX_UPLOAD_BYTES: raise HTTPException(413,"Workbook must be 100 MB or smaller.")
    try:
        raw=await read_upload_limited(file)
        return await run_in_threadpool(require_legal_store().detention_reconciliation,raw,file.filename,month,project,sheet)
    except HTTPException: raise
    except ValueError as exc: raise HTTPException(400,str(exc)) from exc
    finally: await file.close()


@app.post("/api/legal/detention/reconcile-export")
async def legal_detention_reconcile_export(month: str, project: list[str] = Query([]), sheet: str = "", file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"): raise HTTPException(400,"Upload an .xlsx comparison workbook")
    try:
        raw=await read_upload_limited(file)
        payload=await run_in_threadpool(require_legal_store().detention_reconciliation_export,raw,file.filename,month,project,sheet)
        return Response(payload,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":'attachment; filename="detention-comparison-issues.xlsx"'})
    except HTTPException: raise
    except ValueError as exc: raise HTTPException(400,str(exc)) from exc
    finally: await file.close()


@app.post("/api/legal/detention/reconcile-sheets")
async def legal_detention_reconcile_sheets(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"): raise HTTPException(400,"Upload an .xlsx comparison workbook")
    try:
        raw=await read_upload_limited(file)
        sheets=await run_in_threadpool(LegalStore.detention_workbook_sheets,raw)
        if not sheets:raise HTTPException(400,"The workbook does not contain any worksheets.")
        return {"sheets":sheets,"selected":sheets[0]}
    except HTTPException:raise
    except ValueError as exc:raise HTTPException(400,str(exc)) from exc
    finally:await file.close()


@app.get("/api/legal/export/{dataset}")
def legal_export(dataset: str):
    try: payload=require_legal_store().export(dataset)
    except ValueError as exc: raise HTTPException(400, str(exc)) from exc
    return Response(payload, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{dataset}-filtered.xlsx"'})


if STATIC_DIR.exists():
    app.mount("/", StaticFiles(directory=STATIC_DIR, html=True), name="dashboard")
