from __future__ import annotations

import io
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from .legal_platform import format_excel_dates
from .file_security import validate_xlsx_archive

OPEN_STATUS_PATTERN = "open|pend"
CLOSED_STATUS_PATTERN = "closed"
COMPLETED_STATUS_PATTERN = "completed"
IN_PROCESS_STATUS_PATTERN = "process"
UNCOMPLETED_STATUS_PATTERN = "uncompleted"


import pandas as pd
import polars as pl


SHEETS = {"assessment": "Assessments", "services": "Legal Services", "deportation": "Deportation"}

COLS = {
    "assessment": {
        "id": "Assessment ID", "beneficiary": "Beneficiary ID", "project": "Projects - المشروع",
        "location": "Project Location", "date": "Date of Assessment تاريخ التقييم",
        "gender": "Gender النوع الاجتماعي", "age_group": "UNHCR Age Group",
        "age_gender": "Age Gender Group", "nationality": "Nationality الجنسية",
        "community": "Community Type", "status": "Assessment Status حالة التقييم",
        "legal_need": "Type of Legal Service Needed",
        "document": "Type of Documents to be issued  نوع الوثائق التي يجب اصدارها",
        "detained": "Is the beneficiary detained هل المستفيد موقوف",
        "detainee_status": "Detainee current status حالة المعتقل الحالية",
    },
    "services": {
        "id": "Service ID", "beneficiary": "Beneficiary ID", "assessment_id": "Assessment ID",
        "project": "Projects - المشروع", "location": "Project Location",
        "date": "Date of Service Provision /  تاريخ بدء الخدمة",
        "completed_date": "Date Service Completed تاريخ انجاز الخدمة",
        "gender": "Gender النوع الاجتماعي", "age_group": "UNHCR Age Group",
        "age_gender": "Age Gender Group", "nationality": "Nationality الجنسية",
        "community": "Community Type", "service_type": "Type of Service Provided / نوع الخدمة",
        "document": "Type of Document نوع الوثيقة", "status": "Service Status حالة الخدمة",
    },
    "deportation": {
        "id": "PN ID", "project": "Projects - المشروع", "location": "Project Location",
        "date": "Date of deporting تاريخ الترحيل", "gender": "Gender النوع الاجتماعي",
        "age_group": "UNHCR Age Group", "age_gender": "Age Gender Group",
        "nationality": "Nationality الجنسية", "community": "Community Type / نوع المجتمع",
        "authority": "Deporting Authority السلطة التي قامت بالترحيل",
        "reason": "Reason of deporting سبب الترحيل", "destination": "Deported to رحل الى",
    },
}

FILTERS = {
    "assessment": ["project", "location", "gender", "age_group", "nationality", "legal_need", "document", "community", "status", "detained", "detainee_status", "year", "quarter", "month"],
    "services": ["project", "location", "gender", "age_group", "nationality", "legal_need", "document", "community", "service_type", "status", "year", "quarter", "month", "completed_year", "completed_quarter", "completed_month"],
    "deportation": ["project", "location", "gender", "age_group", "nationality", "community", "authority", "reason", "destination", "year", "quarter", "month"],
}

CHARTS = {
    "assessment": [
        ("project", "Project"), ("location", "Project location"), ("age_gender", "Gender / age group"),
        ("nationality", "Nationality"), ("community", "Community type"), ("status", "Assessment status"),
        ("legal_need", "Legal service needed"), ("document", "Document needed"),
        ("detained", "Beneficiary detained"), ("detainee_status", "Detainee current status"),
    ],
    "services": [
        ("project", "Project"), ("location", "Project location"), ("age_gender", "Gender / age group"),
        ("service_type", "Type of service provided"), ("status", "Service status"),
        ("document", "Type of document"), ("nationality", "Nationality"), ("community", "Community type"),
        ("legal_need", "Assessment legal-service need"),
    ],
    "deportation": [
        ("age_gender", "Gender / age group"), ("community", "Community type"),
        ("reason", "Reason for deportation"), ("authority", "Deporting authority"),
        ("nationality", "Nationality"), ("destination", "Deported to"),
        ("project", "Project"), ("location", "Project location"),
    ],
}


def _clean_id(value: Any) -> str | None:
    if pd.isna(value): return None
    if isinstance(value, float) and value.is_integer(): return str(int(value))
    value = str(value).strip()
    return value or None


def _text(value: Any) -> str | None:
    if pd.isna(value): return None
    value = " ".join(str(value).split())
    arabic = re.compile(r"[\u0600-\u06ff\u0750-\u077f\u08a0-\u08ff\ufb50-\ufdff\ufe70-\ufeff]")
    english_answers: list[str] = []
    for answer in value.split(","):
        match = arabic.search(answer)
        english = answer[:match.start()] if match else answer
        english = re.sub(r"[\s\-/–—]+$", "", english.strip())
        if english and english not in english_answers:
            english_answers.append(english)
    return ",".join(english_answers) or None


def _date_series(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=True)


@dataclass
class DataStore:
    source_name: str
    frames: dict[str, pl.DataFrame]
    options: dict[str, dict[str, pl.DataFrame]]
    quality: list[dict[str, Any]]
    loaded_at: str
    raw_sheets: dict[str, pd.DataFrame] = field(default_factory=dict)
    raw_sheet_names: dict[str, str] = field(default_factory=dict)
    raw_sheet_warnings: list[str] = field(default_factory=list)
    raw_metadata_cache: dict[str, dict[str, Any]] = field(default_factory=dict)

    @classmethod
    def from_path(cls, path: Path) -> "DataStore":
        return cls.from_bytes(path.read_bytes(), path.name)

    @classmethod
    def from_bytes(cls, raw: bytes, source_name: str) -> "DataStore":
        validate_xlsx_archive(raw)
        excel = pd.ExcelFile(io.BytesIO(raw), engine="openpyxl")
        missing = [s for s in SHEETS.values() if s not in excel.sheet_names]
        if missing: raise ValueError(f"Missing required sheets: {', '.join(missing)}")
        frames: dict[str, pl.DataFrame] = {}
        options: dict[str, dict[str, pl.DataFrame]] = {"assessment": {}, "services": {}, "deportation": {}}
        quality: list[dict[str, Any]] = []
        today = pd.Timestamp(date.today())

        raw_sheets: dict[str, pd.DataFrame] = {}
        raw_sheet_names: dict[str, str] = {}
        raw_sheet_warnings: list[str] = []
        for index, sheet_name in enumerate(excel.sheet_names):
            sheet_id = f"sheet-{index}"
            try:
                raw_sheet = pd.read_excel(excel, sheet_name=sheet_name, dtype=object)
                raw_sheet.columns = [str(column).strip() or f"Column {position + 1}" for position, column in enumerate(raw_sheet.columns)]
                if raw_sheet.empty or not len(raw_sheet.columns):
                    raw_sheet_warnings.append(f"{sheet_name}: skipped because it is empty.")
                    continue
                raw_sheets[sheet_id] = raw_sheet
                raw_sheet_names[sheet_id] = sheet_name
            except Exception as exc:
                raw_sheet_warnings.append(f"{sheet_name}: skipped because it could not be read ({exc}).")

        assessment_map: dict[str, list[str]] = {}
        for page, sheet in SHEETS.items():
            raw_df = next(frame for sheet_id, frame in raw_sheets.items() if raw_sheet_names[sheet_id] == sheet).copy()
            mapping = COLS[page]
            absent = [col for col in mapping.values() if col not in raw_df.columns]
            if absent: raise ValueError(f"{sheet} is missing columns: {', '.join(absent)}")
            df = pd.DataFrame({key: raw_df[col] for key, col in mapping.items()})
            for col in df.columns:
                if col not in {"date", "completed_date"}: df[col] = df[col].map(_clean_id if col.endswith("id") or col in {"id", "beneficiary"} else _text)
            date_fields = [c for c in ["date", "completed_date"] if c in df]
            for col in date_fields: df[col] = _date_series(df[col])
            primary_date = df["date"]
            df["report_date"] = primary_date
            df["year"] = primary_date.dt.year.astype("Int64").astype(str).replace("<NA>", None)
            df["quarter"] = primary_date.dt.to_period("Q").astype(str).replace("NaT", None)
            df["month"] = primary_date.dt.to_period("M").astype(str).replace("NaT", None)
            df["invalid_date"] = primary_date.isna() | (primary_date > today) | (primary_date.dt.year < 2023)
            if page == "services":
                completed_date = df["completed_date"]
                df["completed_year"] = completed_date.dt.year.astype("Int64").astype(str).replace("<NA>", None)
                df["completed_quarter"] = completed_date.dt.to_period("Q").astype(str).replace("NaT", None)
                df["completed_month"] = completed_date.dt.to_period("M").astype(str).replace("NaT", None)

            for key in ["id"] + (["beneficiary"] if "beneficiary" in df else []):
                missing_count = int(df[key].isna().sum())
                dup_count = int(df[key].notna().sum() - df[key].dropna().nunique())
                quality.append({"page": page, "severity": "Critical" if key == "id" and missing_count else "High", "check": f"Missing {key.replace('_',' ')}", "count": missing_count, "rate": missing_count / max(len(df), 1), "impact": "Rows cannot be counted reliably at the intended grain."})
                quality.append({"page": page, "severity": "High", "check": f"Duplicate {key.replace('_',' ')}", "count": dup_count, "rate": dup_count / max(len(df), 1), "impact": "Distinct counting prevents inflation; duplicates remain flagged."})
            invalid = int(df["invalid_date"].sum())
            quality.append({"page": page, "severity": "High" if invalid else "Low", "check": "Invalid, future, or implausible reporting date", "count": invalid, "rate": invalid / max(len(df), 1), "impact": "Excluded from time trends and the default YTD view."})

            if page == "assessment":
                for field in ["legal_need", "document"]:
                    exploded = df[["id", "beneficiary", field]].copy()
                    exploded[field] = exploded[field].fillna("").str.split(",")
                    exploded = exploded.explode(field)
                    exploded[field] = exploded[field].map(_text)
                    exploded = exploded.dropna(subset=[field, "id"]).drop_duplicates(["id", field])
                    options[page][field] = pl.from_pandas(exploded, include_index=False)
                assessment_map = df.set_index("id")["legal_need"].fillna("").str.split(",").to_dict()
            elif page == "services":
                df["legal_need"] = df["assessment_id"].map(assessment_map)
                service_need = df[["id", "beneficiary", "legal_need"]].explode("legal_need")
                service_need["legal_need"] = service_need["legal_need"].map(_text)
                service_need = service_need.dropna(subset=["id", "legal_need"]).drop_duplicates(["id", "legal_need"])
                options[page]["legal_need"] = pl.from_pandas(service_need, include_index=False)
                unmatched = int(df["assessment_id"].notna().sum() - df["assessment_id"].isin(assessment_map).sum())
                quality.append({"page": page, "severity": "High" if unmatched else "Low", "check": "Service Assessment ID not matched", "count": unmatched, "rate": unmatched / max(len(df), 1), "impact": "Unmatched services cannot inherit assessment legal need."})
                completed_status = df["status"].fillna("").str.lower().str.contains(COMPLETED_STATUS_PATTERN) & ~df["status"].fillna("").str.lower().str.contains(UNCOMPLETED_STATUS_PATTERN)
                missing_completion = int((completed_status & df["completed_date"].isna()).sum())
                invalid_completion = int((df["completed_date"].notna() & ((df["completed_date"] > today) | (df["completed_date"].dt.year < 2023))).sum())
                quality.append({"page": page, "severity": "High" if missing_completion else "Low", "check": "Completed service without completion date", "count": missing_completion, "rate": missing_completion / max(int(completed_status.sum()), 1), "impact": "Completion-period reporting is incomplete."})
                quality.append({"page": page, "severity": "High" if invalid_completion else "Low", "check": "Invalid or future completion date", "count": invalid_completion, "rate": invalid_completion / max(int(df["completed_date"].notna().sum()), 1), "impact": "Excluded from completion trends."})
                df["legal_need"] = df["legal_need"].map(lambda x: ",".join(x) if isinstance(x, list) else None)
            frames[page] = pl.from_pandas(df, include_index=False)
        return cls(source_name, frames, options, quality, pd.Timestamp.now().isoformat(timespec="seconds"), raw_sheets, raw_sheet_names, raw_sheet_warnings)

    def metadata(self) -> dict[str, Any]:
        result: dict[str, Any] = {"source": self.source_name, "loadedAt": self.loaded_at, "pages": {}}
        for page, frame in self.frames.items():
            values: dict[str, list[str]] = {}
            for field in FILTERS[page]:
                if field in self.options[page]:
                    vals = self.options[page][field].get_column(field).drop_nulls().unique().sort().to_list()
                elif field in frame.columns:
                    vals = frame.get_column(field).drop_nulls().unique().sort().to_list()
                else: vals = []
                values[field] = [str(v) for v in vals]
            dimensions = list(dict.fromkeys([field for field, _ in CHARTS[page]] + ["gender", "age_group", "year", "quarter", "month"]))
            result["pages"][page] = {"rows": frame.height, "filters": values, "dimensions": dimensions}
        common = ["project", "location", "gender", "age_group", "nationality", "community", "year", "quarter"]
        result["pages"]["executive"] = {"rows": sum(f.height for f in self.frames.values()), "filters": {field: sorted({str(v) for frame in self.frames.values() if field in frame.columns for v in frame.get_column(field).drop_nulls().unique().to_list()}) for field in common}}
        result["dataExplorer"] = {
            "sheets": [self._raw_sheet_metadata(sheet_id, frame) for sheet_id, frame in self.raw_sheets.items()],
            "warnings": self.raw_sheet_warnings,
        }
        return result

    def _raw_sheet_metadata(self, sheet_id: str, frame: pd.DataFrame) -> dict[str, Any]:
        if sheet_id in self.raw_metadata_cache:
            return self.raw_metadata_cache[sheet_id]
        columns = []
        for column in frame.columns:
            series = frame[column]
            non_null = series.dropna()
            sample = non_null.head(200)
            inferred = "text"
            if not sample.empty:
                numeric = pd.to_numeric(sample, errors="coerce")
                if numeric.notna().mean() >= .9:
                    inferred = "number"
                elif isinstance(sample.iloc[0], (pd.Timestamp, date)):
                    inferred = "date"
                elif re.search(r"\b(date|time|day|month|year)\b", str(column), re.IGNORECASE):
                    dates = pd.to_datetime(sample, errors="coerce", dayfirst=True)
                    if dates.notna().mean() >= .9:
                        inferred = "date"
            unique_values = non_null.map(self._display_value).drop_duplicates()
            values = sorted((str(value) for value in unique_values.tolist()), key=str.casefold) if len(unique_values) <= 100 else []
            columns.append({"name": str(column), "type": inferred, "values": values})
        result = {"id": sheet_id, "name": self.raw_sheet_names[sheet_id], "rows": len(frame), "columns": columns}
        self.raw_metadata_cache[sheet_id] = result
        return result

    @staticmethod
    def _display_value(value: Any) -> Any:
        if pd.isna(value):
            return None
        if isinstance(value, (pd.Timestamp, date)):
            return value.isoformat()
        if hasattr(value, "item"):
            value = value.item()
        return value if isinstance(value, (str, int, float, bool)) else str(value)

    def explorer_query(self, sheet_id: str, search: str = "", filters: list[dict[str, Any]] | None = None,
                       sort_column: str | None = None, sort_direction: str = "asc", page: int = 1,
                       page_size: int = 100, columns: list[str] | None = None) -> dict[str, Any]:
        if sheet_id not in self.raw_sheets:
            raise ValueError("Unknown worksheet")
        frame = self.raw_sheets[sheet_id]
        working = frame.copy()
        if search.strip():
            needle = search.strip().lower()
            mask = pd.Series(False, index=working.index)
            for column in working.columns:
                mask |= working[column].fillna("").astype(str).str.lower().str.contains(needle, regex=False)
            working = working[mask]
        for item in filters or []:
            column = item.get("column")
            operator = item.get("operator")
            value = item.get("value")
            value2 = item.get("value2")
            if column not in working.columns:
                continue
            series = working[column]
            if operator == "blank":
                working = working[series.isna() | series.astype(str).str.strip().eq("")]
            elif operator == "not_blank":
                working = working[series.notna() & ~series.astype(str).str.strip().eq("")]
            elif operator == "equals":
                working = working[series.fillna("").astype(str).str.lower().eq(str(value or "").lower())]
            elif operator == "contains":
                working = working[series.fillna("").astype(str).str.contains(str(value or ""), case=False, regex=False)]
            elif operator in {"gte", "lte", "between"}:
                numeric = pd.to_numeric(series, errors="coerce")
                try:
                    lower = float(value) if value not in (None, "") else None
                    upper = float(value2) if value2 not in (None, "") else None
                except (TypeError, ValueError):
                    continue
                if operator == "gte" and lower is not None:
                    working = working[numeric >= lower]
                elif operator == "lte" and lower is not None:
                    working = working[numeric <= lower]
                elif operator == "between":
                    if lower is not None: working = working[numeric >= lower]
                    if upper is not None: working = working[pd.to_numeric(working[column], errors="coerce") <= upper]
            elif operator in {"date_on", "date_before", "date_after", "date_between"}:
                parsed = pd.to_datetime(series, errors="coerce").dt.normalize()
                first = pd.to_datetime(value, errors="coerce")
                second = pd.to_datetime(value2, errors="coerce")
                if operator == "date_on" and not pd.isna(first): working = working[parsed == first.normalize()]
                elif operator == "date_before" and not pd.isna(first): working = working[parsed <= first.normalize()]
                elif operator == "date_after" and not pd.isna(first): working = working[parsed >= first.normalize()]
                elif operator == "date_between":
                    if not pd.isna(first): working = working[parsed >= first.normalize()]
                    if not pd.isna(second): working = working[pd.to_datetime(working[column], errors="coerce").dt.normalize() <= second.normalize()]
        if sort_column in working.columns:
            working = working.sort_values(sort_column, ascending=sort_direction != "desc", na_position="last", kind="stable")
        selected = [column for column in (columns or list(frame.columns)) if column in frame.columns]
        if not selected:
            selected = list(frame.columns)
        matched = len(working)
        start = (max(page, 1) - 1) * page_size
        page_frame = working.iloc[start:start + page_size][selected]
        rows = [{str(column): self._display_value(value) for column, value in row.items()} for row in page_frame.to_dict("records")]
        return {"sheetId": sheet_id, "totalRows": len(frame), "matchedRows": matched, "page": max(page, 1),
                "pageSize": page_size, "columns": selected, "rows": rows}

    def explorer_export(self, sheet_id: str, search: str, filters: list[dict[str, Any]], sort_column: str | None,
                        sort_direction: str, columns: list[str], export_format: str) -> bytes:
        if sheet_id not in self.raw_sheets:
            raise ValueError("Unknown worksheet")
        result = self.explorer_query(sheet_id, search, filters, sort_column, sort_direction, 1, max(len(self.raw_sheets[sheet_id]), 1), columns)
        frame = pd.DataFrame(result["rows"], columns=result["columns"])
        for column in frame.columns:
            if frame[column].dtype == object:
                frame[column] = frame[column].map(lambda value: "'" + value if isinstance(value, str) and re.match(r"^[=+\-@]", value) else value)
        output = io.BytesIO()
        if export_format == "xlsx":
            with pd.ExcelWriter(output, engine="openpyxl", date_format="YYYY-MMMM-DD", datetime_format="YYYY-MMMM-DD") as writer:
                frame.to_excel(writer, index=False, sheet_name="Filtered data")
                format_excel_dates(writer.book)
        else:
            return frame.to_csv(index=False).encode("utf-8-sig")
        return output.getvalue()

    def _filtered(self, page: str, filters: dict[str, list[str]], default_ytd: bool) -> pl.DataFrame:
        frame = self.frames[page]
        if default_ytd and not filters.get("year"):
            frame = frame.filter((pl.col("year") == "2026") & (~pl.col("invalid_date")))
        for field, selections in filters.items():
            if not selections: continue
            if field in self.options[page]:
                ids = self.options[page][field].filter(pl.col(field).is_in(selections)).get_column("id").unique().to_list()
                frame = frame.filter(pl.col("id").is_in(ids))
            elif field in frame.columns:
                frame = frame.filter(pl.col(field).is_in(selections))
        return frame

    @staticmethod
    def _metric_col(page: str, measure: str) -> str:
        return "beneficiary" if measure == "beneficiaries" and page != "deportation" else "id"

    def dashboard(self, page: str, filters: dict[str, list[str]], measure: str, default_ytd: bool = True) -> dict[str, Any]:
        if page == "executive": return self.executive(filters, default_ytd)
        df = self._filtered(page, filters, default_ytd)
        metric = self._metric_col(page, measure)
        total = df.get_column(metric).drop_nulls().n_unique() if metric in df.columns else df.height

        def distinct_where(expr: pl.Expr) -> int:
            return df.filter(expr).get_column(metric).drop_nulls().n_unique()

        if page == "assessment":
            closed = distinct_where(pl.col("status").str.to_lowercase().str.contains(CLOSED_STATUS_PATTERN))
            open_cases = distinct_where(pl.col("status").str.to_lowercase().str.contains(OPEN_STATUS_PATTERN))
            detained = distinct_where(pl.col("detained").str.to_lowercase().str.contains("yes|نعم"))
            kpis = [("Assessments" if measure == "records" else "Unique beneficiaries", total), ("Open caseload", open_cases), ("Closed", closed), ("Closure rate", closed / total if total else 0), ("Detained", detained), ("Projects", df.get_column("project").drop_nulls().n_unique()), ("Locations", df.get_column("location").drop_nulls().n_unique())]
        elif page == "services":
            completed = distinct_where(pl.col("status").str.to_lowercase().str.contains(COMPLETED_STATUS_PATTERN) & ~pl.col("status").str.to_lowercase().str.contains(UNCOMPLETED_STATUS_PATTERN))
            in_process = distinct_where(pl.col("status").str.to_lowercase().str.contains(IN_PROCESS_STATUS_PATTERN))
            uncompleted = distinct_where(pl.col("status").str.to_lowercase().str.contains(UNCOMPLETED_STATUS_PATTERN))
            kpis = [("Services" if measure == "records" else "Unique beneficiaries", total), ("Completed", completed), ("In process", in_process), ("Uncompleted", uncompleted), ("Completion rate", completed / total if total else 0), ("Projects", df.get_column("project").drop_nulls().n_unique()), ("Locations", df.get_column("location").drop_nulls().n_unique())]
        else:
            kpis = [("Deportation records", total), ("Destinations", df.get_column("destination").drop_nulls().n_unique()), ("Nationalities", df.get_column("nationality").drop_nulls().n_unique()), ("Authorities", df.get_column("authority").drop_nulls().n_unique()), ("Projects", df.get_column("project").drop_nulls().n_unique()), ("Locations", df.get_column("location").drop_nulls().n_unique())]

        charts = []
        for field, title in CHARTS[page]:
            if field in self.options[page]:
                ids = df.get_column("id").unique().to_list()
                joined = self.options[page][field].filter(pl.col("id").is_in(ids))
                agg = joined.group_by(field).agg(pl.col(metric if metric in joined.columns else "id").n_unique().alias("count")).sort("count", descending=True).head(12)
            else:
                agg = df.drop_nulls([field, metric]).group_by(field).agg(pl.col(metric).n_unique().alias("count")).sort("count", descending=True).head(12)
            rows = [{"label": str(r[field]), "count": int(r["count"]), "percent": (r["count"] / total if total else 0)} for r in agg.to_dicts()]
            charts.append({"id": field, "title": title, "kind": "bar", "rows": rows, "multiChoice": field in self.options[page]})

        flow = []
        if page == "deportation":
            flow_df = df.drop_nulls(["nationality", "destination", metric]).group_by(["nationality", "destination"]).agg(pl.col(metric).n_unique().alias("count")).sort("count", descending=True)
            flow = [{"source": str(r["nationality"]), "target": str(r["destination"]), "count": int(r["count"])} for r in flow_df.to_dicts()]

        trend_df = df.filter(~pl.col("invalid_date") & pl.col("month").is_not_null()).group_by("month").agg(pl.col(metric).n_unique().alias("count")).sort("month")
        trend = [{"label": r["month"], "count": int(r["count"]), "percent": r["count"] / total if total else 0} for r in trend_df.to_dicts()]
        open_trend = []
        closed_trend = []
        if page == "assessment":
            valid_dates = ~pl.col("invalid_date") & pl.col("month").is_not_null()
            open_by_month = df.filter(valid_dates & pl.col("status").str.to_lowercase().str.contains(OPEN_STATUS_PATTERN)).group_by("month").agg(pl.col(metric).n_unique().alias("count")).sort("month")
            closed_by_month = df.filter(valid_dates & pl.col("status").str.to_lowercase().str.contains(CLOSED_STATUS_PATTERN)).group_by("month").agg(pl.col(metric).n_unique().alias("count")).sort("month")
            open_trend = [{"label": r["month"], "count": int(r["count"]), "percent": r["count"] / total if total else 0} for r in open_by_month.to_dicts()]
            closed_trend = [{"label": r["month"], "count": int(r["count"]), "percent": r["count"] / total if total else 0} for r in closed_by_month.to_dicts()]
        completion_trend = []
        if page == "services":
            valid = df.filter(pl.col("completed_date").is_not_null() & (pl.col("completed_date") <= pl.lit(date.today())) & (pl.col("completed_date").dt.year() >= 2023))
            comp = valid.group_by("completed_month").agg(pl.col(metric).n_unique().alias("count")).sort("completed_month")
            completion_trend = [{"label": r["completed_month"], "count": int(r["count"]), "percent": r["count"] / total if total else 0} for r in comp.to_dicts()]
        return {"page": page, "measure": measure, "total": total, "kpis": [{"label": l, "value": v, "format": "percent" if "rate" in l.lower() else "number"} for l, v in kpis], "trend": trend, "openTrend": open_trend, "closedTrend": closed_trend, "completionTrend": completion_trend, "flow": flow, "charts": charts, "filteredRows": df.height}

    def executive(self, filters: dict[str, list[str]], default_ytd: bool = True) -> dict[str, Any]:
        frames = {p: self._filtered(p, {k: v for k, v in filters.items() if k in self.frames[p].columns}, default_ytd) for p in self.frames}
        a, s, d = frames["assessment"], frames["services"], frames["deportation"]
        assessment_count = a.get_column("id").drop_nulls().n_unique()
        service_count = s.get_column("id").drop_nulls().n_unique()
        completed = s.filter(pl.col("status").str.to_lowercase().str.contains(COMPLETED_STATUS_PATTERN) & ~pl.col("status").str.to_lowercase().str.contains(UNCOMPLETED_STATUS_PATTERN)).get_column("id").drop_nulls().n_unique()
        open_count = a.filter(pl.col("status").str.to_lowercase().str.contains(OPEN_STATUS_PATTERN)).get_column("id").drop_nulls().n_unique()
        assessed = a.get_column("beneficiary").drop_nulls().n_unique()
        served = s.get_column("beneficiary").drop_nulls().n_unique()
        assessed_ids = set(a.get_column("beneficiary").drop_nulls().unique().to_list())
        served_ids = set(s.get_column("beneficiary").drop_nulls().unique().to_list())
        coverage = len(assessed_ids & served_ids) / assessed if assessed else 0
        kpis = [("Assessments", assessment_count), ("Beneficiaries assessed", assessed), ("Open caseload", open_count), ("Services", service_count), ("Beneficiaries served", served), ("Service coverage", coverage), ("Completed services", completed), ("Completion rate", completed/service_count if service_count else 0), ("Deportations", d.get_column("id").drop_nulls().n_unique())]
        trend_df = a.filter(~pl.col("invalid_date")).group_by("month").agg(pl.col("id").n_unique().alias("count")).sort("month")
        trend = [{"label": r["month"], "count": int(r["count"]), "percent": r["count"] / assessment_count if assessment_count else 0} for r in trend_df.to_dicts()]
        charts=[]
        for field,title in [("project","Assessment caseload by project"),("location","Assessment caseload by location"),("age_gender","Population profile"),("community","Community type")]:
            agg=a.drop_nulls([field,"id"]).group_by(field).agg(pl.col("id").n_unique().alias("count")).sort("count",descending=True).head(12)
            charts.append({"id":field,"title":title,"kind":"bar","multiChoice":False,"rows":[{"label":str(r[field]),"count":int(r["count"]),"percent":r["count"]/assessment_count if assessment_count else 0} for r in agg.to_dicts()]})
        return {"page":"executive","measure":"records","total":assessment_count,"kpis":[{"label":l,"value":v,"format":"percent" if ("rate" in l.lower() or "coverage" in l.lower()) else "number"} for l,v in kpis],"trend":trend,"completionTrend":[],"charts":charts,"filteredRows":sum(x.height for x in frames.values())}

    def quality_summary(self) -> dict[str, Any]:
        return {"rows": self.quality, "source": self.source_name, "loadedAt": self.loaded_at}

    def studio(self, page: str, row_dimension: str, column_dimension: str | None, filters: dict[str, list[str]], measure: str, default_ytd: bool = True) -> dict[str, Any]:
        if page not in self.frames: raise ValueError("Unknown source sheet")
        allowed = set(FILTERS[page]) | {field for field, _ in CHARTS[page]} | {"age_gender", "month", "quarter", "year"}
        dimensions = [row_dimension] + ([column_dimension] if column_dimension else [])
        if any(not dim or dim not in allowed for dim in dimensions): raise ValueError("Unsupported studio dimension")
        base = self._filtered(page, filters, default_ytd)
        metric = self._metric_col(page, measure)
        working = base.select([c for c in {"id", metric, *[d for d in dimensions if d in base.columns]} if c in base.columns])
        for dim in dimensions:
            if dim in self.options[page]:
                working = working.join(self.options[page][dim].select("id", dim), on="id", how="inner")
        working = working.drop_nulls(dimensions + [metric])
        total = base.get_column(metric).drop_nulls().n_unique()
        grouped = working.group_by(dimensions).agg(pl.col(metric).n_unique().alias("count")).sort("count", descending=True)
        if column_dimension:
            top_rows = grouped.group_by(row_dimension).agg(pl.col("count").sum().alias("total")).sort("total", descending=True).head(20).get_column(row_dimension).to_list()
            top_cols = grouped.group_by(column_dimension).agg(pl.col("count").sum().alias("total")).sort("total", descending=True).head(12).get_column(column_dimension).to_list()
            grouped = grouped.filter(pl.col(row_dimension).is_in(top_rows) & pl.col(column_dimension).is_in(top_cols))
        else:
            grouped = grouped.head(30)
        cells = [{"row": str(r[row_dimension]), "column": str(r[column_dimension]) if column_dimension else "Total", "count": int(r["count"]), "percent": r["count"] / total if total else 0} for r in grouped.to_dicts()]
        return {"page": page, "rowDimension": row_dimension, "columnDimension": column_dimension, "measure": measure, "total": total, "cells": cells}

    def export_csv(self, page: str, filters: dict[str, list[str]], default_ytd: bool = True) -> bytes:
        df = self._filtered(page, filters, default_ytd)
        safe = [c for c in df.columns if c not in {"invalid_date"}]
        text_columns = [column for column in safe if df.schema[column] == pl.String]
        escaped = df.select(safe).with_columns([
            pl.when(pl.col(column).str.contains(r"^[=+\-@]"))
            .then(pl.concat_str([pl.lit("'"), pl.col(column)]))
            .otherwise(pl.col(column))
            .alias(column)
            for column in text_columns
        ])
        return escaped.write_csv().encode("utf-8-sig")

    def export_xlsx(self, page: str, filters: dict[str, list[str]], default_ytd: bool = True) -> bytes:
        df=self._filtered(page,filters,default_ytd);columns=[column for column in df.columns if column!="invalid_date"]
        workbook=Workbook();sheet=workbook.active;sheet.title="Filtered data";sheet.append(columns)
        for record in df.select(columns).to_dicts():
            sheet.append([("'"+value if isinstance(value,str) and re.match(r"^[=+\-@]",value) else value) for value in (record.get(column) for column in columns)])
        sheet.freeze_panes="A2";sheet.auto_filter.ref=sheet.dimensions;sheet.row_dimensions[1].height=32
        for cell in sheet[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="2563EB");cell.alignment=Alignment(wrap_text=True,vertical="center")
        for index,column in enumerate(columns,1):sheet.column_dimensions[get_column_letter(index)].width=min(38,max(12,len(column)+2))
        output=io.BytesIO();workbook.save(output);return output.getvalue()
