from __future__ import annotations

import io
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path
from threading import Lock
from typing import Any, Callable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .file_security import safe_spreadsheet_value, validate_xlsx_archive


MANDATORY = ("beneficiaries", "assessments", "legalservices")
OPTIONAL = ("followupslogbooks", "legalfees", "awareness", "deportationrecords")
FILES = MANDATORY + OPTIONAL
INDICATOR_AGE_GROUPS = ("00-04", "05-11", "12-17", "18-39", "40-59", "60+")
INDICATOR_REPORT_ROWS = (
    ("UNHCR 2026 - AMAL CAMP", "AMAL Camp"),
    ("UNHCR 2026 - Erbil", "Urban"), ("UNHCR 2026 - Erbil", "Kawrgawsk Camp"),
    ("UNHCR 2026 - Erbil", "Basirma Camp"), ("UNHCR 2026 - Erbil", "Darashakran Camp"),
    ("UNHCR 2026 - Erbil", "Qushtapa Camp"),
    ("UNHCR 2026 - SULI", "Sulaymaniyah Urban"),
    ("UNHCR 2026 - SULI", "Pshdar Urban (Refugees) + Rania"),
    ("UNHCR 2026 - SULI", "Arbat Camp (Refugees)"),
    ("UNHCR 2026 - Mosul & Kirkuk", "Ninewa نينوى"),
    ("UNHCR 2026 - Mosul & Kirkuk", "Kirkuk كركوك"),
    ("UNHCR 2026 - Baghdad", "Baghdad بغداد"),
    ("UNHCR 2026 - Gov", "Anbar أنبار"), ("UNHCR 2026 - Gov", "Al-Muthanna المثنى"),
    ("UNHCR 2026 - Gov", "Al-Qadisiyyah القادسية"), ("UNHCR 2026 - Gov", "Babil بابل"),
    ("UNHCR 2026 - Gov", "Basra بصرة"), ("UNHCR 2026 - Gov", "Dhi Qar ذي قار"),
    ("UNHCR 2026 - Gov", "Diyala ديالى"), ("UNHCR 2026 - Gov", "Karbala كربلاء"),
    ("UNHCR 2026 - Gov", "Maysan ميسان"), ("UNHCR 2026 - Gov", "Najaf نجف"),
    ("UNHCR 2026 - Gov", "Salah Al-Din صلاح الدين"), ("UNHCR 2026 - Gov", "Wassit واسط"),
)
INDICATOR_CIVIL_DOCUMENTS = (
    "unified national card", "iraqi nationality certificate", "pds card", "birth certificate",
    "proof of birth", "proof of custody", "proof of kinship", "proof of guardianship",
    "marriage certificate", "proof of marriage", "marriage attestation", "divorce certificate",
    "death certificate", "proof of death", "proof of curatorship", "housing card", "passport", "civil id",
)
REPRESENTATION_DOCUMENT_EXCEPTIONS = (
    "custody certificate", "divorce certificate", "marriage attestation",
    "marriage certificate", "proof of kinship", "proof of guardianship",
    "proof of marriage",
)
DISPLAY_NAMES = {
    "beneficiaries": "Beneficiaries", "assessments": "Assessments", "legalservices": "Legal Services",
    "followupslogbooks": "Follow-ups & Logbooks", "legalfees": "Legal Fees", "awareness": "Awareness",
    "deportationrecords": "Deportation Records",
}
REQUIRED_COLUMNS = {
    "beneficiaries": ("Case ID", "Name (Filter Color Red)"),
    "assessments": ("Assessment ID", "Beneficiary ID"),
    "legalservices": ("Service ID", "Assessment ID", "Beneficiary ID"),
}
SPLIT_NAME = re.compile(r"^Name\s*/.*:\s*(Title|First|Middle|Last)\s*$", re.I)
SECURED_FILE = "Secured documents Files"
DATE_HINT = re.compile(r"\b(date|dob|created on|added on|paid date)\b", re.I)
VERSION_SUFFIX = re.compile(r"\s*\((\d+)\)\s*$")
ACTIONS = {
    "Possible duplicate name": "Verify the potential duplicate with the responsible lawyer. Where duplication is confirmed, notify the IM Officer to arrange deletion of the redundant case. The lawyer must update the physical file and PR record to retain the correct Case ID.",
    "Possible duplicate contact and name": "Verify the matching contact number and identity details with the responsible lawyer. Where duplication is confirmed, notify the IM Officer to arrange deletion of the redundant case.",
    "Invalid contact number": "Verify the phone number against the case file and correct it in the platform.",
    "Case without assessment": "The responsible lawyer must complete and link the required assessment in the platform.",
    "Invalid age": "Verify age and date of birth, then correct the inaccurate or missing value.",
    "Marital status below 18": "Review the age and marital-status entries and document any confirmed child-marriage concern.",
    "Spouse below 18": "Verify the spouse date of birth and review the case for child-protection concerns.",
    "Check Community Type vs Nationality": "Verify the Community Type, Nationality, and Project against the case file, then correct the beneficiary classification in the platform.",
    "Beneficiary has multiple assessments": "Review the assessment timeline and confirm that each assessment is intentional and correctly linked.",
    "Current and previous month duplicate": "Compare the selected-month legal service with the beneficiary's earlier services and confirm the repeat activity is intentional.",
    "Selected month with previous assessment": "Compare the selected-month assessment with the beneficiary's earlier assessments and confirm the repeat activity is intentional.",
    "Assessment without services": "Confirm the action plan and create or link the required legal service.",
    "Pending assessment": "Review the pending reason, owner, and next action; update the assessment when resolved.",
    "Open counselling-only assessment": "Confirm whether counselling completes the need or whether assistance/representation must be added.",
    "Detention/immigration inconsistency": "Verify detention and immigration-charge answers and correct the inconsistent response.",
    "Blank legal service need": "Complete the legal-service need before progressing the assessment.",
    "Detained beneficiary has counselling only": "Escalate for legal assistance or representation review and document the decision.",
    "Adult representation without counselling": "Confirm that legal counselling was provided and link or record it.",
    "Representation while not detained": "Review eligibility and the recorded detention/community information.",
    "Detained beneficiary below 10 years": "Review the detention details and immediately follow the child-protection referral process for this beneficiary.",
    "Detention Governorate mismatch": "Verify the detention governorate against the assessment project and project location, then correct the inconsistent entry.",
    "Assessment date after today": "Verify the future date against the source record and correct it if it was entered in error.",
    "Legal service date after today": "Verify the future date against the source record and correct it if it was entered in error.",
    "Type of document in Assessments vs Services": "Compare the document types requested in the assessment with the document types recorded on linked legal services, then correct the missing side.",
    "Type of Legal Service in Assessment vs Services": "Compare the legal service types requested in the assessment with linked legal services and create or correct the missing service type.",
    "Duplicate service": "Compare the service records and correct or remove the duplicate in the source platform.",
    "Missing Type of Document": "Check the service record and complete Type of Document before reporting.",
    "Orphaned assessment relationship": "Correct the Assessment ID or link the service to an existing assessment.",
    "Possible duplicate participant name": "Check whether the matching name belongs to the same person before making any correction.",
    "Duplicate participant in session": "Verify the participant and session topic, then remove or correct the confirmed duplicate record.",
}
REGISTERED_RULES = {
    "beneficiaries": ("Possible duplicate name","Possible duplicate contact and name","Invalid contact number","Case without assessment","Invalid age","Marital status below 18","Spouse below 18","Check Community Type vs Nationality"),
    "assessments": ("Beneficiary has multiple assessments","Selected month with previous assessment","Assessment without services","Pending assessment","Open counselling-only assessment","Blank legal service need","Detained beneficiary has counselling only","Adult representation without counselling","Type of Legal Service in Assessment vs Services","Detention/immigration inconsistency","Representation while not detained","Type of document in Assessments vs Services","Detained beneficiary below 10 years","Detention Governorate mismatch","Assessment date after today"),
    "legalservices": ("Duplicate service","Current and previous month duplicate","Orphaned assessment relationship","Missing Type of Document","Legal service date after today"),
    "awareness": ("Duplicate participant in session","Invalid contact number","Possible duplicate participant name"),
}
REVIEW_EXPORT_PROJECT_SHEETS = {
    "unhcr 2026 - erbil": "North Iraq",
    "unhcr 2026 - suli": "North Iraq",
    "unhcr 2026 - mosul & kirkuk": "North Iraq",
    "unhcr 2026 - amal camp": "AMAL Camp",
    "unhcr 2026 - baghdad": "South Iraq",
    "unhcr 2026 - gov": "South Iraq",
}
ASSESSMENT_PROJECT_GOVERNORATES = {
    "unhcr 2026 - suli": "sulaymaniyah",
    "unhcr 2026 - erbil": "erbil",
}
DETENTION_ASSESSMENT_RULES = frozenset((
    "Detained beneficiary has counselling only", "Detention/immigration inconsistency",
    "Detained beneficiary below 10 years", "Detention Governorate mismatch",
))
AMAL_HIDDEN_ASSESSMENT_RULES = DETENTION_ASSESSMENT_RULES | frozenset((
    "Representation while not detained",
))


def _find(columns: list[str], *needles: str) -> str | None:
    lowered = [(column, column.strip().lower()) for column in columns]
    for needle in needles:
        wanted = needle.lower()
        for original, lower in lowered:
            if lower == wanted:
                return original
    for needle in needles:
        wanted = needle.lower()
        for original, lower in lowered:
            if wanted in lower:
                return original
    return None


def clean_id(value: Any) -> str:
    if pd.isna(value): return ""
    text = str(value).strip()
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


def age_from_date(value: Any) -> int | None:
    parsed=pd.to_datetime(value,errors="coerce",dayfirst=True)
    if pd.isna(parsed): return None
    born=parsed.date();today=date.today()
    if born > today: return None
    return today.year-born.year-((today.month,today.day)<(born.month,born.day))


def normalize_name(value: Any) -> str:
    if pd.isna(value): return ""
    text = unicodedata.normalize("NFKC", str(value)).lower()
    text = re.sub(r"[\u064b-\u065f\u0670\u06d6-\u06ed\u0640]", "", text)
    text = text.translate(str.maketrans("أإآٱةىؤئكی۱۲۳۴۵۶۷۸۹۰١٢٣٤٥٦٧٨٩٠", "ااااهيويكي12345678901234567890"))
    return re.sub(r"[^\w\u0600-\u06ff]+", "", text, flags=re.UNICODE)


def split_multi_value(value: Any) -> list[str]:
    """Split the comma/semicolon-separated multi-select values used in Legal CSVs."""
    return [item.strip() for item in re.split(r"[,;]", clean_id(value)) if item.strip()]


def normalize_document_label(value: Any) -> str:
    """Match document labels despite punctuation and bilingual display suffixes."""
    text=unicodedata.normalize("NFKC",clean_id(value)).casefold()
    english=re.sub(r"[^a-z0-9]+"," ",text).strip()
    if english:return english
    text=re.sub(r"[\u064b-\u065f\u0670\u06d6-\u06ed\u0640]","",text)
    return re.sub(r"[^\u0600-\u06ff0-9]+"," ",text).strip()


def normalize_legal_service_type(value: Any) -> str:
    text=unicodedata.normalize("NFKC",clean_id(value)).casefold()
    if re.search(r"counselling|counseling|استشار",text):return "legal counselling"
    if re.search(r"assistance|مساعد",text):return "legal assistance"
    if re.search(r"representation|تمثيل",text):return "legal representation"
    return re.sub(r"[^a-z0-9\u0600-\u06ff]+"," ",text).strip()


def replace_legal_assistance(value: Any) -> Any:
    """Use the agreed Legal Representation label throughout the imported platform data."""
    if not isinstance(value,str):return value
    value=re.sub(r"legal assistance\s*-\s*مساعدة", "Legal Representation - تمثيل", value, flags=re.I)
    return re.sub(r"legal assistance", "Legal Representation", value, flags=re.I)


def normalize_governorate(value: Any) -> str:
    """Return one canonical Iraqi governorate for bilingual and spelling variants."""
    if pd.isna(value): return ""
    text=unicodedata.normalize("NFKC",str(value)).casefold()
    text=re.sub(r"[\u064b-\u065f\u0670\u06d6-\u06ed\u0640]","",text)
    text=re.sub(r"[^a-z\u0600-\u06ff]+"," ",text).strip()
    aliases={
        "baghdad":("baghdad","بغداد"),"diyala":("diyala","ديالى","ديالا"),
        "erbil":("erbil","arbil","hawler","اربيل","أربيل","هولير"),
        "sulaymaniyah":("sulaymaniyah","sulaymaniah","sulaimani","suli","السليمانية","سليمانية"),
        "duhok":("duhok","dohuk","دهوك"),"ninewa":("ninewa","ninawa","nineveh","نينوى"),
        "kirkuk":("kirkuk","كركوك"),"anbar":("anbar","الانبار","الأنبار"),
        "salah al din":("salah al din","salahaddin","salah ad din","صلاح الدين"),
        "basra":("basra","البصرة","بصرة"),"maysan":("maysan","ميسان"),
        "dhi qar":("dhi qar","thi qar","ذي قار"),"muthanna":("muthanna","المثنى"),
        "qadisiyah":("qadisiyah","qadissiya","diwaniyah","القادسية","الديوانية"),
        "babil":("babil","babylon","بابل"),"karbala":("karbala","كربلاء"),
        "najaf":("najaf","النجف"),"wasit":("wasit","واسط"),
    }
    for canonical,variants in aliases.items():
        if any(re.search(rf"(?<![a-z\u0600-\u06ff]){re.escape(variant)}(?![a-z\u0600-\u06ff])",text) for variant in variants):
            return canonical
    return re.sub(r"\s+"," ",text)


def normalize_comparison_date(value: Any) -> str:
    """Normalize dates, including raw Excel serial numbers, for reconciliation."""
    if pd.isna(value) or str(value).strip()=="": return ""
    if isinstance(value,(int,float)) and not isinstance(value,bool) and 1 <= float(value) <= 80000:
        parsed=pd.Timestamp("1899-12-30")+pd.to_timedelta(float(value),unit="D")
    else:
        parsed=pd.to_datetime(value,errors="coerce",dayfirst=True)
    return parsed.strftime("%Y-%m-%d") if not pd.isna(parsed) else str(value).strip().casefold()


def phone_digits(value: Any) -> str:
    if pd.isna(value): return ""
    return re.sub(r"\D", "", str(value).translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")))


def normalize_phone_value(value: Any) -> Any:
    """Keep source values intact except for adding the missing local zero to 10-digit numbers."""
    if pd.isna(value): return value
    digits = phone_digits(value)
    return f"0{digits}" if len(digits) == 10 else value


def display_value(value: Any) -> Any:
    if pd.isna(value): return ""
    # Legal Platform CSV dates are entered and reviewed as day/month/year.
    if isinstance(value, (pd.Timestamp, date)): return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer(): return int(value)
    return value


def format_excel_dates(workbook: Any) -> None:
    """Make every recognised date an Excel date with one consistent display format."""
    date_pattern=re.compile(r"^(?:\d{4}-\d{1,2}-\d{1,2}(?:[T\s].*)?|\d{1,2}[/-]\d{1,2}[/-]\d{4}(?:\s.*)?)$")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value=cell.value
                if isinstance(value,(pd.Timestamp,date)):
                    cell.value=pd.Timestamp(value).date()
                    cell.number_format="DD/MM/YYYY"
                elif isinstance(value,str) and date_pattern.match(value.strip()):
                    parsed=pd.to_datetime(value,errors="coerce",dayfirst=True)
                    if not pd.isna(parsed):
                        cell.value=parsed.date()
                        cell.number_format="DD/MM/YYYY"


def versioned_dataset_name(filename: str) -> tuple[str, int] | None:
    """Return the canonical dataset and numeric copy version for a CSV name."""
    stem = Path(filename).stem.strip().lower()
    match = VERSION_SUFFIX.search(stem)
    version = int(match.group(1)) if match else 0
    if match: stem = stem[:match.start()]
    canonical = re.sub(r"[\s_-]+", "", stem)
    return (canonical, version) if canonical in FILES else None


def _safe_export(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    result.columns = [safe_spreadsheet_value(column) for column in result.columns]
    for column in result.columns:
        if result[column].dtype == object:
            result[column] = result[column].map(safe_spreadsheet_value)
    return result


@dataclass
class LegalStore:
    frames: dict[str, pd.DataFrame]
    source: str
    warnings: list[str]
    dates: dict[str, list[str]] = field(default_factory=dict)
    flags: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    review_exclusions: list[dict[str, Any]] = field(default_factory=list)
    _metadata_cache: dict[str, Any] | None = field(default=None, init=False, repr=False)
    _deportation_dashboard_cache: dict[str, Any] | None = field(default=None, init=False, repr=False)
    _explorer_filter_cache: dict[str, dict[str, Any]] = field(default_factory=dict, init=False, repr=False)
    _case_filter_cache: dict[str, Any] | None = field(default=None, init=False, repr=False)
    _review_cache: dict[tuple[str, str, int, bool], dict[str, Any]] = field(default_factory=dict, init=False, repr=False)
    _lawyer_cache: dict[tuple[tuple[str, tuple[str, ...]], ...], dict[str, Any]] = field(default_factory=dict, init=False, repr=False)
    _intelligence_cache: dict[tuple[Any, ...], dict[str, Any]] = field(default_factory=dict, init=False, repr=False)
    _indicator_cache: dict[tuple[Any, ...], dict[str, Any]] = field(default_factory=dict, init=False, repr=False)
    _relation_cache: dict[tuple[str, str], dict[str, list[int]]] = field(default_factory=dict, init=False, repr=False)
    _search_cache: dict[str, pd.Series] = field(default_factory=dict, init=False, repr=False)
    _cache_lock: Lock = field(default_factory=Lock, init=False, repr=False)

    @classmethod
    def from_folder(cls, path: Path, progress: Callable[[int], None] | None = None) -> "LegalStore":
        selected: dict[str, tuple[int, Path]] = {}
        for file in path.glob("*.csv"):
            parsed = versioned_dataset_name(file.name)
            if not parsed: continue
            name, version = parsed
            current = selected.get(name)
            if current is None or version > current[0] or (version == current[0] and file.stat().st_mtime > current[1].stat().st_mtime):
                selected[name] = (version, file)
        payload: dict[str, bytes] = {}
        total = max(len(selected), 1)
        for index, (name, (_, file)) in enumerate(selected.items(), start=1):
            payload[name] = file.read_bytes()
            if progress: progress(round(index / total * 15))
        return cls.from_files(payload, path.name, progress)

    @classmethod
    def from_files(cls, payload: dict[str, bytes], source: str, progress: Callable[[int], None] | None = None) -> "LegalStore":
        missing = [name for name in MANDATORY if name not in payload]
        if missing: raise ValueError("Missing mandatory files: " + ", ".join(f"{x}.csv" for x in missing))
        frames: dict[str, pd.DataFrame] = {}
        dates: dict[str, list[str]] = {}
        warnings = [f"Optional file not loaded: {name}.csv" for name in OPTIONAL if name not in payload]
        supported = [(name, raw) for name, raw in payload.items() if name in FILES]
        for index, (name, raw) in enumerate(supported, start=1):
            try:
                df = pd.read_csv(io.BytesIO(raw), dtype=object, encoding="utf-8-sig", keep_default_na=True, low_memory=False)
            except UnicodeDecodeError:
                df = pd.read_csv(io.BytesIO(raw), dtype=object, encoding="cp1252", keep_default_na=True, low_memory=False)
            df.columns = [str(c).strip() or f"Column {i + 1}" for i, c in enumerate(df.columns)]
            absent = [c for c in REQUIRED_COLUMNS.get(name, ()) if c not in df.columns]
            if absent: raise ValueError(f"{name}.csv is missing required columns: {', '.join(absent)}")
            if name=="beneficiaries":
                consolidated=_find(list(df.columns),"Name (Filter Color Red)")
                parts=[]
                for part in ("First","Middle","Last"):
                    column=next((column for column in df.columns if SPLIT_NAME.match(column) and column.rstrip().lower().endswith(part.lower())),None)
                    if column:parts.append(column)
                if consolidated and parts:
                    composed=df[parts].fillna("").astype(str).apply(lambda row:" ".join(value.strip() for value in row if value.strip()),axis=1)
                    blank=df[consolidated].isna()|df[consolidated].astype(str).str.strip().eq("")
                    df.loc[blank,consolidated]=composed[blank]
            phone_column = _find(list(df.columns), "Contact Number" if name == "beneficiaries" else "Phone Number") if name in {"beneficiaries", "awareness"} else None
            if phone_column:
                df[phone_column] = df[phone_column].map(normalize_phone_value)
            drop = [c for c in df.columns if SPLIT_NAME.match(c) or (name == "legalservices" and SECURED_FILE in c and not c.rstrip().endswith(": URL"))]
            df = df.drop(columns=drop)
            # Keep the source CSV untouched while using the agreed terminology in
            # every platform view and every export generated from these frames.
            df = df.apply(lambda column: column.map(replace_legal_assistance) if column.dtype == object else column)
            date_cols = [c for c in df.columns if DATE_HINT.search(c)]
            request_date=_find(list(df.columns),"Date of the Request") if name=="assessments" else None
            for column in date_cols:
                # Assessment Date of the Request is the sole source field entered as month/day/year.
                parsed = pd.to_datetime(df[column], errors="coerce", dayfirst=column != request_date, format="mixed")
                supplied=df[column].notna() & df[column].astype(str).str.strip().ne("")
                # Keep malformed spouse dates available to the review rule instead of
                # converting them to missing values during general date normalization.
                if parsed.notna().any() and not (name=="beneficiaries" and "spouse dob" in column.casefold() and parsed[supplied].isna().any()): df[column] = parsed
            frames[name] = df
            dates[name] = date_cols
            if progress: progress(15 + round(index / max(len(supported), 1) * 55))
        result = cls(frames, source, warnings, dates)
        result.flags = result._build_flags()
        if progress: progress(78)
        # Build the expensive shared review contexts during import so opening a review page is immediate.
        review_datasets=[dataset for dataset in ("beneficiaries", "assessments", "legalservices", "awareness") if dataset in result.frames]
        for index, dataset in enumerate(review_datasets, start=1):
            result.review(dataset)
            if progress: progress(78 + round(index / max(len(review_datasets), 1) * 22))
        if progress: progress(95)
        return result

    def metadata(self) -> dict[str, Any]:
        if self._metadata_cache is not None: return self._metadata_cache
        all_flags=[row for rows in self.flags.values() for row in rows if not row.get("overviewExcluded")]
        severity=pd.Series([row["severity"] for row in all_flags],dtype=object).value_counts().to_dict()
        rules=pd.Series([row["rule"] for row in all_flags],dtype=object).value_counts().head(12).to_dict()
        lawyer_names:set[str]=set()
        for df in self.frames.values():
            lawyer=_find(list(df.columns),"Lawyers","Lawyer")
            if lawyer: lawyer_names.update(x for x in df[lawyer].fillna("").astype(str).str.strip() if x)
        assessments=self.frames["assessments"]
        services=self.frames["legalservices"]
        assessment_id=_find(list(assessments.columns),"Assessment ID")
        assessment_location=_find(list(assessments.columns),"Project Location","Project location")
        assessment_lawyer=_find(list(assessments.columns),"Lawyers","Lawyer")
        assessment_status=_find(list(assessments.columns),"Assessment Status")
        def grouped_counts(frame:pd.DataFrame,column:str|None,identifier:str|None)->list[dict[str,Any]]:
            if not column:return []
            values=frame[column].fillna("Unspecified").astype(str).str.strip().replace("","Unspecified")
            work=frame.assign(_overview_group=values)
            grouped=work.groupby("_overview_group")[identifier].agg(lambda items:items.map(clean_id).replace("",pd.NA).nunique()) if identifier else work.groupby("_overview_group").size()
            return [{"label":str(label),"count":int(count)} for label,count in grouped.sort_values(ascending=False).items()]
        overview_charts={
            "assessmentsByLocation":grouped_counts(assessments,assessment_location,assessment_id),
            "assessmentsByLawyer":grouped_counts(assessments,assessment_lawyer,assessment_id),
            "assessmentStatus":grouped_counts(assessments,assessment_status,assessment_id),
        }
        service_type=_find(list(services.columns),"Type of Service Provided","Type of Service")
        service_status=_find(list(services.columns),"Service Status","Status")
        service_id=_find(list(services.columns),"Service ID")
        representation=services[services[service_type].fillna("").astype(str).str.contains("representation",case=False,na=False)] if service_type else services.iloc[0:0]
        overview_charts["representationServiceStatus"]=grouped_counts(representation,service_status,service_id)
        detained_column=_find(list(assessments.columns),"Is the beneficiary detained")
        assessment_date=_find(list(assessments.columns),"Date of Assessment")
        detained_2026=assessments.iloc[0:0].copy()
        if detained_column and assessment_date:
            dates=pd.to_datetime(assessments[assessment_date],errors="coerce",dayfirst=True)
            detained=assessments[detained_column].fillna("").astype(str).str.contains(r"\byes\b|نعم",case=False,regex=True)
            detained_2026=assessments[detained&dates.ge(pd.Timestamp("2026-01-01"))&dates.lt(pd.Timestamp("2027-01-01"))].copy()
        current_status=_find(list(detained_2026.columns),"Detainee current status")
        release_date=_find(list(detained_2026.columns),"Date of the released or deported","Date of Released","Date of Release")
        release_rows=detained_2026[detained_2026[current_status].fillna("").astype(str).str.contains("released",case=False,na=False)] if current_status else detained_2026.iloc[0:0]
        if release_date:
            release_dates=pd.to_datetime(release_rows[release_date],errors="coerce",dayfirst=True)
            release_rows=release_rows[release_dates.ge(pd.Timestamp("2026-01-01"))&release_dates.lt(pd.Timestamp("2027-01-01"))]
        assessed_months=pd.to_datetime(detained_2026[assessment_date],errors="coerce",dayfirst=True).dt.to_period("M").value_counts() if assessment_date else pd.Series(dtype="int64")
        released_months=pd.to_datetime(release_rows[release_date],errors="coerce",dayfirst=True).dt.to_period("M").value_counts() if release_date else pd.Series(dtype="int64")
        periods=sorted(set(assessed_months.index).union(released_months.index))
        detention_trend=[{"month":period.strftime("%Y-%m"),"detainedAssessments":int(assessed_months.get(period,0)),"released":int(released_months.get(period,0))} for period in periods]
        governorate=_find(list(detained_2026.columns),"Detention Governorate")
        governorate_names={"anbar":"Al-Anbar","basra":"Al-Basrah","babil":"Babil","baghdad":"Baghdad","dhi qar":"Dhi Qar","diyala":"Diyala","duhok":"Dohuk","erbil":"Erbil","karbala":"Karbala","kirkuk":"Kirkuk","maysan":"Maysan","muthanna":"Al-Muthanna","najaf":"An-Najaf","ninewa":"Ninawa","qadisiyah":"Al-Qadisiyah","salah al din":"Salah al-Din","sulaymaniyah":"Al-Sulaimaniyah","wasit":"Wasit"}
        detention_map=[]
        if governorate:
            grouped:dict[str,dict[str,Any]]={}
            for raw,count in detained_2026[governorate].fillna("").astype(str).str.strip().value_counts().items():
                canonical=normalize_governorate(raw)
                label=governorate_names.get(canonical)
                if label:grouped[label]={"label":label,"count":int(count),"detained":int(count),"released":0,"values":[raw]}
            for raw,count in release_rows[governorate].fillna("").astype(str).str.strip().value_counts().items():
                canonical=normalize_governorate(raw)
                label=governorate_names.get(canonical)
                if not label:continue
                item=grouped.setdefault(label,{"label":label,"count":0,"detained":0,"released":0,"values":[]})
                item["released"]+=int(count)
                if raw not in item["values"]:item["values"].append(raw)
            detention_map=list(grouped.values())
        deportations=self.frames.get("deportationrecords",pd.DataFrame())
        deportation_governorate=_find(list(deportations.columns),"Governorate","Project Location","Project location")
        deportation_id=_find(list(deportations.columns),"PN ID","Deportation ID")
        deportation_rows=[]
        if deportation_governorate:
            grouped=deportations.assign(_governorate=deportations[deportation_governorate].map(normalize_governorate))
            grouped=grouped[grouped["_governorate"].ne("")]
            counts=grouped.groupby("_governorate")[deportation_id].agg(lambda values:values.map(clean_id).replace("",pd.NA).nunique()) if deportation_id else grouped.groupby("_governorate").size()
            deportation_rows=[{"label":governorate_names.get(str(name),str(name).title()),"count":int(count)} for name,count in counts.sort_values(ascending=False).items()]
        completed_representation=int(representation[service_status].fillna("").astype(str).str.contains("complete|closed|provided|done|اكتملت",case=False,regex=True).sum()) if service_status else 0
        representation_completion=completed_representation/max(len(representation),1)
        service_date=_find(list(services.columns),"Date of Service Provision","Date of Service")
        assessment_monthly=pd.to_datetime(assessments[assessment_date],errors="coerce",dayfirst=True).dt.to_period("M").value_counts() if assessment_date else pd.Series(dtype="int64")
        service_monthly=pd.to_datetime(services[service_date],errors="coerce",dayfirst=True).dt.to_period("M").value_counts() if service_date else pd.Series(dtype="int64")
        activity_periods=sorted(period for period in set(assessment_monthly.index).union(service_monthly.index) if pd.Period("2026-01",freq="M")<=period<pd.Period("2027-01",freq="M"))
        activity_trend=[{"month":period.strftime("%Y-%m"),"assessments":int(assessment_monthly.get(period,0)),"services":int(service_monthly.get(period,0))} for period in activity_periods]
        # Legal Assistance is reported as representation activity in the Overview.
        representation_activity=services[services[service_type].fillna("").astype(str).str.contains("representation|assistance",case=False,regex=True,na=False)] if service_type else services.iloc[0:0]
        representation_trend=[]
        if service_date:
            service_dates=pd.to_datetime(representation_activity[service_date],errors="coerce",dayfirst=True)
            monthly_activity=representation_activity.assign(_month=service_dates.dt.to_period("M"))
            monthly_activity=monthly_activity[monthly_activity["_month"].notna()]
            if service_id:
                monthly_counts=monthly_activity.groupby("_month")[service_id].agg(lambda values:values.map(clean_id).replace("",pd.NA).nunique())
            else:monthly_counts=monthly_activity.groupby("_month").size()
            representation_trend=[{"month":period.strftime("%Y-%m"),"representation":int(count)} for period,count in monthly_counts.sort_index().items() if pd.Period("2026-01",freq="M")<=period<pd.Period("2027-01",freq="M")]
        service_location=_find(list(services.columns),"Project Location","Project location")
        location_rows:dict[str,dict[str,Any]]={}
        def add_location_rows(frame:pd.DataFrame,column:str|None,metric:str)->None:
            if not column:return
            for location,count in frame[column].fillna("Unspecified").astype(str).str.strip().replace("","Unspecified").value_counts().items():
                row=location_rows.setdefault(str(location),{"location":str(location),"assessments":0,"representationServices":0,"detained":0,"released":0,"completionRate":0.0})
                row[metric]+=int(count)
        add_location_rows(assessments,assessment_location,"assessments")
        add_location_rows(representation,service_location,"representationServices")
        add_location_rows(detained_2026,assessment_location,"detained")
        add_location_rows(release_rows,assessment_location,"released")
        for location,row in location_rows.items():
            if service_location and service_status and row["representationServices"]:
                scoped=representation[representation[service_location].fillna("Unspecified").astype(str).str.strip().replace("","Unspecified").eq(location)]
                row["completionRate"]=float(scoped[service_status].fillna("").astype(str).str.contains("complete|closed|provided|done|اكتملت",case=False,regex=True).sum()/len(scoped))
        location_performance=sorted(location_rows.values(),key=lambda row:(row["assessments"],row["representationServices"],row["detained"]),reverse=True)
        latest_activity=activity_trend[-1] if activity_trend else None
        previous_activity=activity_trend[-2] if len(activity_trend)>1 else None
        change=(latest_activity["assessments"]-previous_activity["assessments"]) if latest_activity and previous_activity else None
        highest_detained=max(location_performance,key=lambda row:row["detained"],default=None)
        insight=(f"Assessments {'increased' if change>=0 else 'decreased'} by {abs(change):,} compared with {previous_activity['month']}." if change is not None else "Monthly assessment movement will appear once two reporting months are available.")
        if highest_detained and highest_detained["detained"]:insight+=f" {highest_detained['location']} has the highest 2026 detention caseload ({highest_detained['detained']:,})."
        gender=_find(list(assessments.columns),"Gender")
        age_group=_find(list(assessments.columns),"UNHCR Age Group","Age group")
        def missing_count(column:str|None)->int:return int(assessments[column].isna().sum()+assessments[column].fillna("").astype(str).str.strip().eq("").sum()) if column else 0
        quality=[{"label":"Missing lawyer","count":missing_count(assessment_lawyer)},{"label":"Missing project location","count":missing_count(assessment_location)},{"label":"Missing assessment status","count":missing_count(assessment_status)},{"label":"Missing gender or age group","count":int(((assessments[gender].isna()|assessments[gender].fillna("").astype(str).str.strip().eq(""))|(assessments[age_group].isna()|assessments[age_group].fillna("").astype(str).str.strip().eq(""))).sum()) if gender and age_group else 0}]
        project_values:set[str]=set()
        for frame in self.frames.values():
            project_column=_find(list(frame.columns),"Projects -","Project")
            if project_column:
                project_values.update(value for value in frame[project_column].fillna("").astype(str).str.strip() if value)
        # Detention eligibility is driven by the Assessments dataset. Auxiliary
        # files can legitimately retain records from other projects.
        amal_only=self._amal_only_assessment_projects()
        self._metadata_cache = {
            "ready": True, "source": self.source, "warnings": self.warnings,
            "availability": {name: name in self.frames for name in FILES},
            "features":{"awareness":any(re.search(r"\bamal\b",value,flags=re.I) for value in project_values),"detention":not amal_only,"deportation":"deportationrecords" in self.frames and not amal_only},
            "sheets": [{"id": name, "name": DISPLAY_NAMES[name], "rows": len(df), "columns": [str(c) for c in df.columns]} for name, df in self.frames.items()],
            "months": self._months(),
            "reviewCounts": {name: sum(not row.get("overviewExcluded",False) for row in rows) for name, rows in self.flags.items()},
            "overview": {"beneficiaries":len(self.frames["beneficiaries"]),"assessments":len(self.frames["assessments"]),"services":len(self.frames["legalservices"]),
                         "followups":len(self.frames["followupslogbooks"]) if "followupslogbooks" in self.frames else None,
                         "fees":len(self.frames["legalfees"]) if "legalfees" in self.frames else None,
                         "awareness":len(self.frames["awareness"]) if "awareness" in self.frames else None,
                         "deportations":len(self.frames["deportationrecords"]) if "deportationrecords" in self.frames else None,
                         "lawyers":len(lawyer_names),"totalFlags":len(all_flags),"severity":severity,"rules":rules,
                         "charts":overview_charts,"representationCompletionRate":representation_completion,"activityTrend":activity_trend,"representationTrend":representation_trend,"locationPerformance":location_performance,"insight":insight,"dataQuality":quality,"deportationsByGovernorate":deportation_rows,"detention2026":{"trend":detention_trend,"map":detention_map}},
        }
        return self._metadata_cache

    def deportation_dashboard(self, filters:dict[str,list[str]]|None=None) -> dict[str, Any]:
        if not filters and self._deportation_dashboard_cache is not None:return self._deportation_dashboard_cache
        source=self.frames.get("deportationrecords")
        if source is None: raise ValueError("deportationrecords.csv is not loaded.")
        source_ident=_find(list(source.columns),"PN ID","Deportation ID")
        source_date=_find(list(source.columns),"Date Of Deportation Knowledge","Date of Deportation Knowledge","Date of deporting","Deportation Date")
        filter_columns=[column for column in source.columns if source[column].nunique(dropna=True)<=200]
        filter_options={"__reviewStyle":["true"]}
        for column in filter_columns:
            if DATE_HINT.search(str(column)):
                months=pd.to_datetime(source[column],errors="coerce",dayfirst=True).dt.to_period("M").astype(str)
                filter_options[str(column)]=sorted(month for month in months.unique() if month!="NaT")
            else:
                filter_options[str(column)]=sorted({str(value).strip() for value in source[column].dropna() if str(value).strip()})
        df=source
        for column,values in (filters or {}).items():
            if not values:continue
            if column not in df.columns:continue
            if DATE_HINT.search(str(column)):
                months=pd.to_datetime(df[column],errors="coerce",dayfirst=True).dt.to_period("M").astype(str)
                df=df[months.isin(values)]
            else:df=df[df[column].fillna("").astype(str).isin(values)]
        ident=_find(list(df.columns),"PN ID","Deportation ID")
        date_col=source_date
        fields=[("Governorate","Governorate","Project Location","Project location"),("Destination","Destination","Country of destination","Deported to"),("Nationality","Nationality"),("Authority","Authority","Detaining Authority"),("Project","Project","Projects -")]
        metric=df[ident].map(clean_id).replace("",pd.NA) if ident else pd.Series(df.index.astype(str),index=df.index)
        total=int(metric.nunique())
        charts=[]
        for title,*hints in fields:
            column=_find(list(df.columns),*hints)
            if not column: continue
            values=df[column].fillna("").astype(str).str.strip()
            working=pd.DataFrame({"label":values,"metric":metric});working=working[working.label.ne("")]
            counts=working.groupby("label").metric.nunique().sort_values(ascending=False).head(12)
            charts.append({"id":column,"title":f"Deportations by {title.lower()}","kind":"bar","multiChoice":False,"rows":[{"label":str(label),"count":int(count),"percent":int(count)/total if total else 0} for label,count in counts.items()]})
        trend=[]
        if date_col:
            months=pd.to_datetime(df[date_col],errors="coerce",dayfirst=True).dt.to_period("M")
            working=df.assign(_month=months); working=working[working._month.notna()]
            working=working.assign(_metric=metric.loc[working.index]);counts=working.groupby("_month")._metric.nunique()
            trend=[{"label":str(month),"count":int(count),"percent":int(count)/total if total else 0} for month,count in counts.items()]
        kpis=[("Deportation records",total),("Destinations",next((len(chart["rows"]) for chart in charts if "destination" in chart["title"]),0)),("Nationalities",next((len(chart["rows"]) for chart in charts if "nationality" in chart["title"]),0)),("Authorities",next((len(chart["rows"]) for chart in charts if "authority" in chart["title"]),0))]
        result={"page":"deportation","measure":"records","total":total,"filteredRows":len(df),"kpis":[{"label":label,"value":value,"format":"number"} for label,value in kpis],"trend":trend,"charts":charts,"flow":[],"filterOptions":filter_options}
        if not filters:self._deportation_dashboard_cache=result
        return result

    def _months(self) -> list[str]:
        values: set[str] = set()
        for name in ("assessments", "legalservices"):
            if name not in self.frames: continue
            column = _find(list(self.frames[name].columns), "Date of Assessment", "Date of Service Provision")
            if column:
                series = pd.to_datetime(self.frames[name][column], errors="coerce")
                values.update(series.dropna().dt.strftime("%Y-%m"))
        return sorted(values)

    @staticmethod
    def _flag(rows: list[dict[str, Any]], dataset: str, rule: str, severity: str, index: Any, row: pd.Series, detail: str) -> None:
        primary = _find(list(row.index), "Case ID", "Assessment ID", "Service ID", "Awareness ID", "Beneficiary ID")
        name=_find(list(row.index),"Name (Filter Color Red)","Participant Name","Name / الأسم")
        case_id=_find(list(row.index),"Case ID","Beneficiary ID")
        assessment_id=_find(list(row.index),"Assessment ID")
        service_id=_find(list(row.index),"Service ID")
        lawyer=_find(list(row.index),"Lawyers","Lawyer","Created By","Created by")
        project=_find(list(row.index),"Projects -","Project")
        location=_find(list(row.index),"Project Location","Project location")
        phone=_find(list(row.index),"Contact Number","Phone Number")
        dob=_find(list(row.index),"Date of Birth","DoB")
        awareness_id=_find(list(row.index),"Awareness ID")
        session_topic=_find(list(row.index),"Session Topic","Topic")
        spouse_name=_find(list(row.index),"Spouse name")
        spouse_dob=_find(list(row.index),"Spouse DoB")
        spouse_age=age_from_date(row.get(spouse_dob,"")) if spouse_dob else None
        marital_status=_find(list(row.index),"Marital Statues","Marital Status")
        assessment_date=_find(list(row.index),"Date of Assessment")
        identification_date=_find(list(row.index),"Date of Identification")
        awareness_date=_find(list(row.index),"Date of Session","Added On")
        created_on=_find(list(row.index),"Created On")
        assessment_status=_find(list(row.index),"Assessment Status")
        legal_service_needed=_find(list(row.index),"Type of Legal Service Needed")
        detained=_find(list(row.index),"Is the beneficiary detained")
        immigration_charge=_find(list(row.index),"Is it an immigration related charge")
        service_type=_find(list(row.index),"Type of Service Provided")
        document_type=_find(list(row.index),"Type of Document")
        court_verdict_detail=_find(list(row.index),"Please specify the Court Verdict")
        other_document_detail=_find(list(row.index),'Type of Document if "Other" please specify')
        legal_concern_specified=_find(list(row.index),"Legal Concern Specified")
        legal_concern=_find(list(row.index),"Legal Concern")
        detention_governorate=_find(list(row.index),"Detention Governorate")
        service_date=_find(list(row.index),"Date of Service Provision")
        rows.append({"dataset": dataset, "rule": rule, "severity": severity, "row": int(index) + 2,
                     "recordId": clean_id(row.get(primary, "")) if primary else "", "name":clean_id(row.get(name,"")) if name else "",
                     "caseId":clean_id(row.get(case_id,"")) if case_id else "", "assessmentId":clean_id(row.get(assessment_id,"")) if assessment_id else "",
                     "serviceId":clean_id(row.get(service_id,"")) if service_id else "", "detail": detail,
                     "lawyer":clean_id(row.get(lawyer,"")) if lawyer else "", "project":clean_id(row.get(project,"")) if project else "",
                     "location":clean_id(row.get(location,"")) if location else "", "phone":clean_id(row.get(phone,"")) if phone else "",
                     "dateOfBirth":display_value(row.get(dob,"")) if dob else "", "beneficiaryAge":age_from_date(row.get(dob,"")) if dob else None, "awarenessId":clean_id(row.get(awareness_id,"")) if awareness_id else "",
                     "sessionTopic":clean_id(row.get(session_topic,"")) if session_topic else "",
                     "spouseName":clean_id(row.get(spouse_name,"")) if spouse_name else "", "spouseDateOfBirth":display_value(row.get(spouse_dob,"")) if spouse_dob else "", "spouseAge":spouse_age,
                     "maritalStatus":clean_id(row.get(marital_status,"")) if marital_status else "",
                     "assessmentDate":display_value(row.get(assessment_date,"")) if assessment_date else "", "identificationDate":display_value(row.get(identification_date,"")) if identification_date else "", "awarenessDate":display_value(row.get(awareness_date,"")) if awareness_date else "", "createdOn":display_value(row.get(created_on,"")) if created_on else "", "assessmentStatus":clean_id(row.get(assessment_status,"")) if assessment_status else "", "legalServiceNeeded":clean_id(row.get(legal_service_needed,"")) if legal_service_needed else "", "beneficiaryDetained":clean_id(row.get(detained,"")) if detained else "", "immigrationRelatedCharge":clean_id(row.get(immigration_charge,"")) if immigration_charge else "", "serviceTypeProvided":clean_id(row.get(service_type,"")) if service_type else "", "typeOfDocument":clean_id(row.get(document_type,"")) if document_type else "", "courtVerdictDetail":clean_id(row.get(court_verdict_detail,"")) if court_verdict_detail else "", "otherDocumentDetail":clean_id(row.get(other_document_detail,"")) if other_document_detail else "", "legalConcernSpecified":clean_id(row.get(legal_concern_specified,"")) if legal_concern_specified else "", "legalConcern":clean_id(row.get(legal_concern,"")) if legal_concern else "", "detentionGovernorate":clean_id(row.get(detention_governorate,"")) if detention_governorate else "",
                     "serviceDate":display_value(row.get(service_date,"")) if service_date else "",
                     "action":ACTIONS.get(rule,"Review the source record, verify the information, and document the correction.")})

    def _build_flags(self) -> dict[str, list[dict[str, Any]]]:
        return {
            "beneficiaries": self._beneficiary_flags(), "assessments": self._assessment_flags(),
            "legalservices": self._service_flags(),
            **({"awareness": self._awareness_flags()} if "awareness" in self.frames else {}),
        }

    def set_review_exclusions(self, exclusions: Any) -> None:
        normalized=[]
        for entry in exclusions:
            if isinstance(entry, tuple):
                rule, value=entry; item={"dataset":"beneficiaries","rule":str(rule).strip(),"identifierType":"caseId","identifierValue":clean_id(value)}
            else:
                item={"dataset":str(entry.get("dataset", "beneficiaries")).strip(),"rule":str(entry.get("rule", "")).strip(),"identifierType":str(entry.get("identifierType", "caseId")).strip(),"identifierValue":str(entry.get("identifierValue", entry.get("caseId", ""))).strip()}
                if item["identifierType"] == "awarenessName": item["identifierValue"]=" ".join(item["identifierValue"].casefold().split())
                else: item["identifierValue"]=clean_id(item["identifierValue"])
            if all(item.values()): normalized.append(item)
        if normalized == self.review_exclusions:
            return
        self.review_exclusions = normalized
        # Exclusions affect only the duplicate-name result.  Defer that
        # calculation to the next duplicate-table request instead of making
        # the local save wait for all review data to be rebuilt.
        for key in list(self._review_cache):
            del self._review_cache[key]

    def _excluded_case_ids(self, rule: str) -> set[str]:
        return {row["identifierValue"] for row in self.review_exclusions if row["dataset"] == "beneficiaries" and row["rule"] == rule and row["identifierType"] == "caseId"}

    def _is_excluded(self, row: dict[str, Any]) -> bool:
        fields={"caseId":clean_id(row.get("caseId", "")),"assessmentId":clean_id(row.get("assessmentId", "")),"serviceId":clean_id(row.get("serviceId", "")),"awarenessId":clean_id(row.get("awarenessId", "")),"awarenessName":" ".join(str(row.get("name", "")).casefold().split())}
        return any(entry["dataset"] == row.get("dataset") and entry["rule"] == row.get("rule") and fields.get(entry["identifierType"], "") == entry["identifierValue"] for entry in self.review_exclusions)

    def _beneficiary_flags(self) -> list[dict[str, Any]]:
        df = self.frames["beneficiaries"]; out: list[dict[str, Any]] = self._name_match_flags(excluded_case_ids=self._excluded_case_ids("Possible duplicate name"))
        out += self._contact_name_match_flags(excluded_case_ids=self._excluded_case_ids("Possible duplicate contact and name"))
        phone = _find(list(df.columns), "Contact Number")
        if phone:
            ignored_prefixes = ("54","55","15","93","94","95","96","98","99","62","46","89","6939","4915","2376","9054","2951","4916")
            for i, value in df[phone].items():
                digits = phone_digits(value)
                if not digits or len(digits) == 1 or len(digits) == 11 or digits.startswith(ignored_prefixes): continue
                self._flag(out, "beneficiaries", "Invalid contact number", "Medium", i, df.loc[i], f"{len(digits)} digits")
        total = _find(list(df.columns), "# total assessments")
        if total:
            for i in df.index[pd.to_numeric(df[total], errors="coerce").fillna(0).eq(0)]: self._flag(out,"beneficiaries","Case without assessment","High",i,df.loc[i],"# total assessments is 0")
        age = _find(list(df.columns), "Age")
        if age:
            numeric = pd.to_numeric(df[age], errors="coerce")
            bad = df[age].isna() | numeric.isna() | numeric.lt(0) | numeric.gt(110)
            for i in df.index[bad]: self._flag(out,"beneficiaries","Invalid age","High",i,df.loc[i],"Age is blank, non-numeric, below 0, or above 110")
            marital = _find(list(df.columns), "Marital Statues", "Marital Status")
            if marital:
                partnered = df[marital].fillna("").astype(str).str.lower().str.contains("married|divorc|separat|widow|متزوج|منفصل|ارمل", regex=True)
                dob_column=_find(list(df.columns),"Date of Birth","DoB")
                current_ages=df[dob_column].map(age_from_date) if dob_column else pd.Series(None,index=df.index,dtype=object)
                effective_ages=pd.Series([int(current_ages[i]) if pd.notna(current_ages[i]) else (int(numeric[i]) if pd.notna(numeric[i]) else None) for i in df.index],index=df.index)
                under_18=effective_ages.map(lambda value:value is not None and value < 18)
                for i in df.index[partnered & under_18]: self._flag(out,"beneficiaries","Marital status below 18","Warning",i,df.loc[i],f"Current age is {effective_ages[i]} based on date of birth" if pd.notna(current_ages[i]) else f"Recorded age is {effective_ages[i]}; date of birth is unavailable")
                spouse_dob = _find(list(df.columns), "Spouse DoB")
                if spouse_dob:
                    spouse_age=df[spouse_dob].map(age_from_date)
                    spouse_under_18=spouse_age.map(lambda value:value is not None and value < 18)
                    for i in df.index[partnered & spouse_under_18]: self._flag(out,"beneficiaries","Spouse below 18","Warning",i,df.loc[i],f"Spouse is {spouse_age[i]} years old based on date of birth")
        spouse_dob=_find(list(df.columns),"Spouse DoB")
        if spouse_dob:
            spouse_values=df[spouse_dob]
            spouse_supplied=spouse_values.notna() & spouse_values.astype(str).str.strip().ne("")
            parsed_spouse_dob=pd.to_datetime(spouse_values,errors="coerce",dayfirst=True)
            for i in df.index[spouse_supplied & parsed_spouse_dob.isna()]:
                self._flag(out,"beneficiaries","Invalid age","High",i,df.loc[i],"Spouse DoB is not a valid date")
            for i in df.index[spouse_supplied & parsed_spouse_dob.notna() & parsed_spouse_dob.gt(pd.Timestamp.today().normalize())]:
                self._flag(out,"beneficiaries","Invalid age","High",i,df.loc[i],"Spouse DoB is later than the current date")
        community = _find(list(df.columns), "Community Type")
        nationality = _find(list(df.columns), "Nationality")
        project = _find(list(df.columns), "Project")
        for i, row in df.iterrows():
            community_value = clean_id(row.get(community, "")) if community else ""
            nationality_value = clean_id(row.get(nationality, "")) if nationality else ""
            project_value = clean_id(row.get(project, "")) if project else ""
            normalized_community = normalize_name(community_value)
            normalized_nationality = normalize_name(nationality_value)
            normalized_project = re.sub(r"\s+", " ", project_value).strip().casefold()
            problems: list[str] = []
            if not community_value: problems.append("Community Type is blank")
            if not nationality_value: problems.append("Nationality is blank")
            if not project_value: problems.append("Project is blank")
            is_idp = "idp" in normalized_community or "نازح" in normalized_community
            is_non_syrian_refugee = "nonsyrianrefugee" in normalized_community or ("غيرسوري" in normalized_community and "لاج" in normalized_community)
            is_syrian_refugee = not is_non_syrian_refugee and ("syrianrefugee" in normalized_community or ("سوري" in normalized_community and "لاج" in normalized_community))
            is_iraqi = "iraq" in normalized_nationality or "عراق" in normalized_nationality
            is_syrian = "syria" in normalized_nationality or "سوري" in normalized_nationality
            if community_value and not (is_idp or is_syrian_refugee or is_non_syrian_refugee):
                problems.append(f"Unexpected Community Type: {community_value}")
            if is_idp:
                if not is_iraqi: problems.append("IDP Community Type requires Iraqi nationality")
                if normalized_project != "unhcr 2026 - amal camp": problems.append("IDP Community Type is permitted only for AMAL Camp")
            elif is_syrian_refugee and not is_syrian:
                problems.append("Syrian Refugee Community Type requires Syrian nationality")
            elif is_non_syrian_refugee and (is_iraqi or is_syrian):
                problems.append("Non-Syrian Refugee Community Type cannot have Iraqi or Syrian nationality")
            if problems:
                self._flag(out, "beneficiaries", "Check Community Type vs Nationality", "High", i, row, "; ".join(problems))
        return out

    def _name_match_flags(self, compare_chars: int = 15, allow_variations: bool = False, exact_only: bool = False, excluded_case_ids: set[str] | None = None) -> list[dict[str, Any]]:
        df=self.frames["beneficiaries"]; name=_find(list(df.columns),"Name (Filter Color Red)");project=_find(list(df.columns),"Projects -","Project");case_id=_find(list(df.columns),"Case ID")
        if not name or not project: return []
        bounded_chars=max(10,min(30,compare_chars))
        normalized=df[name].map(normalize_name)
        excluded = {clean_id(case) for case in (excluded_case_ids or set()) if clean_id(case)}
        project_groups={
            "unhcr 2026 - erbil":"North - Erbil",
            "unhcr 2026 - mosul & kirkuk":"North - Mosul & Kirkuk",
            "unhcr 2026 - suli":"North - SULI",
            "unhcr 2026 - baghdad":"South (Baghdad + Gov)",
            "unhcr 2026 - gov":"South (Baghdad + Gov)",
            "unhcr 2026 - amal camp":"AMAL",
        }
        normalized_projects=df[project].fillna("").astype(str).map(lambda value:re.sub(r"\s+"," ",value).strip().casefold())
        south_projects={"unhcr 2026 - baghdad","unhcr 2026 - gov"}
        eligible={
            int(i):("South Iraq" if normalized_projects[i] in south_projects else (normalized_projects[i] or f"Unspecified project {i}"),text)
            for i,text in normalized.items() if text and (not case_id or clean_id(df.loc[i, case_id]) not in excluded)
        } if exact_only else {int(i):(project_groups.get(normalized_projects[i],""),text[:bounded_chars]) for i,text in normalized.items() if len(text)>=bounded_chars and normalized_projects[i] in project_groups and (not case_id or clean_id(df.loc[i, case_id]) not in excluded)}
        matches:dict[int,set[int]]=defaultdict(set)
        if exact_only or not allow_variations:
            groups:dict[tuple[str,str],list[int]]=defaultdict(list)
            for i,(project_group,prefix) in eligible.items(): groups[(project_group,prefix)].append(i)
            for indexes in groups.values():
                if len(indexes)<2: continue
                for i in indexes: matches[i].update(j for j in indexes if j!=i)
        else:
            buckets:dict[tuple[str,str],list[tuple[int,str]]]=defaultdict(list)
            for i,(project_group,prefix) in eligible.items():
                bucket=(project_group,prefix[:2])
                for j,other in buckets[bucket]:
                    if SequenceMatcher(None,prefix,other).ratio()>=0.90:
                        matches[i].add(j);matches[j].add(i)
                buckets[bucket].append((i,prefix))
        out=[]
        mode="with small spelling differences allowed" if allow_variations else "exactly"
        components:dict[int,str]={};visited:set[int]=set()
        for start in matches:
            if start in visited: continue
            stack=[start];members:set[int]=set()
            while stack:
                current=stack.pop()
                if current in members: continue
                members.add(current);stack.extend(matches[current])
            visited.update(members);group_key=f"{eligible[start][0]}:{min(members)}"
            for member in members: components[member]=group_key
        for i, peers in matches.items():
            project_group=eligible[i][0]
            detail=f"{project_group}: full normalized name exactly matches {len(peers)} other record(s)" if exact_only else f"{project_group}: first {bounded_chars} normalized characters match {mode} with {len(peers)} other record(s)"
            self._flag(out,"beneficiaries","Possible duplicate name","High",i,df.loc[i],detail)
            out[-1]["duplicateGroup"]=components[i]
            out[-1]["nameMatchMode"]="exact" if any(normalized[peer] == normalized[i] for peer in peers) else "variation"
            out[-1]["duplicateSimilarity"]=round(max(SequenceMatcher(None,normalized[i],normalized[peer]).ratio() for peer in peers)*100)
        return out

    def _contact_name_match_flags(self, excluded_case_ids: set[str] | None = None) -> list[dict[str, Any]]:
        """Flag likely duplicate beneficiaries with the same contact and a near-identical name."""
        df=self.frames["beneficiaries"]
        name=_find(list(df.columns),"Name (Filter Color Red)")
        project=_find(list(df.columns),"Projects -","Project")
        contact=_find(list(df.columns),"Contact Number")
        case_id=_find(list(df.columns),"Case ID")
        marker_columns=[column for column in (
            _find(list(df.columns),"# UNHCR"),
            _find(list(df.columns),"ID Number"),
            _find(list(df.columns),"Individual Number ASSISIT"),
            _find(list(df.columns),"Spouse name"),
        ) if column]
        if not name or not project or not contact or not marker_columns:
            return []
        excluded={clean_id(case) for case in (excluded_case_ids or set()) if clean_id(case)}
        project_groups={
            "unhcr 2026 - erbil":"North - Erbil",
            "unhcr 2026 - mosul & kirkuk":"North - Mosul & Kirkuk",
            "unhcr 2026 - suli":"North - SULI",
            "unhcr 2026 - baghdad":"South (Baghdad + Gov)",
            "unhcr 2026 - gov":"South (Baghdad + Gov)",
            "unhcr 2026 - amal camp":"AMAL",
        }
        normalized_names=df[name].map(normalize_name)
        normalized_projects=df[project].fillna("").astype(str).map(lambda value:re.sub(r"\s+"," ",value).strip().casefold())
        contacts=df[contact].map(phone_digits)
        eligible:dict[int,tuple[str,str,str]]={}
        for i in df.index:
            normalized_name=normalized_names[i]
            project_group=project_groups.get(normalized_projects[i],"")
            contact_number=contacts[i]
            has_marker=any(clean_id(df.loc[i,column]) for column in marker_columns)
            is_excluded=case_id and clean_id(df.loc[i,case_id]) in excluded
            if normalized_name and contact_number and project_group and has_marker and not is_excluded:
                eligible[int(i)]=(project_group,contact_number,normalized_name)
        groups:dict[tuple[str,str],list[int]]=defaultdict(list)
        for i,(project_group,contact_number,_) in eligible.items():
            groups[(project_group,contact_number)].append(i)
        matches:dict[int,set[int]]=defaultdict(set)
        for indexes in groups.values():
            if len(indexes)<2:
                continue
            for position,i in enumerate(indexes):
                for j in indexes[position+1:]:
                    if SequenceMatcher(None,eligible[i][2],eligible[j][2]).ratio() >= .90:
                        matches[i].add(j)
                        matches[j].add(i)
        components:dict[int,str]={};visited:set[int]=set()
        for start in matches:
            if start in visited:
                continue
            stack=[start];members:set[int]=set()
            while stack:
                current=stack.pop()
                if current in members:
                    continue
                members.add(current)
                stack.extend(matches[current])
            visited.update(members)
            group_key=f"contact-name:{eligible[start][0]}:{eligible[start][1]}:{min(members)}"
            for member in members:
                components[member]=group_key
        out=[]
        for i,peers in matches.items():
            project_group,contact_number,normalized_name=eligible[i]
            strongest=max(round(SequenceMatcher(None,normalized_name,eligible[peer][2]).ratio()*100) for peer in peers)
            self._flag(out,"beneficiaries","Possible duplicate contact and name","High",i,df.loc[i],f"{project_group}: contact number {contact_number} matches {len(peers)} other record(s); strongest normalized-name similarity is {strongest}%")
            out[-1]["duplicateGroup"]=components[i]
            out[-1]["nameMatchMode"]="contact-and-name"
            out[-1]["duplicateSimilarity"]=strongest
        return out

    def _assessment_reconciliation_flags(self, df: pd.DataFrame, services: pd.DataFrame) -> list[dict[str, Any]]:
        assessment_id=_find(list(df.columns),"Assessment ID");service_assessment=_find(list(services.columns),"Assessment ID")
        if not assessment_id or not service_assessment:return []
        assessment_date=_find(list(df.columns),"Date of Assessment")
        assessment_dates=pd.to_datetime(df[assessment_date],errors="coerce",dayfirst=True) if assessment_date else pd.Series(pd.NaT,index=df.index)
        assessment_documents=_find(list(df.columns),"Type of Documents to be issued");service_documents=_find(list(services.columns),"Type of Document")
        service_needed=_find(list(df.columns),"Type of Legal Service Needed");service_provided=_find(list(services.columns),"Type of Service Provided")
        linked_documents:dict[str,list[str]]=defaultdict(list);linked_services:dict[str,list[str]]=defaultdict(list)
        for _,service in services.iterrows():
            linked_id=clean_id(service.get(service_assessment,""))
            if not linked_id:continue
            if service_documents:linked_documents[linked_id].extend(split_multi_value(service.get(service_documents,"")))
            if service_provided:linked_services[linked_id].extend(split_multi_value(service.get(service_provided,"")))
        out=[]
        for i,row in df.iterrows():
            current_id=clean_id(row.get(assessment_id,""))
            if not current_id or pd.isna(assessment_dates[i]) or assessment_dates[i] < pd.Timestamp("2026-01-01"):continue
            requested_documents=split_multi_value(row.get(assessment_documents,"")) if assessment_documents else []
            delivered_documents=linked_documents.get(current_id,[])
            requested_document_map={normalize_document_label(item):item for item in requested_documents if normalize_document_label(item)}
            delivered_document_map={normalize_document_label(item):item for item in delivered_documents if normalize_document_label(item)}
            requested_display=", ".join(dict.fromkeys(requested_documents));delivered_display=", ".join(dict.fromkeys(delivered_documents))
            missing_in_services=[raw for key,raw in requested_document_map.items() if key not in delivered_document_map]
            missing_in_assessment=[raw for key,raw in delivered_document_map.items() if key not in requested_document_map]
            for finding,missing in (("Missing Type of Document in Services",missing_in_services),("Missing Type of Document in Assessment",missing_in_assessment)):
                if not missing:continue
                self._flag(out,"assessments","Type of document in Assessments vs Services","Medium",i,row,f"{finding}: {', '.join(missing)}")
                out[-1].update({"comparisonFinding":finding,"assessmentDocuments":requested_display,"serviceDocuments":delivered_display,"missingValues":", ".join(missing)})
            requested_types=split_multi_value(row.get(service_needed,"")) if service_needed else []
            provided_types=linked_services.get(current_id,[])
            requested_type_map={normalize_legal_service_type(item):item for item in requested_types if normalize_legal_service_type(item)}
            provided_type_keys={normalize_legal_service_type(item) for item in provided_types if normalize_legal_service_type(item)}
            missing_types=[raw for key,raw in requested_type_map.items() if key not in provided_type_keys]
            if missing_types:
                self._flag(out,"assessments","Type of Legal Service in Assessment vs Services","Medium",i,row,f"Missing linked legal service type(s): {', '.join(missing_types)}")
                out[-1].update({"requestedServiceTypes":", ".join(dict.fromkeys(requested_types)),"providedServiceTypes":", ".join(dict.fromkeys(provided_types)),"missingValues":", ".join(missing_types)})
        return out

    def _future_date_flags(self, dataset: str, rule: str, fields: tuple[str, ...]) -> list[dict[str, Any]]:
        df=self.frames[dataset]; today=pd.Timestamp(date.today()); columns=[(label,_find(list(df.columns),label)) for label in fields]; out=[]
        for i,row in df.iterrows():
            future=[]
            for label,column in columns:
                if not column: continue
                value=pd.to_datetime(row.get(column),errors="coerce",dayfirst=True)
                if not pd.isna(value) and value.normalize()>today:
                    future.append(f"{label}: {value.strftime('%d/%m/%Y')}")
            if future:self._flag(out,dataset,rule,"Medium",i,row,"; ".join(future))
        return out

    def _assessment_flags(self) -> list[dict[str, Any]]:
        df = self.frames["assessments"]; services = self.frames["legalservices"]; out: list[dict[str, Any]] = self._assessment_reconciliation_flags(df,services)
        out += self._future_date_flags("assessments","Assessment date after today",("Date of Assessment","Date of the released or deported","Date of Detention","Date of Assessment Closure","Date of the Request"))
        beneficiary = _find(list(df.columns), "Beneficiary ID"); assessment = _find(list(df.columns), "Assessment ID")
        assessment_date=_find(list(df.columns),"Date of Assessment"); status=_find(list(df.columns), "Assessment Status")
        if beneficiary:
            ids=df[beneficiary].map(clean_id)
            assessment_months=pd.to_datetime(df[assessment_date],errors="coerce",dayfirst=True).dt.to_period("M") if assessment_date else pd.Series(pd.NaT,index=df.index)
            same_month_counts=pd.Series(list(zip(ids,assessment_months)),index=df.index)
            same_month_counts=same_month_counts[(ids.ne("")) & assessment_months.notna()].value_counts()
            open_status=status and df[status].fillna("").astype(str).str.contains("open",case=False,regex=False)
            open_counts=ids[open_status & ids.ne("")].value_counts() if isinstance(open_status,pd.Series) else pd.Series(dtype="int64")
            for i in df.index:
                reasons=[]
                month_count=int(same_month_counts.get((ids[i],assessment_months[i]),0)) if ids[i] and pd.notna(assessment_months[i]) else 0
                if month_count>=2: reasons.append(f"{month_count} assessments in {assessment_months[i]}")
                open_count=int(open_counts.get(ids[i],0)) if ids[i] else 0
                if isinstance(open_status,pd.Series) and bool(open_status[i]) and open_count>=2: reasons.append(f"{open_count} Open assessments")
                if reasons:
                    self._flag(out,"assessments","Beneficiary has multiple assessments","Medium",i,df.loc[i],"; ".join(reasons))
                    out[-1]["duplicateGroup"]=f"assessment-beneficiary:{ids[i]}"
        total = _find(list(df.columns), "# Total Services")
        if total:
            for i in df.index[pd.to_numeric(df[total],errors="coerce").fillna(0).eq(0)]: self._flag(out,"assessments","Assessment without services","High",i,df.loc[i],"# Total Services is 0")
        elif assessment:
            service_assessment = _find(list(services.columns), "Assessment ID")
            linked = set(services[service_assessment].map(clean_id)) if service_assessment else set()
            for i in df.index[~df[assessment].map(clean_id).isin(linked)]: self._flag(out,"assessments","Assessment without services","High",i,df.loc[i],"No linked legal service")
        need = _find(list(df.columns), "Type of Legal Service Needed")
        if status:
            st = df[status].fillna("").astype(str).str.lower()
            for i in df.index[st.str.contains("pend")]: self._flag(out,"assessments","Pending assessment","Medium",i,df.loc[i],"Assessment status is Pending")
            if need:
                only_c = df[need].fillna("").astype(str).str.lower().str.contains("counselling") & ~df[need].fillna("").astype(str).str.lower().str.contains("assistance|representation")
                for i in df.index[st.str.contains("open") & only_c]: self._flag(out,"assessments","Open counselling-only assessment","Medium",i,df.loc[i],"Open assessment needs counselling only")
        detained = _find(list(df.columns), "Is the beneficiary detained")
        immigration = _find(list(df.columns), "Is it an immigration related charge")
        if detained and immigration:
            yes = df[detained].fillna("").astype(str).str.lower().str.contains("yes|نعم")
            no = df[detained].fillna("").astype(str).str.lower().str.contains("^no|لا")
            charge = df[immigration].fillna("").astype(str).str.strip(); charge_yes = charge.str.lower().str.contains("yes|نعم")
            community=_find(list(df.columns),"Community Type"); date_col=_find(list(df.columns),"Date of Assessment")
            refugee=df[community].fillna("").astype(str).str.contains("refugee",case=False,regex=False) if community else pd.Series(False,index=df.index)
            dates=pd.to_datetime(df[date_col],errors="coerce") if date_col else pd.Series(pd.NaT,index=df.index)
            eligible=refugee & dates.notna() & dates.dt.year.ge(2026)
            for i in df.index[eligible & yes & ~charge_yes]: self._flag(out,"assessments","Detention/immigration inconsistency","High",i,df.loc[i],f"Refugee assessment dated {dates[i].strftime('%d/%m/%Y')}: Detained is Yes but immigration charge is blank or No")
            for i in df.index[eligible & no & charge.ne("")]: self._flag(out,"assessments","Detention/immigration inconsistency","High",i,df.loc[i],f"Refugee assessment dated {dates[i].strftime('%d/%m/%Y')}: Detained is No but immigration charge is populated")
        date_of_birth=_find(list(df.columns),"Date of Birth","DoB")
        if detained and date_of_birth:
            detained_yes=df[detained].fillna("").astype(str).str.contains(r"\byes\b|نعم",case=False,regex=True)
            ages=df[date_of_birth].map(age_from_date)
            for i in df.index[detained_yes & ages.map(lambda age: age is not None and age < 10)]:
                self._flag(out,"assessments","Detained beneficiary below 10 years","High",i,df.loc[i],f"Detained beneficiary is {ages[i]} years old")
        detention_governorate=_find(list(df.columns),"Detention Governorate");project=_find(list(df.columns),"Projects -","Project");location=_find(list(df.columns),"Project Location")
        if detention_governorate and (project or location):
            for i,row in df.iterrows():
                if not detained or not re.search(r"\byes\b|نعم",clean_id(row.get(detained,"")),re.I): continue
                if project and re.search(r"\bamal\b",clean_id(row.get(project,"")),re.I): continue
                detained_place=normalize_governorate(row.get(detention_governorate,""))
                project_key=re.sub(r"\s+"," ",clean_id(row.get(project,""))).casefold() if project else ""
                expected={ASSESSMENT_PROJECT_GOVERNORATES[project_key]} if project_key in ASSESSMENT_PROJECT_GOVERNORATES else {normalize_governorate(row.get(column,"")) for column in (project,location) if column and normalize_governorate(row.get(column,""))}
                if expected and not detained_place:
                    self._flag(out,"assessments","Detention Governorate mismatch","Medium",i,row,f"Detention Governorate is missing; expected one of: {', '.join(sorted(expected))}")
                elif detained_place and expected and detained_place not in expected:
                    self._flag(out,"assessments","Detention Governorate mismatch","Medium",i,row,f"Detention Governorate {clean_id(row.get(detention_governorate,''))} does not match Project/Project Location ({', '.join(sorted(expected))})")
        if need:
            blank = df[need].isna() | df[need].astype(str).str.strip().eq("")
            for i in df.index[blank]: self._flag(out,"assessments","Blank legal service need","High",i,df.loc[i],"Type of Legal Service Needed is blank")
            if assessment:
                service_assessment=_find(list(services.columns),"Assessment ID"); provided=_find(list(services.columns),"Type of Service Provided")
                delivered:dict[str,str]={}
                if service_assessment and provided:
                    delivered=services.assign(_aid=services[service_assessment].map(clean_id)).groupby("_aid")[provided].apply(lambda x:" ".join(x.dropna().astype(str)).lower()).to_dict()
                documents_to_issue=_find(list(df.columns),"Type of Documents to be issued")
                for i,row in df.iterrows():
                    wanted=str(row.get(need,"") or "").lower(); actual=delivered.get(clean_id(row.get(assessment,"")),"")
                    if detained and immigration and re.search("yes|نعم",str(row.get(detained,"")),re.I) and re.search("yes|نعم",str(row.get(immigration,"")),re.I) and "counselling" in actual and not re.search("assistance|representation",actual): self._flag(out,"assessments","Detained beneficiary has counselling only","High",i,row,"Detained immigration case has counselling only")
                    age_col=_find(list(df.columns),"Age")
                    created_column=_find(list(df.columns),"Created On")
                    created_on=pd.to_datetime(row.get(created_column),errors="coerce",dayfirst=True) if created_column else pd.NaT
                    age_is_adult=age_col and pd.to_numeric(pd.Series([row.get(age_col)]),errors="coerce").iloc[0]>18
                    requests_representation="representation" in wanted
                    requests_counselling="counselling" in wanted
                    if age_is_adult and requests_representation and requests_counselling and not pd.isna(created_on) and created_on >= pd.Timestamp("2026-01-01") and "counselling" not in actual: self._flag(out,"assessments","Adult representation without counselling","Medium",i,row,"Adult assessment requests representation and counselling, but linked legal services have no counselling")
                    community=_find(list(df.columns),"Community Type")
                    non_idp=community and not re.search("idp|نازح",str(row.get(community,"")),re.I)
                    not_detained=detained and re.search("^no|لا",str(row.get(detained,"")),re.I)
                    requested_documents={document.strip() for document in re.split(r"[,;]",clean_id(row.get(documents_to_issue,""))) if document.strip()} if documents_to_issue else set()
                    only_exempt_documents=bool(requested_documents) and all(any(exception in document.casefold() for exception in REPRESENTATION_DOCUMENT_EXCEPTIONS) for document in requested_documents)
                    if non_idp and not_detained and not only_exempt_documents and not pd.isna(created_on) and created_on >= pd.Timestamp("2026-01-01") and re.search("assistance|representation",actual):
                        self._flag(out,"assessments","Representation while not detained","Medium",i,row,"Non-IDP has Assistance/Representation, was created in 2026 or later, and is not detained")
                        out[-1]["typeOfDocument"]=clean_id(row.get(documents_to_issue,"")) if documents_to_issue else ""
        return out

    def _assessment_month_flags(self, comparison_month: str | None = None) -> tuple[list[dict[str,Any]],str,list[str]]:
        df=self.frames["assessments"]; beneficiary=_find(list(df.columns),"Beneficiary ID"); date_col=_find(list(df.columns),"Date of Assessment"); created_col=_find(list(df.columns),"Created On")
        if not beneficiary or not date_col:return [],"",[]
        dates=pd.to_datetime(df[date_col],errors="coerce");months=dates.dt.to_period("M");available=sorted(months.dropna().astype(str).unique().tolist())
        active=comparison_month if comparison_month in available else (available[-1] if available else "")
        if not active:return [],active,available
        selected=pd.Period(active,freq="M");ids=df[beneficiary].map(clean_id);created=pd.to_datetime(df[created_col],errors="coerce",dayfirst=True) if created_col else pd.Series(pd.NaT,index=df.index);earlier:dict[str,list[tuple[pd.Timestamp,pd.Period]]]=defaultdict(list)
        for i in df.index[(months<selected)&ids.ne("")]:earlier[ids[i]].append((dates[i],months[i]))
        out=[]
        for i in df.index[(months==selected)&ids.isin(earlier.keys())]:
            created_on=created[i]
            if pd.isna(created_on) or created_on < pd.Timestamp("2026-08-01"):
                continue
            if created_on.to_period("M")==selected:
                continue
            history=sorted(earlier[ids[i]],key=lambda item:item[0]);history_months={month for _,month in history}
            immediate_previous=selected-1
            has_older_history=any(month<immediate_previous for month in history_months)
            grace=history_months=={immediate_previous} and created_on.to_period("M")==selected+1 and created_on.day<=4
            if grace and not has_older_history:
                continue
            history_dates=[item[0] for item in history]
            self._flag(out,"assessments","Selected month with previous assessment","High",i,df.loc[i],f"Selected month {active}; {len(history_dates)} earlier assessment(s), from {history_dates[0].strftime('%d/%m/%Y')} to {history_dates[-1].strftime('%d/%m/%Y')}; created on {created_on.strftime('%d/%m/%Y')}")
        return out,active,available

    def _amal_only_assessment_projects(self) -> bool:
        df=self.frames["assessments"]; project=_find(list(df.columns),"Projects -","Project")
        if not project:return False
        projects={clean_id(value).casefold() for value in df[project] if clean_id(value)}
        return bool(projects) and all("amal" in value for value in projects)

    @staticmethod
    def _is_detention_column(column: str) -> bool:
        normalized=column.casefold()
        return "detain" in normalized or "immigration related charge" in normalized

    def _service_month_flags(self, comparison_month: str | None = None) -> tuple[list[dict[str, Any]],str,list[str]]:
        df=self.frames["legalservices"]; beneficiary=_find(list(df.columns),"Beneficiary ID"); date_col=_find(list(df.columns),"Date of Service Provision"); created_col=_find(list(df.columns),"Created On")
        if not beneficiary or not date_col:return [],"",[]
        dates=pd.to_datetime(df[date_col],errors="coerce",dayfirst=True);months=dates.dt.to_period("M");available=sorted(months.dropna().astype(str).unique().tolist())
        active=comparison_month if comparison_month in available else (available[-1] if available else "")
        if not active:return [],active,available
        selected=pd.Period(active,freq="M");ids=df[beneficiary].map(clean_id);created=pd.to_datetime(df[created_col],errors="coerce",dayfirst=True) if created_col else pd.Series(pd.NaT,index=df.index);earlier:dict[str,list[tuple[pd.Timestamp,pd.Period]]]=defaultdict(list)
        for i in df.index[(months<selected)&ids.ne("")]:earlier[ids[i]].append((dates[i],months[i]))
        out=[]
        for i in df.index[(months==selected)&ids.isin(earlier.keys())]:
            created_on=created[i]
            if pd.isna(created_on) or created_on < pd.Timestamp("2026-08-01"):
                continue
            if created_on.to_period("M")==selected:
                continue
            history=sorted(earlier[ids[i]],key=lambda item:item[0]);history_months={month for _,month in history}
            immediate_previous=selected-1
            has_older_history=any(month<immediate_previous for month in history_months)
            grace=history_months=={immediate_previous} and created_on.to_period("M")==selected+1 and created_on.day<=4
            if grace and not has_older_history:
                continue
            history_dates=[item[0] for item in history]
            self._flag(out,"legalservices","Current and previous month duplicate","High",i,df.loc[i],f"Selected month {active}; {len(history_dates)} earlier service(s), from {history_dates[0].strftime('%d/%m/%Y')} to {history_dates[-1].strftime('%d/%m/%Y')}; created on {created_on.strftime('%d/%m/%Y')}")
        return out,active,available

    def _service_flags(self) -> list[dict[str, Any]]:
        df=self.frames["legalservices"]; out=[]; sid=_find(list(df.columns),"Service ID")
        out += self._future_date_flags("legalservices","Legal service date after today",("Date of Service Provision","Date Service Completed","Date of Issuance"))
        beneficiary=_find(list(df.columns),"Beneficiary ID"); aid=_find(list(df.columns),"Assessment ID"); provided=_find(list(df.columns),"Type of Service Provided"); document_type=_find(list(df.columns),"Type of Document")
        if beneficiary and aid and provided and document_type:
            keys={}
            for i,row in df.iterrows():
                key=(clean_id(row.get(beneficiary,"")),clean_id(row.get(aid,"")),normalize_legal_service_type(row.get(provided,"")),normalize_document_label(row.get(document_type,"")))
                if all(key): keys.setdefault(key,[]).append(i)
            for key,indexes in keys.items():
                if len(indexes)<2:continue
                duplicate_group=f"service:{'|'.join(key)}"
                for i in indexes:
                    self._flag(out,"legalservices","Duplicate service","High",i,df.loc[i],f"Same Beneficiary ID, Assessment ID, Type of Service Provided, and Type of Document occurs {len(indexes)} times")
                    out[-1]["duplicateGroup"]=duplicate_group
        assessment_id=_find(list(self.frames["assessments"].columns),"Assessment ID")
        if aid and assessment_id:
            valid=set(self.frames["assessments"][assessment_id].map(clean_id))
            for i in df.index[~df[aid].map(clean_id).isin(valid)]: self._flag(out,"legalservices","Orphaned assessment relationship","High",i,df.loc[i],"Assessment ID does not exist in assessments.csv")
        if document_type:
            blank=df[document_type].isna()|df[document_type].astype(str).str.strip().eq("")
            for i in df.index[blank]: self._flag(out,"legalservices","Missing Type of Document","Medium",i,df.loc[i],"Type of Document is blank")
        return out

    def _awareness_flags(self) -> list[dict[str, Any]]:
        df=self.frames["awareness"]; out=[]; name=_find(list(df.columns),"Participant Name")
        if name:
            vals=df[name].map(normalize_name); name_counts=vals[vals.ne("")].value_counts(); topic=_find(list(df.columns),"Session Topic","Topic")
            topics=df[topic].fillna("").astype(str).str.strip().str.lower() if topic else pd.Series("",index=df.index)
            pairs=pd.Series(list(zip(vals,topics)),index=df.index);pair_counts=pairs[vals.ne("") & topics.ne("")].value_counts()
            for i in df.index[vals.map(name_counts).fillna(0).ge(2)]:
                same_session=int(pair_counts.get((vals[i],topics[i]),0)) if topics[i] else 0
                if same_session >= 2:
                    self._flag(out,"awareness","Duplicate participant in session","High",i,df.loc[i],f"Name and session topic occur {same_session} times")
                    out[-1]["duplicateGroup"]=f"awareness:{vals[i]}|{topics[i]}"
                else:
                    self._flag(out,"awareness","Possible duplicate participant name","Minor",i,df.loc[i],f"Name occurs {int(name_counts[vals[i]])} times across different sessions")
                    out[-1]["overviewExcluded"]=True
        phone=_find(list(df.columns),"Phone Number")
        if phone:
            for i,v in df[phone].items():
                digits=phone_digits(v)
                if not digits or len(digits) == 1 or len(digits) == 11: continue
                self._flag(out,"awareness","Invalid contact number","Medium",i,df.loc[i],f"{len(digits)} digits")
        return out

    def review(self, dataset: str, search: str="", rule: str="", page: int=1, page_size: int=100,
               severity: str="", lawyer: str="", project: str="", location: str="", date: str="", comparison_month:str="",
               name_compare_chars:int=15,allow_name_variations:bool=False,exact_matches_only:bool=False) -> dict[str, Any]:
        bounded_chars=max(10,min(30,name_compare_chars));cache_key=(dataset,comparison_month or "",bounded_chars,bool(allow_name_variations))
        with self._cache_lock:
            context=self._review_cache.get(cache_key)
            if context is None:
                rows=[dict(item) for item in self.flags.get(dataset,[])]
                if dataset=="beneficiaries":
                    rows=[r for r in rows if r["rule"] not in {"Possible duplicate name","Possible duplicate contact and name"}]
                    rows += self._name_match_flags(bounded_chars,allow_name_variations,excluded_case_ids=self._excluded_case_ids("Possible duplicate name"))
                    rows += self._contact_name_match_flags(excluded_case_ids=self._excluded_case_ids("Possible duplicate contact and name"))
                    rows=[row for row in rows if not self._is_excluded(row)]
                active_month="";available_months=[]
                if dataset=="assessments":
                    month_rows,active_month,available_months=self._assessment_month_flags(comparison_month);rows += [dict(item) for item in month_rows]
                    if self._amal_only_assessment_projects():
                        rows=[item for item in rows if item["rule"] not in AMAL_HIDDEN_ASSESSMENT_RULES]
                elif dataset=="legalservices":
                    month_rows,active_month,available_months=self._service_month_flags(comparison_month);rows += [dict(item) for item in month_rows]
                if dataset in {"assessments","legalservices"}:
                    beneficiaries=self.frames["beneficiaries"];case_col=_find(list(beneficiaries.columns),"Case ID");name_col=_find(list(beneficiaries.columns),"Name (Filter Color Red)")
                    phone_col=_find(list(beneficiaries.columns),"Contact Number");project_col=_find(list(beneficiaries.columns),"Project");location_col=_find(list(beneficiaries.columns),"Project Location")
                    if case_col:
                        wanted=[column for column in (case_col,name_col,phone_col,project_col,location_col) if column]
                        bmap={clean_id(row.get(case_col,"")):row for row in beneficiaries[wanted].to_dict("records")}
                        for item in rows:
                            source=bmap.get(item.get("caseId",""))
                            if not source: continue
                            if not item.get("name") and name_col:item["name"]=clean_id(source.get(name_col,""))
                            for key,column in (("phone",phone_col),("project",project_col),("location",location_col)):
                                if not item.get(key) and column:item[key]=clean_id(source.get(column,""))
                observed=pd.Series([r["rule"] for r in rows],dtype=object).value_counts().to_dict();registered_rules=REGISTERED_RULES.get(dataset,())
                if dataset=="assessments" and self._amal_only_assessment_projects(): registered_rules=tuple(name for name in registered_rules if name not in AMAL_HIDDEN_ASSESSMENT_RULES)
                rule_counts={name:int(observed.get(name,0)) for name in registered_rules}
                date_field={"beneficiaries":"identificationDate","assessments":"assessmentDate","legalservices":"serviceDate","awareness":"awarenessDate"}.get(dataset,"")
                options={key:sorted({r.get(key,"") for r in rows if r.get(key,"")}) for key in ("severity","lawyer","project","location")}
                options["date"]=sorted({pd.to_datetime(r.get(date_field,""),errors="coerce",dayfirst=True).strftime("%Y-%m") for r in rows if date_field and not pd.isna(pd.to_datetime(r.get(date_field,""),errors="coerce",dayfirst=True))})
                name_records=0;eligible_name_records=0
                if dataset=="beneficiaries":
                    name_column=_find(list(self.frames[dataset].columns),"Name (Filter Color Red)")
                    project_column=_find(list(self.frames[dataset].columns),"Projects -","Project")
                    if name_column:
                        normalized_names=self.frames[dataset][name_column].map(normalize_name)
                        name_records=int(normalized_names.ne("").sum())
                        if project_column:
                            recognized_projects={"unhcr 2026 - erbil","unhcr 2026 - mosul & kirkuk","unhcr 2026 - suli","unhcr 2026 - baghdad","unhcr 2026 - gov","unhcr 2026 - amal camp"}
                            projects=self.frames[dataset][project_column].fillna("").astype(str).map(lambda value:re.sub(r"\s+"," ",value).strip().casefold())
                            eligible_name_records=int((normalized_names.str.len().ge(bounded_chars) & projects.isin(recognized_projects)).sum())
                context={"rows":rows,"ruleCounts":rule_counts,"filterOptions":options,"availableMonths":available_months,"activeComparisonMonth":active_month,"nameRecordCount":name_records,"eligibleNameRecordCount":eligible_name_records}
                self._review_cache[cache_key]=context
        rows=[row for row in context["rows"] if not self._is_excluded(row)]
        if exact_matches_only and rule == "Possible duplicate name":
            rows=self._name_match_flags(exact_only=True, excluded_case_ids=self._excluded_case_ids("Possible duplicate name"))
        elif rule:
            rows=[r for r in rows if r["rule"]==rule]
        for key,selection in (("severity",severity),("lawyer",lawyer),("project",project),("location",location)):
            if selection: rows=[r for r in rows if r.get(key)==selection]
        if date:
            date_field={"beneficiaries":"identificationDate","assessments":"assessmentDate","legalservices":"serviceDate","awareness":"awarenessDate"}.get(dataset,"")
            rows=[r for r in rows if date_field and not pd.isna(pd.to_datetime(r.get(date_field,""),errors="coerce",dayfirst=True)) and pd.to_datetime(r.get(date_field,""),errors="coerce",dayfirst=True).strftime("%Y-%m")==date]
        if search:
            needle=search.lower(); rows=[r for r in rows if needle in " ".join(map(str,r.values())).lower()]
        if rule in {"Marital status below 18", "Spouse below 18"}:
            date_key = "spouseDateOfBirth" if rule == "Spouse below 18" else "dateOfBirth"
            rows.sort(key=lambda item: str(item.get(date_key, "")), reverse=True)
        start=(page-1)*page_size
        rules=sorted({r["rule"] for r in self.flags.get(dataset,[])})
        return {"dataset":dataset,"total":len(rows),"page":page,"pageSize":page_size,"rules":list(REGISTERED_RULES.get(dataset,rules)),"ruleCounts":context["ruleCounts"],"filterOptions":context["filterOptions"],"availableMonths":context["availableMonths"],"activeComparisonMonth":context["activeComparisonMonth"],"nameRecordCount":context["nameRecordCount"],"eligibleNameRecordCount":context["eligibleNameRecordCount"],"nameCompareCharsApplied":bounded_chars,"allowNameVariationsApplied":bool(allow_name_variations),"rows":rows[start:start+page_size]}

    def review_export(self,dataset:str,comparison_month:str="",name_compare_chars:int=15,allow_name_variations:bool=False,exact_matches_only:bool=False,selected_rules:list[str]|None=None,severity:str="",lawyer:str="",project:str="",location:str="",date:str="",search:str="",ignore_court_verdict:bool=False)->bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
        from openpyxl.utils import get_column_letter
        flags=list(self.flags.get(dataset,[]))
        if dataset=="beneficiaries":
            flags=[r for r in flags if r["rule"] not in {"Possible duplicate name","Possible duplicate contact and name"}]
            flags += self._name_match_flags(name_compare_chars,allow_name_variations,exact_only=exact_matches_only,excluded_case_ids=self._excluded_case_ids("Possible duplicate name"))
            flags += self._contact_name_match_flags(excluded_case_ids=self._excluded_case_ids("Possible duplicate contact and name"))
            flags=[row for row in flags if not self._is_excluded(row)]
        if dataset=="assessments":
            flags+=self._assessment_month_flags(comparison_month)[0]
            if self._amal_only_assessment_projects(): flags=[row for row in flags if row["rule"] not in AMAL_HIDDEN_ASSESSMENT_RULES]
        elif dataset=="legalservices":flags+=self._service_month_flags(comparison_month)[0]
        flags=[row for row in flags if not self._is_excluded(row)]
        if selected_rules is not None: flags=[row for row in flags if row.get("rule") in selected_rules]
        for key,selection in (("severity",severity),("lawyer",lawyer),("project",project),("location",location)):
            if selection: flags=[row for row in flags if row.get(key)==selection]
        if date:
            date_field={"beneficiaries":"identificationDate","assessments":"assessmentDate","legalservices":"serviceDate","awareness":"awarenessDate"}.get(dataset,"")
            flags=[row for row in flags if date_field and not pd.isna(pd.to_datetime(row.get(date_field,""),errors="coerce",dayfirst=True)) and pd.to_datetime(row.get(date_field,""),errors="coerce",dayfirst=True).strftime("%Y-%m")==date]
        if search:
            needle=search.lower();flags=[row for row in flags if needle in " ".join(map(str,row.values())).lower()]
        if ignore_court_verdict and dataset=="legalservices":
            flags=[row for row in flags if not (row.get("rule")=="Duplicate service" and re.search(r"court verdict|\bother\b|اخرى",str(row.get("typeOfDocument", "")),flags=re.I))]
        frame=self.frames[dataset]; review_columns=["Review Finding","Recommended Action","Review Detail"];columns=review_columns+list(frame.columns)
        workbook=Workbook()
        palette=("FCE8E6","FFF4D6","E7F0FF","E5F5EA","F0E8FA","FFECDD","E3F4F4","F7E8F1")
        rule_colors={rule:palette[i%len(palette)] for i,rule in enumerate(sorted({x["rule"] for x in flags}))}
        duplicate_colors={key:palette[i%len(palette)] for i,key in enumerate(sorted({x.get("duplicateGroup","") for x in flags if x.get("duplicateGroup")}))}
        thin=Side(style="thin",color="D9E2EC");name_source=_find(list(frame.columns),"Name (Filter Color Red)")
        def write_sheet(sheet:Any,items:list[dict[str,Any]])->None:
            sheet.append([safe_spreadsheet_value(column) for column in columns])
            for flag in items:
                source=frame.iloc[flag["row"]-2] if 0 <= flag["row"]-2 < len(frame) else pd.Series(dtype=object)
                values=[flag["rule"],flag["action"],flag["detail"]]+[display_value(source.get(c,"")) for c in frame.columns]
                sheet.append([safe_spreadsheet_value(value) for value in values]);row_number=sheet.max_row;sheet.row_dimensions[row_number].height=24
                sheet.cell(row_number,1).fill=PatternFill("solid",fgColor=rule_colors[flag["rule"]])
                if flag.get("duplicateGroup"):
                    exact=flag.get("nameMatchMode")=="exact";color="FDE8E8" if exact else duplicate_colors[flag["duplicateGroup"]]
                    sheet.cell(row_number,8).fill=PatternFill("solid",fgColor=color)
                    if exact:sheet.cell(row_number,8).font=Font(bold=True,color="991B1B")
                    if name_source:
                        source_cell=sheet.cell(row_number,len(review_columns)+1+list(frame.columns).index(name_source));source_cell.fill=PatternFill("solid",fgColor=color)
                        if exact:source_cell.font=Font(bold=True,color="991B1B")
                for cell in sheet[row_number]:
                    cell.alignment=Alignment(vertical="center",wrap_text=True);cell.border=Border(bottom=thin)
            for cell in sheet[1]:
                cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="2454C6");cell.alignment=Alignment(vertical="center",wrap_text=True)
            sheet.row_dimensions[1].height=34
            sheet.freeze_panes="A2";sheet.auto_filter.ref=f"A1:{get_column_letter(len(columns))}{len(items)+1}"
            for index,name in enumerate(columns,1):sheet.column_dimensions[get_column_letter(index)].width=min(32,max(13,len(str(name))+2))
        if dataset in {"beneficiaries","assessments","legalservices"}:
            grouped={name:[] for name in ("North Iraq","AMAL Camp","South Iraq")};unclassified=[]
            for flag in flags:
                project_key=re.sub(r"\s+"," ",str(flag.get("project","")).strip()).casefold()
                sheet_name=REVIEW_EXPORT_PROJECT_SHEETS.get(project_key)
                (grouped[sheet_name] if sheet_name else unclassified).append(flag)
            workbook.remove(workbook.active)
            for name,items in grouped.items(): write_sheet(workbook.create_sheet(name),items)
            if unclassified: write_sheet(workbook.create_sheet("Unclassified"),unclassified)
        else:
            sheet=workbook.active;sheet.title="Review findings";write_sheet(sheet,flags)
        output=io.BytesIO();workbook.save(output);return output.getvalue()

    def explorer_filters(self,dataset:str)->dict[str,Any]:
        if dataset not in self.frames: raise ValueError("Dataset not loaded")
        if dataset in self._explorer_filter_cache:return self._explorer_filter_cache[dataset]
        frame=self.frames[dataset];columns=[]
        for column in frame.columns:
            values=sorted({str(value).strip() for value in frame[column].dropna().unique() if str(value).strip()})
            if 0 < len(values) <= 200: columns.append({"name":str(column),"values":values})
        result={"columns":columns};self._explorer_filter_cache[dataset]=result;return result

    def studio(self,dataset:str,row_dimension:str,column_dimension:str="",filters:dict[str,list[str]]|None=None,measure:str="records")->dict[str,Any]:
        if dataset not in self.frames: raise ValueError("Selected source file is not loaded.")
        frame=self.frames[dataset].copy()
        if dataset=="assessments" and self._amal_only_assessment_projects():
            hidden_columns=[column for column in frame.columns if self._is_detention_column(column)]
            frame=frame.drop(columns=hidden_columns)
        if row_dimension not in frame.columns or (column_dimension and column_dimension not in frame.columns): raise ValueError("Choose valid source columns.")
        for column,values in (filters or {}).items():
            if column in frame.columns and values: frame=frame[frame[column].fillna("").astype(str).isin(values)]
        rows=frame[row_dimension].fillna("Not provided").astype(str).str.strip().replace("","Not provided")
        columns=frame[column_dimension].fillna("Not provided").astype(str).str.strip().replace("","Not provided") if column_dimension else pd.Series("Total",index=frame.index)
        identifier=_find(list(frame.columns),"Beneficiary ID","Case ID") if measure=="beneficiaries" else None
        values=frame[identifier].map(clean_id) if identifier else pd.Series(frame.index.astype(str),index=frame.index)
        grouped=pd.DataFrame({"row":rows,"column":columns,"value":values}).groupby(["row","column"])["value"].nunique(); total=int(values.nunique())
        return {"page":dataset,"rowDimension":row_dimension,"columnDimension":column_dimension or None,"measure":measure,"total":total,"cells":[{"row":str(row),"column":str(column),"count":int(count),"percent":int(count)/total if total else 0} for (row,column),count in grouped.items()]}

    def analytics_dashboard(self,dataset:str,filters:dict[str,list[str]]|None=None,search:str="",page:int=1,page_size:int=100,sort_column:str="",sort_direction:str="asc")->dict[str,Any]:
        if dataset not in {"assessments","legalservices","beneficiaries","awareness"} or dataset not in self.frames: raise ValueError("Selected Analytics Studio section is not loaded.")
        frame=self.frames[dataset].copy(); original_columns=list(frame.columns)
        amal_only_assessments=dataset=="assessments" and self._amal_only_assessment_projects()
        if amal_only_assessments:
            hidden_columns=[column for column in frame.columns if self._is_detention_column(column)]
            frame=frame.drop(columns=hidden_columns)
            original_columns=[column for column in original_columns if column not in hidden_columns]
        specs={
          "assessments":{"id":("Assessment ID",),"beneficiary":("Beneficiary ID",),"date":("Date of Assessment",),"charts":[("Project",("Projects -","Project")),("Project location",("Project Location",)),("Gender / age group",("Age Gender Group","UNHCR Age Group")),("Nationality",("Nationality",)),("Community type",("Community Type",)),("Assessment status",("Assessment Status",)),("Legal service needed",("Type of Legal Service Needed",)),("Document needed",("Type of Documents to be issued",)),("Beneficiary detained",("Is the beneficiary detained",)),("Detainee current status",("Detainee current status",))]},
          "legalservices":{"id":("Service ID",),"beneficiary":("Beneficiary ID",),"date":("Date of Service Provision",),"charts":[("Project",("Projects -","Project")),("Project location",("Project Location",)),("Gender / age group",("Age Gender Group","UNHCR Age Group")),("Type of service provided",("Type of Service Provided",)),("Service status",("Service Status",)),("Type of document",("Type of Document",)),("Nationality",("Nationality",)),("Community type",("Community Type",)),("Assessment legal-service need",("_assessment_need",))]},
          "beneficiaries":{"id":("Case ID","Beneficiary ID"),"beneficiary":("Case ID","Beneficiary ID"),"date":("Date of Identification",),"charts":[("Project",("Project",)),("Project location",("Project Location",)),("Gender / age group",("Age Gender Group","UNHCR Age Group")),("Nationality",("Nationality",)),("Community type",("Community Type",)),("Marital status",("Marital Statues","Marital Status")),("Vulnerability",("Type of vulnerabilities",)),("Protection category",("Protection Category",))]},
          "awareness":{"id":("Awareness ID",),"beneficiary":("Participant Name",),"date":("Date of Session","Added On"),"charts":[("Session topic - sessions",("Session Topic",)),("Session topic - participants",("Session Topic",)),("Gender / age group",("Gender Age Group","UNHCR Age Group")),("Nationality",("Nationality",)),("Community type",("Community Type",)),("Governorate",("Governorate",)),("Lawyer",("Lawyer",))]},
        }[dataset]
        if dataset=="legalservices" and "assessments" in self.frames:
            aid=_find(list(frame.columns),"Assessment ID"); assessment=self.frames["assessments"]; other_id=_find(list(assessment.columns),"Assessment ID"); need=_find(list(assessment.columns),"Type of Legal Service Needed")
            if aid and other_id and need: frame["_assessment_need"]=frame[aid].map(clean_id).map(dict(zip(assessment[other_id].map(clean_id),assessment[need].fillna("").astype(str))))
        date_col=_find(list(frame.columns),*specs["date"]); id_col=_find(list(frame.columns),*specs["id"]); beneficiary_col=_find(list(frame.columns),*specs["beneficiary"])
        if date_col:
            dates=pd.to_datetime(frame[date_col],errors="coerce",dayfirst=True); frame["Year"]=dates.dt.year.astype("Int64").astype(str).replace("<NA>","");frame["Quarter"]=dates.dt.to_period("Q").astype(str);frame["Month"]=dates.dt.to_period("M").astype(str)
        chart_columns=[]; warnings=[]
        for title,hints in specs["charts"]:
            column=_find(list(frame.columns),*hints)
            if column: chart_columns.append((title,column))
            else: warnings.append(f"{title}: source column not available")
        # The Analytics Studio drawer is intentionally broader than the visible
        # dashboard charts.  Keep every source field available there while the
        # section toolbar exposes just the four most useful quick filters.
        filter_columns=list(dict.fromkeys(original_columns+(["Year","Quarter","Month"] if date_col else [])))
        filter_options={column:sorted({str(value).strip() for value in frame[column].dropna() if str(value).strip()})[:500] for column in filter_columns}
        filtered=frame
        for column,values in (filters or {}).items():
            if column in filtered.columns and values: filtered=filtered[filtered[column].fillna("").astype(str).isin(values)]
        if search:
            mask=filtered.astype(str).apply(lambda column:column.str.contains(search,case=False,na=False,regex=False)).any(axis=1);filtered=filtered[mask]
        metric=filtered[id_col].map(clean_id) if id_col else pd.Series(filtered.index.astype(str),index=filtered.index); total=int(metric.replace("",pd.NA).nunique())
        charts=[]
        for title,column in chart_columns:
            chart_metric=(filtered[beneficiary_col].fillna("").astype(str).str.strip() if dataset=="awareness" and title=="Session topic - participants" and beneficiary_col else metric)
            working=pd.DataFrame({"label":filtered[column].fillna("").astype(str).str.strip(),"metric":chart_metric});working=working[working.label.ne("") & working.metric.ne("")]
            counts=working.groupby("label").metric.nunique().sort_values(ascending=False).head(12);charts.append({"id":column,"title":title,"kind":"bar","multiChoice":False,"rows":[{"label":label,"count":int(count),"percent":int(count)/total if total else 0} for label,count in counts.items()]})
        trend=[]
        if date_col:
            working=pd.DataFrame({"month":frame.loc[filtered.index,"Month"],"metric":metric});counts=working[working.month.ne("")].groupby("month").metric.nunique().sort_index();trend=[{"label":month,"count":int(count),"percent":int(count)/total if total else 0} for month,count in counts.items()]
        status_col=_find(list(filtered.columns),"Assessment Status" if dataset=="assessments" else "Service Status")
        distinct=lambda mask:int(metric[mask].replace("",pd.NA).nunique())
        if dataset=="assessments":
            status=filtered[status_col].fillna("").astype(str).str.lower() if status_col else pd.Series("",index=filtered.index);open_count=distinct(status.str.contains("open|pend"));closed=distinct(status.str.contains("closed"));kpis=[("Assessments",total),("Unique beneficiaries",filtered[beneficiary_col].map(clean_id).nunique() if beneficiary_col else 0),("Open caseload",open_count),("Closed",closed)]
            if not amal_only_assessments:
                detained_col=_find(list(filtered.columns),"Is the beneficiary detained");immigration_col=_find(list(filtered.columns),"Is it an immigration related charge","Immigration related charge");detained=filtered[detained_col].fillna("").astype(str).str.contains(r"\byes\b|نعم",case=False,regex=True) if detained_col else pd.Series(False,index=filtered.index);immigration=filtered[immigration_col].fillna("").astype(str).str.contains(r"\byes\b|نعم",case=False,regex=True) if immigration_col else pd.Series(False,index=filtered.index);kpis.append(("Detention cases with immigration charges",distinct(detained&immigration)))
            kpis.append(("Closure rate",closed/total if total else 0))
        elif dataset=="legalservices":
            status=filtered[status_col].fillna("").astype(str).str.lower() if status_col else pd.Series("",index=filtered.index);completed=distinct(status.str.contains("completed")&~status.str.contains("uncompleted"));uncompleted=distinct(status.str.contains("uncompleted|not completed|incomplete"));kpis=[("Services",total),("Unique beneficiaries",filtered[beneficiary_col].map(clean_id).nunique() if beneficiary_col else 0),("Completed",completed),("Uncompleted services",uncompleted),("In process",distinct(status.str.contains("process"))),("Completion rate",completed/total if total else 0)]
        elif dataset=="beneficiaries":
            assessed=_find(list(filtered.columns),"# total assessments");served=_find(list(filtered.columns),"# of Services");kpis=[("Beneficiaries",total),("Assessed cases",int(pd.to_numeric(filtered[assessed],errors="coerce").fillna(0).gt(0).sum()) if assessed else 0),("Served cases",int(pd.to_numeric(filtered[served],errors="coerce").fillna(0).gt(0).sum()) if served else 0)]
        else:kpis=[("Awareness records",total),("Participant names",filtered[beneficiary_col].fillna("").astype(str).str.strip().replace("",pd.NA).nunique() if beneficiary_col else 0),("Session topics",filtered[_find(list(filtered.columns),"Session Topic")].nunique() if _find(list(filtered.columns),"Session Topic") else 0)]
        if sort_column in filtered.columns: filtered=filtered.sort_values(sort_column,ascending=sort_direction!="desc",kind="stable")
        start=(max(page,1)-1)*page_size;table=filtered.iloc[start:start+page_size][original_columns].copy()
        records=[{"__rowKey":str(index),**{column:display_value(value) for column,value in row.items()}} for index,row in table.iterrows()]
        return {"dataset":dataset,"total":total,"matchedRows":len(filtered),"page":page,"pageSize":page_size,"kpis":[{"label":label,"value":float(value),"format":"percent" if "rate" in label.lower() else "number"} for label,value in kpis],"trend":trend,"charts":charts,"filterOptions":filter_options,"columns":original_columns,"rows":records,"warnings":warnings}

    def explorer(self, dataset: str, search: str="", page: int=1, page_size: int=100, filter_column: str="", filter_value: str="", filters:dict[str,list[str]]|None=None, sort_column:str="", sort_direction:str="asc") -> dict[str, Any]:
        if dataset not in self.frames: raise ValueError("Dataset not loaded")
        frame=self.frames[dataset].copy()
        if dataset=="deportationrecords":
            date_col=_find(list(frame.columns),"Date of deporting","Date of deportation")
            if date_col:
                dates=pd.to_datetime(frame[date_col],errors="coerce",dayfirst=True)
                frame["Month"]=dates.dt.to_period("M").astype(str).replace("NaT","")
        if search:
            searchable=self._search_cache.get(dataset)
            if searchable is None:
                searchable=frame.fillna("").astype(str).agg(" ".join,axis=1).str.lower();self._search_cache[dataset]=searchable
            frame=frame[searchable.str.contains(search.lower(),regex=False)]
        if filter_column in frame.columns and filter_value:
            frame=frame[frame[filter_column].fillna("").astype(str).str.contains(filter_value,case=False,regex=False)]
        for column,selections in (filters or {}).items():
            if column in frame.columns and selections: frame=frame[frame[column].fillna("").astype(str).isin(selections)]
        if sort_column in frame.columns:
            values=frame[sort_column]
            if DATE_HINT.search(sort_column): values=pd.to_datetime(values,errors="coerce",dayfirst=True)
            elif pd.api.types.is_numeric_dtype(values): values=pd.to_numeric(values,errors="coerce")
            else: values=values.fillna("").astype(str).str.casefold()
            frame=frame.assign(_explorer_sort=values).sort_values("_explorer_sort",ascending=sort_direction!="desc",kind="stable",na_position="last").drop(columns="_explorer_sort")
        start=(page-1)*page_size
        rows=[{"__rowKey":str(index),**{str(k):display_value(v) for k,v in row.items()}} for index,row in frame.iloc[start:start+page_size].iterrows()]
        return {"dataset":dataset,"total":len(frame),"page":page,"pageSize":page_size,"columns":[str(c) for c in frame.columns],"rows":rows}

    def explorer_export(self, dataset: str, search: str="", filters:dict[str,list[str]]|None=None, export_format: str="xlsx") -> bytes:
        if dataset not in self.frames: raise ValueError("Dataset not loaded")
        frame=self.frames[dataset].copy();source_columns=list(frame.columns)
        date_hints={"assessments":("Date of Assessment",),"legalservices":("Date of Service Provision",),"beneficiaries":("Date of Identification",),"awareness":("Date of Session","Added On")}
        date_col=_find(list(frame.columns),*date_hints.get(dataset,()))
        if date_col:
            dates=pd.to_datetime(frame[date_col],errors="coerce",dayfirst=True);frame["Year"]=dates.dt.year.astype("Int64").astype(str).replace("<NA>","");frame["Quarter"]=dates.dt.to_period("Q").astype(str);frame["Month"]=dates.dt.to_period("M").astype(str)
        if dataset=="legalservices" and "assessments" in self.frames:
            aid=_find(list(frame.columns),"Assessment ID");assessment=self.frames["assessments"];other_id=_find(list(assessment.columns),"Assessment ID");need=_find(list(assessment.columns),"Type of Legal Service Needed")
            if aid and other_id and need:frame["_assessment_need"]=frame[aid].map(clean_id).map(dict(zip(assessment[other_id].map(clean_id),assessment[need].fillna("").astype(str))))
        if search:
            needle=search.lower();mask=pd.Series(False,index=frame.index)
            for column in frame.columns:mask|=frame[column].fillna("").astype(str).str.lower().str.contains(needle,regex=False)
            frame=frame[mask]
        for column,selections in (filters or {}).items():
            if column in frame.columns and selections:frame=frame[frame[column].fillna("").astype(str).isin(selections)]
        frame=_safe_export(frame[source_columns])
        output=io.BytesIO()
        if export_format=="csv":
            return frame.to_csv(index=False).encode("utf-8-sig")
        if export_format!="xlsx":raise ValueError("Unsupported export format")
        with pd.ExcelWriter(output,engine="openpyxl",date_format="DD/MM/YYYY",datetime_format="DD/MM/YYYY") as writer:
            frame.to_excel(writer,index=False,sheet_name="Filtered data")
            sheet=writer.book["Filtered data"];sheet.freeze_panes="A2";sheet.auto_filter.ref=sheet.dimensions
            for cell in sheet[1]:
                cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="1D4ED8")
            for index,column in enumerate(frame.columns,1):
                width=max([len(str(column))]+[len(str(value)) for value in frame[column].head(250)])+2
                sheet.column_dimensions[get_column_letter(index)].width=min(45,max(12,width))
            format_excel_dates(writer.book)
        return output.getvalue()

    def case_filters(self) -> dict[str, Any]:
        if self._case_filter_cache is not None:return self._case_filter_cache
        labels={"beneficiaries":"Beneficiary","assessments":"Assessment","legalservices":"Legal service","followupslogbooks":"Follow-up","legalfees":"Legal fee"};groups=[]
        for dataset in labels:
            if dataset not in self.frames: continue
            columns=[]
            for column in self.frames[dataset].columns:
                series=self.frames[dataset][column].dropna()
                if series.empty or series.nunique(dropna=True)>200:continue
                raw=[str(value).strip() for value in series.unique() if str(value).strip()]
                normalized:dict[str,str]={}
                for value in raw: normalized.setdefault(value.casefold(),value)
                values=sorted(normalized.values(),key=str.casefold)
                if 0 < len(values) <= 200: columns.append({"key":f"{dataset}::{column}","name":str(column),"values":values})
            groups.append({"dataset":dataset,"label":labels[dataset],"columns":columns})
        self._case_filter_cache={"groups":groups};return self._case_filter_cache

    def _case_ids_for_rows(self,dataset:str,rows:pd.DataFrame)->set[str]:
        if rows.empty:return set()
        if dataset=="beneficiaries":
            column=_find(list(rows.columns),"Case ID");return set(rows[column].map(clean_id)) if column else set()
        if dataset=="assessments":
            column=_find(list(rows.columns),"Beneficiary ID");return set(rows[column].map(clean_id)) if column else set()
        if dataset=="legalservices":
            beneficiary=_find(list(rows.columns),"Beneficiary ID")
            if beneficiary:return set(rows[beneficiary].map(clean_id))
            assessment=_find(list(rows.columns),"Assessment ID");assessment_ids=set(rows[assessment].map(clean_id)) if assessment else set()
        elif dataset=="followupslogbooks":
            service=_find(list(rows.columns),"Service ID");service_ids=set(rows[service].map(clean_id)) if service else set();services=self._related_set("legalservices","Service ID",service_ids)
            return self._case_ids_for_rows("legalservices",services)
        elif dataset=="legalfees":
            service=_find(list(rows.columns),"Legal Service ID","Service ID");service_ids=set(rows[service].map(clean_id)) if service else set();services=self._related_set("legalservices","Service ID",service_ids)
            return self._case_ids_for_rows("legalservices",services)
        else:return set()
        assessments=self._related_set("assessments","Assessment ID",assessment_ids)
        return self._case_ids_for_rows("assessments",assessments)

    def _case_candidate_ids(self,query:str,filters:dict[str,list[str]]|None=None)->set[str]:
        b=self.frames["beneficiaries"];bid=_find(list(b.columns),"Case ID");candidate_ids=set(b[bid].map(clean_id))
        for key,selections in (filters or {}).items():
            if not selections:continue
            dataset,column=(key.split("::",1) if "::" in key else ("beneficiaries",key));frame=self.frames.get(dataset)
            if frame is None or column not in frame.columns:continue
            series=frame[column].fillna("").astype(str);mask=pd.Series(False,index=frame.index)
            for selection in selections:mask|=series.str.contains(str(selection),case=False,regex=False)
            candidate_ids&=self._case_ids_for_rows(dataset,frame[mask])
        needle=query.strip()
        if needle:
            mask=b[bid].map(clean_id).str.contains(needle,case=False,regex=False)
            for hint in ("Name (Filter Color Red)","Contact Number","ID Number","# UNHCR","Individual Number ASSISIT"):
                column=_find(list(b.columns),hint)
                if column:mask|=b[column].fillna("").astype(str).str.contains(needle,case=False,regex=False)
            search_ids=set(b.loc[mask,bid].map(clean_id))
            for dataset,id_hint in (("assessments","Assessment ID"),("legalservices","Service ID")):
                frame=self.frames[dataset];column=_find(list(frame.columns),id_hint)
                if column:
                    rows=frame[frame[column].map(clean_id).str.contains(needle,case=False,regex=False)]
                    search_ids|=self._case_ids_for_rows(dataset,rows)
            candidate_ids&=search_ids
        return candidate_ids

    @staticmethod
    def _prefixed_frame(frame:pd.DataFrame,dataset:str)->pd.DataFrame:
        result=frame.copy();result.columns=[f"{dataset}::{column}" for column in result.columns];return result

    def _connected_case_frame(self,candidate_ids:set[str])->pd.DataFrame:
        b=self.frames["beneficiaries"];b_id=_find(list(b.columns),"Case ID")
        scoped=b[b[b_id].map(clean_id).isin(candidate_ids)].copy()
        result=self._prefixed_frame(scoped,"beneficiaries")
        a=self.frames["assessments"];a_beneficiary=_find(list(a.columns),"Beneficiary ID")
        if a_beneficiary:
            left=f"beneficiaries::{b_id}";right=f"assessments::{a_beneficiary}"
            result["__case_key"]=result[left].map(clean_id)
            assessment=self._prefixed_frame(a,"assessments");assessment["__case_key"]=assessment[right].map(clean_id)
            result=result.merge(assessment,on="__case_key",how="left",sort=False,suffixes=("", "__duplicate"))
        services=self.frames["legalservices"];service_assessment=_find(list(services.columns),"Assessment ID");assessment_id=_find(list(a.columns),"Assessment ID")
        if service_assessment and assessment_id:
            service=self._prefixed_frame(services,"legalservices");service["__assessment_key"]=service[f"legalservices::{service_assessment}"].map(clean_id)
            result["__assessment_key"]=result[f"assessments::{assessment_id}"].map(clean_id)
            result=result.merge(service,on="__assessment_key",how="left",sort=False,suffixes=("", "__duplicate"))
        service_id=_find(list(services.columns),"Service ID")
        if service_id:
            result["__service_key"]=result[f"legalservices::{service_id}"].map(clean_id)
            child:pd.DataFrame|None=None
            for dataset,hints in (("followupslogbooks",("Service ID",)),("legalfees",("Legal Service ID","Service ID"))):
                if dataset not in self.frames:continue
                frame=self.frames[dataset];foreign=_find(list(frame.columns),*hints)
                if not foreign:continue
                part=self._prefixed_frame(frame,dataset);part["__service_key"]=part[f"{dataset}::{foreign}"].map(clean_id);part["__child_ordinal"]=part.groupby("__service_key",sort=False).cumcount()
                child=part if child is None else child.merge(part,on=["__service_key","__child_ordinal"],how="outer",sort=False)
            if child is not None:result=result.merge(child,on="__service_key",how="left",sort=False)
        lawyer_series=pd.Series("",index=result.index,dtype="object")
        for dataset in ("legalservices","assessments","beneficiaries"):
            source=self.frames.get(dataset);lawyer=_find(list(source.columns),"Lawyers","Lawyer") if source is not None else None
            if lawyer:
                values=result.get(f"{dataset}::{lawyer}",pd.Series("",index=result.index)).fillna("").astype(str).str.strip()
                lawyer_series=lawyer_series.mask(lawyer_series.eq(""),values)
        result.insert(0,"Lawyer",lawyer_series.mask(lawyer_series.eq(""),"Unassigned"))
        return result.drop(columns=[column for column in result.columns if column.startswith("__") or column.endswith("__duplicate")],errors="ignore")

    def _case_columns(self,frame:pd.DataFrame)->list[dict[str,str]]:
        labels={"beneficiaries":"Beneficiary","assessments":"Assessment","legalservices":"Service","followupslogbooks":"Follow-up","legalfees":"Legal fee"}
        return [{"key":column,"label":column.split("::",1)[-1],"dataset":"Summary" if column=="Lawyer" else labels.get(column.split("::",1)[0],"Other")} for column in frame.columns]

    def _default_case_columns(self,available:list[dict[str,str]])->list[str]:
        keys=[item["key"] for item in available];chosen=["Lawyer"]
        hints={"beneficiaries":("Case ID","Name (Filter Color Red)","Project","Project Location"),"assessments":("Assessment ID","Date of Assessment","Assessment Status"),"legalservices":("Service ID","Type of Service","Date of Service","Service Status"),"followupslogbooks":("Follow-up ID","Status"),"legalfees":("Legal Fee ID","Status","Amount")}
        for dataset,wanted in hints.items():
            dataset_keys=[key for key in keys if key.startswith(dataset+"::")]
            for hint in wanted:
                match=next((key for key in dataset_keys if hint.casefold() in key.split("::",1)[1].casefold()),None)
                if match and match not in chosen:chosen.append(match)
        return chosen

    def case(self,query:str,filters:dict[str,list[str]]|None=None,limit:int=20,view_mode:str="cards",page:int=1,page_size:int=100,sort_column:str="",sort_direction:str="asc",selected_columns:list[str]|None=None) -> dict[str, Any]:
        b=self.frames["beneficiaries"];bid=_find(list(b.columns),"Case ID");candidate_ids=self._case_candidate_ids(query,filters)
        found=b[b[bid].map(clean_id).isin(candidate_ids)]
        if view_mode=="table":
            case_sort=sort_column.split("::",1)[-1] if sort_column else bid
            if sort_column=="Lawyer":case_sort=_find(list(found.columns),"Lawyers","Lawyer") or bid
            elif case_sort not in found.columns:case_sort=_find(list(found.columns),case_sort) or bid
            if case_sort in found.columns:
                values=found[case_sort];numeric=pd.to_numeric(values,errors="coerce")
                key=numeric if numeric.notna().sum()>=max(1,len(values)//2) else (pd.to_datetime(values,errors="coerce") if DATE_HINT.search(case_sort) else values.fillna("").astype(str).str.casefold())
                found=found.assign(__sort=key).sort_values("__sort",ascending=sort_direction!="desc",na_position="last",kind="stable").drop(columns="__sort")
            total_cases=len(found);page=max(1,page);page_size=max(1,min(100,page_size));start=(page-1)*page_size;found=found.iloc[start:start+page_size]
        else:
            total_cases=len(found);found=found.head(limit if query.strip() or filters or limit>20 else min(limit,12))
        cases=[]
        for _, row in found.iterrows():
            case_id=clean_id(row[bid]); assessments=self._related("assessments","Beneficiary ID",case_id)
            assessment_nodes=[]; total_services=total_followups=total_fees=0
            for _,assessment_row in assessments.iterrows():
                assessment_id=clean_id(assessment_row.get(_find(list(assessments.columns),"Assessment ID"),"")); service_rows=self._related("legalservices","Assessment ID",assessment_id); service_nodes=[]
                for _,service_row in service_rows.iterrows():
                    service_id=clean_id(service_row.get(_find(list(service_rows.columns),"Service ID"),"")); followups=[];fees=[]
                    if "followupslogbooks" in self.frames: followups=[self._row(x) for _,x in self._related("followupslogbooks","Service ID",service_id).iterrows()]
                    if "legalfees" in self.frames: fees=[self._row(x) for _,x in self._related("legalfees","Legal Service ID",service_id).iterrows()]
                    total_followups+=len(followups);total_fees+=len(fees);service_nodes.append({"service":self._row(service_row),"followups":followups,"fees":fees})
                total_services+=len(service_nodes);assessment_nodes.append({"assessment":self._row(assessment_row),"services":service_nodes})
            lawyers=[]
            for node in assessment_nodes:
                for service_node in node["services"]:
                    lawyer=_find(list(service_node["service"]),"Lawyers","Lawyer")
                    if lawyer and clean_id(service_node["service"].get(lawyer,"")):lawyers.append(clean_id(service_node["service"].get(lawyer,"")))
                lawyer=_find(list(node["assessment"]),"Lawyers","Lawyer")
                if lawyer and clean_id(node["assessment"].get(lawyer,"")):lawyers.append(clean_id(node["assessment"].get(lawyer,"")))
            beneficiary=self._row(row);lawyer=_find(list(beneficiary),"Lawyers","Lawyer")
            if lawyer and clean_id(beneficiary.get(lawyer,"")):lawyers.append(clean_id(beneficiary.get(lawyer,"")))
            item={"beneficiary":beneficiary,"lawyers":list(dict.fromkeys(lawyers)),"assessments":assessment_nodes,"counts":{"assessments":len(assessment_nodes),"services":total_services,"followups":total_followups,"fees":total_fees}}
            cases.append(item)
        return {"query":query,"cases":cases,"rows":[],"columns":[],"availableColumns":[],"totalRows":total_cases if view_mode=="table" else 0,"totalCases":total_cases,"page":page if view_mode=="table" else 1,"pageSize":page_size}

    def case_export(self, query: str, filters: dict[str, list[str]] | None = None, case_ids: list[str] | None = None) -> bytes:
        bulk_export=not query.strip() and not case_ids
        candidate_ids=self._case_candidate_ids(query,filters)
        if case_ids:
            candidate_ids &= {clean_id(case_id) for case_id in case_ids if clean_id(case_id)}
        def unique_records(frame:pd.DataFrame,*id_hints:str)->pd.DataFrame:
            identifier=_find(list(frame.columns),*id_hints)
            return frame.drop_duplicates(subset=[identifier]) if identifier else frame.drop_duplicates()
        beneficiaries=self.frames["beneficiaries"]
        beneficiary_id=_find(list(beneficiaries.columns),"Case ID")
        assessments=unique_records(self._related_set("assessments","Beneficiary ID",candidate_ids),"Assessment ID")
        assessment_id=_find(list(assessments.columns),"Assessment ID")
        assessment_ids=set(assessments[assessment_id].map(clean_id)) if assessment_id else set()
        services_frame=self.frames["legalservices"]
        service_beneficiary=_find(list(services_frame.columns),"Beneficiary ID")
        service_assessment=_find(list(services_frame.columns),"Assessment ID")
        service_mask=pd.Series(False,index=services_frame.index)
        if service_beneficiary:service_mask|=services_frame[service_beneficiary].map(clean_id).isin(candidate_ids)
        if service_assessment:service_mask|=services_frame[service_assessment].map(clean_id).isin(assessment_ids)
        services=unique_records(services_frame[service_mask],"Service ID")
        service_id=_find(list(services.columns),"Service ID")
        service_ids=set(services[service_id].map(clean_id)) if service_id else set()
        def service_children(dataset:str,*hints:str)->pd.DataFrame:
            frame=self.frames.get(dataset)
            if frame is None:return pd.DataFrame()
            foreign_key=_find(list(frame.columns),*hints)
            identifier=("Follow-ups & Logbook ID","Follow-up ID") if dataset=="followupslogbooks" else ("Fee ID","Legal Fee ID")
            return unique_records(frame[frame[foreign_key].map(clean_id).isin(service_ids)],*identifier) if foreign_key else frame.iloc[0:0]
        sheets=[
            ("Beneficiaries",unique_records(beneficiaries[beneficiaries[beneficiary_id].map(clean_id).isin(candidate_ids)],"Case ID") if beneficiary_id else beneficiaries.iloc[0:0]),
            ("Assessments",assessments),
            ("Services",services),
            ("Follow-ups",service_children("followupslogbooks","Service ID")),
            ("Fees",service_children("legalfees","Legal Service ID","Service ID")),
        ]
        output=io.BytesIO()
        with pd.ExcelWriter(output,engine="openpyxl",date_format="DD/MM/YYYY",datetime_format="DD/MM/YYYY") as writer:
            thin=Side(style="thin",color="D9E2EC");border=Border(bottom=thin)
            for sheet_name,frame in sheets:
                if sheet_name in {"Follow-ups","Fees"} and frame.empty:
                    continue
                safe=_safe_export(frame.copy())
                safe.to_excel(writer,index=False,sheet_name=sheet_name)
                sheet=writer.book[sheet_name]
                if not len(safe.columns):continue
                sheet.freeze_panes="A2";sheet.auto_filter.ref=f"A1:{get_column_letter(len(safe.columns))}{max(sheet.max_row,1)}"
                date_columns=[index for index,column in enumerate(safe.columns,1) if pd.api.types.is_datetime64_any_dtype(safe[column])]
                for column in date_columns:
                    for row in range(2,sheet.max_row+1):
                        sheet.cell(row,column).number_format="DD/MM/YYYY"
                for cell in sheet[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="174EA6");cell.alignment=Alignment(wrap_text=True,vertical="center");cell.border=border
                sheet.row_dimensions[1].height=34
                for index,column in enumerate(safe.columns,1):
                    values=[len(str(value)) for value in safe.iloc[:300,index-1].fillna("")] if len(safe) else [0]
                    sheet.column_dimensions[get_column_letter(index)].width=min(38,max(12,len(str(column))+2,max(values,default=0)+2))
                if bulk_export:
                    # Avoid per-cell style/height work for large all-case exports.
                    # The workbook is still filtered, frozen, and has styled headers.
                    continue
                for row in range(2,sheet.max_row+1):
                    if row%2==0:
                        for cell in sheet[row]:cell.fill=PatternFill("solid",fgColor="F4F8FC")
                    for cell in sheet[row]:cell.alignment=Alignment(vertical="top",wrap_text=True);cell.border=border
                # Excel does not auto-fit wrapped text. Estimate its line count
                # from the final column width so case data is visible on open.
                for row in range(2,sheet.max_row+1):
                    lines=1
                    for cell in sheet[row]:
                        width=max(1,int(sheet.column_dimensions[get_column_letter(cell.column)].width or 12)-1)
                        text=str(cell.value or "")
                        lines=max(lines,sum(max(1,(len(part)+width-1)//width) for part in text.splitlines() or [""]))
                    sheet.row_dimensions[row].height=min(180,max(21,15*lines+6))
            format_excel_dates(writer.book)
        return output.getvalue()

    def _related(self,dataset:str,column_hint:str,value:str)->pd.DataFrame: return self._related_set(dataset,column_hint,{value})
    def _related_set(self,dataset:str,column_hint:str,values:set[str])->pd.DataFrame:
        df=self.frames[dataset];cache_key=(dataset,column_hint.lower());index=self._relation_cache.get(cache_key)
        if index is None:
            column=_find(list(df.columns),column_hint)
            if not column:return df.iloc[0:0]
            index=defaultdict(list)
            for row_index,value in zip(df.index,df[column].map(clean_id)):index[value].append(row_index)
            self._relation_cache[cache_key]=dict(index);index=self._relation_cache[cache_key]
        selected=[row_index for value in values for row_index in index.get(value,[])]
        return df.loc[selected] if selected else df.iloc[0:0]
    @staticmethod
    def _row(row:pd.Series)->dict[str,Any]: return {str(k):display_value(v) for k,v in row.items()}

    def lawyer_summary(self, filters:dict[str,list[str]]|None=None) -> dict[str, Any]:
        filters=filters or {};cache_key=tuple(sorted((key,tuple(sorted(values))) for key,values in filters.items() if values))
        if cache_key in self._lawyer_cache:return self._lawyer_cache[cache_key]
        filter_hints={"lawyer":("Lawyers","Lawyer"),"createdBy":("Original Created By","Created By","Created by"),"project":("Project",),"location":("Project Location","Location"),"assessmentMonth":("Date of Assessment",)}
        def filtered(dataset:str)->pd.DataFrame:
            frame=self.frames[dataset]
            for key,selections in filters.items():
                column=_find(list(frame.columns),*filter_hints.get(key,(key,)))
                if column and selections:
                    if key=="assessmentMonth":
                        months=pd.to_datetime(frame[column],errors="coerce",dayfirst=True).dt.strftime("%Y-%m")
                        frame=frame[months.isin(selections)]
                    else:frame=frame[frame[column].fillna("").astype(str).str.strip().isin(selections)]
            return frame
        options={key:set() for key in filter_hints}
        for frame in self.frames.values():
            for key,hints in filter_hints.items():
                column=_find(list(frame.columns),*hints)
                if column:
                    if key=="assessmentMonth":options[key].update(value for value in pd.to_datetime(frame[column],errors="coerce",dayfirst=True).dt.strftime("%Y-%m").dropna().unique())
                    else:options[key].update(str(value).strip() for value in frame[column].dropna().unique() if str(value).strip())
        result=[]
        for dataset in ("beneficiaries","assessments","legalservices","followupslogbooks","awareness"):
            if dataset not in self.frames: continue
            df=filtered(dataset); lawyer=_find(list(df.columns),"Lawyers","Lawyer")
            if not lawyer: continue
            id_hint={"beneficiaries":"Case ID","assessments":"Assessment ID","legalservices":"Service ID","followupslogbooks":"Follow-ups & Logbook ID","awareness":"Awareness ID"}[dataset]
            ident=_find(list(df.columns),id_hint)
            grouped=df.assign(_lawyer=df[lawyer].fillna("Unassigned").astype(str).str.strip()).groupby("_lawyer")
            for person, group in grouped: result.append({"lawyer":person or "Unassigned","metric":DISPLAY_NAMES[dataset],"count":int(group[ident].map(clean_id).nunique()) if ident else len(group)})
        assessments=filtered("assessments")
        assessment_date=_find(list(assessments.columns),"Date of Assessment")
        lawyer=_find(list(assessments.columns),"Lawyers","Lawyer"); breakdowns=[]
        if lawyer:
            for hint,label in (("Gender","Gender"),("Community Type","Community type"),("Nationality","Nationality"),("Assessment Status","Assessment status"),("Is the beneficiary detained","Detention"),("Is it an immigration related charge","Immigration charge")):
                column=_find(list(assessments.columns),hint)
                if column:
                    grouped=assessments.assign(_lawyer=assessments[lawyer].fillna("Unassigned"),_value=assessments[column].fillna("Blank")).groupby(["_lawyer","_value"]).size()
                    breakdowns.extend({"lawyer":str(person),"dimension":label,"value":str(answer),"count":int(count)} for (person,answer),count in grouped.items())
        services=filtered("legalservices");charts=[]
        def add_chart(frame:pd.DataFrame,title:str,dimension:str,hints:tuple[str,...],kind:str="bar"):
            column=_find(list(frame.columns),*hints)
            if not column:return
            counts=frame[column].fillna("Blank").astype(str).str.strip().replace("","Blank").value_counts().head(12)
            charts.append({"title":title,"dimension":dimension,"kind":kind,"items":[{"label":str(label),"count":int(count)} for label,count in counts.items()]})
        add_chart(assessments,"Assessment status","Assessment status",("Assessment Status",),"donut")
        service_type=_find(list(services.columns),"Type of Service Provided","Type of Legal Service Needed")
        representation=services[services[service_type].fillna("").astype(str).str.contains("representation",case=False,regex=False)] if service_type else services.iloc[0:0]
        add_chart(representation,"Representation service status","Service status",("Service Status","Status"),"donut")
        if service_type:
            service_labels=services[service_type].fillna("").astype(str)
            service_type_items=[]
            for label,pattern in (("Legal Representation",r"assistance|representation"),("Legal Counselling",r"counselling|counseling")):
                count=int(service_labels.str.contains(pattern,case=False,regex=True).sum())
                if count:service_type_items.append({"label":label,"count":count})
            charts.append({
                "title":"Service type",
                "dimension":"Legal assistance and representation",
                "kind":"donut",
                "items":service_type_items,
            })
        beneficiary_id=_find(list(assessments.columns),"Beneficiary ID");assessment_id=_find(list(assessments.columns),"Assessment ID");service_id=_find(list(services.columns),"Service ID")
        total_beneficiaries=int(assessments[beneficiary_id].map(clean_id).nunique()) if beneficiary_id else 0
        total_assessments=int(assessments[assessment_id].map(clean_id).nunique()) if assessment_id else len(assessments)
        total_services=int(services[service_id].map(clean_id).nunique()) if service_id else len(services)
        monthly_assessments=[]
        if lawyer and assessment_date:
            dated=assessments.copy()
            dated["_assessment_date"]=pd.to_datetime(dated[assessment_date],errors="coerce",dayfirst=True)
            dated=dated[dated["_assessment_date"].notna()]
            dated["_lawyer"]=dated[lawyer].fillna("Unassigned").astype(str).str.strip().replace("","Unassigned")
            dated["_month"]=dated["_assessment_date"].dt.to_period("M").astype(str)
            if len(dated):
                month_count=int(dated["_month"].nunique())
                if assessment_id:
                    grouped=dated.assign(_assessment_id=dated[assessment_id].map(clean_id)).groupby(["_lawyer","_month"])["_assessment_id"].nunique()
                    lawyer_totals=dated.assign(_assessment_id=dated[assessment_id].map(clean_id)).groupby("_lawyer")["_assessment_id"].nunique()
                else:
                    grouped=dated.groupby(["_lawyer","_month"]).size()
                    lawyer_totals=dated.groupby("_lawyer").size()
                averages=(lawyer_totals/month_count).to_dict()
                monthly_assessments=[{"lawyer":str(person),"month":str(month),"count":int(count),"average":float(averages.get(person,0))} for (person,month),count in grouped.items()]
        rep_status=_find(list(representation.columns),"Service Status","Status") if len(representation) else None
        completed=int(representation[rep_status].fillna("").astype(str).str.contains("complete|closed|provided|done",case=False,regex=True).sum()) if rep_status else 0
        kpis=[
            {"label":"Distinct beneficiaries","value":total_beneficiaries,"detail":"Across filtered assessments"},
            {"label":"Assessments","value":total_assessments,"detail":f"{total_assessments/max(total_beneficiaries,1):.1f} per beneficiary"},
            {"label":"Legal services","value":total_services,"detail":f"{total_services/max(total_assessments,1):.1f} per assessment"},
            {"label":"Representation services","value":len(representation),"detail":f"{len(representation)/max(total_services,1)*100:.1f}% of services"},
            {"label":"Representation completed","value":completed,"detail":f"{completed/max(len(representation),1)*100:.1f}% completion rate"},
        ]
        payload={"rows":result,"monthlyAssessments":monthly_assessments,"breakdowns":breakdowns,"charts":charts,"kpis":kpis,"filterOptions":{key:sorted(values) for key,values in options.items()},"activeFilters":filters,"availability":{name:name in self.frames for name in FILES}};self._lawyer_cache[cache_key]=payload;return payload

    def representation_case_load(self, filters:dict[str,list[str]]|None=None, status:str="open") -> dict[str,Any]:
        """Aggregate representation services by lawyer, document type, and the applicable service month."""
        filters=filters or {}
        services=self.frames.get("legalservices",pd.DataFrame()).copy()
        if services.empty:return {"status":status,"months":[],"rows":[]}
        filter_hints={"lawyer":("Lawyers","Lawyer"),"createdBy":("Original Created By","Created By","Created by"),"project":("Projects - المشروع","Project"),"location":("Project Location","Project location","Governorate")}
        service_type=_find(list(services.columns),"Type of Service Provided","Type of Legal Service Needed")
        service_status=_find(list(services.columns),"Service Status","Status")
        date_column=_find(list(services.columns),"Date of Service Provision") if status=="open" else _find(list(services.columns),"Date Service Completed","Date of Service Close")
        if not service_type or not service_status or not date_column:return {"status":status,"months":[],"rows":[]}
        for key,selections in filters.items():
            if not selections:continue
            if key=="assessmentMonth":
                months=pd.to_datetime(services[date_column],errors="coerce",dayfirst=True).dt.strftime("%Y-%m")
                services=services[months.isin(selections)]
                continue
            column=_find(list(services.columns),*filter_hints.get(key,(key,)))
            if column:services=services[services[column].fillna("").astype(str).str.strip().isin(selections)]
        representation=services[service_type].fillna("").astype(str).str.contains("representation",case=False,regex=False)
        state=services[service_status].fillna("").astype(str)
        state_match=state.str.contains(r"in[ -]?(?:progress|process)",case=False,regex=True) if status=="open" else state.str.contains("closed|completed",case=False,regex=True)
        working=services[representation&state_match].copy()
        working["_month"]=pd.to_datetime(working[date_column],errors="coerce",dayfirst=True).dt.strftime("%Y-%m")
        working=working[working["_month"].notna()]
        lawyer=_find(list(working.columns),"Lawyers","Lawyer")
        document=_find(list(working.columns),"Type of Document")
        service_id=_find(list(working.columns),"Service ID")
        if not lawyer:return {"status":status,"months":[],"rows":[]}
        working["_lawyer"]=working[lawyer].fillna("").astype(str).str.strip().replace("","Unassigned")
        working["_document"]=working[document].fillna("").astype(str).str.strip().replace("","Blank") if document else "Blank"
        working["_service"]=working[service_id].map(clean_id) if service_id else working.index.astype(str)
        beneficiary=_find(list(working.columns),"Beneficiary ID","Case ID")
        assessment=_find(list(working.columns),"Assessment ID")
        provision_date=_find(list(working.columns),"Date of Service Provision")
        close_date=_find(list(working.columns),"Date Service Completed","Date of Service Close")
        details:dict[tuple[str,str,str],list[dict[str,str]]]=defaultdict(list)
        for _,record in working.iterrows():
            key=(str(record["_lawyer"]),str(record["_document"]),str(record["_month"]))
            details[key].append({"serviceId":str(record["_service"]),"beneficiaryId":str(record[beneficiary]) if beneficiary else "","assessmentId":str(record[assessment]) if assessment else "","lawyer":key[0],"document":key[1],"status":str(record[service_status]),"provisionDate":str(record[provision_date]) if provision_date else "","closeDate":str(record[close_date]) if close_date else "","month":key[2]})
        grouped=working.groupby(["_lawyer","_document","_month"])["_service"].nunique()
        rows=[]
        for (person,doc,month),count in grouped.items():
            key=(str(person),str(doc),str(month));rows.append({"lawyer":key[0],"document":key[1],"month":key[2],"count":int(count),"services":details[key]})
        return {"status":status,"months":sorted(working["_month"].unique()),"rows":rows}

    def intelligence(self, page:str, filters:dict[str,list[str]]|None=None) -> dict[str,Any]:
        """Cross-dataset summaries with distinct source denominators and no join multiplication."""
        filters=filters or {};cache_key=(page,tuple(sorted((key,tuple(sorted(values))) for key,values in filters.items() if values)))
        if cache_key in self._intelligence_cache:return self._intelligence_cache[cache_key]
        filter_hints={"lawyer":("Lawyers","Lawyer"),"createdBy":("Original Created By","Created By","Created by"),"project":("Projects - المشروع","Project"),"location":("Project Location","Project location","Governorate"),"assessmentMonth":("Date of Assessment",)}
        def scoped(name:str)->pd.DataFrame:
            frame=self.frames.get(name,pd.DataFrame()).copy()
            for key,selections in filters.items():
                column=_find(list(frame.columns),*filter_hints.get(key,(key,)))
                if column and selections:
                    if key=="assessmentMonth":
                        months=pd.to_datetime(frame[column],errors="coerce",dayfirst=True).dt.strftime("%Y-%m")
                        frame=frame[months.isin(selections)]
                    else:frame=frame[frame[column].fillna("").astype(str).str.strip().isin(selections)]
            return frame
        frames={name:scoped(name) for name in FILES}
        assessments=frames["assessments"]
        assessment_date=_find(list(assessments.columns),"Date of Assessment")
        def distinct(name:str,*hints:str)->int:
            frame=frames.get(name,pd.DataFrame());column=_find(list(frame.columns),*hints)
            return int(frame[column].map(clean_id).replace("",pd.NA).nunique()) if column else len(frame)
        counts={
            "beneficiaries":distinct("beneficiaries","Case ID"),"assessments":distinct("assessments","Assessment ID"),
            "services":distinct("legalservices","Service ID"),"followups":distinct("followupslogbooks","Follow-ups & Logbook ID","Follow-up ID"),
            "fees":distinct("legalfees","Fee ID"),"deportations":distinct("deportationrecords","PN ID"),"awareness":distinct("awareness","Awareness ID"),
        }
        services=frames["legalservices"];service_status=_find(list(services.columns),"Service Status","Status")
        completed_mask=services[service_status].fillna("").astype(str).str.contains("complete|closed|provided|done|اكتملت",case=False,regex=True) if service_status else pd.Series(False,index=services.index)
        completed_services=int(services.loc[completed_mask,_find(list(services.columns),"Service ID")].map(clean_id).nunique()) if len(services) and _find(list(services.columns),"Service ID") else int(completed_mask.sum())
        service_type=_find(list(services.columns),"Type of Service Provided","Type of Legal Service Needed")
        representation=services[services[service_type].fillna("").astype(str).str.contains("representation",case=False,regex=False)] if service_type else services.iloc[0:0]
        representation_completed=int(completed_mask.reindex(representation.index,fill_value=False).sum())
        representation_in_process=int(representation[service_status].fillna("").astype(str).str.contains(r"in[ -]?(?:progress|process)",case=False,regex=True).sum()) if service_status else 0
        fees=frames["legalfees"];amount_col=_find(list(fees.columns),"Amount Spent","Total Cost")
        amounts=pd.to_numeric(fees[amount_col].astype(str).str.replace(r"[^0-9.\-]","",regex=True),errors="coerce").fillna(0) if amount_col else pd.Series(0,index=fees.index,dtype=float)
        total_fees=float(amounts.sum())
        followups=frames["followupslogbooks"]
        next_follow=_find(list(followups.columns),"Date of next follow up");follow_complete=_find(list(followups.columns),"Service completed")
        overdue=0
        if next_follow:
            next_dates=pd.to_datetime(followups[next_follow],errors="coerce",dayfirst=True)
            open_mask=~followups[follow_complete].fillna("").astype(str).str.contains(r"\byes\b|complete|نعم",case=False,regex=True) if follow_complete else pd.Series(True,index=followups.index)
            overdue=int((next_dates.lt(pd.Timestamp.today().normalize())&open_mask).sum())
        assessment_id=_find(list(assessments.columns),"Assessment ID");beneficiary_id=_find(list(assessments.columns),"Beneficiary ID")
        detained_column=_find(list(assessments.columns),"Is the beneficiary detained")
        detained_mask=assessments[detained_column].fillna("").astype(str).str.contains(r"\byes\b|نعم",case=False,regex=True) if detained_column else pd.Series(False,index=assessments.index)
        detention_count=int(assessments.loc[detained_mask,assessment_id].map(clean_id).replace("",pd.NA).nunique()) if assessment_id else int(detained_mask.sum())
        amal_project_exists=any(
            bool(project_column and frame[project_column].fillna("").astype(str).str.contains(r"\bamal\b",case=False,regex=True).any())
            for frame in frames.values()
            for project_column in [_find(list(frame.columns),"Projects - المشروع","Project")]
        )
        selected_projects=[str(value).strip() for value in filters.get("project",[]) if str(value).strip()]
        scoped_projects={
            str(value).strip()
            for frame in frames.values()
            for project_column in [_find(list(frame.columns),"Projects - المشروع","Project")]
            if project_column
            for value in frame[project_column].dropna()
            if str(value).strip()
        }
        amal_only=(bool(selected_projects) and all(re.search(r"\bamal\b",value,re.I) for value in selected_projects)) or (not selected_projects and bool(scoped_projects) and all(re.search(r"\bamal\b",value,re.I) for value in scoped_projects))
        service_assessment=_find(list(services.columns),"Assessment ID")
        assessment_ids=set(assessments[assessment_id].map(clean_id)) if assessment_id else set()
        linked_assessments=set(services[service_assessment].map(clean_id)) if service_assessment else set()
        missing_services=len(assessment_ids-linked_assessments)
        funnel=[
            {"label":"Beneficiaries","value":counts["beneficiaries"]},{"label":"Assessments","value":distinct("assessments","Assessment ID")},
            {"label":"Assessments with services","value":len(assessment_ids&linked_assessments)},{"label":"Services delivered","value":counts["services"]},
            {"label":"Completed services","value":completed_services},{"label":"Follow-ups recorded","value":counts["followups"]},
        ]
        event_specs={
            "assessments":("Assessments",("Date of Assessment",),("Assessment ID",)),"legalservices":("Services",("Date of Service Provision",),("Service ID",)),
            "followupslogbooks":("Follow-ups",("Date of follow-up",),("Follow-ups & Logbook ID","Follow-up ID")),"deportationrecords":("Deportations",("Date of deporting",),("PN ID",)),
            "awareness":("Awareness",("Date of Session","Added On"),("Awareness ID",)),
        }
        monthly:dict[str,dict[str,Any]]={}
        for name,(label,date_hints,id_hints) in event_specs.items():
            frame=frames.get(name,pd.DataFrame());date_col=_find(list(frame.columns),*date_hints);id_col=_find(list(frame.columns),*id_hints)
            if not date_col:continue
            work=frame.assign(_month=pd.to_datetime(frame[date_col],errors="coerce",dayfirst=True).dt.strftime("%Y-%m"))
            work=work[work["_month"].notna()]
            series=work.groupby("_month")[id_col].nunique() if id_col else work.groupby("_month").size()
            for month,value in series.items():monthly.setdefault(str(month),{"month":str(month)})[label]=int(value)
        monthly_rows=[{"month":month,**{label:int(row.get(label,0)) for label,_,_ in event_specs.values()}} for month,row in sorted(monthly.items())]
        geography:dict[str,dict[str,Any]]={}
        for name,label in (("beneficiaries","Beneficiaries"),("assessments","Assessments"),("legalservices","Services"),("deportationrecords","Deportations"),("awareness","Awareness")):
            frame=frames.get(name,pd.DataFrame());column=_find(list(frame.columns),"Governorate","Project Location","Project location")
            if not column:continue
            for raw,count in frame[column].fillna("").astype(str).value_counts().items():
                place=normalize_governorate(raw)
                if place:geography.setdefault(place,{"label":place.title(),"Beneficiaries":0,"Assessments":0,"Services":0,"Deportations":0,"Awareness":0})[label]+=int(count)
        breakdowns=[]
        for title,name,hints in (("Service mix","legalservices",("Type of Service Provided",)),("Nationality","assessments",("Nationality",))):
            frame=frames.get(name,pd.DataFrame());column=_find(list(frame.columns),*hints)
            if column:
                raw_values=frame[column].fillna("Blank").astype(str).str.strip().replace("","Blank")
                if title=="Service mix":
                    raw_values=raw_values.map(lambda value:"Legal Representation" if re.search(r"assistance|representation",value,re.I) else "Legal Counselling" if re.search(r"counselling|counseling",value,re.I) else value)
                values=raw_values.value_counts().head(8)
                breakdowns.append({"title":title,"total":int(len(raw_values)),"items":[{"label":str(label),"value":int(value)} for label,value in values.items()]})
        document_type=_find(list(representation.columns),"Type of Documents to be issued","Type of Document")
        if document_type:
            document_values=representation[document_type].fillna("Blank").astype(str).str.strip().replace("","Blank")
            document_counts=document_values.value_counts().head(8)
            breakdowns.append({"title":"Representation document type","total":int(len(document_values)),"items":[{"label":str(label),"value":int(value)} for label,value in document_counts.items()]})
        for title,name,hints in (("Population","assessments",("Community Type",)),("Project","assessments",("Projects - المشروع","Project"))):
            frame=frames.get(name,pd.DataFrame());column=_find(list(frame.columns),*hints)
            if column:
                raw_values=frame[column].fillna("Blank").astype(str).str.strip().replace("","Blank")
                values=raw_values.value_counts().head(8)
                breakdowns.append({"title":title,"total":int(len(raw_values)),"items":[{"label":str(label),"value":int(value)} for label,value in values.items()]})
        lawyer_names=set()
        for frame in frames.values():
            column=_find(list(frame.columns),"Lawyers","Lawyer","Created by")
            if column:lawyer_names.update(x for x in frame[column].fillna("").astype(str).str.strip() if x)
        lawyer_rows=[]
        assessment_month_count=0
        if assessment_date and len(assessments):assessment_month_count=int(pd.to_datetime(assessments[assessment_date],errors="coerce",dayfirst=True).dt.to_period("M").nunique())
        assessment_project=_find(list(assessments.columns),"Projects - المشروع","Project")
        projects=sorted({str(value).strip() for value in assessments[assessment_project].dropna() if str(value).strip()}) if assessment_project else ["Unassigned project"]
        for project in projects:
          for person in sorted(lawyer_names):
            row={"project":project,"lawyer":person}
            for name,label,id_hints in (("assessments","assessments",("Assessment ID",)),("legalservices","services",("Service ID",)),("followupslogbooks","followups",("Follow-ups & Logbook ID","Follow-up ID")),("legalfees","fees",("Fee ID",)),("deportationrecords","deportations",("PN ID",)),("awareness","awareness",("Awareness ID",))):
                frame=frames.get(name,pd.DataFrame());column=_find(list(frame.columns),"Lawyers","Lawyer","Created by");id_col=_find(list(frame.columns),*id_hints)
                selected=frame[frame[column].fillna("").astype(str).str.strip().eq(person)] if column else frame.iloc[0:0]
                project_column=_find(list(selected.columns),"Projects - المشروع","Project")
                if project_column:selected=selected[selected[project_column].fillna("").astype(str).str.strip().eq(project)]
                row[label]=int(selected[id_col].map(clean_id).nunique()) if id_col else len(selected)
            lawyer_assessments=assessments[assessments[_find(list(assessments.columns),"Lawyers","Lawyer")].fillna("").astype(str).str.strip().eq(person)] if _find(list(assessments.columns),"Lawyers","Lawyer") else assessments.iloc[0:0]
            if assessment_project:lawyer_assessments=lawyer_assessments[lawyer_assessments[assessment_project].fillna("").astype(str).str.strip().eq(project)]
            lawyer_services=services[services[_find(list(services.columns),"Lawyers","Lawyer")].fillna("").astype(str).str.strip().eq(person)] if _find(list(services.columns),"Lawyers","Lawyer") else services.iloc[0:0]
            service_project=_find(list(lawyer_services.columns),"Projects - المشروع","Project")
            if service_project:lawyer_services=lawyer_services[lawyer_services[service_project].fillna("").astype(str).str.strip().eq(project)]
            lawyer_status=_find(list(lawyer_services.columns),"Service Status","Status")
            lawyer_completed=int(lawyer_services[lawyer_status].fillna("").astype(str).str.contains("complete|closed|provided|done|اكتملت",case=False,regex=True).sum()) if lawyer_status else 0
            fee_lawyer=_find(list(fees.columns),"Lawyers","Lawyer","Created by")
            lawyer_fee_rows=fees[fees[fee_lawyer].fillna("").astype(str).str.strip().eq(person)] if fee_lawyer else fees.iloc[0:0]
            fee_project=_find(list(lawyer_fee_rows.columns),"Projects - المشروع","Project")
            if fee_project:lawyer_fee_rows=lawyer_fee_rows[lawyer_fee_rows[fee_project].fillna("").astype(str).str.strip().eq(project)]
            lawyer_amounts=pd.to_numeric(lawyer_fee_rows[amount_col].astype(str).str.replace(r"[^0-9.\-]","",regex=True),errors="coerce").fillna(0) if amount_col else pd.Series(dtype=float)
            row.update({"monthlyAverage":len(lawyer_assessments)/max(assessment_month_count,1),"completedServices":lawyer_completed,"completionRate":lawyer_completed/max(row["services"],1),"averageCost":float(lawyer_amounts.sum())/max(lawyer_completed,1)})
            if row["assessments"] or row["services"] or row["followups"] or row["fees"] or row["awareness"]:lawyer_rows.append(row)
        lawyer_rows.sort(key=lambda row:(str(row["project"]),-int(row["assessments"]),str(row["lawyer"])))
        risks=[
            {"label":"Assessments without services","value":missing_services,"severity":"high" if missing_services else "low"},
            {"label":"Overdue follow-ups","value":overdue,"severity":"high" if overdue else "low"},
            {"label":"Data-quality findings","value":sum(len(items) for items in self.flags.values()),"severity":"medium"},
            {"label":"Deportation records","value":counts["deportations"],"severity":"medium" if counts["deportations"] else "low"},
        ]
        insights=[]
        if monthly_rows:
            latest=monthly_rows[-1];previous=monthly_rows[-2] if len(monthly_rows)>1 else None
            if previous:
                change=latest["Assessments"]-previous["Assessments"]
                insights.append({"title":"Assessment momentum","detail":f"{abs(change)} {'more' if change>=0 else 'fewer'} assessments than {previous['month']}.","tone":"positive" if change>=0 else "attention"})
        insights.extend([
            {"title":"Service conversion","detail":f"{completed_services/max(counts['services'],1)*100:.1f}% of services are completed.","tone":"positive" if completed_services/max(counts['services'],1)>=.7 else "attention"},
            {"title":"Follow-up attention","detail":f"{overdue} follow-ups have a past next-action date and are not marked completed.","tone":"attention" if overdue else "positive"},
            {"title":"Detention added","detail":f"{counts['deportations']} deportation records are included separately from detention-related assessments.","tone":"neutral"},
        ])
        lawyer_summary=self.lawyer_summary(filters)
        kpis=[
            {"label":"Beneficiaries","value":counts["beneficiaries"],"format":"number"},{"label":"Assessments","value":counts["assessments"],"format":"number"},
            {"label":"Legal services","value":counts["services"],"format":"number"},{"label":"Completed services","value":completed_services,"format":"number"},
            {"label":"Representation services","value":len(representation),"format":"number"},{"label":"Representation completed","value":representation_completed,"format":"number"},
            {"label":"In-Process Legal Representation","value":representation_in_process,"format":"number"},
            {"label":"Follow-ups","value":counts["followups"],"format":"number"},{"label":"Legal fees","value":total_fees,"format":"currency"},
            {"label":"Deportations","value":counts["deportations"],"format":"number"},{"label":"Awareness participants","value":counts["awareness"],"format":"number"},
        ]
        if page=="lawyer-intelligence":
            kpis=[item for item in kpis if item["label"] not in {"Deportations","Awareness participants"}]
            if not amal_only:kpis.append({"label":"Detentions","value":detention_count,"format":"number"})
            if amal_project_exists:kpis.append({"label":"Awareness participants","value":counts["awareness"],"format":"number"})
        payload={"page":page,"period":"2026 onward","kpis":kpis,"funnel":funnel,"monthly":monthly_rows,"geography":list(geography.values()),"breakdowns":breakdowns,"lawyers":lawyer_rows,"lawyerSummary":{"rows":lawyer_summary["rows"],"monthlyAssessments":lawyer_summary["monthlyAssessments"],"charts":lawyer_summary["charts"]},"risks":risks,"insights":insights,
        "finance":{"total":total_fees,"averagePerCompletedService":total_fees/max(completed_services,1),"records":counts["fees"]},
        "filterOptions":{key:sorted({str(value).strip() for frame in self.frames.values() for hint in hints for column in [_find(list(frame.columns),hint)] if column for value in frame[column].dropna().unique() if str(value).strip()}) for key,hints in filter_hints.items()},
        "activeFilters":filters,"availability":{name:name in self.frames for name in FILES}}
        self._intelligence_cache[cache_key]=payload;return payload

    def detention_cases(self, search:str="", page:int=1, page_size:int=100, filters:dict[str,list[str]]|None=None, sort_column:str="", sort_direction:str="asc") -> dict[str,Any]:
        frame=self.frames["assessments"];detained=_find(list(frame.columns),"Is the beneficiary detained")
        if not detained:return {"total":0,"page":page,"pageSize":page_size,"columns":[],"rows":[],"kpis":[],"filterOptions":{}}
        yes=frame[detained].fillna("").astype(str).str.lower().str.contains(r"\byes\b|نعم",regex=True)
        scoped=frame[yes].copy()
        column_specs=(
            ("Project",("Projects","Project")),("Project location",("Project Location",)),("Beneficiary ID",("Beneficiary ID",)),
            ("Assessment ID",("Assessment ID",)),("Assessment date",("Date of Assessment",)),("Name",("Name /", "Name")),
            ("Gender",("Gender",)),("Age",("Age",)),("Gender / age group",("Age Gender Group","Gender Age Group")),
            ("Community type",("Community Type",)),("Nationality",("Nationality",)),("Detained",("Is the beneficiary detained",)),
            ("Transferred from NES",("transferred from NES",)),("Detention date",("Date of Detention",)),
            ("Detention governorate",("Detention Governorate",)),("Detaining authority",("Detaining Authority",)),
            ("Place of detention",("Place of detention",)),("Reasons for detention",("Reasons for Detention",)),
            ("Possible charges",("Possible Charges",)),("Immigration charge",("immigration related charge",)),
            ("Current status",("Detainee current status",)),("Release type",("Type of Released",)),
            ("Release/deportation date",("Date of the released or deported","Date of Released","Date of Release")),("Comments",("Comments on detention",)),
            ("Lawyer",("Lawyers", "Lawyer")),
        )
        selected:list[tuple[str,str]]=[]
        for label,hints in column_specs:
            column=_find(list(frame.columns),*hints)
            if column and column not in {item[1] for item in selected}:selected.append((label,column))
        detail_specs=(
            ("Lawyers",("Lawyers","Lawyer")),
            ("Projects - المشروع",("Projects","Project")),
            ("Project Location",("Project Location",)),
            ("Date of Assessment تاريخ التقييم",("Date of Assessment",)),
            ("Beneficiary ID",("Beneficiary ID",)),
            ("Name / الأسم",("Name /", "Name")),
            ("DoB / تأريخ الولاده",("DoB /","Date of Birth","DoB")),
            ("Age",("Age",)),
            ("Gender النوع الاجتماعي",("Gender",)),
            ("Date of Detention تاريخ الاحتجاز",("Date of Detention",)),
            ("Detention Governorate / محافظة الاحتجاز",("Detention Governorate",)),
            ("Detaining Authority جهة الاحتجاز",("Detaining Authority",)),
            ("Place of detention - مكان الاحتجاز",("Place of detention",)),
            ("Reasons for Detention أسباب الاحتجاز",("Reasons for Detention",)),
            ("Possible Charges التهم المحتملة",("Possible Charges",)),
            ("Nationality الجنسية",("Nationality",)),
            ("Name of the reporting person اسم الشخص المبلغ",("Name of the reporting person",)),
            ("Relationship to the detainee العلاقة بالمحتجز",("Relationship to the detainee",)),
            ("Phone number of the reporter رقم هاتف المبلغ",("Phone number of the reporter",)),
            ("Type of Legal Service Needed",("Type of Legal Service Needed",)),
            ("Detainee current status حالة المعتقل الحالية",("Detainee current status",)),
            ("Type of Released نوع الافراج",("Type of Released",)),
            ("Date of the released or deported تاريخ الافراج او الترحيل",("Date of the released or deported",)),
            ("Reasons for Detention if Other - أسباب الاحتجاز اذا كانت اخرى",("Reasons for Detention if Other",)),
            ("Is it an immigration related charge? هل هو معتقل على اساس قانون الاقامة ؟",("immigration related charge",)),
            ("Comments on detention التعليقات على الاحتجاز",("Comments on detention",)),
            ('Detaining Authority "If Other" - " جهة الاحتجاز "إذا كانت أخرى',("Detaining Authority \"If Other\"","Detaining Authority If Other")),
        )
        detail_selected=[(label,_find(list(frame.columns),*hints)) for label,hints in detail_specs]
        columns_by_label=dict(selected)
        filter_labels=("Gender / age group","Community type","Nationality","Project","Project location","Detention governorate","Detaining authority","Immigration charge","Current status","Lawyer")
        options={label:sorted({str(value).strip() for value in scoped[column].dropna().unique() if str(value).strip()}) for label,column in selected if label in filter_labels}
        assessment_date=columns_by_label.get("Assessment date")
        release_date=columns_by_label.get("Release/deportation date")
        month_options=lambda column: sorted({str(value) for value in pd.to_datetime(scoped[column],errors="coerce",dayfirst=True).dt.to_period("M").dropna()},reverse=True) if column else []
        options["Date of Assessment"]=month_options(assessment_date)
        options["Date of the released"]=month_options(release_date)
        options["Type of Released"]=sorted({str(value).strip() for value in scoped[columns_by_label["Release type"]].dropna().unique() if str(value).strip()}) if columns_by_label.get("Release type") else []
        for label,values in (filters or {}).items():
            if label=="month":continue
            if label=="Date of Assessment" and assessment_date and values:
                scoped=scoped[pd.to_datetime(scoped[assessment_date],errors="coerce",dayfirst=True).dt.to_period("M").astype(str).isin(values)]
                continue
            if label=="Date of the released" and release_date and values:
                scoped=scoped[pd.to_datetime(scoped[release_date],errors="coerce",dayfirst=True).dt.to_period("M").astype(str).isin(values)]
                continue
            column=columns_by_label.get("Release type" if label=="Type of Released" else label)
            if column and values:scoped=scoped[scoped[column].fillna("").astype(str).isin(values)]
        trend_scope=scoped.copy()
        month_values=(filters or {}).get("month",[])
        if assessment_date and month_values:
            months=pd.to_datetime(scoped[assessment_date],errors="coerce").dt.to_period("M").astype(str)
            scoped=scoped[months.isin(month_values)]
        if search:
            needle=search.lower();mask=pd.Series(False,index=scoped.index)
            for column in dict.fromkeys(column for _,column in detail_selected if column):mask|=scoped[column].fillna("").astype(str).str.lower().str.contains(needle,regex=False)
            scoped=scoped[mask]
        status_col=columns_by_label.get("Current status");immigration_col=columns_by_label.get("Immigration charge");beneficiary_col=columns_by_label.get("Beneficiary ID")
        active=int(scoped[status_col].fillna("").astype(str).str.contains("detain|detention|custody|held|لازال",case=False,regex=True).sum()) if status_col else len(scoped)
        immigration_count=int(scoped[immigration_col].fillna("").astype(str).str.contains(r"\byes\b|نعم",case=False,regex=True).sum()) if immigration_col else 0
        distinct=int(scoped[beneficiary_col].map(clean_id).nunique()) if beneficiary_col else len(scoped)
        trend=[]
        released_scope=trend_scope.copy()
        if status_col:
            released_scope=released_scope[released_scope[status_col].fillna("").astype(str).str.contains("released",case=False,regex=False)]
        else: released_scope=released_scope.iloc[0:0]
        if release_date:
            released_dates=pd.to_datetime(released_scope[release_date],errors="coerce")
            released_scope=released_scope[released_dates.notna()]
            if month_values: released_scope=released_scope[released_dates.dt.to_period("M").astype(str).isin(month_values)]
        else: released_scope=released_scope.iloc[0:0]
        if search and not released_scope.empty:
            release_search=pd.Series(False,index=released_scope.index)
            for column in dict.fromkeys(column for _,column in detail_selected if column):release_search|=released_scope[column].fillna("").astype(str).str.lower().str.contains(search.lower(),regex=False)
            released_scope=released_scope[release_search]
        monthly=[]
        if assessment_date:
            parsed=pd.to_datetime(trend_scope[assessment_date],errors="coerce");monthly.append(parsed.dropna().dt.to_period("M"))
        if release_date:
            trend_released=trend_scope[trend_scope[status_col].fillna("").astype(str).str.contains("released",case=False,regex=False)] if status_col else trend_scope.iloc[0:0]
            parsed_release=pd.to_datetime(trend_released[release_date],errors="coerce");monthly.append(parsed_release.dropna().dt.to_period("M"))
        periods=pd.concat(monthly,ignore_index=True) if monthly else pd.Series(dtype="period[M]")
        if not periods.empty:
            calendar=([pd.Period(value,freq="M") for value in sorted(set(month_values))] if month_values else pd.period_range(periods.min(),periods.max(),freq="M"))
            assessed_counts=pd.to_datetime(trend_scope[assessment_date],errors="coerce").dt.to_period("M").value_counts() if assessment_date else pd.Series(dtype="int64")
            released_counts=pd.to_datetime(trend_released[release_date],errors="coerce").dt.to_period("M").value_counts() if release_date else pd.Series(dtype="int64")
            trend=[
                {"month":period.strftime("%Y-%m"),"detainedAssessments":int(assessed_counts.get(period,0)),"released":int(released_counts.get(period,0))}
                for period in calendar
                if int(assessed_counts.get(period,0))>0 or int(released_counts.get(period,0))>0
            ]
        charts=[]
        for label in ("Gender / age group","Community type","Nationality","Project","Project location"):
            column=columns_by_label.get(label)
            if not column:continue
            counts=scoped[column].fillna("Blank").astype(str).str.strip().replace("","Blank").value_counts().head(20)
            charts.append({"id":label,"title":label,"items":[{"label":str(item),"count":int(count)} for item,count in counts.items()]})
        governorate_column=columns_by_label.get("Detention governorate")
        governorate_aliases=(
            (("sulayman",),"Al-Sulaimaniyah"),(("ninewa","ninawa","mosul"),"Ninawa"),
            (("qadis","diwani"),"Al-Qadisiyah"),(("muthanna",),"Al-Muthanna"),
            (("thi qar","dhi qar","nasir"),"Dhi Qar"),(("salah",),"Salah al-Din"),
            (("anbar",),"Al-Anbar"),(("basra",),"Al-Basrah"),(("karbala","kerbala"),"Karbala"),
            (("najaf",),"An-Najaf"),(("wasit","wassit"),"Wasit"),(("baghdad",),"Baghdad"),
            (("babil","babylon"),"Babil"),(("maysan","missan"),"Maysan"),(("diyala",),"Diyala"),
            (("kirkuk",),"Kirkuk"),(("erbil","arbil"),"Erbil"),(("dohuk","duhok"),"Dohuk"),
        )
        map_groups:dict[str,dict[str,Any]]={}
        def add_map_counts(source:pd.DataFrame,metric:str)->None:
            if not governorate_column:return
            for raw,count in source[governorate_column].fillna("").astype(str).str.strip().value_counts().items():
                if not raw:continue
                lowered=raw.lower();canonical=next((name for aliases,name in governorate_aliases if any(alias in lowered for alias in aliases)),"")
                if not canonical:continue
                group=map_groups.setdefault(canonical,{"label":canonical,"count":0,"detained":0,"released":0,"values":[]})
                group[metric]+=int(count)
                if metric=="detained":group["count"]+=int(count)
                if raw not in group["values"]:group["values"].append(raw)
        add_map_counts(scoped,"detained");add_map_counts(released_scope,"released")
        sort_source=dict(detail_selected).get(sort_column)
        if sort_source:
            values=scoped[sort_source]
            if DATE_HINT.search(sort_source): values=pd.to_datetime(values,errors="coerce",dayfirst=True)
            elif pd.api.types.is_numeric_dtype(values): values=pd.to_numeric(values,errors="coerce")
            else: values=values.fillna("").astype(str).str.casefold()
            scoped=scoped.assign(_detention_sort=values).sort_values("_detention_sort",ascending=sort_direction!="desc",kind="stable",na_position="last").drop(columns="_detention_sort")
        start=(page-1)*page_size;page_frame=scoped.iloc[start:start+page_size]
        rows=[]
        beneficiary_names:dict[str,str]={}
        if beneficiary_col:
            beneficiaries=self.frames["beneficiaries"];case_column=_find(list(beneficiaries.columns),"Case ID");name_column=_find(list(beneficiaries.columns),"Name (Filter Color Red)")
            if case_column and name_column:beneficiary_names=dict(zip(beneficiaries[case_column].map(clean_id),beneficiaries[name_column].fillna("").astype(str)))
        for row_index,record in page_frame.iterrows():
            row={"__rowKey":str(row_index),**{label:display_value(record.get(column,"")) if column else "" for label,column in detail_selected}};row["caseId"]=clean_id(record.get(beneficiary_col,"")) if beneficiary_col else ""
            if not row.get("Name / الأسم") and row["caseId"] in beneficiary_names:row["Name / الأسم"]=beneficiary_names[row["caseId"]]
            rows.append(row)
        return {"total":len(scoped),"page":page,"pageSize":page_size,"columns":[label for label,_ in detail_selected],"rows":rows,"filterOptions":options,"trend":trend,"charts":charts,"map":{"items":list(map_groups.values())},"kpis":[{"label":"Detention assessments","value":len(scoped)},{"label":"Distinct beneficiaries","value":distinct},{"label":"Currently detained","value":active},{"label":"Immigration-related","value":immigration_count}]}

    @staticmethod
    def detention_workbook_sheets(raw:bytes) -> list[str]:
        try:
            validate_xlsx_archive(raw)
            return list(pd.ExcelFile(io.BytesIO(raw)).sheet_names)
        except Exception as exc:raise ValueError("The comparison file could not be read. Upload a valid .xlsx workbook.") from exc

    def detention_reconciliation(self, raw:bytes, filename:str, month:str, project:str|list[str]="", sheet_name:str="") -> dict[str,Any]:
        months=list(dict.fromkeys(item.strip() for item in month.split(",") if item.strip()))
        projects=list(dict.fromkeys(item.strip() for item in ([project] if isinstance(project,str) else project) if item.strip()))
        if not months or any(not re.fullmatch(r"\d{4}-\d{2}",item) for item in months): raise ValueError("Select at least one valid assessment month before comparing.")
        try:
            validate_xlsx_archive(raw)
            workbook=pd.ExcelFile(io.BytesIO(raw));available_sheets=list(workbook.sheet_names)
            selected_sheet=sheet_name or (available_sheets[0] if available_sheets else "")
            if selected_sheet not in available_sheets:raise ValueError(f"Worksheet not found: {selected_sheet}")
            external=pd.read_excel(workbook,sheet_name=selected_sheet,dtype=object)
        except Exception as exc: raise ValueError("The comparison file could not be read. Upload a valid .xlsx workbook.") from exc
        external.columns=[str(column).strip() for column in external.columns]
        frame=self.frames["assessments"]
        platform_fields={
            "Beneficiary ID":("Beneficiary ID",),"Name":("Name /", "Name"),"Date of birth":("DoB /","Date of Birth","DoB"),
            "Gender":("Gender",),"Date of detention":("Date of Detention",),
            "Detention governorate":("Detention Governorate",),"Detaining authority":("Detaining Authority",),
            "Reasons for detention":("Reasons for Detention",),"Possible charges":("Possible Charges",),
            "Nationality":("Nationality",),
            "Legal service needed":("Type of Legal Service Needed",),"Current status":("Detainee current status",),
            "Release type":("Type of Released",),"Release/deportation date":("Date of the released or deported",),
        }
        external_fields={
            "Beneficiary ID":("Beneficiary ID (Platform Case ID)",),"Name":("Arabic Name",),"Date of birth":("Date of Birth",),
            "Gender":("Sex",),"Date of detention":("Date of Arrest",),
            "Detention governorate":("Detention Governorate",),"Detaining authority":("Detaining Authority",),
            "Reasons for detention":("Reason of Arrest",),"Possible charges":("Charges",),"Nationality":("Nationality",),
            "Legal service needed":("Type of Service by INTERSOS",),
            "Current status":("Detainee Current Status",),"Release type":("Type of Release",),
            "Release/deportation date":("Date of Release/Deportation",),
        }
        platform_columns={label:_find(list(frame.columns),*hints) for label,hints in platform_fields.items()}
        external_columns={label:_find(list(external.columns),*hints) for label,hints in external_fields.items()}
        missing=[external_fields[label][0] for label,column in external_columns.items() if not column]
        platform_project=_find(list(frame.columns),"Projects","Project")
        external_project=_find(list(external.columns),"Projects - المشروع","Projects","Project")
        identification_date=_find(list(external.columns),"Identification Date")
        assessment_date=_find(list(frame.columns),"Date of Assessment")
        if not external_columns["Beneficiary ID"]: raise ValueError("The workbook is missing Beneficiary ID (Platform Case ID).")
        if not identification_date: raise ValueError("The workbook is missing Identification Date | تاريخ تحديد الحالة.")
        if not assessment_date: raise ValueError("The platform Assessments data has no Date of Assessment column.")
        platform_month=pd.to_datetime(frame[assessment_date],errors="coerce",dayfirst=True).dt.to_period("M").astype(str)
        detained=_find(list(frame.columns),"Is the beneficiary detained")
        platform=frame[platform_month.isin(months)].copy()
        if detained: platform=platform[platform[detained].fillna("").astype(str).str.contains(r"\byes\b|نعم",case=False,regex=True)]
        if projects:
            if not platform_project: raise ValueError("The platform Assessments data has no Project column.")
            platform=platform[platform[platform_project].fillna("").astype(str).str.strip().isin(projects)]
        external_month=pd.to_datetime(external[identification_date],errors="coerce",dayfirst=True).dt.to_period("M").astype(str)
        external=external[external_month.isin(months)].copy()
        if projects and external_project: external=external[external[external_project].fillna("").astype(str).str.strip().isin(projects)]
        id_platform=platform_columns["Beneficiary ID"]
        id_external=external_columns["Beneficiary ID"]
        platform["__id"]=platform[id_platform].map(clean_id) if id_platform else ""
        external["__id"]=external[id_external].map(clean_id)
        platform_missing_id=platform[platform["__id"].eq("")].copy()
        external_missing_id=external[external["__id"].eq("")].copy()
        platform=platform[platform["__id"].ne("")];external=external[external["__id"].ne("")]
        def comparable(value:Any,label:str)->str:
            if pd.isna(value) or str(value).strip()=="": return ""
            if label=="Detention governorate": return normalize_governorate(value)
            if "date" in label.lower(): return normalize_comparison_date(value)
            if label=="Name": return normalize_name(value)
            if label=="Gender":
                # Assessment values commonly contain an English value followed by its Arabic translation.
                # The reconciliation compares the English portion only.
                value=re.sub(r"[\u0600-\u06ff\u0750-\u077f\u08a0-\u08ff]+"," ",str(value))
                return re.sub(r"[^a-z]+"," ",value.casefold()).strip()
            if label=="Legal service needed":
                normalized=re.sub(r"\s+"," ",unicodedata.normalize("NFKC",str(value)).strip().casefold())
                return "legal representation" if "legal representation" in normalized else normalized
            return re.sub(r"\s+"," ",unicodedata.normalize("NFKC",str(value)).strip().casefold())
        external_by_id={key:group for key,group in external.groupby("__id",sort=False)}
        platform_by_id={key:group for key,group in platform.groupby("__id",sort=False)}
        results=[];matched=0
        name_platform=platform_columns.get("Name");name_external=external_columns.get("Name")
        lawyer_platform=_find(list(frame.columns),"Lawyer")
        lawyer_external=_find(list(external.columns),"Lawyer")
        def preferred_lawyer(source:Any=None,target:Any=None)->str:
            assessment_lawyer=display_value(source.get(lawyer_platform,"")) if source is not None and lawyer_platform else ""
            excel_lawyer=display_value(target.get(lawyer_external,"")) if target is not None and lawyer_external else ""
            return str(assessment_lawyer or excel_lawyer or "")
        beneficiaries=self.frames["beneficiaries"];beneficiary_case=_find(list(beneficiaries.columns),"Case ID");beneficiary_name=_find(list(beneficiaries.columns),"Name (Filter Color Red)")
        beneficiary_names=dict(zip(beneficiaries[beneficiary_case].map(clean_id),beneficiaries[beneficiary_name].fillna("").astype(str))) if beneficiary_case and beneficiary_name else {}
        assessment_case_ids=set(frame[id_platform].map(clean_id)) if id_platform else set()
        beneficiary_case_ids=set(beneficiaries[beneficiary_case].map(clean_id)) if beneficiary_case else set()
        def case_available(case_id:str)->bool:
            # The case review needs both the assessment history and beneficiary profile.
            return bool(case_id) and case_id in assessment_case_ids and case_id in beneficiary_case_ids
        note_external=_find(list(external.columns),"Note","ملاحظات")
        for case_id,source_group in platform_by_id.items():
            target_group=external_by_id.get(case_id)
            source=source_group.iloc[-1]
            name=display_value(source.get(name_platform,"")) if name_platform else ""
            if not name:name=beneficiary_names.get(case_id,"")
            if target_group is None:
                results.append({"beneficiaryId":case_id,"caseAvailable":case_available(case_id),"name":name,"lawyer":preferred_lawyer(source),"note":"Case ID available in Assessments but missing from Excel","differences":[{"field":"Case ID","assessment":"Present","excel":"Missing"}]});continue
            best_note=None;best_difference=None;best_target=None
            for _,target in target_group.iterrows():
                differences=[]
                for label in platform_fields:
                    if label=="Beneficiary ID":continue
                    left_column=platform_columns.get(label);right_column=external_columns.get(label)
                    if not left_column or not right_column:continue
                    left=comparable(source.get(left_column,""),label);right=comparable(target.get(right_column,""),label)
                    if left!=right:differences.append(label)
                if best_difference is None or len(differences)<len(best_difference):
                    best_difference=differences;best_note=display_value(target.get(note_external,"")) if note_external else "";best_target=target
            if not best_difference: matched+=1;continue
            note="Different: "+", ".join(best_difference)
            if best_note:note+=f" · Workbook note: {best_note}"
            difference_values=[]
            for label in best_difference:
                left_column=platform_columns.get(label);right_column=external_columns.get(label)
                difference_values.append({"field":label,"assessment":display_value(source.get(left_column,"")) if left_column else "","excel":display_value(best_target.get(right_column,"")) if best_target is not None and right_column else ""})
            results.append({"beneficiaryId":case_id,"caseAvailable":case_available(case_id),"name":name or display_value(target_group.iloc[0].get(name_external,"")),"lawyer":preferred_lawyer(source,best_target),"note":best_note or "Field difference (no workbook note)","differences":difference_values})
        for case_id,target_group in external_by_id.items():
            if case_id in platform_by_id:continue
            target=target_group.iloc[-1]
            results.append({"beneficiaryId":case_id,"caseAvailable":case_available(case_id),"name":display_value(target.get(name_external,"")) if name_external else "","lawyer":preferred_lawyer(None,target),"note":"Case ID available in Excel but missing from Assessments","differences":[{"field":"Case ID","assessment":"Missing","excel":"Present"}]})
        for _,source in platform_missing_id.iterrows():
            results.append({"beneficiaryId":"","caseAvailable":False,"name":display_value(source.get(name_platform,"")) if name_platform else "","lawyer":preferred_lawyer(source),"note":"Case ID missing in Assessments","differences":[{"field":"Case ID","assessment":"Missing","excel":"Not comparable"}]})
        for _,target in external_missing_id.iterrows():
            results.append({"beneficiaryId":"","caseAvailable":False,"name":display_value(target.get(name_external,"")) if name_external else "","lawyer":preferred_lawyer(None,target),"note":"Case ID missing in Excel","differences":[{"field":"Case ID","assessment":"Not comparable","excel":"Missing"}]})
        compared_fields=[label for label in platform_fields if label!="Beneficiary ID" and platform_columns.get(label) and external_columns.get(label)]
        warnings=[f"Column not compared because it is missing: {name}" for name in missing]
        for label in ("Date of birth","Detention governorate"):
            if not platform_columns.get(label):warnings.append(f"Platform Assessments column not found, so it was not compared: {label}")
        return {"month":", ".join(months),"months":months,"project":", ".join(projects),"projects":projects,"filename":filename,"sheet":selected_sheet,"sheets":available_sheets,"platformRecords":len(platform_by_id),"comparisonRecords":len(external_by_id),"missingCaseIds":{"assessments":len(platform_missing_id),"excel":len(external_missing_id)},"matched":matched,"unmatched":len(results),"comparedFields":compared_fields,"rows":results,"warnings":warnings}

    def detention_reconciliation_export(self, raw:bytes, filename:str, month:str, project:str|list[str]="", sheet_name:str="") -> bytes:
        comparison=self.detention_reconciliation(raw,filename,month,project,sheet_name)
        output=io.BytesIO()
        palette=("E8F1FB","FDF0E7","EAF6EE","F3ECFA","FFF7D9","E8F5F5")
        with pd.ExcelWriter(output,engine="openpyxl",date_format="DD/MM/YYYY",datetime_format="DD/MM/YYYY") as writer:
            sheet=writer.book.create_sheet("Comparison issues")
            sheet.merge_cells("A1:G1");sheet["A1"]="INTERSOS | Detention Excel comparison issues"
            sheet["A1"].font=Font(bold=True,color="FFFFFF",size=14);sheet["A1"].fill=PatternFill("solid",fgColor="0B5C95");sheet["A1"].alignment=Alignment(horizontal="left",vertical="center")
            sheet.merge_cells("A2:G2");sheet["A2"]=f"Assessment month(s): {comparison['month']}  |  Project(s): {comparison['project'] or 'All'}  |  Workbook: {comparison['filename']}"
            sheet["A2"].font=Font(italic=True,color="526777");sheet["A2"].alignment=Alignment(wrap_text=True)
            headers=("Lawyer","Note group","Case ID","Name","Different field","Assessment value","Excel value")
            for column,header in enumerate(headers,1):
                cell=sheet.cell(4,column,header);cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="1D4ED8");cell.alignment=Alignment(wrap_text=True,vertical="center")
            row_number=5
            notes=list(dict.fromkeys(row["note"] for row in comparison["rows"]))
            for group_index,note in enumerate(notes):
                fill=PatternFill("solid",fgColor=palette[group_index%len(palette)])
                for issue in (item for item in comparison["rows"] if item["note"]==note):
                    for difference in issue["differences"] or [{"field":"Record","assessment":"","excel":""}]:
                        values=(issue.get("lawyer","") or "—",note,issue.get("beneficiaryId","") or "—",display_value(issue.get("name","")) or "—",difference["field"],display_value(difference["assessment"]) or "—",display_value(difference["excel"]) or "—")
                        for column,value in enumerate(values,1):
                            cell=sheet.cell(row_number,column,value);cell.fill=fill;cell.alignment=Alignment(wrap_text=True,vertical="top")
                            cell.border=Border(bottom=Side(style="thin",color="D5DFE8"))
                        row_number+=1
            if row_number==5:
                sheet.merge_cells("A5:G5");sheet["A5"]="No comparison issues found.";sheet["A5"].alignment=Alignment(horizontal="center")
            sheet.freeze_panes="A5";sheet.auto_filter.ref=f"A4:G{max(5,row_number-1)}";sheet.row_dimensions[1].height=26;sheet.row_dimensions[2].height=30
            for column,width in enumerate((22,44,18,28,25,30,30),1):sheet.column_dimensions[get_column_letter(column)].width=width
            format_excel_dates(writer.book)
        return output.getvalue()

    def export(self,dataset:str)->bytes:
        if dataset not in self.frames: raise ValueError("Dataset not loaded")
        frame=_safe_export(self.frames[dataset])
        output=io.BytesIO()
        with pd.ExcelWriter(output,engine="openpyxl",date_format="DD/MM/YYYY",datetime_format="DD/MM/YYYY") as writer:
            frame.to_excel(writer,index=False,sheet_name="Data");sheet=writer.book["Data"];sheet.freeze_panes="A2";sheet.auto_filter.ref=sheet.dimensions
            for cell in sheet[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="2563EB");cell.alignment=Alignment(wrap_text=True)
            for index,column in enumerate(frame.columns,1):sheet.column_dimensions[get_column_letter(index)].width=min(38,max(12,len(str(column))+2))
            format_excel_dates(writer.book)
        return output.getvalue()
