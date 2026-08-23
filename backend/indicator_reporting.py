from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any, Callable

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .legal_platform import _find, clean_id


AGE_GROUPS = ("00-04", "05-11", "12-17", "18-39", "40-59", "60+")
REPORT_ROWS = (
    ("UNHCR 2026 - AMAL CAMP", "AMAL Camp"),
    ("UNHCR 2026 - Erbil", "Urban"), ("UNHCR 2026 - Erbil", "Kawrgawsk Camp"),
    ("UNHCR 2026 - Erbil", "Basirma Camp"), ("UNHCR 2026 - Erbil", "Darashakran Camp"),
    ("UNHCR 2026 - Erbil", "Qushtapa Camp"),
    ("UNHCR 2026 - SULI", "Sulaymaniyah Urban"),
    ("UNHCR 2026 - SULI", "Pshdar Urban (Refugees) + Rania"),
    ("UNHCR 2026 - SULI", "Arbat Camp (Refugees)"),
    ("UNHCR 2026 - Mosul & Kirkuk", "Ninewa نينوى"),
    ("UNHCR 2026 - Mosul & Kirkuk", "Kirkuk كركوك"),
    ("UNHCR 2026 - Baghdad", "Baghdad بغداد"),
    ("UNHCR 2026 - Gov", "Anbar أنبار"), ("UNHCR 2026 - Gov", "Al-Muthanna المثنى"),
    ("UNHCR 2026 - Gov", "Al-Qadisiyyah القادسية"), ("UNHCR 2026 - Gov", "Babil بابل"),
    ("UNHCR 2026 - Gov", "Basra بصرة"), ("UNHCR 2026 - Gov", "Dhi Qar ذي قار"),
    ("UNHCR 2026 - Gov", "Diyala ديالى"), ("UNHCR 2026 - Gov", "Karbala كربلاء"),
    ("UNHCR 2026 - Gov", "Maysan ميسان"), ("UNHCR 2026 - Gov", "Najaf نجف"),
    ("UNHCR 2026 - Gov", "Salah Al-Din صلاح الدين"), ("UNHCR 2026 - Gov", "Wassit واسط"),
)
CIVIL_DOCUMENTS = (
    "unified national card", "iraqi nationality certificate", "pds card", "birth certificate",
    "proof of birth", "proof of custody", "proof of kinship", "proof of guardianship",
    "marriage certificate", "proof of marriage", "marriage attestation", "divorce certificate",
    "death certificate", "proof of death", "proof of curatorship", "housing card", "passport", "civil id",
)


def _norm(value: Any) -> str:
    if pd.isna(value): return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value)).strip().casefold())


def _lookup(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u0600-\u06ff]+", "", _norm(value))


# Keep historical exports reportable after the project-location rename.
LOCATION_ALIASES = {
    _lookup("UNHCR Mosul"): _lookup("Ninewa نينوى"),
    _lookup("UNHCR Kirkuk"): _lookup("Kirkuk كركوك"),
}


def _contains(value: Any, *terms: str) -> bool:
    text = _norm(value)
    return any(term.casefold() in text for term in terms)


def _yes(value: Any) -> bool:
    text = _norm(value)
    return text.startswith("yes") or text.startswith("نعم")


def _no(value: Any) -> bool:
    text = _norm(value)
    return text.startswith("no") or text.startswith("لا")


def _completed(value: Any) -> bool:
    return _contains(value, "completed", "اكتملت", "اکتملت")


def _released(value: Any) -> bool:
    return _contains(value, "released", "الافراج", "الإفراج")


def _only_counselling(value: Any) -> bool:
    services = [_norm(part) for part in str(value or "").split(",") if _norm(part)]
    return bool(services) and all("legal counselling" in service for service in services)


def _single_counselling(value: Any) -> bool:
    services = [_norm(part) for part in str(value or "").split(",") if _norm(part)]
    return len(services) == 1 and "legal counselling" in services[0]


def _assistance_or_representation(value: Any) -> bool:
    return _contains(value, "legal assistance", "legal representation")


def _population(value: Any) -> str:
    text = _norm(value)
    if "non-syrian refugee" in text: return "non-syrian-refugee"
    if "syrian refugee" in text: return "syrian-refugee"
    if "idp" in text or "نازح" in text: return "idp"
    return "other"


def _community_filter_label(value: Any) -> str:
    """Return the English-only community label used by the reporting filter."""
    population = _population(value)
    if population == "idp": return "IDP"
    if population == "syrian-refugee": return "Syrian Refugee"
    if population == "non-syrian-refugee": return "Non-Syrian Refugee"
    return ""


def _gender(value: Any) -> str:
    text = _norm(value)
    if text.startswith("male"): return "male"
    if text.startswith("female"): return "female"
    return ""


def _age_group(value: Any) -> str:
    compact = re.sub(r"[()\s]", "", _norm(value).replace("–", "-").replace("—", "-"))
    return compact if compact in AGE_GROUPS else ""


def _common_columns(frame: pd.DataFrame) -> dict[str, str | None]:
    names = list(frame.columns)
    return {
        "project": _find(names, "Projects -", "Projects", "Project"),
        "location": _find(names, "Project Location", "Project location"),
        "gender": _find(names, "Gender النوع الاجتماعي", "Gender الجنس", "Gender"),
        "age": _find(names, "UNHCR Age Group"),
        "community": _find(names, "Community Type /", "Community Type"),
    }


def build_indicator_report(frames: dict[str, pd.DataFrame], from_date: str = "", to_date: str = "", projects: list[str] | None = None, locations: list[str] | None = None, years: list[str] | None = None, quarters: list[str] | None = None, months: list[str] | None = None, community_types: list[str] | None = None) -> dict[str, Any]:
    projects, locations, years, quarters, months, community_types = projects or [], locations or [], years or [], quarters or [], months or [], community_types or []
    # The source values often append Arabic text (for example, "Syrian Refugee
    # لاجئ-سوري"). Expose the reporting populations as clean English labels
    # while retaining a population-aware filter against the original data.
    community_options=sorted({_community_filter_label(value) for frame in frames.values() for column in [_common_columns(frame).get("community")] if column for value in frame[column].dropna() if _community_filter_label(value)})
    if community_types:
        selected_communities=set(community_types)
        frames={name:frame[frame[_common_columns(frame)["community"]].map(_community_filter_label).isin(selected_communities)].copy() if _common_columns(frame).get("community") else frame.copy() for name,frame in frames.items()}
    start = pd.to_datetime(from_date, errors="coerce") if from_date else None
    end = pd.to_datetime(to_date, errors="coerce") if to_date else None
    if from_date and pd.isna(start): raise ValueError("From date is invalid.")
    if to_date and pd.isna(end): raise ValueError("To date is invalid.")
    if start is not None: start = start.normalize()
    if end is not None: end = end.normalize() + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)
    if start is not None and end is not None and start > end: raise ValueError("From date must be before To date.")

    canonical = {(_lookup(project), _lookup(location)): (project, location) for project, location in REPORT_ROWS}
    for (project_key, location_key), pair in list(canonical.items()):
        for legacy_location, current_location in LOCATION_ALIASES.items():
            if location_key == current_location:
                canonical[(project_key, legacy_location)] = pair
    visible_rows = [pair for pair in REPORT_ROWS if (not projects or pair[0] in projects) and (not locations or pair[1] in locations)]
    beneficiary_frame = frames.get("beneficiaries", pd.DataFrame())
    beneficiary_id_column = _find(list(beneficiary_frame.columns), "Case ID", "Beneficiary ID")
    beneficiary_name_column = _find(list(beneficiary_frame.columns), "Name (Filter Color Red)", "Beneficiary Name", "Name")
    beneficiary_names = {clean_id(row.get(beneficiary_id_column, "")): clean_id(row.get(beneficiary_name_column, "")) for _, row in beneficiary_frame.iterrows()} if beneficiary_id_column and beneficiary_name_column else {}

    def mapped_pair(row: pd.Series, cols: dict[str, str | None]) -> tuple[str, str] | None:
        project = row.get(cols["project"], "") if cols["project"] else ""
        location = row.get(cols["location"], "") if cols["location"] else ""
        location_text = _norm(location)
        if "pshdar urban" in location_text or location_text == "rania":
            project, location = "UNHCR 2026 - SULI", "Pshdar Urban (Refugees) + Rania"
        elif "diyala" in location_text or "ديالى" in location_text:
            project, location = "UNHCR 2026 - Gov", "Diyala ديالى"
        return canonical.get((_lookup(project), _lookup(location)))

    def in_period(value: Any) -> bool:
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
        if pd.isna(parsed) or (start is not None and parsed < start) or (end is not None and parsed > end): return False
        month_key, quarter_key, year_key = parsed.strftime("%Y-%m"), f"{parsed.year}-Q{parsed.quarter}", str(parsed.year)
        return (not years or year_key in years) and (not quarters or quarter_key in quarters) and (not months or month_key in months)

    def in_countifs_period(value: Any) -> bool:
        """Match Excel COUNTIFS date criteria, including a blank date when no lower bound is set."""
        parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
        if pd.isna(parsed):
            return not _norm(value) and start is None and not years and not quarters and not months
        if (start is not None and parsed < start) or (end is not None and parsed > end): return False
        month_key, quarter_key, year_key = parsed.strftime("%Y-%m"), f"{parsed.year}-Q{parsed.quarter}", str(parsed.year)
        return (not years or year_key in years) and (not quarters or quarter_key in quarters) and (not months or month_key in months)

    def select(frame: pd.DataFrame, key_column: str | None, predicate: Callable[[pd.Series], bool], monthly: bool = False) -> list[pd.Series]:
        chosen, seen = [], set()
        completed_column = _find(list(frame.columns), "Date Service Completed") if monthly else None
        for _, row in frame.iterrows():
            if not predicate(row): continue
            key = clean_id(row.get(key_column, "")) if key_column else ""
            if not key: continue
            if monthly:
                parsed = pd.to_datetime(row.get(completed_column, ""), errors="coerce", dayfirst=True)
                if pd.isna(parsed): continue
                key = f"{parsed.strftime('%Y-%m')}|{key}"
            if key in seen: continue
            seen.add(key); chosen.append(row)
        return chosen

    def countifs_rows(frame: pd.DataFrame, predicate: Callable[[pd.Series], bool]) -> list[pd.Series]:
        """COUNTIFS counts matching rows; it does not de-duplicate record identifiers."""
        return [row for _, row in frame.iterrows() if predicate(row)]

    def unique_rows(frame: pd.DataFrame, key_column: str | None, predicate: Callable[[pd.Series], bool]) -> list[pd.Series]:
        chosen, seen = [], set()
        for _, row in frame.iterrows():
            if not predicate(row): continue
            key = clean_id(row.get(key_column, "")) if key_column else ""
            if key in seen: continue
            seen.add(key); chosen.append(row)
        return chosen

    amal_project = "UNHCR 2026 - AMAL CAMP"
    # AMAL is an IDP-only project.  A project filter that includes AMAL therefore
    # presents only IDP reporting; without AMAL, IDP reporting is not relevant.
    refugee_community_types = {"Syrian Refugee", "Non-Syrian Refugee"}
    show_refugee_columns = not (projects and amal_project in projects) and "IDP" not in community_types
    # With no project filter, the report covers every available project,
    # including AMAL; retain the IDP section and workbook sheet in that case.
    show_idp_columns = (not projects or amal_project in projects) and not (refugee_community_types & set(community_types))
    all_populations = tuple(
        ([(("syrian-refugee", "Syrian Refugees")), (("non-syrian-refugee", "Non-Syrian Refugees"))] if show_refugee_columns else [])
        + ([("idp", "IDP")] if show_idp_columns else [])
    )
    populations = {
        "all": all_populations,
        "refugee": (("syrian-refugee", "Syrian Refugees"), ("non-syrian-refugee", "Non-Syrian Refugees")),
        "idp": (("idp", "IDP"),),
    }

    def matrix(records: list[pd.Series], frame: pd.DataFrame, population_id: str, population_group: str, source_label: str = "") -> dict[str, Any]:
        if population_group == "idp":
            matrix_rows = [pair for pair in visible_rows if pair[0] == "UNHCR 2026 - AMAL CAMP"]
        elif population_group == "refugee":
            matrix_rows = [pair for pair in visible_rows if pair[0] != "UNHCR 2026 - AMAL CAMP"]
        else:
            matrix_rows = visible_rows
        cols = _common_columns(frame); beneficiary_column = _find(list(frame.columns), "Beneficiary ID", "Case ID", "PN ID"); assessment_column = _find(list(frame.columns), "Assessment ID"); name_column = _find(list(frame.columns), "Name (Filter Color Red)", "Beneficiary Name", "Name")
        counts = {pair: [0] * 12 for pair in matrix_rows}; ids = {pair: [[] for _ in range(12)] for pair in matrix_rows}; assessment_ids = {pair: [[] for _ in range(12)] for pair in matrix_rows}; unclassified = unknown = 0
        for row in records:
            if _population(row.get(cols["community"], "")) != population_id: continue
            pair = mapped_pair(row, cols)
            if pair is None: unknown += 1; continue
            if pair not in counts: continue
            sex, age = _gender(row.get(cols["gender"], "")), _age_group(row.get(cols["age"], ""))
            if not sex or not age: unclassified += 1; continue
            index = AGE_GROUPS.index(age) + (0 if sex == "male" else 6)
            counts[pair][index] += 1
            beneficiary_id = clean_id(row.get(beneficiary_column, "")) if beneficiary_column else ""
            assessment_id = clean_id(row.get(assessment_column, "")) if assessment_column else ""
            name = clean_id(row.get(name_column, "")) if name_column else ""
            if not name: name = beneficiary_names.get(beneficiary_id, "")
            source = source_label or ("assessments.csv" if assessment_column else "source record")
            if beneficiary_id:
                ids[pair][index].append(f"{beneficiary_id}  |  Assessment ID: {assessment_id}  |  Name: {name}  |  Source: {source}")
            assessment_ids[pair][index].append(assessment_id)
        rows, totals, total_ids = [], [0] * 13, [[] for _ in range(13)]
        for project, location in matrix_rows:
            values = counts[(project, location)] + [sum(counts[(project, location)])]
            totals = [a + b for a, b in zip(totals, values)]
            row_ids = ids[(project, location)]
            row_ids = row_ids + [[beneficiary_id for cell_ids in row_ids for beneficiary_id in cell_ids]]
            row_assessment_ids = assessment_ids[(project, location)]
            row_assessment_ids = row_assessment_ids + [[assessment_id for cell_ids in row_assessment_ids for assessment_id in cell_ids]]
            total_ids = [existing + incoming for existing, incoming in zip(total_ids, row_ids)]
            rows.append({"project": project, "location": location, "values": values, "beneficiaryIds": row_ids, "assessmentIds": row_assessment_ids})
        return {"rows": rows, "totals": totals, "total": totals[-1], "totalBeneficiaryIds": total_ids, "warnings": {"unclassified": unclassified, "unknownLocation": unknown}}

    def indicator(identifier: str, title: str, source: str, date_field: str, rule: str, population_group: str, records: list[pd.Series], frame: pd.DataFrame) -> dict[str, Any]:
        source_file = {"Assessments": "assessments.csv", "Legal Services": "legalservices.csv", "Deportation Records": "deportationrecords.csv", "Awareness": "awareness.csv", "Assessments (this year)": "assessments.csv - this year", "Legal Services (carry-over)": "legalservices.csv - carry-over"}.get(source, source)
        sections = [{"id": pid, "label": label, **matrix(records, frame, pid, population_group, source_file)} for pid, label in populations[population_group]]
        return {"id": identifier, "title": title, "source": source, "dateField": date_field, "rule": rule, "population": population_group, "total": sum(section["total"] for section in sections), "sections": sections, "children": [], "contributions": {}}

    def add_sections(left: list[dict[str, Any]], right: list[dict[str, Any]]) -> list[dict[str, Any]]:
        result = []
        for a, b in zip(left, right):
            rows = [{"project": ar["project"], "location": ar["location"], "values": [x + y for x, y in zip(ar["values"], br["values"])], "beneficiaryIds": [x + y for x, y in zip(ar.get("beneficiaryIds", [[] for _ in ar["values"]]), br.get("beneficiaryIds", [[] for _ in br["values"]]))], "assessmentIds": [x + y for x, y in zip(ar.get("assessmentIds", [[] for _ in ar["values"]]), br.get("assessmentIds", [[] for _ in br["values"]]))]} for ar, br in zip(a["rows"], b["rows"])]
            totals = [x + y for x, y in zip(a["totals"], b["totals"])]
            total_ids = [x + y for x, y in zip(a.get("totalBeneficiaryIds", [[] for _ in totals]), b.get("totalBeneficiaryIds", [[] for _ in totals]))]
            result.append({"id": a["id"], "label": a["label"], "rows": rows, "totals": totals, "total": totals[-1], "totalBeneficiaryIds": total_ids, "warnings": {"unclassified": a["warnings"]["unclassified"] + b["warnings"]["unclassified"], "unknownLocation": a["warnings"]["unknownLocation"] + b["warnings"]["unknownLocation"]}})
        return result

    def parent(identifier: str, title: str, rule: str, children: list[dict[str, Any]], carryover: bool = False, contributions: dict[str, Any] | None = None) -> dict[str, Any]:
        sections = add_sections(children[0]["sections"], children[1]["sections"])
        return {"id": identifier, "title": title, "source": "Assessments + Legal Services" if carryover else "Assessments", "dateField": "Date of Assessment / Date Service Completed" if carryover else "Date of Assessment", "rule": rule, "population": "refugee", "total": sum(section["total"] for section in sections), "sections": sections, "children": children, "contributions": contributions or {}}

    assessments, services = frames["assessments"], frames["legalservices"]
    acols, scols = list(assessments.columns), list(services.columns)
    aid = _find(acols, "Assessment ID"); adate = _find(acols, "Date of Assessment"); aneed = _find(acols, "Type of Legal Service Needed")
    aprovided = _find(acols, "Legal Service (Type of Service Provided", "Type of Service Provided"); adetained = _find(acols, "Is the beneficiary detained")
    aimmigration = _find(acols, "Is it an immigration related charge"); astatus = _find(acols, "Detainee current status")
    arelease = _find(acols, "Date of the released or deported"); adoc = _find(acols, "Type of Documents to be issued")
    sid = _find(scols, "Service ID"); sbid = _find(scols, "Beneficiary ID"); stype = _find(scols, "Type of Service Provided")
    sstatus = _find(scols, "Service Status"); sprovision = _find(scols, "Date of Service Provision"); scompleted = _find(scols, "Date Service Completed")
    sdetained = _find(scols, "Is the beneficiary detained"); sdoc = _find(scols, "Type of Document")

    reached = select(assessments, aid, lambda row: in_period(row.get(adate, "")) and bool(_norm(row.get(aneed, ""))))
    representation_base = select(assessments, aid, lambda row: in_period(row.get(adate, "")) and _assistance_or_representation(row.get(aprovided, "")) and _contains(row.get(aneed, ""), "legal representation"))
    carryover = select(services, sbid, lambda row: _carryover_match(row, sprovision, scompleted, sstatus, stype, in_countifs_period), monthly=True)
    monthly_carryover: dict[str, int] = {}
    for row in carryover:
        parsed = pd.to_datetime(row.get(scompleted, ""), errors="coerce", dayfirst=True)
        if pd.isna(parsed): continue
        month = parsed.strftime("%Y-%m"); monthly_carryover[month] = monthly_carryover.get(month, 0) + 1

    representation_children = []
    for suffix, label, detained_test in (("detainee", "Detainee", _yes), ("other", "Other", _no)):
        base = [row for row in representation_base if detained_test(row.get(adetained, ""))]
        carried = [row for row in carryover if detained_test(row.get(sdetained, ""))]
        base_result = indicator("tmp", "", "Assessments (this year)", "Date of Assessment", "", "refugee", base, assessments)
        carry_result = indicator("tmp", "", "Legal Services (carry-over)", "Date Service Completed", "", "refugee", carried, services)
        sections = add_sections(base_result["sections"], carry_result["sections"])
        representation_children.append({"id": f"legal-representation-{suffix}", "title": f"# of persons provided with legal representation ({label})", "source": "Assessments + Legal Services", "dateField": "Date of Assessment / Date Service Completed", "rule": "Unique Assessment IDs in the reporting period plus eligible pre-2026 services, counted once per Beneficiary ID per completion month.", "population": "refugee", "total": sum(section["total"] for section in sections), "sections": sections, "children": [], "contributions": {"assessmentPeriod": base_result["total"], "carryOver": carry_result["total"]}})
    representation = parent("legal-representation", "# of persons provided with legal representation (attributes- detainee and other)", "Detainee plus Other: unique Assessment IDs and monthly unique Beneficiary IDs from eligible carry-over services.", representation_children, True, {"assessmentPeriod": sum(child["contributions"]["assessmentPeriod"] for child in representation_children), "carryOver": sum(child["contributions"]["carryOver"] for child in representation_children), "monthlyCarryOver": monthly_carryover})

    counselling_base = countifs_rows(assessments, lambda row: in_countifs_period(row.get(adate, "")) and _only_counselling(row.get(aprovided, "")) and _single_counselling(row.get(aneed, "")))
    counselling_children = [indicator(f"legal-counselling-{suffix}", f"# of persons provided with legal counselling ({label})", "Assessments", "Date of Assessment", "Only Counselling provided and Legal Counselling needed.", "refugee", [row for row in counselling_base if detained_test(row.get(adetained, ""))], assessments) for suffix, label, detained_test in (("detainee", "Detainee", _yes), ("other", "Other", _no))]
    counselling = parent("legal-counselling", "# of persons provided with legal counselling (attributes- detainee and other)", "Detainee plus Other; no carry-over.", counselling_children)

    identified = countifs_rows(assessments, lambda row: in_countifs_period(row.get(adate, "")) and _yes(row.get(adetained, "")) and _yes(row.get(aimmigration, "")))
    released = countifs_rows(assessments, lambda row: in_countifs_period(row.get(arelease, "")) and _yes(row.get(adetained, "")) and _yes(row.get(aimmigration, "")) and _released(row.get(astatus, "")) and _assistance_or_representation(row.get(aprovided, "")))
    scivil = _find(scols, "Is Civil Documents")
    secured = countifs_rows(services, lambda row: in_countifs_period(row.get(scompleted, "")) and _yes(row.get(scivil, "")) and _completed(row.get(sstatus, "")) and not _contains(row.get(stype, ""), "legal counselling"))
    uid = countifs_rows(services, lambda row: in_countifs_period(row.get(scompleted, "")) and _contains(row.get(sdoc, ""), "unified national card") and not _contains(row.get(stype, ""), "legal counselling"))

    def is_amal_camp(row: pd.Series, frame: pd.DataFrame) -> bool:
        columns = _common_columns(frame)
        return _norm(row.get(columns["project"], "")) == _norm(amal_project) and _norm(row.get(columns["location"], "")) == _norm("AMAL Camp")

    civil_counselling = indicator("civil-counselling", "# of persons receiving legal counselling for civil documentation", "Assessments", "Date of Assessment", "Same criteria as legal counselling (Other), restricted to AMAL Camp.", "idp", [row for row in counselling_base if _no(row.get(adetained, "")) and is_amal_camp(row, assessments)], assessments)
    civil_representation_base = [row for row in representation_base if _no(row.get(adetained, "")) and is_amal_camp(row, assessments)]
    civil_representation_carryover = [row for row in carryover if _no(row.get(sdetained, "")) and is_amal_camp(row, services)]
    civil_representation_assessments = indicator("tmp", "", "Assessments (this year)", "Date of Assessment", "", "idp", civil_representation_base, assessments)
    civil_representation_services = indicator("tmp", "", "Legal Services (carry-over)", "Date Service Completed", "", "idp", civil_representation_carryover, services)
    civil_representation_sections = add_sections(civil_representation_assessments["sections"], civil_representation_services["sections"])
    civil_representation = {"id": "civil-representation", "title": "# of persons receiving legal representation for civil documentation", "source": "Assessments + Legal Services", "dateField": "Date of Assessment / Date Service Completed", "rule": "Same criteria as legal representation (Other), restricted to AMAL Camp.", "population": "idp", "total": sum(section["total"] for section in civil_representation_sections), "sections": civil_representation_sections, "children": [], "contributions": {"assessmentPeriod": civil_representation_assessments["total"], "carryOver": civil_representation_services["total"]}}
    civil_representation_document_rows = countifs_rows(services, lambda row: in_countifs_period(row.get(scompleted, "")) and _completed(row.get(sstatus, "")) and _contains(row.get(stype, ""), "legal representation") and is_amal_camp(row, services))

    legal_assistance_total = {
        "id": "06-1-1-legal-assistance",
        "title": "06.1.1 Number of people who received legal assistance",
        "source": "Assessments + Legal Services",
        "dateField": "Date of Assessment / Date Service Completed",
        "rule": "Sum of legal counselling and legal representation, including Detainee and Other attributes.",
        "population": "refugee",
        "total": counselling["total"] + representation["total"],
        "sections": add_sections(counselling["sections"], representation["sections"]),
        "children": [],
        "contributions": {"assessmentPeriod": counselling["total"] + representation["contributions"]["assessmentPeriod"], "carryOver": representation["contributions"]["carryOver"]},
    }
    deport = frames.get("deportationrecords", pd.DataFrame())
    ddate = _find(list(deport.columns), "Date of Deportation Knowledge", "Date of deporting")
    deported = indicator("deported", "# of persons deported from detention (with immigration related charges)", "Deportation Records", "Date of Deportation Knowledge - تاريخ العلم بالترحيل", "All rows matching the reporting dimensions and Date of Deportation Knowledge.", "refugee", countifs_rows(deport, lambda row: in_countifs_period(row.get(ddate, ""))), deport)
    awareness = frames.get("awareness", pd.DataFrame())
    wdate = _find(list(awareness.columns), "Date of Session")
    awareness_rows = countifs_rows(awareness, lambda row: in_countifs_period(row.get(wdate, "")))

    refugee_indicators = [
        indicator("detention-immigration", "# of persons identified in detention (with immigration related charges)", "Assessments", "Date of Assessment", "Detained and immigration-related charge are Yes.", "refugee", identified, assessments),
        deported,
        indicator("released-immigration", "# of persons successfully released from detention based on (with immigration related charges)", "Assessments", "Date of release/deportation", "Detained, immigration-related, Released, with Legal Assistance or Representation.", "refugee", released, assessments),
        legal_assistance_total,
        counselling,
        representation,
    ]
    idp_indicators = [
        civil_counselling,
        indicator("secured-civil-documentation", "# of secured civil documentation", "Legal Services", "Date Service Completed", "Civil Documents is Yes, service is Completed, and type is not Legal Counselling.", "idp", secured, services),
        indicator("uid-secured", "# of persons who received UIDs", "Legal Services", "Date Service Completed", "Unified National Card and type is not Legal Counselling.", "idp", uid, services),
        civil_representation,
        indicator("legal-awareness-participants", "# of girls, boys, women, and men participating in legal awareness sessions", "Awareness", "Date of Session", "All awareness-session participant records in the selected reporting period, counted by gender and age group.", "idp", awareness_rows, awareness),
    ]
    reached_indicators = [indicator("individuals-reached", "# of individuals receiving legal assistance, representation or counselling", "Assessments", "Date of Assessment", "Type of Legal Service Needed is not blank.", "all", reached, assessments)]
    groups = []
    if show_refugee_columns: groups.append({"id": "refugee", "label": "Refugee", "indicators": refugee_indicators})
    if show_idp_columns: groups.append({"id": "idp", "label": "IDP", "indicators": idp_indicators})
    groups.append({"id": "individual-beneficiaries-reached", "label": "# of individual beneficiaries reached", "indicators": reached_indicators})
    north_projects = {"UNHCR 2026 - Erbil", "UNHCR 2026 - SULI", "UNHCR 2026 - Mosul & Kirkuk"}

    def narrative_location(value: Any) -> str:
        """Keep the English location label and remove appended Arabic text."""
        text = str(value or "").strip()
        return re.split(r"\s+(?=[\u0600-\u06ffØÙÚÛ])", text, maxsplit=1)[0].strip()

    def narrative_topic(value: Any) -> str:
        """Keep an English session topic while dropping its appended Arabic translation."""
        text = str(value or "").strip()
        return re.split(r"\s*(?:/|-|–|—)?\s*(?=[\u0600-\u06ff])", text, maxsplit=1)[0].rstrip(" /-–—").strip()

    def demographic_values(sections: list[dict[str, Any]], row_filter: Callable[[dict[str, Any]], bool] | None = None) -> list[int]:
        values = [0] * 12
        for section in sections:
            rows = section["rows"] if row_filter else [{"values": section["totals"]}]
            for row in rows:
                if row_filter and not row_filter(row): continue
                values = [left + right for left, right in zip(values, row["values"][:12])]
        return values

    def demographic_text(values: list[int]) -> str:
        boys, men = sum(values[:3]), sum(values[3:6])
        girls, women = sum(values[6:9]), sum(values[9:12])
        label = lambda value, singular, plural: f"{value:,} {singular if value == 1 else plural}"
        groups = [(girls, "girl", "girls"), (women, "woman", "women"), (boys, "boy", "boys"), (men, "man", "men")]
        visible = [label(value, singular, plural) for value, singular, plural in groups if value]
        if not visible: return "no classified gender or age data"
        return list_text(visible)

    def list_text(values: list[str]) -> str:
        if len(values) < 2: return values[0] if values else ""
        if len(values) == 2: return f"{values[0]} and {values[1]}"
        return ", ".join(values[:-1]) + f", and {values[-1]}"

    if len(months) == 1:
        try: narrative_period = pd.Timestamp(f"{months[0]}-01").strftime("%B %Y")
        except (TypeError, ValueError): narrative_period = str(months[0])
    else: narrative_period = "the selected reporting period" if months else "the reporting period"

    def legal_assistance_narrative(entry: dict[str, Any]) -> str:
        if not entry["total"]: return ""
        sections = [section for section in entry["sections"] if section["id"] in {"syrian-refugee", "non-syrian-refugee"}]

        def bucket(label: str, predicate: Callable[[dict[str, Any]], bool]) -> str:
            total_values = demographic_values(sections, predicate); total = sum(total_values)
            if not total: return ""
            nationality_parts = []
            for section, nationality in ((next((value for value in sections if value["id"] == "syrian-refugee"), None), "Syrian refugees"), (next((value for value in sections if value["id"] == "non-syrian-refugee"), None), "non-Syrian refugees")):
                if not section: continue
                values = demographic_values([section], predicate); subtotal = sum(values)
                if subtotal:
                    nationality_label = nationality[:-1] if subtotal == 1 and nationality.endswith("s") else nationality
                    nationality_parts.append(f"{subtotal:,} {nationality_label} ({demographic_text(values)})")
            return f"{label}: A total of {total:,} individuals, including {list_text(nationality_parts)}."

        north_predicate = lambda row: row["project"] in north_projects
        south_predicate = lambda row: row["project"] not in north_projects and row["project"] != amal_project
        parts = [f"Through {narrative_period}, a total of {entry['total']:,} individuals ({demographic_text(demographic_values(sections))}) had access to legal assistance services through legal counselling and legal representation in courts and other governmental entities, with the following age, gender, location, and nationality breakdown:"]
        north_total = sum(demographic_values(sections, north_predicate))
        if north_total:
            parts.append(f"North Iraq (Erbil, Sulaymaniyah, Mosul and Kirkuk): A total of {north_total:,} cases.")
            for sentence in (bucket("Erbil", lambda row: row["project"] == "UNHCR 2026 - Erbil"), bucket("Sulaymaniyah", lambda row: row["project"] == "UNHCR 2026 - SULI"), bucket("Mosul", lambda row: row["project"] == "UNHCR 2026 - Mosul & Kirkuk" and "ninewa" in _norm(row["location"])), bucket("Kirkuk", lambda row: row["project"] == "UNHCR 2026 - Mosul & Kirkuk" and "kirkuk" in _norm(row["location"]))):
                if sentence: parts.append(sentence)
        carry_over = entry.get("contributions", {}).get("carryOver", 0)
        if carry_over: parts.append(f"Of the individuals supported through {narrative_period}, {carry_over:,} were carry-over cases from 2025 that were successfully closed during the reporting period.")
        south_total = sum(demographic_values(sections, south_predicate))
        if south_total:
            parts.append(f"South Iraq (Centre and Southern Governorates): A total of {south_total:,} individuals were supported with legal counselling and representation.")
            for sentence in (bucket("Baghdad", lambda row: row["project"] == "UNHCR 2026 - Baghdad"), bucket("Southern Governorates", lambda row: row["project"] not in north_projects and row["project"] not in {"UNHCR 2026 - Baghdad", amal_project})):
                if sentence: parts.append(sentence)
        service_totals = (("Legal representation for detention cases", representation_children[0]["total"]), ("Legal representation for complex civil documentation", representation_children[1]["total"]), ("Legal counselling on civil documentation", counselling["total"]))
        percentages = [f"{label}: {(value / entry['total'] * 100):.1f}%" for label, value in service_totals]
        parts.append("The primary services were legal representation for detained refugees and asylum seekers, complex civil-documentation representation, and legal counselling. Service shares were: " + "; ".join(percentages) + ".")
        return " ".join(parts).replace(" North Iraq (Erbil, Sulaymaniyah, Mosul and Kirkuk):", "\nNorth Iraq (Erbil, Sulaymaniyah, Mosul and Kirkuk):").replace(" South Iraq (Centre and Southern Governorates):", "\nSouth Iraq (Centre and Southern Governorates):")

    def civil_representation_narrative(entry: dict[str, Any]) -> str:
        if not entry["total"]: return ""
        counts: dict[str, int] = {}
        for row in civil_representation_document_rows:
            document = narrative_location(row.get(sdoc, "")) if sdoc else ""
            if document: counts[document] = counts.get(document, 0) + 1
        top_documents = [document for document, _ in sorted(counts.items(), key=lambda value: (-value[1], value[0]))[:5]]
        remark = f"A total of {entry['total']:,} individuals ({demographic_text(demographic_values(entry['sections']))}) received legal representation."
        if top_documents: remark += f" The top types of cases included {list_text(top_documents)}."
        carry_over = entry.get("contributions", {}).get("carryOver", 0)
        if carry_over: remark += f"\n\nOf the individuals supported in {narrative_period}, {carry_over:,} were carry-over cases from 2025 that were successfully closed during the reporting period."
        return remark

    def legal_awareness_narrative(entry: dict[str, Any]) -> str:
        if not entry["total"]: return ""
        values = demographic_values(entry["sections"])
        gender_groups = (
            ("Girls", sum(values[6:9])),
            ("Boys", sum(values[:3])),
            ("Women", sum(values[9:12])),
            ("Men", sum(values[3:6])),
        )
        gender_summary = ", ".join(f"{label}: {value:,} ({value / entry['total'] * 100:.1f}%)" for label, value in gender_groups if value)
        awareness_id_column = _find(list(awareness.columns), "Awareness ID")
        topic_column = _find(list(awareness.columns), "Session Topic")
        session_ids = {clean_id(row.get(awareness_id_column, "")) for row in awareness_rows if awareness_id_column and clean_id(row.get(awareness_id_column, ""))}
        topics: dict[str, dict[str, Any]] = {}
        awareness_columns = _common_columns(awareness)
        for row in awareness_rows:
            topic = narrative_topic(row.get(topic_column, "")) if topic_column else ""
            topic = topic or "Unspecified topic"
            detail = topics.setdefault(topic, {"participants": 0, "sessionIds": set(), "genderValues": [0] * 12})
            detail["participants"] += 1
            session_id = clean_id(row.get(awareness_id_column, "")) if awareness_id_column else ""
            if session_id: detail["sessionIds"].add(session_id)
            gender, age = _gender(row.get(awareness_columns["gender"], "")), _age_group(row.get(awareness_columns["age"], ""))
            if gender and age:
                detail["genderValues"][AGE_GROUPS.index(age) + (0 if gender == "male" else 6)] += 1
        def topic_gender_summary(values: list[int]) -> str:
            groups = (("Girls", sum(values[6:9])), ("Women", sum(values[9:12])), ("Boys", sum(values[:3])), ("Men", sum(values[3:6])))
            return ", ".join(f"{label}: {value:,}" for label, value in groups if value)
        topic_summary = [f"• {topic} - {len(detail['sessionIds']):,} awareness session{'s' if len(detail['sessionIds']) != 1 else ''} | Participants: {detail['participants']:,}{f' | {topic_gender_summary(detail["genderValues"])}' if topic_gender_summary(detail['genderValues']) else ''}" for topic, detail in sorted(topics.items(), key=lambda item: (-item[1]["participants"], item[0]))]
        sessions = len(session_ids)
        sections = [
            "LEGAL AWARENESS SESSIONS",
            f"Reporting period: {narrative_period}  |  Participants: {entry['total']:,}  |  Awareness sessions: {sessions:,}",
            "",
            f"PARTICIPANT PROFILE  |  {gender_summary}",
        ]
        if topic_summary: sections.extend(["", "SESSION-TOPIC BREAKDOWN", *topic_summary])
        return "\n".join(sections)

    def attach_narrative(entry: dict[str, Any]) -> None:
        narrative_rows: list[dict[str, Any]] = []
        section_groups = (
            ("Refugees", [section for section in entry["sections"] if section["id"] in {"syrian-refugee", "non-syrian-refugee"}]),
            ("IDP", [section for section in entry["sections"] if section["id"] == "idp"]),
        )
        for population, sections in section_groups:
            if not sections:
                continue
            combined: dict[tuple[str, str], dict[str, Any]] = {}
            for section in sections:
                for row in section["rows"]:
                    value = row["values"][-1]
                    if not value:
                        continue
                    location = narrative_location(row["location"])
                    key = (row["project"], location)
                    combined.setdefault(key, {"project": row["project"], "location": location, "total": 0})["total"] += value
            active = list(combined.values())
            north = [row for row in active if row["project"] in north_projects]
            amal = [row for row in active if row["project"] == amal_project]
            south = [row for row in active if row not in north and row not in amal]
            total_achievement = sum(section["total"] for section in sections)
            remark = ""
            if total_achievement:
                parts = [f'A total of {total_achievement:,} individuals were reported for {entry["title"]} among {population}.']
                for label, rows in (("North Iraq (Erbil, Sulaymaniyah, Mosul and Kirkuk)", north), ("South Iraq (Baghdad and the Centre and Southern Governorates)", south), ("Al-Amal Centre", amal)):
                    if rows:
                        total = sum(row["total"] for row in rows)
                        locations_text = "; ".join(f'{row["location"]}: {row["total"]:,}' for row in rows)
                        parts.append(f'{label}: {total:,} individuals ({locations_text}).')
                contributions = entry.get("contributions", {})
                if isinstance(contributions.get("assessmentPeriod"), (int, float)):
                    parts.append(f'Of the total, {contributions["assessmentPeriod"]:,} were counted from assessments in the reporting period and {contributions.get("carryOver", 0):,} were completed carry-over legal-service cases.')
                remark = " ".join(parts)
            narrative_rows.append({"indicator": entry["title"], "population": population, "totalAchievement": total_achievement, "remarks": remark, "locations": active})
        if entry["id"] == "06-1-1-legal-assistance":
            special_remark = legal_assistance_narrative(entry)
            for row in narrative_rows: row["remarks"] = special_remark
        elif entry["id"] == "civil-representation":
            special_remark = civil_representation_narrative(entry)
            for row in narrative_rows: row["remarks"] = special_remark
        elif entry["id"] == "legal-awareness-participants":
            special_remark = legal_awareness_narrative(entry)
            for row in narrative_rows: row["remarks"] = special_remark
        entry["narrative"] = {"remark": "\n\n".join(row["remarks"] for row in narrative_rows if row["remarks"]), "rows": narrative_rows}
        for child in entry.get("children", []): attach_narrative(child)
    for group in groups:
        for item in group["indicators"]: attach_narrative(item)
    project_order = list(dict.fromkeys(project for project, _ in REPORT_ROWS))
    report_dates = []
    for frame, columns in ((assessments, (adate, arelease)), (services, (sprovision, scompleted))):
        for column in columns:
            if column: report_dates.extend(pd.to_datetime(frame[column], errors="coerce", dayfirst=True).dropna().tolist())
    for dataset, hints in (("deportationrecords", ("Date of Deportation Knowledge", "Date of deporting")), ("awareness", ("Date of Session",))):
        if dataset in frames:
            column = _find(list(frames[dataset].columns), *hints)
            if column: report_dates.extend(pd.to_datetime(frames[dataset][column], errors="coerce", dayfirst=True).dropna().tolist())
    month_options = sorted({date.strftime("%Y-%m") for date in report_dates}, reverse=True)
    quarter_options = sorted({f"{date.year}-Q{date.quarter}" for date in report_dates}, reverse=True)
    year_options = sorted({str(date.year) for date in report_dates}, reverse=True)
    return {"fromDate": from_date, "toDate": to_date, "ageGroups": list(AGE_GROUPS), "filterOptions": {"projects": project_order, "locations": list(dict.fromkeys(location for _, location in REPORT_ROWS)), "locationsByProject": {project: [location for row_project, location in REPORT_ROWS if row_project == project] for project in project_order}, "years": year_options, "quarters": quarter_options, "months": month_options, "communityTypes": community_options}, "activeFilters": {"projects": projects, "locations": locations, "years": years, "quarters": quarters, "months": months, "communityTypes": community_types}, "groups": groups}


def _carryover_match(row: pd.Series, provision_column: str | None, completed_column: str | None, status_column: str | None, type_column: str | None, in_period: Callable[[Any], bool]) -> bool:
    provision = pd.to_datetime(row.get(provision_column, ""), errors="coerce", dayfirst=True)
    completed_status = _norm(row.get(status_column, "")) in {"completed اكتملت", "completed اکتملت"}
    return not pd.isna(provision) and provision <= pd.Timestamp("2025-12-31") and in_period(row.get(completed_column, "")) and completed_status and not _contains(row.get(type_column, ""), "legal counselling")


def build_indicator_workbook(report: dict[str, Any], monthly_reports: list[tuple[str, dict[str, Any]]] | None = None) -> bytes:
    """Create a population-separated, fully labelled indicator workbook."""
    workbook = Workbook()
    blue, pale_blue, pale_orange, pale_green = "1687D9", "D9EAF7", "FCE8C3", "DCEED9"
    thin = Side(style="thin", color="B8C7D4")
    sheet_names = {
        "refugee": "Refugee",
        "idp": "IDP",
        "individual-beneficiaries-reached": "Individual Beneficiaries",
    }
    groups = [group for group in report["groups"] if group["indicators"]]

    for group_index, group in enumerate(groups):
        sheet = workbook.active if group_index == 0 else workbook.create_sheet()
        sheet.title = sheet_names.get(group["id"], group["label"][:31])
        sheet.sheet_view.showGridLines = False
        row_number = 1
        for item in group["indicators"]:
            for entry in [item, *item.get("children", [])]:
                sections = entry["sections"]
                # IDP reporting is limited to the single Al-Amal location: keep
                # the useful per-row Total column, but omit its redundant grand-total row.
                show_grand_total = group["id"] != "idp"
                last_column = 2 + 13 * len(sections) + 1
                sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=last_column)
                title = sheet.cell(row_number, 1, entry["title"])
                title.fill = PatternFill("solid", fgColor=blue); title.font = Font(color="FFFFFF", bold=True, size=12); title.alignment = Alignment(vertical="center")
                sheet.row_dimensions[row_number].height = 24
                row_number += 1
                column = 3
                for section in sections:
                    sheet.merge_cells(start_row=row_number, start_column=column, end_row=row_number, end_column=column + 12)
                    cell = sheet.cell(row_number, column, section["label"])
                    fill = pale_green if section["id"] == "idp" else pale_blue if section["id"] == "syrian-refugee" else "E8DFF5"
                    cell.fill = PatternFill("solid", fgColor=fill); cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
                    column += 13
                row_number += 1
                column = 3
                for _section in sections:
                    sheet.merge_cells(start_row=row_number, start_column=column, end_row=row_number, end_column=column + 5)
                    sheet.cell(row_number, column, "Male").fill = PatternFill("solid", fgColor=pale_blue)
                    sheet.merge_cells(start_row=row_number, start_column=column + 6, end_row=row_number, end_column=column + 11)
                    sheet.cell(row_number, column + 6, "Female").fill = PatternFill("solid", fgColor=pale_orange)
                    sheet.cell(row_number, column + 12, "Activity")
                    column += 13
                row_number += 1
                headers = ["Project", "Project Location"]
                for _section in sections: headers.extend([*AGE_GROUPS, *AGE_GROUPS, "Activity"])
                headers.append("Total")
                for index, value in enumerate(headers, 1):
                    cell = sheet.cell(row_number, index, value)
                    if 3 <= index <= 2 + 13 * len(sections):
                        section_offset = (index - 3) % 13
                        if section_offset < 6:
                            cell.fill = PatternFill("solid", fgColor=pale_blue)
                        elif section_offset < 12:
                            cell.fill = PatternFill("solid", fgColor=pale_orange)
                    elif index == last_column:
                        cell.fill = PatternFill("solid", fgColor="E7EEF4")
                header_row = row_number; row_number += 1
                base_rows = sections[0]["rows"] if sections else []
                for index, base in enumerate(base_rows):
                    values: list[Any] = [base["project"], base["location"]]
                    total = 0
                    for section in sections:
                        section_values = section["rows"][index]["values"]
                        values.extend([value if value else None for value in section_values[:12]])
                        values.append(None)
                        total += section_values[-1]
                    values.append(total)
                    for column, value in enumerate(values, 1):
                        cell = sheet.cell(row_number, column, value)
                        if column == last_column:
                            cell.fill = PatternFill("solid", fgColor="E7EEF4")
                    # Keep locations such as Pshdar Urban + Rania and projects
                    # such as Mosul & Kirkuk on the same single-line row height.
                    sheet.row_dimensions[row_number].height = 21
                    for column in (1, 2):
                        sheet.cell(row_number, column).alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
                    row_number += 1
                data_end_row = row_number - 1
                if show_grand_total:
                    totals: list[Any] = ["Total", "All selected locations"]
                    for section in sections:
                        totals.extend([value if value else None for value in section["totals"][:12]])
                        totals.append(None)
                    totals.append(sum(section["total"] for section in sections))
                    for column, value in enumerate(totals, 1):
                        cell = sheet.cell(row_number, column, value); cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="E7EEF4")
                table_end_row = row_number
                for row in sheet.iter_rows(min_row=header_row - 2, max_row=table_end_row, min_col=1, max_col=last_column):
                    for cell in row:
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        cell.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True) if cell.row <= data_end_row and cell.row > header_row and cell.column in (1, 2) else Alignment(horizontal="center", vertical="center", wrap_text=True)
                row_number += 2 if show_grand_total else 1
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 22
        for column in range(3, sheet.max_column + 1): sheet.column_dimensions[get_column_letter(column)].width = 9
    if not groups:
        workbook.active.title = "Indicator Report"
    if monthly_reports:
        analysis = workbook.create_sheet("Analysis")
        analysis.sheet_view.showGridLines = False
        analysis.merge_cells("A1:H1")
        analysis["A1"] = "Monthly indicator analysis"
        analysis["A1"].fill = PatternFill("solid", fgColor=blue)
        analysis["A1"].font = Font(color="FFFFFF", bold=True, size=14)
        analysis["A1"].alignment = Alignment(vertical="center")
        analysis.row_dimensions[1].height = 28
        selected = report.get("activeFilters", {})
        scope = ", ".join(value for values in selected.values() for value in values) or "All selected reporting data"
        analysis["A2"] = f"Applied filters: {scope}"
        analysis["A2"].font = Font(italic=True, color="5E7184")
        analysis.merge_cells("A2:H2")

        def entries(source: dict[str, Any]) -> dict[str, dict[str, Any]]:
            return {
                entry["id"]: entry
                for group in source["groups"]
                for item in group["indicators"]
                for entry in [item, *item.get("children", [])]
            }

        month_labels = [pd.Timestamp(f"{month}-01").strftime("%b %Y") for month, _ in monthly_reports]
        month_entries = [(month, entries(month_report)) for month, month_report in monthly_reports]
        row_number = 4
        for group in groups:
            for item in group["indicators"]:
                for entry in [item, *item.get("children", [])]:
                    analysis.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=max(3, len(month_labels) + 2))
                    title_cell = analysis.cell(row_number, 1, entry["title"])
                    title_cell.fill = PatternFill("solid", fgColor="E7F2FC")
                    title_cell.font = Font(color="116BB2", bold=True)
                    row_number += 1
                    population_rows: list[tuple[str, set[str]]] = []
                    section_ids = {section["id"] for section in entry["sections"]}
                    if section_ids & {"syrian-refugee", "non-syrian-refugee"}:
                        population_rows.append(("Refugees", {"syrian-refugee", "non-syrian-refugee"}))
                    if "idp" in section_ids:
                        population_rows.append(("IDP", {"idp"}))
                    headers = ["Month", *[population for population, _ in population_rows]]
                    for column, value in enumerate(headers, 1):
                        cell = analysis.cell(row_number, column, value)
                        cell.fill = PatternFill("solid", fgColor="D9EAF7")
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")
                    header_row = row_number
                    row_number += 1
                    for month_index, (month, monthly_entry_map) in enumerate(month_entries):
                        analysis.cell(row_number, 1, month_labels[month_index])
                        for column, (_, section_ids) in enumerate(population_rows, 2):
                            monthly_entry = monthly_entry_map.get(entry["id"])
                            value = sum(section["total"] for section in monthly_entry["sections"] if section["id"] in section_ids) if monthly_entry else 0
                            analysis.cell(row_number, column, value if value else None)
                        row_number += 1
                    data_end = row_number - 1
                    total_column = len(population_rows) + 1
                    analysis.cell(row_number, 1, "Total")
                    for column in range(2, total_column + 1):
                        analysis.cell(row_number, column, f"=SUM({get_column_letter(column)}{header_row + 1}:{get_column_letter(column)}{data_end})")
                    analysis.cell(row_number, 1).font = Font(bold=True)
                    for column in range(2, total_column + 1): analysis.cell(row_number, column).font = Font(bold=True)
                    if population_rows and month_labels:
                        chart = LineChart()
                        chart.title = "Monthly trend"
                        chart.style = 13
                        chart.y_axis.title = "Individuals"
                        chart.x_axis.title = "Month"
                        data = Reference(analysis, min_col=2, max_col=total_column, min_row=header_row, max_row=data_end)
                        categories = Reference(analysis, min_col=1, min_row=header_row + 1, max_row=data_end)
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(categories)
                        chart.height, chart.width = 7, 14
                        chart.legend.position = "b"
                        chart.anchor = f"{get_column_letter(total_column + 2)}{header_row - 1}"
                        analysis.add_chart(chart)
                    for row in analysis.iter_rows(min_row=header_row, max_row=row_number, min_col=1, max_col=total_column):
                        for cell in row:
                            cell.border = Border(bottom=thin)
                            cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left", vertical="center")
                    row_number += 3
        analysis.column_dimensions["A"].width = 56
        for column in range(2, len(month_labels) + 3): analysis.column_dimensions[get_column_letter(column)].width = 14
    buffer = BytesIO(); workbook.save(buffer)
    return buffer.getvalue()


def build_narrative_workbook(report: dict[str, Any]) -> bytes:
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Narrative Report"; sheet.sheet_view.showGridLines = False; sheet.freeze_panes = "A2"
    selected_months = report.get("activeFilters", {}).get("months", [])
    if len(selected_months) == 1:
        try: achievement_period = pd.Timestamp(f"{selected_months[0]}-01").strftime("%B %Y")
        except (TypeError, ValueError): achievement_period = str(selected_months[0])
    else: achievement_period = "Selected Period" if selected_months else "Reporting Period"
    headers = ["Indicators", "Population", f"Total Achievement - {achievement_period}", "Remarks"]
    sheet.append(headers)
    for cell in sheet[1]: cell.fill = PatternFill("solid", fgColor="1687D9"); cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    for group in report["groups"]:
        for item in group["indicators"]:
            for entry in [item, *item.get("children", [])]:
                for row in entry.get("narrative", {}).get("rows", []): sheet.append([row["indicator"], row["population"], row["totalAchievement"], row["remarks"]])
    thin = Side(style="thin", color="D6E0E8")
    for row in sheet.iter_rows(min_row=2):
        for cell in row: cell.border = Border(bottom=thin); cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[2].number_format = "#,##0"
    for column, width in {"A":55,"B":24,"C":27,"D":115}.items(): sheet.column_dimensions[column].width = width
    buffer=BytesIO(); workbook.save(buffer); return buffer.getvalue()
