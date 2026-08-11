"""Build 14 bilingual technology guides and synthetic practice projects."""
from pathlib import Path
from textwrap import dedent
import shutil

import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, Preformatted, SimpleDocTemplate, Spacer, Table, TableStyle

ROOT = Path(__file__).resolve().parents[1]
PDF_DIR = ROOT / "output" / "pdf"
EXAMPLES = ROOT / "learning-examples"
FONT = r"C:\Windows\Fonts\arial.ttf"
FONT_BOLD = r"C:\Windows\Fonts\arialbd.ttf"

GUIDES = [
 ("web-foundations", "Web Foundations", "HTML, CSS, JavaScript", "Build a safe analytics card that changes after a button click.", "HTML gives a page structure, CSS gives it appearance, and JavaScript gives it behaviour.", "<button id=\"add\">Add</button>\n<p id=\"total\">12 records</p>\n<script>let n=12; add.onclick=()=>total.textContent=`${++n} records`</script>", "Open index.html in a browser.", "web"),
 ("typescript", "TypeScript", "TypeScript", "Describe a dashboard record so mistakes are caught before the browser runs.", "TypeScript adds types to JavaScript. Types make autocomplete, refactoring, and API data safer.", "type Summary = { month: string; total: number };\nconst row: Summary = { month: 'January', total: 12 };\nconsole.log(row.total);", "npm install then npm run check", "ts"),
 ("react", "React", "React, React DOM", "Create a reusable component with changing state.", "React builds an interface from components and updates the screen when state changes.", "import {useState} from 'react';\nexport function Counter(){\n const [n,setN]=useState(12);\n return <button onClick={()=>setN(n+1)}>{n}</button>;\n}", "npm install then npm run dev", "react"),
 ("vite", "Vite", "Vite", "Run a development server and create a production build.", "Vite serves source files quickly in development and bundles an optimised dist folder for release.", "{\n  \"scripts\": { \"dev\": \"vite\", \"build\": \"vite build\" }\n}", "npm install; npm run dev; npm run build", "vite"),
 ("python-foundations", "Python Foundations", "Python", "Filter synthetic records and calculate a result.", "Python uses readable syntax and is well suited to APIs, automation, spreadsheets, and data work.", "records=[{'status':'open'},{'status':'closed'}]\ndef open_count(rows):\n    return sum(row['status']=='open' for row in rows)\nprint(open_count(records))", "python app.py", "python"),
 ("fastapi-uvicorn", "FastAPI and Uvicorn", "FastAPI, Uvicorn", "Create a local API route that returns a synthetic summary.", "FastAPI maps URLs to Python functions. Uvicorn runs that API server and FastAPI provides /docs automatically.", "from fastapi import FastAPI\napp=FastAPI()\n@app.get('/summary')\ndef summary(): return {'total':12,'source':'synthetic'}", "pip install -r requirements.txt; python -m uvicorn main:app --reload", "fastapi"),
 ("excel-openpyxl", "Excel Files with OpenPyXL", "OpenPyXL", "Create and read a synthetic Excel workbook.", "OpenPyXL lets Python create, read, and validate .xlsx sheets and cells without opening Excel.", "from openpyxl import Workbook\nbook=Workbook(); sheet=book.active\nsheet.append(['Month','Total']); sheet.append(['Jan',12])\nbook.save('demo.xlsx')", "pip install -r requirements.txt; python workbook.py", "excel"),
 ("data-analysis", "Data Analysis", "Pandas, Polars, PyArrow", "Group synthetic data by month using safe tabular-data tools.", "Pandas is flexible, Polars is fast for column expressions, and PyArrow provides typed columnar interchange.", "import polars as pl\nframe=pl.DataFrame({'month':['Jan','Jan'],'total':[4,8]})\nprint(frame.group_by('month').agg(pl.col('total').sum()))", "pip install -r requirements.txt; python analysis.py", "data"),
 ("plotly", "Interactive Charts with Plotly", "Plotly, react-plotly.js", "Render a small interactive chart from synthetic totals.", "Plotly creates browser charts with tooltips, zoom, selection, and image export.", "const data=[{type:'bar',x:['Jan','Feb'],y:[12,18]}];\nPlotly.newPlot('chart',data,{title:'Synthetic totals'});", "npm install then npm run dev", "plotly"),
 ("testing", "Testing", "Vitest, Pytest", "Write an automated check for a calculation.", "Tests document expected behaviour and catch regressions before users do.", "def total(values): return sum(values)\ndef test_total():\n    assert total([2,3]) == 5", "pip install -r requirements.txt; pytest", "testing"),
 ("exports", "PNG and PDF Export", "html2canvas, jsPDF, svg2pdf.js", "Export a small browser report as a PDF.", "html2canvas captures HTML, jsPDF creates a PDF, and svg2pdf.js preserves vector SVG when applicable.", "const canvas=await html2canvas(report);\nconst pdf=new jsPDF();\npdf.addImage(canvas.toDataURL(),'PNG',10,10,180,90);\npdf.save('report.pdf');", "npm install then npm run dev", "exports"),
 ("animation-icons", "UI Animation and Icons", "Rive, Lucide React", "Add an accessible icon and understand an animation asset.", "Icons should clarify actions; Rive animations should be optional and never distract from important data.", "import {Download} from 'lucide-react';\n<button aria-label=\"Download report\"><Download/> Download</button>", "npm install then npm run dev", "icons"),
 ("pywebview", "Desktop Applications with PyWebView", "PyWebView", "Open a local web page in a native desktop window.", "PyWebView wraps local web content in a native window and can provide a careful Python-to-JavaScript bridge.", "import webview\nwebview.create_window('Learning demo','index.html')\nwebview.start()", "pip install -r requirements.txt; python app.py", "pywebview"),
 ("windows-packaging", "Windows Packaging", "PyInstaller, Inno Setup", "Bundle a small Python app and understand an installer workflow.", "PyInstaller bundles Python and required files; Inno Setup installs the bundle for Windows users.", "pyinstaller --windowed --onedir --name LearningApp app.py\nISCC installer.iss", "pip install pyinstaller; pyinstaller --windowed app.py", "packaging"),
]

TOPICS = {
 "web-foundations": [
  "How browsers request, parse, and display a page", "Semantic HTML: headings, sections, forms, tables, and accessibility",
  "CSS selectors, cascade, inheritance, and specificity", "The box model, spacing, colours, typography, Flexbox, and Grid",
  "JavaScript values, variables, arrays, objects, and functions", "The DOM: finding, creating, changing, and removing elements",
  "Events: click, input, submit, keyboard, and event propagation", "Fetch, JSON, promises, async/await, loading, and error states",
  "Responsive design and browser developer tools", "Accessible, secure, and maintainable frontend habits"],
 "typescript": [
  "Why static type checking helps JavaScript projects", "Primitive types, inference, annotations, and literal types",
  "Arrays, tuples, object types, interfaces, and type aliases", "Optional properties, readonly data, unions, and intersections",
  "Typed functions, callbacks, promises, and async results", "Narrowing with typeof, in, equality, and discriminated unions",
  "Generics for reusable containers and functions", "Modules, imports, exports, and declaration files",
  "tsconfig, strict mode, compilation, and source maps", "Typing API responses without pretending untrusted data is safe"],
 "react": [
  "Components, JSX, rendering, and composition", "Props, one-way data flow, and reusable component interfaces",
  "State with useState and immutable updates", "Events, forms, controlled inputs, and validation",
  "Effects with useEffect and cleanup", "Lists, keys, conditional rendering, and empty states",
  "Derived values with useMemo and stable callbacks", "Sharing state, context, and custom hooks",
  "Loading, error, success, and accessibility patterns", "Component testing and avoiding common React anti-patterns"],
 "vite": [
  "What a development server and bundler do", "Creating a project and understanding index.html as the entry point",
  "ES modules, dependency resolution, and hot module replacement", "npm scripts and development versus production modes",
  "Environment variables and the VITE_ public prefix", "Static assets, public files, imports, and URLs",
  "TypeScript and React plugin integration", "Development proxy configuration for a local API",
  "Production builds, chunks, source maps, and preview", "Diagnosing missing modules, stale caches, and build failures"],
 "python-foundations": [
  "Running Python, indentation, comments, and expressions", "Numbers, strings, booleans, None, and type conversion",
  "Lists, tuples, dictionaries, sets, and comprehensions", "Conditions, for loops, while loops, and control flow",
  "Functions, parameters, return values, and scope", "Modules, imports, packages, and virtual environments",
  "Files, paths, context managers, JSON, and CSV", "Exceptions, validation, logging, and useful error messages",
  "Classes, dataclasses, and when objects are useful", "Testing, debugging, style, and safe handling of sensitive data"],
 "fastapi-uvicorn": [
  "HTTP requests, responses, methods, paths, and status codes", "Creating a FastAPI app and running it with Uvicorn",
  "Path parameters, query parameters, and request bodies", "Pydantic models, validation, serialization, and schemas",
  "GET, POST, file upload, and multipart form data", "Exceptions, status codes, middleware, and CORS",
  "Dependency injection and separating routes from business logic", "Async versus sync endpoints and blocking work",
  "Automatic OpenAPI documentation and manual API testing", "Security, upload limits, local-only hosting, and production checks"],
 "excel-openpyxl": [
  "Workbook, worksheet, row, column, and cell concepts", "Creating workbooks and appending typed rows",
  "Opening files safely in normal, read-only, and data-only modes", "Reading headers and validating expected sheet structure",
  "Dates, numbers, formulas, blank cells, and type surprises", "Styles, number formats, widths, filters, and frozen panes",
  "Iterating efficiently and avoiding unnecessary memory use", "Writing formulas versus reading cached formula values",
  "In-memory workbooks with BytesIO and upload validation", "Corrupt files, unsafe archives, privacy, and export testing"],
 "data-analysis": [
  "Rows, columns, schemas, data types, nulls, and identifiers", "Creating Pandas and Polars DataFrames",
  "Selecting, filtering, sorting, renaming, and deriving columns", "Group-by aggregation, counts, distinct counts, and percentages",
  "Joins, relationship keys, duplicates, and unmatched records", "Dates, strings, categories, and data cleaning",
  "Pandas eager operations and Polars expression style", "PyArrow schemas and efficient interchange formats",
  "Data-quality checks and reproducible transformations", "Performance, memory, correctness, and avoiding misleading metrics"],
 "plotly": [
  "Choosing a chart that matches the analytical question", "Traces, x/y data, layout, configuration, and rendering",
  "Bar, line, scatter, pie, histogram, and table trade-offs", "Labels, hover templates, legends, colours, and accessibility",
  "Sorting, top-N categories, percentages, and missing values", "Responsive sizing, margins, axes, and long category names",
  "Click and selection events for cross-filtering", "React Plotly components and immutable updates",
  "Exporting SVG, PNG, and presentation-safe charts", "Testing chart transformations separately from rendering"],
 "testing": [
  "Why tests matter and what a useful test proves", "Arrange, Act, Assert and descriptive test names",
  "Pytest discovery, assertions, fixtures, and parametrization", "Vitest suites, expectations, modules, and mocking",
  "Unit, integration, API, UI, and end-to-end test boundaries", "Testing success, empty, invalid, and failure cases",
  "Temporary files and synthetic test data", "Mocking time, network calls, and external services carefully",
  "Coverage, regression tests, and avoiding brittle tests", "Reading failures and debugging the smallest reproducible case"],
 "exports": [
  "Raster images, vectors, PDF pages, resolution, and aspect ratio", "Capturing a selected HTML element with html2canvas",
  "Creating pages, text, images, and metadata with jsPDF", "Converting SVG graphics with svg2pdf.js",
  "Pixel dimensions, DPI, scaling, and readable output", "Page size, orientation, margins, and multi-page reports",
  "Temporarily applying an export-safe theme", "Fonts, Arabic text, embedding, and glyph support",
  "Loading states, download names, errors, and browser limits", "Visual export testing and privacy review"],
 "animation-icons": [
  "Icons as communication rather than decoration", "Lucide React imports, sizing, colour, stroke, and labels",
  "Accessible names, tooltips, focus, and keyboard use", "Rive files, artboards, state machines, and animations",
  "Rendering Rive with WebGL and fallback behaviour", "Inputs, triggers, hover, click, and application state",
  "Performance, lazy loading, and resource cleanup", "Reduced-motion preferences and optional animation",
  "Poster images and graceful failure when assets are missing", "Testing meaningful UI without depending on animation timing"],
 "pywebview": [
  "How a native window can host a local web application", "Creating windows, choosing a URL, sizing, and lifecycle",
  "Starting a local API before opening the window", "Choosing a free localhost port and waiting for readiness",
  "The JavaScript-Python bridge and strict API boundaries", "Native open/save/folder dialogs",
  "Downloads, external links, permissions, and WebView2", "Threads, shutdown, cleanup, and preventing orphan servers",
  "Development browser mode versus packaged desktop mode", "Security, debugging, logs, and platform differences"],
 "windows-packaging": [
  "Why packaging differs from simply running source code", "Virtual environments, locked versions, and reproducible builds",
  "PyInstaller analysis, hidden imports, data files, and hooks", "One-folder versus one-file packaging trade-offs",
  "Windowed applications, icons, manifests, and version resources", "Testing the dist folder on a clean Windows machine",
  "Inno Setup sections, files, shortcuts, uninstall, and upgrades", "Application versioning and safe update manifests",
  "Code signing, hashes, release integrity, and SmartScreen", "Release checklists, rollback, support logs, and automation"],
}

NOTES = {
 "web-foundations": [
  "A browser requests files over HTTP, parses HTML into the DOM, parses CSS into style rules, runs JavaScript, then combines layout and paint instructions into pixels.",
  "Semantic elements describe purpose: main contains primary content, nav contains navigation, button performs an action, and label names a form control. Meaning improves accessibility and maintenance.",
  "Selectors choose elements. The cascade resolves competing rules using origin, importance, specificity, and source order. Inheritance passes selected properties such as font settings to descendants.",
  "Every element has content, padding, border, and margin. Flexbox arranges items in one dimension; Grid handles rows and columns. Responsive rules adapt layouts to screen width.",
  "Values include strings, numbers, booleans, null, arrays, and objects. let stores a changing binding; const prevents reassignment. Functions receive parameters and return reusable results.",
  "The DOM is the browser's object representation of HTML. querySelector finds an element; textContent changes text; createElement builds safe nodes; append attaches them to the page.",
  "Events report user and browser actions. addEventListener connects an event to a handler. preventDefault stops default form navigation; propagation explains how events travel through ancestors.",
  "fetch returns a Promise for an HTTP response. Check response.ok, parse JSON, await asynchronous work inside try/catch, and display separate loading, success, empty, and error states.",
  "Use relative units, flexible containers, media queries, and mobile-first rules. Browser developer tools inspect elements, computed styles, network requests, console errors, and responsive layouts.",
  "Keyboard access, visible focus, labels, sufficient contrast, escaped text, validated URLs, small modules, and clear naming make interfaces safer and easier to use."],
 "typescript": [
  "JavaScript errors often appear only at runtime. TypeScript analyses source before execution and reports incompatible values, missing properties, unsafe calls, and incorrect assumptions.",
  "string, number, boolean, null, and undefined are primitive types. Inference learns a type from an initial value; an annotation explicitly states the intended contract.",
  "Type arrays as Item[]. Tuples fix positions. An interface names an object shape and can extend another; a type alias can also represent unions and primitives.",
  "A question mark makes a property optional; readonly prevents assignment through that type. A union allows alternatives; an intersection combines requirements from multiple types.",
  "Type parameter values and the return value. Promise<Result> describes an asynchronous result. Callback types specify which arguments a function receives and what it returns.",
  "A union must be narrowed before using member-specific operations. typeof, property checks, equality, custom predicates, and discriminant fields prove which alternative is present.",
  "A generic type parameter preserves relationships without choosing a concrete type. function first<T>(items:T[]):T|undefined returns the same element type supplied by the caller.",
  "export exposes a value or type; import consumes it. Declaration files describe JavaScript libraries. Use import type when an import exists only for checking.",
  "tsconfig controls target JavaScript, module handling, included files, and strict checks. Strict mode catches null, implicit-any, function variance, and property initialization problems.",
  "A type assertion does not validate network JSON. Parse unknown input with explicit checks or a schema validator before treating it as a trusted application type."],
 "react": [
  "A component is a function that returns JSX. JSX becomes element descriptions. Composition places components inside one another so a screen is built from focused reusable parts.",
  "Props are read-only inputs supplied by a parent. Data flows downward; callbacks communicate actions upward. Small, explicit prop interfaces make components easier to understand and test.",
  "useState stores data between renders. Call its setter instead of mutating existing objects or arrays; create a new value so React can detect and render the update.",
  "Controlled inputs receive value from state and update it in onChange. Submit handlers validate, prevent default navigation, report field errors, and invoke application actions.",
  "useEffect synchronises with systems outside rendering, such as network requests or subscriptions. Dependencies control reruns; cleanup cancels or disconnects obsolete work.",
  "map renders lists; stable keys identify items across updates. Conditional rendering shows loading, empty, error, or content states instead of assuming data always exists.",
  "useMemo caches an expensive derived value and useCallback caches a function identity. Use them for measured problems, not automatically, because they add complexity.",
  "Lift shared state to the nearest common owner. Context shares broadly needed values. A custom hook extracts reusable stateful behaviour without creating visible UI.",
  "Disable repeated submissions, announce errors, preserve keyboard focus, label controls, and make loading state clear. Error boundaries handle unexpected rendering failures.",
  "Test visible behaviour and user actions. Avoid copying props into state, changing state during render, missing effect dependencies, unstable list keys, and giant all-purpose components."],
 "vite": [
  "A development server serves modules quickly and supports live updates. A production bundler resolves dependencies, transforms syntax, optimises assets, and writes deployable files.",
  "Vite treats index.html as part of the module graph. Its module script imports the application entry file, which then imports components, styles, and assets.",
  "Native ES modules use import and export. Vite pre-bundles dependencies and replaces changed modules through hot module replacement without refreshing the entire page.",
  "npm run dev starts development mode; npm run build creates dist; npm run preview serves that build locally. Production behaviour must be checked separately from development.",
  "Only variables prefixed VITE_ are exposed to browser code through import.meta.env. Never place passwords, private keys, or server secrets in frontend environment variables.",
  "Imported assets receive processed URLs and hashes. Files in public are copied unchanged and referenced from the root. Choose imports when the asset belongs to source code.",
  "Plugins extend Vite's transforms. The React plugin handles JSX and fast refresh; TypeScript is transpiled by Vite while tsc should run separately for type checking.",
  "A development proxy forwards selected paths such as /api to a backend server, avoiding hard-coded origins and simplifying local CORS behaviour.",
  "Build output may split code into chunks loaded on demand. Source maps connect built errors to source. preview verifies asset paths and production-only problems.",
  "Read the first build error, confirm the import path and filename case, inspect package versions, clear only relevant caches, and reproduce with the smallest module."],
 "python-foundations": [
  "Python executes statements in order. Indentation defines blocks. Comments begin with #. An expression produces a value; a statement performs an action such as assignment.",
  "int and float store numbers; str stores text; bool stores truth values; None represents absence. Explicit conversion such as int(text) can fail and should be validated.",
  "Lists are ordered and mutable, tuples are fixed sequences, dictionaries map keys to values, and sets keep unique values. Comprehensions build collections concisely.",
  "if chooses a branch, for iterates items, while repeats while a condition remains true, break exits, and continue skips to the next iteration.",
  "Functions name reusable work. Parameters receive input; return sends output. Default and keyword arguments improve clarity. Local variables normally exist only inside their function.",
  "A module is a Python file; a package groups modules. Imports reuse code. A virtual environment isolates project dependencies from other applications.",
  "Pathlib builds paths safely. with closes files automatically. json handles structured text; csv handles rows and columns. Always specify encoding when reading text.",
  "Raise exceptions when a function cannot fulfil its contract. Catch only errors you can handle, add context, and log operational facts without leaking sensitive data.",
  "A class combines state and behaviour; a dataclass reduces boilerplate for data records. Prefer functions and dictionaries until an object clearly improves the design.",
  "Small functions, type hints, tests, a debugger, and readable formatting support change. Validate files before processing and minimise personal data throughout the pipeline."],
 "fastapi-uvicorn": [
  "HTTP clients send a method, path, headers, and optional body. Servers return status, headers, and body. GET reads; POST submits; 2xx succeeds; 4xx is client error; 5xx is server failure.",
  "FastAPI holds route definitions. Uvicorn is the ASGI server that accepts connections and calls the app. The module:variable notation main:app imports app from main.py.",
  "Path parameters identify a resource, query parameters modify a request, and request bodies carry structured data. Type hints drive parsing, validation, and generated documentation.",
  "Pydantic models declare fields and constraints. FastAPI converts valid input to model instances and returns a detailed 422 response when validation fails.",
  "Use GET for safe retrieval and POST for submitted data. UploadFile streams uploaded content and exposes metadata; multipart handling is required for browser file forms.",
  "HTTPException produces deliberate client errors. Middleware wraps every request. CORS controls which browser origins may call the API; narrow it to known local origins.",
  "Dependencies provide shared validation, authentication, or resources. Keep route functions thin and move calculations into testable service modules.",
  "async def can await non-blocking I/O; normal def is suitable for synchronous work. CPU-heavy data processing still blocks and may require a worker strategy.",
  "FastAPI publishes OpenAPI at /openapi.json and interactive Swagger UI at /docs. Exercise routes with normal, empty, invalid, and oversized requests.",
  "Limit upload size and file type, distrust filenames and workbook structure, bind local tools to 127.0.0.1, avoid detailed internal errors, and test security boundaries."],
 "excel-openpyxl": [
  "An .xlsx workbook contains worksheets; each worksheet contains rows and columns; a cell has a coordinate, value, and optional style or formula.",
  "Workbook creates a file, active selects the default sheet, title renames it, append adds one row, and save writes the ZIP-based .xlsx package.",
  "load_workbook opens an existing file. read_only streams large sheets; data_only reads cached formula results instead of formula text when Excel has saved those results.",
  "Read the first row as headers, normalise harmless whitespace, reject missing required columns, and report unexpected or duplicate headers before analysing rows.",
  "Excel dates may be datetime objects or numbers with formats. Blank cells are None. Numeric identifiers can lose leading zeroes, so define expected column semantics.",
  "Styles control fonts, fills, borders, alignment, and number formats. Column widths, filters, and frozen headers improve a human-facing export but do not change data meaning.",
  "iter_rows(values_only=True) avoids creating unnecessary cell objects. Read only required sheets and columns, especially for large uploads.",
  "Writing '=SUM(B2:B5)' stores a formula. OpenPyXL does not calculate it; Excel or another calculation engine must generate the cached result.",
  "BytesIO keeps an uploaded workbook in memory. Validate ZIP structure and size before parsing, and return clear errors for missing or malformed sheets.",
  "Treat spreadsheets as untrusted input. Reject macros when unsupported, limit archive expansion, exclude sensitive fields from exports, and reopen generated files in tests."],
 "data-analysis": [
  "A schema names columns and their types. Null means missing, not zero. An identifier labels an entity and should not be accidentally converted to a measurement.",
  "pd.DataFrame and pl.DataFrame create tabular objects. Inspect shape, column names, dtypes, null counts, and sample rows before calculating anything.",
  "Select only needed columns, filter with explicit conditions, sort deliberately, rename to canonical names, and derive columns from documented rules.",
  "Group-by partitions rows by key and aggregates each group. Distinct counts require a defined identifier. Percentages need a documented denominator and filter context.",
  "Joins match keys across tables. Validate expected one-to-one or many-to-one relationships, duplicates, null keys, and unmatched rows to prevent silent multiplication or loss.",
  "Parse dates with invalid-value handling, normalise text cautiously, preserve raw values when auditing, and use categorical types for controlled repeated labels.",
  "Pandas operations are generally eager and index-aware. Polars builds column expressions and supports lazy optimisation. Translate logic explicitly instead of assuming identical behaviour.",
  "PyArrow stores typed columnar arrays and tables used by Parquet and inter-library transfer. A stable schema prevents repeated guessing and costly conversions.",
  "Check uniqueness, completeness, validity, consistency, and referential integrity. Make transformation steps deterministic and record rejected or corrected values.",
  "Optimise only after correctness: select early, avoid row loops, choose efficient types, and profile. Never improve speed by weakening validation or changing metric definitions."],
 "plotly": [
  "Start with the question: compare categories, show change over time, display distribution, or examine relationship. Choose the simplest chart that answers it.",
  "A trace contains data and mark type; layout controls titles, axes, legend, and spacing; config controls interactions such as mode bar and responsiveness.",
  "Bars compare categories, lines emphasise ordered change, scatterplots show relationships, histograms show distributions, pie charts suit very few parts, and tables preserve exact values.",
  "Use direct labels where practical, concise hover templates for detail, restrained colours, and a legend only when it removes ambiguity. Do not rely on colour alone.",
  "Sort by analytical purpose, group the tail when justified, disclose top-N filtering, define percentage denominator, and represent missing values separately from zero.",
  "Enable responsive sizing, reserve margins for labels, wrap or shorten long categories, set axis ranges deliberately, and test both narrow and wide screens.",
  "Plotly emits click and selection events containing points. Convert them into filter values, update application state, and clearly show or clear active cross-filters.",
  "react-plotly.js receives data, layout, config, and handlers as props. Produce new objects when values change and separate transformation logic from UI rendering.",
  "Vector SVG stays sharp; PNG needs sufficient dimensions and scale. Apply a high-contrast export theme and verify titles, legends, Arabic fonts, and margins.",
  "Unit-test functions that sort, aggregate, label, and build traces. Visual smoke tests then confirm Plotly renders the transformed structure as intended."],
 "testing": [
  "A useful test protects an important behaviour with deterministic input and a clear expected result. It should fail for the bug and pass for the correct implementation.",
  "Arrange creates input, Act calls the subject, and Assert compares the outcome. A name should state the scenario and expected behaviour.",
  "Pytest discovers test_ files and test_ functions. Plain assert gives readable diffs; fixtures provide setup; parametrize checks many cases with one test body.",
  "Vitest groups tests with describe, defines them with test, and checks results with expect. Mock modules only at real external boundaries.",
  "Unit tests isolate logic; integration tests connect modules; API tests exercise routes; UI tests exercise visible behaviour; end-to-end tests cover complete user flows.",
  "Test typical success plus empty collections, missing fields, malformed values, duplicates, boundary dates, permission errors, and failures from dependencies.",
  "Temporary directories isolate generated files and are cleaned automatically. Synthetic factories create realistic shapes without copying production personal data.",
  "Freeze time when dates affect results. Mock network responses for deterministic failures. Excessive mocking can produce tests that pass while integration is broken.",
  "Coverage locates unexecuted code but does not prove correctness. Add a regression test for every fixed bug and avoid assertions tied to irrelevant implementation details.",
  "Read the first failure, reproduce only that test, inspect the actual value, reduce the input, and distinguish a product defect from an outdated expectation."],
 "exports": [
  "Raster images store pixels and blur when enlarged; vector graphics store shapes and remain sharp. PDF pages can contain both. Aspect ratio prevents stretching.",
  "html2canvas reads the DOM and paints an approximation to canvas. Wait for fonts and images, select only the report element, and account for unsupported CSS.",
  "jsPDF creates a document with page size and units. Add text, vector operations, or images at explicit coordinates, then save or return the resulting bytes.",
  "svg2pdf.js translates SVG nodes into PDF vector commands. Use it for sharp charts, but verify fonts, clipping, transforms, and unsupported SVG features.",
  "Canvas width and height determine pixels. A higher scale improves print sharpness but increases memory and file size. Test the final physical dimensions.",
  "Choose A4 or Letter, portrait or landscape, and stable margins. For multi-page output, calculate available height and split content without cutting labels or rows.",
  "Export from a temporary white, high-contrast theme independent of the active screen theme. Restore the UI in finally even if export fails.",
  "A PDF can display only available or embedded glyphs. Register a font containing Arabic characters and verify shaping and right-to-left order in rendered pages.",
  "Disable duplicate clicks while generating, report progress and errors, choose descriptive safe filenames, revoke temporary URLs, and handle browser memory limits.",
  "Render the final file to images and inspect every page. Check clipping, blank charts, tiny labels, missing fonts, wrong dates, and accidental sensitive fields."],
 "animation-icons": [
  "An icon should reinforce a known action, not force users to guess. Pair unfamiliar icons with text and use a consistent visual vocabulary.",
  "Lucide icons are React components. Import only needed icons, set size and stroke consistently, inherit colour where possible, and avoid decorative inconsistency.",
  "Icon-only buttons require an accessible name. Keep visible focus, keyboard activation, adequate target size, and a tooltip as supplemental—not primary—information.",
  "A .riv file may contain artboards, named animations, and state machines. A state machine transitions according to boolean, number, and trigger inputs.",
  "The Rive React renderer loads the asset and renders to WebGL or canvas. Provide dimensions and a fallback when graphics support or the file fails.",
  "Read state-machine inputs once loaded, update booleans or numbers, and fire triggers for events. Keep application state authoritative rather than hiding logic in animation.",
  "Large assets and continuous rendering consume resources. Lazy-load offscreen animation, pause when hidden, limit instances, and release renderer resources on unmount.",
  "Respect prefers-reduced-motion by pausing nonessential movement or showing a poster. Never make animation required to understand status or complete a task.",
  "A poster image provides immediate and reliable content while Rive loads. Missing assets should leave a usable interface and a quiet diagnostic message.",
  "Test labels, controls, and state transitions independently of animation timing. A user must be able to complete the workflow when animation is disabled."],
 "pywebview": [
  "PyWebView creates a native window containing the platform web engine. The interface remains HTML/CSS/JavaScript while Python manages the process and native integration.",
  "create_window defines title, URL, width, height, and options. webview.start enters the GUI event loop, so initialise required services before or through a coordinated callback.",
  "A desktop web app often starts Uvicorn on localhost in a background thread, waits for a health endpoint, then points the window at the built frontend URL.",
  "Bind a temporary socket to port zero to obtain a free port, start the server, and poll readiness with a timeout. Report a clear error if startup fails.",
  "js_api exposes selected Python methods to JavaScript. Treat every call as untrusted input, expose the smallest surface, validate parameters, and avoid arbitrary filesystem access.",
  "Native dialogs can select files and folders. Constrain extensions where appropriate, handle cancellation, normalise returned paths, and never assume the user selected a valid workbook.",
  "Allow downloads deliberately and open external links in the default browser rather than the trusted app window. WebView2 is the normal Windows rendering engine.",
  "Coordinate server and window shutdown, signal Uvicorn to exit, join threads with limits, and clean temporary files so closing the app does not leave processes behind.",
  "During development, a browser and Vite server provide better debugging. In packaged mode, serve built files and open the same application inside PyWebView.",
  "Bind to localhost, randomise ports, validate bridge calls, log startup stages, hide sensitive values, and test WebView2 availability and file dialogs on supported Windows versions."],
 "windows-packaging": [
  "Source code assumes Python, packages, and assets exist. Packaging collects a controlled runtime so users can start the application without a development environment.",
  "Build from a clean virtual environment with pinned versions. Record the toolchain and regenerate dependencies so a future release can be reproduced.",
  "PyInstaller analyses imports but dynamic imports may need hidden-import entries. add-data bundles frontend files and icons; hooks collect package-specific binaries and metadata.",
  "One-folder builds start faster and are easier to inspect; one-file builds extract at launch and may trigger more security scanning. Test the chosen trade-off.",
  "--windowed hides the console for GUI apps. Provide a real .ico, version metadata, application manifest, and an accessible way to retrieve diagnostic logs.",
  "Copy dist to a clean Windows account or virtual machine without Python. Test startup, upload, charts, exports, dialogs, updates, uninstall, and offline behaviour.",
  "Inno Setup's Setup section defines identity and install location; Files copies the bundle; Icons creates shortcuts; uninstall information and upgrade rules support lifecycle management.",
  "Use semantic versions consistently in the app, installer, and manifest. An updater should verify trusted URLs, signatures or hashes, download safely, and support recovery.",
  "Code signing proves publisher identity and package integrity. Timestamp signatures, publish cryptographic hashes, protect signing keys, and understand that reputation develops over releases.",
  "Automate clean build, tests, package, malware scan, signature verification, install/uninstall smoke tests, checksum publication, rollback retention, and release notes."],
}

LESSON_CODE = {
 "web-foundations": [
  "<!-- Browser starts with this HTML file -->\n<h1>Hello, web!</h1>",
  "<main>\n  <h1>Dashboard</h1>\n  <button type=\"button\">Refresh</button>\n</main>",
  "p { color: navy; }\n.card p { font-weight: bold; }\n#total { font-size: 2rem; }",
  ".card { padding: 16px; border: 1px solid #ccc; }\n.grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }",
  "const project = 'Learning';\nlet total = 12;\nconst rows = [4, 8];\nfunction add(a, b) { return a + b; }",
  "const total = document.querySelector('#total');\ntotal.textContent = '12 records';\nconst note = document.createElement('p');\nnote.textContent = 'Synthetic data';\ndocument.body.append(note);",
  "const button = document.querySelector('#refresh');\nbutton.addEventListener('click', () => {\n  console.log('Refresh clicked');\n});",
  "async function loadSummary() {\n  const response = await fetch('/api/summary');\n  if (!response.ok) throw new Error('Request failed');\n  return response.json();\n}",
  "@media (max-width: 600px) {\n  .grid { grid-template-columns: 1fr; }\n}\n/* Inspect this rule in browser DevTools. */",
  "<label for=\"month\">Month</label>\n<select id=\"month\"><option>January</option></select>\n<button aria-label=\"Download report\">Download</button>"],
 "typescript": [
  "let total: number = 12;\n// total = 'twelve'; // Type error before runtime",
  "const project = 'Learning'; // inferred string\nlet ready: boolean = true;\nconst fixed: 2026 = 2026;",
  "interface Summary { month: string; total: number }\nconst rows: Summary[] = [{month:'Jan', total:12}];",
  "type Result = {ok:true; total:number} | {ok:false; error:string};\nconst value: Result = {ok:true, total:12};",
  "async function load(): Promise<number> {\n  return 12;\n}\nconst show = (value:number): string => `${value}`;",
  "function message(value: string | number) {\n  return typeof value === 'number' ? value.toFixed(0) : value.trim();\n}",
  "function first<T>(items: T[]): T | undefined {\n  return items[0];\n}\nconst month = first(['Jan','Feb']);",
  "// summary.ts\nexport type Summary = {total:number};\nexport const empty: Summary = {total:0};\n// another file: import {empty} from './summary';",
  "// tsconfig.json\n{\n  \"compilerOptions\": { \"strict\": true, \"target\": \"ES2022\" }\n}",
  "function isSummary(x: unknown): x is {total:number} {\n  return typeof x === 'object' && x !== null && 'total' in x;\n}"],
 "react": [
  "function Welcome() {\n  return <h1>Hello, React!</h1>;\n}\n// Render <Welcome /> inside the root.",
  "function Card({title, total}) {\n  return <section><h2>{title}</h2><p>{total}</p></section>;\n}\n<Card title=\"January\" total={12} />",
  "const [total, setTotal] = useState(12);\n<button onClick={() => setTotal(total + 1)}>{total}</button>",
  "const [month,setMonth] = useState('Jan');\n<form onSubmit={e => e.preventDefault()}>\n <input value={month} onChange={e => setMonth(e.target.value)} />\n</form>",
  "useEffect(() => {\n  const controller = new AbortController();\n  loadData(controller.signal);\n  return () => controller.abort();\n}, []);",
  "{loading ? <p>Loading...</p> : rows.length === 0 ? <p>No data</p> :\n rows.map(row => <p key={row.id}>{row.total}</p>)}",
  "const grandTotal = useMemo(\n  () => rows.reduce((sum,row) => sum + row.total, 0),\n  [rows]\n);",
  "function useDocumentTitle(title) {\n  useEffect(() => { document.title = title; }, [title]);\n}\n// Context is for values needed by many descendants.",
  "<button disabled={saving} aria-busy={saving}>\n  {saving ? 'Saving...' : 'Save'}\n</button>\n{error && <p role=\"alert\">{error}</p>}",
  "test('increases total', async () => {\n  render(<Counter />);\n  await user.click(screen.getByRole('button'));\n  expect(screen.getByText('13')).toBeVisible();\n});"],
 "vite": [
  "npm create vite@latest learning-app\ncd learning-app\nnpm install\nnpm run dev",
  "<!-- index.html -->\n<div id=\"root\"></div>\n<script type=\"module\" src=\"/src/main.js\"></script>",
  "// math.js\nexport const add = (a,b) => a+b;\n// main.js\nimport {add} from './math.js';\nconsole.log(add(2,3));",
  "{\n  \"scripts\": {\n    \"dev\": \"vite\",\n    \"build\": \"vite build\",\n    \"preview\": \"vite preview\"\n  }\n}",
  "// .env\nVITE_API_URL=http://127.0.0.1:8000\n// source\nconsole.log(import.meta.env.VITE_API_URL);",
  "import logoUrl from './logo.png';\nconst image = document.createElement('img');\nimage.src = logoUrl;\n// public/manual.pdf is referenced as /manual.pdf",
  "// vite.config.js\nimport {defineConfig} from 'vite';\nimport react from '@vitejs/plugin-react';\nexport default defineConfig({plugins:[react()]});",
  "export default defineConfig({\n  server:{proxy:{'/api':'http://127.0.0.1:8000'}}\n});",
  "npm run build\n# Inspect dist/\nnpm run preview\n# Test the production files locally",
  "npm ls vite\nnpm run build\n# Read the first error, then check its file and import path."],
 "python-foundations": [
  "# hello.py\nprint('Hello, Python!')\n# Run: python hello.py",
  "total = 12          # int\nrate = 0.75          # float\nmonth = 'January'    # str\nready = True         # bool\nmissing = None",
  "months = ['Jan','Feb']\npoint = (10,20)\nsummary = {'month':'Jan','total':12}\nunique_ids = {'A1','A2'}",
  "for total in [4,8,0]:\n    if total > 0:\n        print('Has data')\n    else:\n        print('Empty')",
  "def percentage(part, whole):\n    if whole == 0:\n        return 0\n    return part / whole * 100\nprint(percentage(3, 12))",
  "# utils.py: def add(a,b): return a+b\nfrom pathlib import Path\nprint(Path.cwd())\n# Create isolation: python -m venv .venv",
  "from pathlib import Path\nimport json\ndata = {'total':12}\nPath('summary.json').write_text(json.dumps(data), encoding='utf-8')",
  "try:\n    total = int(input('Total: '))\nexcept ValueError as exc:\n    print('Enter a whole number')",
  "from dataclasses import dataclass\n@dataclass\nclass Summary:\n    month: str\n    total: int\nprint(Summary('Jan',12))",
  "def total(values: list[int]) -> int:\n    return sum(values)\n\ndef test_total():\n    assert total([4,8]) == 12"],
 "fastapi-uvicorn": [
  "GET /api/summary HTTP/1.1\nHost: 127.0.0.1:8000\n\n# Response: 200 OK\n{\"total\":12}",
  "from fastapi import FastAPI\napp = FastAPI()\n@app.get('/health')\ndef health(): return {'status':'ready'}\n# Run: uvicorn main:app --reload",
  "@app.get('/records/{record_id}')\ndef record(record_id: int, details: bool=False):\n    return {'id':record_id,'details':details}",
  "from pydantic import BaseModel, Field\nclass Query(BaseModel):\n    month: str\n    limit: int = Field(10, ge=1, le=100)\n@app.post('/query')\ndef query(body:Query): return body",
  "from fastapi import UploadFile, File\n@app.post('/upload')\nasync def upload(file:UploadFile=File(...)):\n    return {'name':file.filename}",
  "from fastapi import HTTPException\nif not found:\n    raise HTTPException(404, 'Record not found')\n# Configure CORS only for known frontend origins.",
  "from fastapi import Depends\ndef page_size(limit:int=10): return min(limit,100)\n@app.get('/rows')\ndef rows(limit:int=Depends(page_size)): return []",
  "@app.get('/wait')\nasync def wait():\n    result = await asynchronous_io()\n    return result\n# CPU-heavy work is not made non-blocking by async.",
  "# Start server, then open:\nhttp://127.0.0.1:8000/docs\n# Try valid and invalid values and compare response status/details.",
  "# Local desktop binding\nuvicorn main:app --host 127.0.0.1 --port 8000\n# Validate file size/type/content; do not trust filename alone."],
 "excel-openpyxl": [
  "from openpyxl import Workbook\nbook=Workbook(); sheet=book.active\nprint(sheet.title, sheet['A1'].value)",
  "sheet.title='Summary'\nsheet.append(['Month','Total'])\nsheet.append(['Jan',12])\nbook.save('demo.xlsx')",
  "from openpyxl import load_workbook\nbook=load_workbook('demo.xlsx', read_only=True, data_only=True)\nprint(book.sheetnames)",
  "headers=[cell.value for cell in sheet[1]]\nrequired={'Month','Total'}\nmissing=required-set(headers)\nif missing: raise ValueError(f'Missing: {missing}')",
  "from datetime import datetime\nsheet.append([datetime(2026,1,1), 12])\nfor row in sheet.iter_rows(values_only=True): print(row)",
  "from openpyxl.styles import Font\nsheet['A1'].font=Font(bold=True)\nsheet.freeze_panes='A2'\nsheet.auto_filter.ref=sheet.dimensions",
  "for month,total in sheet.iter_rows(min_row=2, values_only=True):\n    print(month,total)\n# values_only avoids unnecessary cell objects.",
  "sheet['B3']='=SUM(B2:B2)'\nbook.save('formula.xlsx')\n# OpenPyXL stores this formula but does not calculate it.",
  "from io import BytesIO\nstream=BytesIO(upload_bytes)\nbook=load_workbook(stream, read_only=True)\n# Validate archive size before parsing.",
  "try:\n    book=load_workbook('upload.xlsx',read_only=True)\nexcept Exception as exc:\n    raise ValueError('Invalid .xlsx workbook') from exc"],
 "data-analysis": [
  "schema = {'Record ID':'string','Month':'string','Total':'integer'}\n# Missing is not zero; identifiers are not measurements.",
  "import pandas as pd\nimport polars as pl\nrows={'month':['Jan','Feb'],'total':[12,8]}\npdf=pd.DataFrame(rows); plf=pl.DataFrame(rows)",
  "clean=(plf.select('month','total')\n .filter(pl.col('total')>=0)\n .sort('month')\n .with_columns((pl.col('total')*2).alias('double')))\nprint(clean)",
  "print(plf.group_by('month').agg(\n pl.col('total').sum().alias('total'),\n pl.len().alias('rows')\n))",
  "people=pl.DataFrame({'id':['A','B'],'name':['One','Two']})\nvisits=pl.DataFrame({'id':['A','A'],'count':[1,2]})\nprint(visits.join(people,on='id',how='left'))",
  "clean=plf.with_columns(\n pl.col('month').str.strip_chars().str.to_uppercase(),\n pl.col('total').cast(pl.Int64, strict=False)\n)",
  "# Pandas\npdf.assign(double=pdf['total']*2)\n# Polars expression\nplf.with_columns((pl.col('total')*2).alias('double'))",
  "import pyarrow as pa\ntable=pa.table({'month':['Jan'],'total':pa.array([12],type=pa.int64())})\nprint(table.schema)",
  "assert plf['month'].null_count()==0\nassert plf['total'].min()>=0\nassert plf.height==plf.unique().height",
  "# Select early and avoid Python row loops.\nresult=(plf.lazy().filter(pl.col('total')>0)\n .group_by('month').agg(pl.col('total').sum()).collect())"],
 "plotly": [
  "// Question: compare totals by month.\n// A bar chart is clearer than a pie chart for this comparison.",
  "const data=[{type:'bar',x:['Jan','Feb'],y:[12,8]}];\nconst layout={title:'Monthly total'};\nPlotly.newPlot('chart',data,layout,{responsive:true});",
  "const bar={type:'bar',x:['A','B'],y:[4,8]};\nconst line={type:'scatter',mode:'lines+markers',x:['Jan','Feb'],y:[4,8]};",
  "const trace={type:'bar',x:['Jan'],y:[12],\n text:['12 records'],textposition:'auto',\n hovertemplate:'%{x}: %{y}<extra></extra>'};",
  "const rows=[{name:'A',value:20},{name:'B',value:8}];\nrows.sort((a,b)=>b.value-a.value);\n// Disclose if only top categories are shown.",
  "const layout={autosize:true,margin:{l:100,r:20,t:50,b:80},\n xaxis:{automargin:true},yaxis:{rangemode:'tozero'}};",
  "chart.on('plotly_click', event => {\n  const selected=event.points[0].x;\n  applyFilter(selected);\n});",
  "<Plot data={traces} layout={layout} config={{responsive:true}}\n onClick={event => setFilter(event.points[0].x)} />",
  "await Plotly.toImage(chart,{format:'png',width:1600,height:900,scale:2});\n// Render and inspect labels, legend, and margins.",
  "test('sorts descending',()=>{\n expect(buildTrace(rows).x).toEqual(['A','B']);\n});\n// Test transformation before visual rendering."],
 "testing": [
  "def total(values): return sum(values)\n\ndef test_total_adds_values():\n    assert total([4,8]) == 12",
  "def test_percentage():\n    # Arrange\n    part,whole=3,12\n    # Act\n    result=part/whole*100\n    # Assert\n    assert result==25",
  "import pytest\n@pytest.mark.parametrize('values,expected',[([],0),([2],2),([2,3],5)])\ndef test_total(values,expected): assert sum(values)==expected",
  "import {expect,test} from 'vitest';\nimport {total} from './math';\ntest('adds values',()=>expect(total([4,8])).toBe(12));",
  "# Unit: total([4,8])\n# Integration: workbook -> calculation\n# API: POST /query\n# UI: choose month and see total\n# E2E: upload -> chart -> export",
  "def test_empty(): assert total([])==0\ndef test_invalid():\n    with pytest.raises(TypeError): total(None)",
  "def test_export(tmp_path):\n    target=tmp_path/'report.txt'\n    target.write_text('Synthetic',encoding='utf-8')\n    assert target.read_text()=='Synthetic'",
  "from unittest.mock import patch\n@patch('service.fetch')\ndef test_network_failure(fetch):\n    fetch.side_effect=TimeoutError()\n    assert service.load()=={'error':'timeout'}",
  "# Coverage finds unexecuted lines; it does not prove assertions are meaningful.\n# Add a regression test that fails before each bug fix.",
  "pytest -q test_math.py::test_total\n# Read expected versus actual, inspect the value, then reduce the input."],
 "exports": [
  "// Raster: PNG pixels; vector: SVG shapes.\n// 1600x900 is 16:9. Enlarging a small raster image causes blur.",
  "const element=document.querySelector('#report');\nconst canvas=await html2canvas(element,{scale:2,backgroundColor:'#fff'});",
  "import {jsPDF} from 'jspdf';\nconst pdf=new jsPDF({orientation:'landscape',unit:'mm',format:'a4'});\npdf.text('Synthetic report',15,15);",
  "import 'svg2pdf.js';\nconst svg=document.querySelector('svg');\nawait pdf.svg(svg,{x:15,y:25,width:260,height:140});",
  "const scale=Math.min(2,window.devicePixelRatio||1);\nconst canvas=await html2canvas(report,{scale});\nconsole.log(canvas.width,canvas.height);",
  "const pageW=297, margin=15;\nconst contentW=pageW-margin*2;\npdf.addImage(image,'PNG',margin,20,contentW,150);\npdf.addPage();",
  "document.body.classList.add('export-theme');\ntry { await createPdf(); }\nfinally { document.body.classList.remove('export-theme'); }",
  "pdf.addFileToVFS('NotoNaskh.ttf',fontBase64);\npdf.addFont('NotoNaskh.ttf','NotoNaskh','normal');\npdf.setFont('NotoNaskh');\n// Render to verify Arabic shaping.",
  "button.disabled=true;status.textContent='Creating PDF...';\ntry{await exportReport()}catch(e){showError(e)}finally{button.disabled=false}",
  "// QA checklist:\n// no clipped text; chart present; readable labels; correct date;\n// no private fields; filename safe; every page visually inspected."],
 "animation-icons": [
  "import {Download} from 'lucide-react';\n<button><Download aria-hidden=\"true\"/> Download report</button>",
  "<Download size={20} strokeWidth={2} color=\"currentColor\" />\n// Import only the icons that are used.",
  "<button aria-label=\"Close dialog\" title=\"Close\">\n  <X aria-hidden=\"true\"/>\n</button>",
  "// A Rive file can contain an artboard, animations, and a state machine.\nconst {RiveComponent}=useRive({src:'/avatar/guardian.riv',stateMachines:'Main'});",
  "function Avatar(){\n const {RiveComponent}=useRive({src:'/avatar.riv',autoplay:true});\n return <RiveComponent aria-label=\"Guardian character\"/>;\n}",
  "const inputs=useStateMachineInput(rive,'Main','active');\nuseEffect(()=>{if(inputs) inputs.value=ready},[inputs,ready]);",
  "useEffect(()=>()=>rive?.cleanup(),[rive]);\n// Lazy-load the component and pause continuous animation when hidden.",
  "const reduce=matchMedia('(prefers-reduced-motion: reduce)').matches;\nreturn reduce ? <img src=\"/poster.webp\" alt=\"Guardian\"/> : <Avatar/>;",
  "return <picture><source srcSet=\"/poster.webp\"/><img src=\"/fallback.png\" alt=\"Guardian\"/></picture>;",
  "test('download remains usable without animation',()=>{\n render(<Page animations={false}/>);\n expect(screen.getByRole('button',{name:/download/i})).toBeEnabled();\n});"],
 "pywebview": [
  "import webview\nwebview.create_window('Learning app','index.html',width=900,height=650)\nwebview.start()",
  "window=webview.create_window('Dashboard','http://127.0.0.1:8000')\n# webview.start() enters the GUI event loop.",
  "threading.Thread(target=server.run,daemon=True).start()\nwait_until_ready('http://127.0.0.1:8000/health')\nwebview.create_window('App','http://127.0.0.1:8000')",
  "with socket.socket() as sock:\n    sock.bind(('127.0.0.1',0))\n    port=sock.getsockname()[1]\nprint(port)",
  "class Api:\n    def choose_month(self,month):\n        if month not in {'Jan','Feb'}: raise ValueError('Invalid month')\n        return {'month':month}\nwebview.create_window('App','index.html',js_api=Api())",
  "result=webview.windows[0].create_file_dialog(\n webview.FileDialog.OPEN,'',False,'',('Excel (*.xlsx)',))\nif not result: print('Cancelled')",
  "webview.settings['ALLOW_DOWNLOADS']=True\nwebview.settings['OPEN_EXTERNAL_LINKS_IN_BROWSER']=True\n# Validate every selected path.",
  "try:\n    webview.start()\nfinally:\n    server.should_exit=True\n    thread.join(timeout=5)",
  "# Development: Vite browser URL for DevTools.\n# Package: built frontend served locally inside Edge WebView2.",
  "# Bind only to 127.0.0.1. Validate bridge inputs.\n# Log startup stages without data values. Check WebView2 before launch."],
 "windows-packaging": [
  "# Development depends on installed Python and packages.\n# A package includes the interpreter, imports, binaries, and application assets.",
  "python -m venv .venv\n.\\.venv\\Scripts\\Activate.ps1\npip install -r requirements.txt\npip freeze > build-versions.txt",
  "python -m PyInstaller --onedir --windowed --name LearningApp `\n --add-data \"frontend/dist;frontend/dist\" app.py",
  "# One-folder: dist/LearningApp/ with visible files, faster startup.\n# One-file: one executable that extracts at startup.",
  "python -m PyInstaller --windowed --icon app.ico app.py\n# Add Windows version resources and keep diagnostic logs accessible.",
  "& '.\\dist\\LearningApp\\LearningApp.exe'\n# Test on a clean Windows VM: launch, dialogs, export, close, relaunch.",
  "[Setup]\nAppName=LearningApp\nAppVersion=1.0.0\nDefaultDirName={autopf}\\LearningApp\n[Files]\nSource: \"dist\\LearningApp\\*\"; DestDir: \"{app}\"; Flags: recursesubdirs",
  "APP_VERSION = '1.2.0'\n# Use the same version in app, installer, filename, and update manifest.\n# Verify update URL, hash/signature, and rollback path.",
  "Get-FileHash .\\LearningApp-Setup.exe -Algorithm SHA256\n# Sign with a protected certificate and timestamp; verify signature before release.",
  "# Release pipeline:\n# clean -> test -> build -> package -> scan -> sign -> install test -> hash -> publish -> retain rollback"],
}

AR = {
 "intro": "\u0634\u0631\u062d \u0628\u0633\u064a\u0637: \u062a\u0639\u0644\u0645 \u0627\u0644\u0641\u0643\u0631\u0629 \u0623\u0648\u0644\u0627\u064b\u060c \u062b\u0645 \u062c\u0631\u0628 \u0627\u0644\u0645\u062b\u0627\u0644 \u0627\u0644\u0635\u063a\u064a\u0631 \u0628\u064a\u0627\u0646\u0627\u062a \u0627\u0635\u0637\u0646\u0627\u0639\u064a\u0629 \u0641\u0642\u0637.",
 "goal": "\u0627\u0644\u0647\u062f\u0641: \u0628\u0646\u0627\u0621 \u0645\u062b\u0627\u0644 \u0635\u063a\u064a\u0631 \u0648\u0641\u0647\u0645\u0647 \u062e\u0637\u0648\u0629 \u0628\u062e\u0637\u0648\u0629.",
 "steps": "\u0627\u0644\u062e\u0637\u0648\u0627\u062a: \u0627\u0641\u062a\u062d \u0645\u062c\u0644\u062f \u0627\u0644\u0645\u062b\u0627\u0644\u060c \u0634\u063a\u0644 \u0627\u0644\u0623\u0645\u0631\u060c \u063a\u064a\u0631 \u0642\u064a\u0645\u0629 \u0648\u0627\u062d\u062f\u0629\u060c \u0648\u0644\u0627\u062d\u0638 \u0627\u0644\u0646\u062a\u064a\u062c\u0629.",
 "mistakes": "\u0623\u062e\u0637\u0627\u0621 \u0634\u0627\u0626\u0639\u0629: \u062a\u0634\u063a\u064a\u0644 \u0627\u0644\u0623\u0645\u0631 \u0645\u0646 \u0645\u062c\u0644\u062f \u062e\u0627\u0637\u0626\u060c \u0646\u0633\u064a\u0627\u0646 \u062a\u062b\u0628\u064a\u062a \u0627\u0644\u062a\u0628\u0639\u064a\u0627\u062a\u060c \u0623\u0648 \u062a\u063a\u064a\u064a\u0631 \u0623\u0634\u064a\u0627\u0621 \u0643\u062b\u064a\u0631\u0629 \u0628\u0645\u0631\u0629 \u0648\u0627\u062d\u062f\u0629.",
 "exercise": "\u062a\u0645\u0631\u064a\u0646: \u0623\u0636\u0641 \u0634\u0647\u0631\u0627\u064b \u0623\u0648 \u0642\u064a\u0645\u0629 \u0622\u0645\u0646\u0629 \u062c\u062f\u064a\u062f\u0629\u060c \u062b\u0645 \u0627\u0634\u0631\u062d \u0645\u0627 \u062d\u062f\u062b \u0628\u0643\u0644\u0645\u0627\u062a\u0643.",
}

def arabic(s): return get_display(arabic_reshaper.reshape(s))

def styles():
    pdfmetrics.registerFont(TTFont("Guide", FONT)); pdfmetrics.registerFont(TTFont("GuideBold", FONT_BOLD))
    base=getSampleStyleSheet()
    return {
      "title":ParagraphStyle("title",parent=base["Title"],fontName="GuideBold",fontSize=23,leading=28,textColor=colors.HexColor("#0B4F6C"),spaceAfter=10),
      "h":ParagraphStyle("h",parent=base["Heading2"],fontName="GuideBold",fontSize=14,leading=18,textColor=colors.HexColor("#0B4F6C"),spaceBefore=10,spaceAfter=5),
      "body":ParagraphStyle("body",parent=base["BodyText"],fontName="Guide",fontSize=10.2,leading=14.5,spaceAfter=6),
      "ar":ParagraphStyle("ar",parent=base["BodyText"],fontName="Guide",fontSize=10.2,leading=15,alignment=2,spaceAfter=7),
      "code":ParagraphStyle("code",fontName="Courier",fontSize=8.1,leading=10.5,backColor=colors.HexColor("#F2F6F8"),borderColor=colors.HexColor("#D9E2EC"),borderWidth=.5,borderPadding=7,spaceAfter=8),
    }

def footer(canvas,doc):
    canvas.saveState(); canvas.setFont("Guide",8); canvas.setFillColor(colors.HexColor("#52616B"))
    canvas.drawString(2*cm,1*cm,"Bilingual Technology Learning Pack - synthetic examples only")
    canvas.drawRightString(A4[0]-2*cm,1*cm,f"Page {doc.page}"); canvas.restoreState()

def pdf_quick(index, guide, st):
    slug,title,tools,goal,why,code,run,kind=guide
    out=PDF_DIR/f"{index:02d}-{slug}.pdf"
    doc=SimpleDocTemplate(str(out),pagesize=A4,leftMargin=2*cm,rightMargin=2*cm,topMargin=1.6*cm,bottomMargin=1.7*cm,title=title)
    parts=[Paragraph(f"{index:02d}. {title}",st["title"]),Paragraph(f"Tools: {tools}",st["body"])]
    for heading,en,ar_text in [
      ("Learning goal",goal,AR["goal"]),("Prerequisites","Windows, VS Code, a terminal, and an internet connection for first-time installs.","\u0627\u0644\u0645\u062a\u0637\u0644\u0628\u0627\u062a: \u0648\u064a\u0646\u062f\u0648\u0632\u060c VS Code\u060c \u0648\u0646\u0627\u0641\u0630\u0629 \u0623\u0648\u0627\u0645\u0631."),("Explain it simply",why,AR["intro"])]:
        parts += [Paragraph(heading,st["h"]),Paragraph(en,st["body"]),Paragraph(arabic(ar_text),st["ar"])]
    parts += [Paragraph("Core example",st["h"]),Preformatted(dedent(code),st["code"]),Paragraph("Build it step by step",st["h"]),Paragraph(f"1. Open <b>learning-examples/{index:02d}-{slug}</b>. 2. Read README.md. 3. Run: <b>{run}</b>. 4. Change one synthetic value and run it again.",st["body"]),Paragraph(arabic(AR["steps"]),st["ar"])]
    parts += [Paragraph("Expected result",st["h"]),Paragraph("You will see a small page, terminal result, API response, chart, test result, file export, or desktop window using synthetic values only.",st["body"]),Paragraph("Common mistakes",st["h"]),Paragraph("• Run commands from the wrong folder.<br/>• Forget dependency installation.<br/>• Change multiple things at the same time.",st["body"]),Paragraph(arabic(AR["mistakes"]),st["ar"])]
    parts += [Paragraph("Practice and dashboard connection",st["h"]),Paragraph("Try the exercise-solution file only after your own attempt. This dashboard uses this technology as part of its local React/Vite frontend, FastAPI/Python backend, data pipeline, exports, or Windows desktop distribution.",st["body"]),Paragraph(arabic(AR["exercise"]),st["ar"])]
    next_text="You completed the path." if index==len(GUIDES) else f"Next: {index+1:02d}. {GUIDES[index][1]}"
    parts += [Spacer(1,5),Table([["Next step",next_text]],colWidths=[3*cm,12*cm],style=TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EAF4F4")),("BOX",(0,0),(-1,-1),.5,colors.HexColor("#9AD1D4")),("FONTNAME",(0,0),(-1,-1),"Guide"),("FONTSIZE",(0,0),(-1,-1),9),("LEFTPADDING",(0,0),(-1,-1),7),("RIGHTPADDING",(0,0),(-1,-1),7),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))]
    doc.build(parts,onFirstPage=footer,onLaterPages=footer)

def pdf(index, guide, st):
    slug,title,tools,goal,why,code,run,kind=guide
    out=PDF_DIR/f"{index:02d}-{slug}.pdf"
    doc=SimpleDocTemplate(str(out),pagesize=A4,leftMargin=2*cm,rightMargin=2*cm,topMargin=1.6*cm,bottomMargin=1.7*cm,title=title)
    parts=[Paragraph(f"{index:02d}. {title}",st["title"]),Paragraph("Detailed beginner handbook - English and Arabic",st["body"]),Paragraph(f"Technologies: {tools}",st["body"])]
    parts += [Paragraph("What you will learn",st["h"]),Paragraph(goal,st["body"]),Paragraph(why,st["body"]),Paragraph(arabic(AR["goal"]),st["ar"]),Paragraph(arabic(AR["intro"]),st["ar"])]
    parts += [Paragraph("How to study",st["h"]),Paragraph("Study one lesson at a time. Type examples yourself, predict the result, run the code, then change one thing. Keep a notebook of new words, errors, root causes, and fixes.",st["body"]),Paragraph("Course map",st["h"]),Paragraph("Setup and mental model; ten core concepts; a guided mini-project; the real dashboard architecture; debugging; exercises and answers.",st["body"]),PageBreak()]

    parts += [Paragraph("Lesson 1 - Setup and mental model",st["title"]),Paragraph("Prerequisites",st["h"]),Paragraph("Windows 10 or 11, VS Code, PowerShell, and permission to install learning dependencies. Use only synthetic data in tutorials and screenshots.",st["body"])]
    parts += [Paragraph("Setup sequence",st["h"]),Paragraph(f"1. Open PowerShell.<br/>2. Go to <b>learning-examples/{index:02d}-{slug}</b>.<br/>3. Read README.md.<br/>4. Run <b>{run}</b>.<br/>5. Read the complete error if it fails.<br/>6. Change one safe value and rerun.",st["body"]),Paragraph(arabic(AR["steps"]),st["ar"])]
    parts += [Paragraph("Mental model",st["h"]),Paragraph(f"Treat {tools} as one layer in a system. Identify the input, the transformation or rule, the output, and possible failures. Understanding that flow is more valuable than memorising syntax.",st["body"]),Paragraph("Checkpoint",st["h"]),Paragraph("Explain: What problem does this technology solve? What is its input? What output should it produce? What can fail?",st["body"]),PageBreak()]

    topics=TOPICS[slug]
    for number,topic in enumerate(topics,1):
        note=NOTES[slug][number-1]
        example=LESSON_CODE[slug][number-1]
        parts += [Paragraph(f"Lesson {number+1} - {topic}",st["title"])]
        parts += [Paragraph("What is it?",st["h"]),Paragraph(note,st["body"])]
        parts += [Paragraph("Why do we need it?",st["h"]),Paragraph(f"Without understanding this idea, code may appear to work while handling data incorrectly or failing on a different input. In {title}, this concept helps create behaviour that is explicit, testable, and easier to change.",st["body"])]
        parts += [Paragraph("Syntax and example",st["h"]),Preformatted(dedent(example),st["code"])]
        parts += [Paragraph("What happens step by step",st["h"]),Paragraph("1. Read the example from the first line downward.<br/>2. Identify the input or starting value.<br/>3. Identify the operation, rule, or declaration.<br/>4. Find the output or visible effect.<br/>5. Predict the result before running it.<br/>6. Run it and compare the real result with your prediction.",st["body"])]
        parts += [Paragraph("Try it yourself",st["h"]),Paragraph(f"Type the example instead of pasting it. Change one name and one synthetic value. Then introduce an empty or incorrect value and observe the result. Explain how this demonstrates <b>{topic}</b>.",st["body"])]
        ar_text="\u0647\u0630\u0627 \u0627\u0644\u062f\u0631\u0633 \u064a\u0634\u0631\u062d: "+topic+". \u0627\u0643\u062a\u0628 \u0627\u0644\u0645\u062b\u0627\u0644 \u0628\u0646\u0641\u0633\u0643\u060c \u062d\u062f\u062f \u0627\u0644\u0645\u062f\u062e\u0644\u0627\u062a \u0648\u0627\u0644\u0639\u0645\u0644\u064a\u0629 \u0648\u0627\u0644\u0646\u062a\u064a\u062c\u0629\u060c \u062b\u0645 \u063a\u064a\u0631 \u0642\u064a\u0645\u0629 \u0648\u0627\u062d\u062f\u0629 \u0644\u062a\u0641\u0647\u0645 \u0627\u0644\u0633\u0644\u0648\u0643."
        parts += [Paragraph(arabic(ar_text),st["ar"]),Paragraph("Quick check",st["h"]),Paragraph("Can you define the concept, point to it in the example, predict the output, and change the example without help? If not, repeat this lesson before continuing.",st["body"]),PageBreak()]

    parts += [Paragraph("Lesson 7 - Guided mini-project",st["title"]),Paragraph("Project goal",st["h"]),Paragraph(goal,st["body"]),Paragraph("Starting example",st["h"]),Preformatted(dedent(code),st["code"])]
    parts += [Paragraph("Read the code",st["h"]),Paragraph("Find imported tools or page elements. Identify the synthetic input. Find the function, event, route, or transformation that changes it. Finally, locate where the result is displayed, returned, saved, or packaged.",st["body"])]
    parts += [Paragraph("Build sequence",st["h"]),Paragraph("1. Run the untouched example.<br/>2. Record the result.<br/>3. Rename one unclear value.<br/>4. Add a synthetic record.<br/>5. Validate empty input.<br/>6. Validate incorrect input.<br/>7. Add a helpful error.<br/>8. Rerun every case.<br/>9. Compare with exercise-solution.<br/>10. Explain every line.",st["body"])]
    parts += [Paragraph("Definition of done",st["h"]),Paragraph("Normal, empty, and invalid cases behave deliberately; no private data is used; and you can explain every line without reading the guide.",st["body"]),PageBreak()]

    parts += [Paragraph("Lesson 8 - The real dashboard",st["title"]),Paragraph("Architecture connection",st["h"]),Paragraph(f"The application combines {tools} with the rest of the stack. The frontend gathers filters and files; the local API validates requests; the data layer transforms workbook rows; charts present aggregates; export tools create files; and the desktop layer packages the application for Windows.",st["body"])]
    parts += [Paragraph("Trace the upload feature",st["h"]),Paragraph("The user selects a workbook. The frontend sends it to the local API. The backend validates the file and sheets. The data layer creates safe tables. The API returns metadata. React updates state. Plotly displays results. Identify exactly where this guide's technology participates.",st["body"])]
    parts += [Paragraph("Privacy and quality",st["h"]),Paragraph("Do not place narrative, contact, or identifying fields in tutorials, logs, screenshots, tests, or exports. Use generated identifiers and synthetic totals. Validate at each boundary and collect only fields required for the calculation.",st["body"]),Paragraph("Architecture exercise",st["h"]),Paragraph("Draw five boxes: UI, API, Data Processing, Export, Desktop Packaging. Add arrows and label the data. Circle the box related to this guide and add two validation checks.",st["body"]),PageBreak()]

    parts += [Paragraph("Lesson 9 - Debugging",st["title"]),Paragraph("Reliable debugging loop",st["h"]),Paragraph("1. Reproduce with the smallest input.<br/>2. Read the first meaningful error.<br/>3. Find the exact file and line.<br/>4. Inspect value and type.<br/>5. Compare expected and actual results.<br/>6. Change one thing.<br/>7. Rerun the same case.<br/>8. Add a regression test.",st["body"])]
    parts += [Paragraph("Common beginner mistakes",st["h"]),Paragraph("- Wrong folder or command.<br/>- Missing dependency or inactive environment.<br/>- Misspelled file, variable, field, route, or import.<br/>- Assuming input is always present.<br/>- Mixing setup, processing, and display in one block.<br/>- Hiding an exception rather than understanding it.<br/>- Using real sensitive data in a test.",st["body"]),Paragraph(arabic(AR["mistakes"]),st["ar"])]
    parts += [Paragraph("Error notebook",st["h"]),Paragraph("Record: command, expected result, actual error, root cause, smallest fix, and test added. This turns errors into reusable knowledge.",st["body"]),PageBreak()]

    parts += [Paragraph("Lesson 10 - Exercises and answers",st["title"]),Paragraph("Exercises",st["h"]),Paragraph("1. Explain the technology in three sentences.<br/>2. Recreate the example without copying.<br/>3. Add one normal synthetic record.<br/>4. Handle empty input.<br/>5. Handle invalid input.<br/>6. Improve unclear names.<br/>7. Add or describe one test.<br/>8. Locate the technology in the dashboard.<br/>9. Record one error and root cause.<br/>10. Add one mini-project feature.",st["body"]),Paragraph(arabic(AR["exercise"]),st["ar"])]
    parts += [Paragraph("Answer guide",st["h"]),Paragraph("A strong solution is observable and explainable: the documented command works; output changes for the new record; empty and invalid cases have deliberate results; names communicate purpose; a test states an expected result; and the architecture explanation identifies the correct boundary.",st["body"])]
    next_text="You completed the learning path." if index==len(GUIDES) else f"Next: {index+1:02d}. {GUIDES[index][1]}"
    parts += [Spacer(1,8),Table([["Next step",next_text]],colWidths=[3*cm,12*cm],style=TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EAF4F4")),("BOX",(0,0),(-1,-1),.5,colors.HexColor("#9AD1D4")),("FONTNAME",(0,0),(-1,-1),"Guide"),("FONTSIZE",(0,0),(-1,-1),9),("PADDING",(0,0),(-1,-1),7)]))]
    doc.build(parts,onFirstPage=footer,onLaterPages=footer)

def files(kind):
    common={"README.md":"# Practice example\n\nUse only synthetic data. Read the matching PDF, run the example, change one value, then compare with exercise-solution.\n","exercise-solution.txt":"Change one synthetic value, run the example again, and write down why the result changed.\n"}
    if kind=="web": common["index.html"]="<main><h1>Safe dashboard demo</h1><p id='total'>12 records</p><button id='add'>Add</button></main><script>let n=12;add.onclick=()=>total.textContent=`${++n} records`</script>"
    elif kind=="ts": common.update({"package.json":"{\"scripts\":{\"check\":\"tsc --noEmit\"},\"devDependencies\":{\"typescript\":\"latest\"}}","index.ts":"type Summary={month:string;total:number};const row:Summary={month:'Jan',total:12};console.log(row);\n"})
    elif kind in {"react","icons"}: common.update({"package.json":"{\"scripts\":{\"dev\":\"vite\"},\"dependencies\":{\"vite\":\"latest\",\"react\":\"latest\",\"react-dom\":\"latest\",\"lucide-react\":\"latest\"}}","index.html":"<div id='root'></div><script type='module' src='/src.jsx'></script>","src.jsx":"import React,{useState}from'react';import{createRoot}from'react-dom/client';function App(){const[n,setN]=useState(12);return <button onClick={()=>setN(n+1)}>Total: {n}</button>}createRoot(document.querySelector('#root')).render(<App/>);\n"})
    elif kind in {"vite","plotly","exports"}: common.update({"package.json":"{\"scripts\":{\"dev\":\"vite\",\"build\":\"vite build\"},\"devDependencies\":{\"vite\":\"latest\"}}","index.html":"<main><h1>Synthetic demo</h1><p>12 records</p></main><script type='module' src='/src.js'></script>","src.js":"console.log('Synthetic learning example is ready.');\n"})
    elif kind=="fastapi": common.update({"requirements.txt":"fastapi\nuvicorn[standard]\n","main.py":"from fastapi import FastAPI\napp=FastAPI()\n@app.get('/summary')\ndef summary(): return {'total':12,'source':'synthetic'}\n"})
    elif kind=="excel": common.update({"requirements.txt":"openpyxl\n","workbook.py":"from openpyxl import Workbook\nb=Workbook();b.active.append(['Month','Total']);b.active.append(['Jan',12]);b.save('demo.xlsx');print('Created demo.xlsx')\n"})
    elif kind=="data": common.update({"requirements.txt":"pandas\npolars\npyarrow\n","analysis.py":"import polars as pl\nf=pl.DataFrame({'month':['Jan','Jan'],'total':[4,8]});print(f.group_by('month').agg(pl.col('total').sum()))\n"})
    elif kind=="testing": common.update({"requirements.txt":"pytest\n","calculator.py":"def total(values): return sum(values)\n","test_calculator.py":"from calculator import total\ndef test_total(): assert total([2,3])==5\n","vitest-example.js":"import{expect,test}from'vitest';test('adds',()=>expect(2+3).toBe(5));\n"})
    elif kind=="pywebview": common.update({"requirements.txt":"pywebview\n","index.html":"<h1>Local desktop demo</h1><p>Synthetic data only.</p>","app.py":"import webview\nwebview.create_window('Learning demo','index.html');webview.start()\n"})
    elif kind=="packaging": common.update({"app.py":"import tkinter as tk\nr=tk.Tk();tk.Label(r,text='Synthetic demo').pack(padx=30,pady=20);r.mainloop()\n","build.ps1":"python -m PyInstaller --noconfirm --windowed --onedir --name LearningApp app.py\n","installer.iss":"[Setup]\nAppName=LearningApp\nAppVersion=1.0\nDefaultDirName={autopf}\\LearningApp\n"})
    else: common["app.py"]="print('Synthetic learning example')\n"
    return common

def build_start_here(st):
    lessons=[
      ("What programming is","A program is a precise list of instructions that a computer executes. Source code is human-readable text. An interpreter or compiler translates it into work the computer can perform.","input -> instructions -> output","Describe the input and output of a calculator."),
      ("Files, folders, and extensions","A file stores content and has a name. A folder organises files. The extension after the final dot suggests the format: .py for Python, .html for HTML, .ts for TypeScript, .json for structured text, and .pdf for documents.","Learning/\n  hello.py\n  index.html\n  data/\n    summary.json","Create this folder structure in a safe learning folder."),
      ("Installing and using VS Code","VS Code is a text editor for code. Open a folder, use the Explorer to create files, use tabs to edit them, save with Ctrl+S, and read highlighted syntax and problems. Extensions add language support but do not replace the language runtime.","File -> Open Folder\nExplorer -> New File\nTerminal -> New Terminal\nCtrl+S -> Save","Open learning-examples/00-start-here in VS Code."),
      ("PowerShell and the terminal","A terminal lets you type commands. PowerShell shows a prompt. The current directory decides which files a command sees. Use pwd to show it, Get-ChildItem to list items, cd to change folder, and clear to clear the screen.","pwd\nGet-ChildItem\ncd .\\learning-examples\\00-start-here\nGet-ChildItem","Navigate to the practice folder and list its files."),
      ("Commands, arguments, and errors","The first word names a command; later words are arguments. Quoted paths preserve spaces. Exit code zero normally means success. Error output often names the missing command, file, package, or source line.","python hello.py\n# command: python\n# argument: hello.py\ncd \"C:\\A Folder With Spaces\"","Run one valid command and one deliberately misspelled command; compare messages."),
      ("Installing Python","Python runs backend and data examples. During installation select the option that adds Python to PATH. Verify with python --version. pip installs packages, while python -m pip ensures pip belongs to the selected Python.","python --version\npython -m pip --version\npython -c \"print('Python works')\"","Run the three verification commands."),
      ("Your first Python program","Create hello.py, type a print call, save, then run it from the same folder. Quotes create text. Parentheses pass the text to print. The terminal shows the program's output.","# hello.py\nprint('Hello! I ran my first Python program.')\n# PowerShell\npython hello.py","Change the message and run the file again."),
      ("Installing Node.js and npm","Node.js runs JavaScript outside the browser and supplies npm, the package manager used by frontend projects. Verify both commands. package.json describes scripts and dependencies for one project.","node --version\nnpm --version\nnpm init -y","Verify Node and npm; do not install packages globally unless a guide requires it."),
      ("Your first browser page","A browser opens an HTML file. HTML tags describe content. Save index.html, double-click it, then refresh after changes. Browser developer tools open with F12 and show elements, console messages, and network activity.","<!doctype html>\n<html>\n<body>\n  <h1>My first page</h1>\n  <p>Hello, browser!</p>\n</body>\n</html>","Change the heading, save, and refresh the browser."),
      ("Dependencies and package managers","A dependency is code your project uses. npm install reads package.json for JavaScript. pip install -r requirements.txt reads Python requirements. Dependencies belong to the project and should use known compatible versions.","npm install\npython -m pip install -r requirements.txt","Find package.json and requirements.txt in the learning examples."),
      ("Python virtual environments","A virtual environment gives one Python project its own installed packages. Create .venv, activate it, install requirements, and deactivate when finished. The prompt normally shows (.venv) while active.","python -m venv .venv\n.\\.venv\\Scripts\\Activate.ps1\npython -m pip install -r requirements.txt\ndeactivate","Create and activate a virtual environment in a disposable practice folder."),
      ("How to learn and debug","Do not memorise everything. Type small examples, predict results, change one value, and explain the result. When an error occurs, read it fully, locate the first relevant file and line, inspect the value, and change one thing at a time.","Expected: 5\nActual: TypeError on line 4\nCause: text was added to a number\nFix: convert or validate input\nTest: normal, empty, invalid","Start an error notebook using these five labels."),
    ]
    out=PDF_DIR/'00-start-here.pdf'
    doc=SimpleDocTemplate(str(out),pagesize=A4,leftMargin=2*cm,rightMargin=2*cm,topMargin=1.6*cm,bottomMargin=1.7*cm,title='Start Here')
    story=[Paragraph("00. Start Here - Absolute Beginner",st["title"]),Paragraph("No previous computer programming knowledge required",st["body"]),Paragraph("This guide explains the words and actions used by every later PDF. Complete it before Web Foundations.",st["body"]),Paragraph(arabic("\u0627\u0628\u062f\u0623 \u0645\u0646 \u0647\u0646\u0627. \u0644\u0627 \u062a\u062d\u062a\u0627\u062c \u0625\u0644\u0649 \u0623\u064a \u0645\u0639\u0631\u0641\u0629 \u0633\u0627\u0628\u0642\u0629 \u0628\u0627\u0644\u0628\u0631\u0645\u062c\u0629. \u0627\u062f\u0631\u0633 \u0647\u0630\u0627 \u0627\u0644\u062f\u0644\u064a\u0644 \u0642\u0628\u0644 \u0628\u0642\u064a\u0629 \u0627\u0644\u0643\u062a\u0628."),st["ar"]),Paragraph("Learning rule",st["h"]),Paragraph("Type every example. Do not paste it. A typing mistake is useful practice because learning to read errors is part of programming.",st["body"]),PageBreak()]
    for i,(title,explanation,example,exercise) in enumerate(lessons,1):
        story += [Paragraph(f"Lesson {i} - {title}",st["title"]),Paragraph("Explanation",st["h"]),Paragraph(explanation,st["body"]),Paragraph("Example",st["h"]),Preformatted(example,st["code"]),Paragraph("Do it step by step",st["h"]),Paragraph("1. Read the example.<br/>2. Type it exactly.<br/>3. Save the file if applicable.<br/>4. Run the command.<br/>5. Compare what happened with the explanation.<br/>6. Change one small value and repeat.",st["body"]),Paragraph("Try it yourself",st["h"]),Paragraph(exercise,st["body"]),Paragraph(arabic("\u0627\u0643\u062a\u0628 \u0627\u0644\u0645\u062b\u0627\u0644 \u0628\u0646\u0641\u0633\u0643\u060c \u062b\u0645 \u0634\u063a\u0644\u0647 \u0648\u063a\u064a\u0631 \u0634\u064a\u0626\u0627\u064b \u0648\u0627\u062d\u062f\u0627\u064b. \u0625\u0630\u0627 \u0638\u0647\u0631 \u062e\u0637\u0623\u060c \u0627\u0642\u0631\u0623 \u0627\u0644\u0631\u0633\u0627\u0644\u0629 \u0643\u0627\u0645\u0644\u0629."),st["ar"]),PageBreak()]
    story += [Paragraph("Ready for Guide 01",st["title"]),Paragraph("Before continuing, you should be able to create a folder and file, open a folder in VS Code, navigate with PowerShell, run a Python file, open an HTML page, verify Python/Node/npm, and explain the difference between source code, a command, a dependency, input, and output.",st["body"]),Paragraph("If any item is unclear, repeat the matching lesson. Then continue to 01 - Web Foundations.",st["body"])]
    doc.build(story,onFirstPage=footer,onLaterPages=footer)
    folder=EXAMPLES/'00-start-here';folder.mkdir(parents=True,exist_ok=True)
    (folder/'hello.py').write_text("print('Hello! I ran my first Python program.')\n",encoding='utf-8')
    (folder/'index.html').write_text("<!doctype html><html><body><h1>My first page</h1><p>Hello, browser!</p></body></html>\n",encoding='utf-8')
    (folder/'README.md').write_text("# Start Here practice\n\nOpen this folder in VS Code. Run `python hello.py`, then open `index.html` in a browser.\n",encoding='utf-8')

def main():
    if PDF_DIR.exists(): shutil.rmtree(PDF_DIR)
    PDF_DIR.mkdir(parents=True); EXAMPLES.mkdir(exist_ok=True); st=styles(); build_start_here(st)
    for i,g in enumerate(GUIDES,1):
        folder=EXAMPLES/f"{i:02d}-{g[0]}"; folder.mkdir(parents=True,exist_ok=True)
        for name,content in files(g[-1]).items(): (folder/name).write_text(content,encoding="utf-8")
        pdf(i,g,st)
    (EXAMPLES/"README.md").write_text("# Learning examples\n\nStudy the numbered PDFs in output/pdf, then run the matching safe synthetic example folder.\n",encoding="utf-8")
    print(f"Created {len(GUIDES) + 1} PDFs and companion folders.")

if __name__=="__main__": main()
